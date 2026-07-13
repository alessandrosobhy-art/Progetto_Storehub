from __future__ import annotations

from app_logging import log_swallowed
import csv
import hashlib
import io
import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta
from typing import Any, Dict, List

from flask import Blueprint, Response, current_app, flash, redirect, render_template, request, session, url_for, send_file, jsonify

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore
    Font = None  # type: ignore
    Alignment = None  # type: ignore

import time

from db_integration import get_user_warehouse_stores, get_warehouse_stores
from orari_repository import list_turni_week
from staff_repository import list_staff
from estrazioni_repository import get_fabbisogno_code_for_site
from delivery_repository import export_weekly_rows
from ilp_finance_repository import DATE_FIELDS, TABLES as FINANCE_TABLES, list_distinct_store_values, list_finance_rows
from finance_store_mapping_repository import ensure_finance_store_mapping_tables, get_assignments_by_code, get_code_catalog_map
from spese_repository import search_spese_range_multi
from versamenti_repository import search_versamenti_range_multi
from rendiconto_verifiche_repository import save_meta
from orari_config_repository import list_orari_causali


estrazioni_bp = Blueprint("estrazioni", __name__, url_prefix="/estrazioni")

_INCLUDED_CAUSALI_NORM = {
    "permesso",
    "ferie",
    "allattamento",
    "malattia",
    "prestito",
    "training",
}


def _require_login():
    if not session.get("uid"):
        nxt = request.full_path if request.query_string else request.path
        return redirect(url_for("login", next=nxt))
    return None


def _is_scheduling_enabled_for_tenant() -> bool:
    try:
        from tenant_config_repository import current_tenant_key, get_tenant

        tenant_key = str(session.get("tenant_key") or current_tenant_key()).strip() or current_tenant_key()
        tenant = get_tenant(tenant_key)
        return bool(tenant.get("scheduling_enabled"))
    except Exception:
        current_app.logger.exception("Errore lettura flag scheduling tenant")
        return False


def _parse_week(value: str) -> tuple[date, date] | None:
    """Parse HTML <input type=week> value (YYYY-Www) into Monday..Sunday."""
    v = (value or "").strip()
    if not v:
        return None
    try:
        year_s, week_s = v.split("-W")
        y = int(year_s)
        w = int(week_s)
        start = date.fromisocalendar(y, w, 1)
        end = start + timedelta(days=6)
        return start, end
    except Exception:
        return None


def _available_stores_for_user() -> list[dict]:
    stores: list[dict] = []
    try:
        user_id = str(session.get("uid") or "").strip()
        role = str(session.get("role") or "").lower()

        try:
            if user_id:
                now = int(time.time())
                cached_for = session.get("role_verified_for")
                cached_at = int(session.get("role_verified_at") or 0)
                need_check = (role == "admin") or (cached_for != user_id) or ((now - cached_at) > 600)
                if need_check:
                    from db_integration import get_profile_role_by_id

                    srv_role = get_profile_role_by_id(user_id)
                    if srv_role:
                        role = str(srv_role).lower()
                        session["role"] = srv_role
                        session["role_verified_for"] = user_id
                        session["role_verified_at"] = now
        except Exception:
            log_swallowed('estrazioni:105')

        if role == "admin":
            stores = get_warehouse_stores() or []
        else:
            assigned = []
            if user_id:
                try:
                    assigned = get_user_warehouse_stores(str(user_id))
                except Exception:
                    assigned = []

            all_stores = get_warehouse_stores() or []
            allowed_codes = {str(row.get("store_code")) for row in assigned if row.get("store_code")}
            if allowed_codes:
                stores = [s for s in all_stores if str((s or {}).get("code")) in allowed_codes]
            else:
                cur_code = session.get("store_code")
                if cur_code:
                    stores = [s for s in all_stores if str((s or {}).get("code")) == str(cur_code)]
                else:
                    stores = []
    except Exception:
        stores = []

    try:
        stores = sorted(
            stores or [],
            key=lambda s: (
                str((s.get("name") if isinstance(s, dict) else getattr(s, "name", "")) or "").strip().lower(),
                str((s.get("code") if isinstance(s, dict) else getattr(s, "code", "")) or "").strip().lower(),
            ),
        )
    except Exception:
        log_swallowed('estrazioni:139')
    return stores


def _fmt_money(v: float) -> str:
    try:
        x = float(v or 0.0)
    except Exception:
        x = 0.0
    s = f"{x:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def _parse_iso_date(value: str) -> date | None:
    v = str(value or "").strip()
    if not v:
        return None
    try:
        return date.fromisoformat(v)
    except Exception:
        return None


def _build_verifica_record_key(kind: str, row: dict) -> str:
    k = str(kind or "").strip().upper()
    payload: dict[str, Any]
    if k == "VERSAMENTI":
        payload = {
            "kind": k,
            "site": str((row or {}).get("site") or "").strip(),
            "data_versamento_iso": str((row or {}).get("data_versamento_iso") or "").strip(),
            "dal_iso": str((row or {}).get("dal_iso") or "").strip(),
            "al_iso": str((row or {}).get("al_iso") or "").strip(),
            "nome": str((row or {}).get("nome") or "").strip(),
            "tipo": str((row or {}).get("tipo") or "").strip(),
            "tessera": str((row or {}).get("tessera") or "").strip(),
            "riferimento": str((row or {}).get("riferimento") or "").strip(),
            "valore_key": str((row or {}).get("valore_key") or "").strip(),
        }
    else:
        payload = {
            "kind": k,
            "site": str((row or {}).get("site") or "").strip(),
            "data_iso": str((row or {}).get("data_iso") or "").strip(),
            "tipo": str((row or {}).get("tipo") or "").strip(),
            "fornitore": str((row or {}).get("fornitore") or "").strip(),
            "documento": str((row or {}).get("documento") or "").strip(),
            "importo_key": str((row or {}).get("importo_key") or "").strip(),
        }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _fmt_ddmmyyyy(d: date) -> str:
    return f"{d.day:02d}/{d.month:02d}/{d.year:04d}"


def _norm_finance_store_name(value: str) -> str:
    s = str(value or "").upper().strip()
    s = s.replace("I LOVE POKE", " ")
    s = s.replace("ILOVE POKE", " ")
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _finance_store_aliases(value: str) -> set[str]:
    raw = str(value or "").strip()
    aliases: set[str] = set()
    norm = _norm_finance_store_name(raw)
    if norm:
        aliases.add(norm)
        parts = [p.strip() for p in re.split(r"\s*-\s*", raw) if p.strip()]
        for p in parts:
            pn = _norm_finance_store_name(p)
            if pn and not pn.isdigit():
                aliases.add(pn)
        for token in norm.split():
            if token and not token.isdigit() and len(token) >= 3:
                aliases.add(token)
    return aliases


def _finance_store_matches(warehouse_name: str, finance_name: str) -> bool:
    w_aliases = _finance_store_aliases(warehouse_name)
    f_aliases = _finance_store_aliases(finance_name)
    if not w_aliases or not f_aliases:
        return False
    for wa in w_aliases:
        for fa in f_aliases:
            if wa == fa or wa in fa or fa in wa:
                return True
    return False


def _finance_store_options_for_user(table_key: str, stores: list[dict], role: str) -> list[dict]:
    # Nelle tabelle ILP_FINANCE il campo Store è descrittivo, spesso vuoto e non
    # coerente con l'anagrafica store dell'app. Qui usiamo quindi solo i valori
    # realmente presenti nel database finance, senza tentare un aggancio ai codici.
    raw_values = list_distinct_store_values(table_key)
    return [{"value": v, "label": v} for v in raw_values]


def _parse_finance_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value or "").strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            log_swallowed('estrazioni:256')
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


def _extract_finance_codes_from_texts(*values: Any) -> list[str]:
    found: list[str] = []
    seen: set[str] = set()
    for raw in values:
        s = str(raw or "").strip()
        if not s:
            continue
        for code in re.findall(r"(?<!\d)(\d{6,20})(?!\d)", s):
            norm = "".join(ch for ch in code if ch.isdigit())
            if norm and norm not in seen:
                seen.add(norm)
                found.append(norm)
    return found


def _build_app_store_catalog(stores: list[dict]) -> list[dict]:
    out: list[dict] = []
    for raw in stores or []:
        code = str((raw or {}).get("code") or "").strip()
        name = str((raw or {}).get("name") or "").strip()
        if not code and not name:
            continue
        aliases = _finance_store_aliases(name)
        norm_name = _norm_finance_store_name(name)
        tokens = {tok for tok in norm_name.split() if len(tok) >= 3 and not tok.isdigit()}
        if code:
            aliases.add(code.upper())
        out.append(
            {
                "code": code,
                "name": name,
                "label": f"{code} - {name}" if code and name else (name or code),
                "aliases": {a for a in aliases if a},
                "tokens": tokens,
            }
        )
    return out


def _resolve_store_code_from_assignments(code: str, record_date: date | None, assignments_by_code: dict[str, list[dict]]) -> str:
    rows = list(assignments_by_code.get(str(code or "").strip(), []) or [])
    if not rows:
        return ""
    if record_date is None:
        return str((rows[0] or {}).get("StoreCode") or "").strip()
    best_code = ""
    best_from: date | None = None
    for row in rows:
        valid_from = _parse_finance_date((row or {}).get("ValidFrom"))
        if valid_from is None:
            continue
        if valid_from <= record_date and (best_from is None or valid_from > best_from):
            best_from = valid_from
            best_code = str((row or {}).get("StoreCode") or "").strip()
    return best_code


def _match_store_from_text(combined_text: str, app_catalog: list[dict]) -> tuple[str, str]:
    text_norm = _norm_finance_store_name(combined_text)
    if not text_norm:
        return "", ""

    best_score = 0
    best_label = ""
    best_code = ""
    text_tokens = {tok for tok in text_norm.split() if len(tok) >= 3 and not tok.isdigit()}

    for item in app_catalog:
        score = 0
        for alias in item.get("aliases") or []:
            alias_norm = _norm_finance_store_name(alias)
            if not alias_norm or len(alias_norm) < 3:
                continue
            if text_norm == alias_norm:
                score = max(score, 100)
            elif alias_norm in text_norm:
                score = max(score, 92 if len(alias_norm) >= 6 else 84)
        item_tokens = set(item.get("tokens") or set())
        if item_tokens and text_tokens:
            overlap = len(item_tokens & text_tokens)
            if overlap:
                coverage = overlap / max(1, len(item_tokens))
                score = max(score, int(60 + coverage * 30))
        if score > best_score:
            best_score = score
            best_code = str(item.get("code") or "").strip()
            best_label = str(item.get("label") or "").strip()

    if best_score >= 78:
        return best_code, best_label
    return "", ""


def _augment_finance_result_with_store_app(result: dict, stores: list[dict], table_key: str) -> dict:
    try:
        ensure_finance_store_mapping_tables()
    except Exception:
        return result

    rows = list((result or {}).get("rows") or [])
    if not rows:
        cols = list((result or {}).get("columns") or [])
        if not any(str((c or {}).get("key") or "") == "StoreApp" for c in cols):
            insert_at = 0
            for idx, col in enumerate(cols):
                if str((col or {}).get("key") or "") in {"Store", "Descrizione", "Note", "NotaVers"}:
                    insert_at = idx + 1
                    break
            cols.insert(insert_at, {"key": "StoreApp", "label": "Store app"})
            result["columns"] = cols
        return result

    codes: set[str] = set()
    for row in rows:
        codes.update(
            _extract_finance_codes_from_texts(
                row.get("Store"),
                row.get("Descrizione"),
                row.get("Note"),
                row.get("NotaVers"),
            )
        )

    catalog_map = get_code_catalog_map(codes)
    assignments_by_code = get_assignments_by_code(codes)
    app_catalog = _build_app_store_catalog(stores)
    app_by_code = {str((s or {}).get("code") or "").strip(): str((s or {}).get("name") or "").strip() for s in stores or []}

    for row in rows:
        record_date = _parse_finance_date(row.get("DataContabile")) or _parse_finance_date(row.get("DataValuta")) or _parse_finance_date(row.get("DataInserimento"))
        extracted_codes = _extract_finance_codes_from_texts(
            row.get("Store"),
            row.get("Descrizione"),
            row.get("Note"),
            row.get("NotaVers"),
        )
        resolved_store_code = ""

        for code in extracted_codes:
            resolved_store_code = _resolve_store_code_from_assignments(code, record_date, assignments_by_code)
            if resolved_store_code:
                break

        combined_hint_parts = [
            str(row.get("Store") or "").strip(),
            str(row.get("Descrizione") or "").strip(),
            str(row.get("Note") or "").strip(),
            str(row.get("NotaVers") or "").strip(),
        ]
        if not resolved_store_code:
            for code in extracted_codes:
                code_info = catalog_map.get(code) or {}
                code_label = str(code_info.get("SourceLabel") or "").strip()
                if code_label:
                    resolved_store_code, _ = _match_store_from_text(code_label, app_catalog)
                    if resolved_store_code:
                        break

        if not resolved_store_code:
            resolved_store_code, _ = _match_store_from_text(" ".join(p for p in combined_hint_parts if p), app_catalog)

        store_name = app_by_code.get(resolved_store_code, "")
        row["StoreAppCode"] = resolved_store_code
        row["StoreAppName"] = store_name
        row["StoreApp"] = f"{resolved_store_code} - {store_name}" if resolved_store_code and store_name else (store_name or resolved_store_code)

    cols = list((result or {}).get("columns") or [])
    if not any(str((c or {}).get("key") or "") == "StoreApp" for c in cols):
        insert_at = len(cols)
        for idx, col in enumerate(cols):
            if str((col or {}).get("key") or "") in {"Store", "Descrizione", "Note", "NotaVers"}:
                insert_at = idx + 1
                break
        cols.insert(insert_at, {"key": "StoreApp", "label": "Store app"})
        result["columns"] = cols
    result["rows"] = rows
    return result


def _pad_int(v: str, width: int) -> str:
    s = "".join([c for c in (v or "").strip() if c.isdigit()])
    if not s:
        return "".zfill(width)
    return s.zfill(width)


def _parse_hhmm(s: str) -> tuple[int, int] | None:
    s = (s or "").strip()
    if not s:
        return None
    try:
        hh, mm = s.split(":")
        return int(hh), int(mm)
    except Exception:
        return None


def _duration_hhmm(start_hhmm: str, end_hhmm: str) -> str:
    a = _parse_hhmm(start_hhmm)
    b = _parse_hhmm(end_hhmm)
    if not a or not b:
        return ""
    sh, sm = a
    eh, em = b
    start_m = sh * 60 + sm
    end_m = eh * 60 + em
    if end_m < start_m:
        end_m += 24 * 60
    mins = max(0, end_m - start_m)
    hh = mins // 60
    mm = mins % 60
    return f"{hh:02d}:{mm:02d}"


def _segment_minutes(start_hhmm: str, end_hhmm: str) -> int:
    a = _parse_hhmm(start_hhmm)
    b = _parse_hhmm(end_hhmm)
    if not a or not b:
        return 0
    sh, sm = a
    eh, em = b
    start_m = sh * 60 + sm
    end_m = eh * 60 + em
    if end_m < start_m:
        end_m += 24 * 60
    return max(0, end_m - start_m)


def _hhmm_add_minutes(start_hhmm: str, mins_to_add: int) -> str:
    a = _parse_hhmm(start_hhmm)
    if not a:
        return start_hhmm
    total = ((a[0] * 60 + a[1]) + int(mins_to_add or 0)) % (24 * 60)
    return f"{total // 60:02d}:{total % 60:02d}"


def _is_auto_extra_eligible(causale: str, start_hhmm: str, end_hhmm: str) -> bool:
    if not ((start_hhmm or "").strip() and (end_hhmm or "").strip()):
        return False
    c = _norm_causale(causale)
    return (not c) or c in {"prestito", "training", "extra"}


def _sort_day_key(value: str) -> str:
    s = (value or "").strip()
    if not s:
        return "9999-99-99"
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except Exception:
        try:
            return datetime.strptime(s, "%d/%m/%Y").date().isoformat()
        except Exception:
            return s


def _build_auto_extra_map(turni: list[dict], contract_hours_by_name: dict[str, int]) -> dict[tuple[int, int], int]:
    by_name: dict[str, dict] = {}

    for row_idx, t in enumerate(turni):
        nom = str((t.get("nominativo") or "")).strip()
        if not nom:
            continue
        key = nom.lower()
        rec = by_name.setdefault(
            key,
            {
                "contract_mins": max(0, int(contract_hours_by_name.get(key) or 0)) * 60,
                "total_mins": 0,
                "explicit_extra_mins": 0,
                "segments": [],
            },
        )
        day_key = _sort_day_key(str(t.get("data") or ""))

        slots = [
            (1, str(t.get("causale") or "").strip(), str(t.get("inizio_1") or "").strip(), str(t.get("fine_1") or "").strip()),
            (2, str(t.get("causale2") or "").strip(), str(t.get("inizio_2") or "").strip(), str(t.get("fine_2") or "").strip()),
        ]
        for slot_no, causale, start_hhmm, end_hhmm in slots:
            mins = _segment_minutes(start_hhmm, end_hhmm)
            if mins <= 0:
                continue
            caus_norm = _norm_causale(causale)
            rec["total_mins"] += mins
            if caus_norm == "extra":
                rec["explicit_extra_mins"] += mins
                continue
            rec["segments"].append(
                {
                    "row_idx": row_idx,
                    "slot_no": slot_no,
                    "day_key": day_key,
                    "mins": mins,
                    "auto_eligible": _is_auto_extra_eligible(causale, start_hhmm, end_hhmm),
                }
            )

    out: dict[tuple[int, int], int] = {}
    for rec in by_name.values():
        contract_mins = int(rec.get("contract_mins") or 0)
        if contract_mins <= 0:
            continue
        target_extra = max(0, int(rec.get("total_mins") or 0) - contract_mins)
        remaining_auto = max(0, target_extra - int(rec.get("explicit_extra_mins") or 0))
        segments = sorted(rec.get("segments") or [], key=lambda x: (x["day_key"], x["slot_no"]))
        for seg in segments:
            if remaining_auto <= 0:
                break
            if not seg.get("auto_eligible"):
                continue
            take = min(int(seg.get("mins") or 0), remaining_auto)
            if take > 0:
                out[(int(seg["row_idx"]), int(seg["slot_no"]))] = take
                remaining_auto -= take
    return out


def _norm_causale(value: str) -> str:
    s = (value or "").strip().lower()
    return "".join(ch for ch in s if ch.isalnum())


def _load_causali_rules() -> dict[str, dict]:
    """Regole causali normalizzate per nome, da richiamare UNA volta per export.

    list_orari_causali() apre una connessione e, a monte, seed_default_orari_config()
    esegue un MERGE su ogni causale/inquadramento di default: e' un'operazione da
    fare una volta per richiesta, non per ogni turno/slot controllato.
    """
    try:
        return {
            _norm_causale(str((r or {}).get("name") or "")): dict(r or {})
            for r in list_orari_causali(active_only=True)
        }
    except Exception:
        return {}


def _should_export_slot(causale: str, start_hhmm: str, end_hhmm: str, rules: dict[str, dict] | None = None) -> bool:
    c = _norm_causale(causale)
    has_times = bool((start_hhmm or "").strip() and (end_hhmm or "").strip())
    # rules=None (retrocompatibilita' per eventuali altri chiamanti): richiede ancora
    # una lettura a chiamata. I chiamanti che processano molte righe (es. l'export
    # HQ) DEVONO passare le regole gia' caricate con _load_causali_rules().
    if rules is None:
        rules = _load_causali_rules()

    if c == "extra":
        return False
    if c and c in rules:
        rule = rules.get(c) or {}
        return bool(rule.get("auto_extra_eligible") or rule.get("justifies_contract_hours") or rule.get("requires_loan_store") or rule.get("counts_training"))
    if c in _INCLUDED_CAUSALI_NORM:
        return True
    if not c and has_times:
        return True
    return False


def _build_csv_rows(*, week_start: date, week_end: date, progressivo: str, sites: list[str]) -> list[list[str]]:
    prog = (progressivo or "").strip()
    if not prog:
        try:
            prog = f"{week_start.isocalendar().week:02d}"
        except Exception:
            prog = "00"

    rows: list[list[str]] = []

    clean_sites = [str(s or "").strip() for s in sites if str(s or "").strip()]

    # Il fetch per store (fabbisogno + staff + turni settimana) e' I/O di rete verso
    # SQL Server: ogni chiamata apre la sua connessione (nessun pooling nel progetto)
    # e con molti store selezionati farlo in serie e' la causa dei timeout su
    # Estrazioni HQ -> Scheduling. Lo eseguiamo in parallelo con un pool limitato;
    # la logica di costruzione righe sotto resta identica e sequenziale.
    def _fetch_site_data(site_code: str) -> tuple[str, str, list[dict]]:
        fab = get_fabbisogno_code_for_site(site_code)
        fab_10 = _pad_int(str(fab or ""), 10)
        staff = list_staff(store_code=site_code) or []
        turni_raw = list_turni_week(store_code=site_code, start_day=week_start, end_day=week_end) or []
        return fab_10, staff, turni_raw

    site_data: dict[str, tuple[str, list[dict], list[dict]]] = {}
    max_workers = min(8, len(clean_sites)) or 1
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        future_to_site = {pool.submit(_fetch_site_data, s): s for s in clean_sites}
        for future in as_completed(future_to_site):
            site_code = future_to_site[future]
            try:
                fab_10, staff, turni_raw = future.result()
            except Exception:
                current_app.logger.exception("Errore fetch dati scheduling per store %s", site_code)
                fab_10, staff, turni_raw = "0000000000", [], []
            site_data[site_code] = (fab_10, staff, turni_raw)

    # Caricata UNA volta per l'intero export, non per ogni turno/slot (vedi
    # _load_causali_rules): era questo, non il fetch per-store, il vero collo di
    # bottiglia dei timeout su esportazioni con molti store/turni.
    causali_rules = _load_causali_rules()

    for site in clean_sites:
        site_code = site
        if not site_code:
            continue

        rows.append(["C01", "C", "000000", prog, _fmt_ddmmyyyy(week_start), _fmt_ddmmyyyy(week_end)])

        fab_10, staff, turni_raw = site_data.get(site_code, ("0000000000", [], []))

        name_to_code: dict[str, str] = {}
        contract_hours_by_name: dict[str, int] = {}
        sched_allowed: set[str] = set()
        for p in staff:
            nom = (p.get("nome_cognome") or p.get("nominativo") or "").strip()
            if not nom:
                continue
            low = nom.lower()
            code = (p.get("codice_dipendente") or p.get("codice") or p.get("matricola") or "").strip()
            try:
                contract_hours_by_name[low] = int(p.get("ore_contrattuali") or 0)
            except Exception:
                contract_hours_by_name[low] = 0
            if bool(p.get("scheduling", False)) and code:
                name_to_code[low] = _pad_int(code, 7)
                sched_allowed.add(low)

        turni: list[dict] = []
        for t in turni_raw:
            nom = (t.get("nominativo") or "").strip()
            if not nom:
                continue
            if nom.lower() not in sched_allowed:
                continue
            turni.append(t)

        auto_extra_remove = _build_auto_extra_map(turni, contract_hours_by_name)

        for row_idx, t in enumerate(turni):
            nom = (t.get("nominativo") or "").strip()
            if not nom:
                continue

            cod_dip = (name_to_code.get(nom.lower()) or "").strip()
            if not cod_dip:
                continue
            d_iso = (t.get("data") or "").strip()
            try:
                d = datetime.fromisoformat(d_iso).date()
            except Exception:
                try:
                    d = datetime.strptime(d_iso, "%d/%m/%Y").date()
                except Exception:
                    continue

            caus1 = (t.get("causale") or "").strip()
            caus2 = (t.get("causale2") or "").strip()
            in1 = (t.get("inizio_1") or "").strip()
            fi1 = (t.get("fine_1") or "").strip()
            in2 = (t.get("inizio_2") or "").strip()
            fi2 = (t.get("fine_2") or "").strip()

            included_slots: list[tuple[str, str, str]] = []
            slot_defs = [
                (1, caus1, in1, fi1),
                (2, caus2, in2, fi2),
            ]
            for slot_no, causale, start_hhmm, end_hhmm in slot_defs:
                if not _should_export_slot(causale, start_hhmm, end_hhmm, causali_rules):
                    continue

                if start_hhmm and end_hhmm:
                    remove_mins = int(auto_extra_remove.get((row_idx, slot_no)) or 0)
                    total_mins = _segment_minutes(start_hhmm, end_hhmm)
                    keep_mins = max(0, total_mins - remove_mins)
                    if keep_mins <= 0:
                        continue
                    if remove_mins > 0:
                        start_hhmm = _hhmm_add_minutes(start_hhmm, remove_mins)
                    included_slots.append((causale, start_hhmm, end_hhmm))
                else:
                    included_slots.append((causale, start_hhmm, end_hhmm))

            const_3 = "000000"
            const_4 = "000487"
            const_9 = "0000000001"

            if not included_slots:
                c1_norm = _norm_causale(caus1)
                c2_norm = _norm_causale(caus2)
                has_any_times = bool(in1 or fi1 or in2 or fi2)
                is_rest_causale = (c1_norm in ("", "off")) and (c2_norm in ("", "off"))
                if is_rest_causale and (not has_any_times):
                    rows.append(
                        [
                            "C02",
                            "A",
                            const_3,
                            const_4,
                            cod_dip,
                            _fmt_ddmmyyyy(d),
                            "0000001",
                            fab_10,
                            const_9,
                            "",
                            "",
                            "",
                            "",
                        ]
                    )
                continue

            for i, (causale, start_hhmm, end_hhmm) in enumerate(included_slots, start=1):
                caus_norm = _norm_causale(causale)
                turno_code = 6 if caus_norm == "riposofestivo" else (2 + (i - 1))
                dur = _duration_hhmm(start_hhmm, end_hhmm) if (start_hhmm and end_hhmm) else ""
                rows.append(
                    [
                        "C02",
                        "A",
                        const_3,
                        const_4,
                        cod_dip,
                        _fmt_ddmmyyyy(d),
                        str(turno_code).zfill(7),
                        fab_10,
                        const_9,
                        start_hhmm,
                        end_hhmm,
                        dur,
                        "",
                    ]
                )

    return rows


@estrazioni_bp.route("/")
def home():
    r = _require_login()
    if r:
        return r
    return redirect(url_for("estrazioni.scheduling"))


@estrazioni_bp.get("/verifica-rendiconto")
def verifica_rendiconto():
    r = _require_login()
    if r:
        return r

    stores = _available_stores_for_user()
    return render_template(
        "estrazioni_verifica_rendiconto.html",
        stores_count=len(stores or []),
    )


@estrazioni_bp.get("/api/verifica-rendiconto")
def api_verifica_rendiconto():
    r = _require_login()
    if r:
        return jsonify({"ok": False, "error": "Sessione scaduta."}), 401

    kind = str(request.args.get("kind") or "VERSAMENTI").strip().upper()
    if kind not in {"VERSAMENTI", "SPESE"}:
        kind = "VERSAMENTI"

    start_iso = str(request.args.get("start") or "").strip()
    end_iso = str(request.args.get("end") or "").strip()
    start_d = _parse_iso_date(start_iso)
    end_d = _parse_iso_date(end_iso)
    if (not start_d) or (not end_d):
        return jsonify({"ok": False, "error": "Intervallo non valido."}), 400
    if end_d < start_d:
        start_d, end_d = end_d, start_d

    stores = _available_stores_for_user()
    store_map = {
        str((s or {}).get("code") or "").strip(): str((s or {}).get("name") or "").strip()
        for s in (stores or [])
    }
    store_codes = [c for c in store_map.keys() if c]

    warnings: List[str] = []
    total = 0.0
    rows_out: List[Dict[str, Any]] = []
    columns: List[Dict[str, Any]] = []

    try:
        if kind == "SPESE":
            res = search_spese_range_multi(store_codes, start=start_d, end=end_d) if store_codes else {"rows": [], "total": 0.0, "warnings": []}
            warnings = list(res.get("warnings") or [])
            total = float(res.get("total") or 0.0)

            columns = [
                {"key": "store", "label": "Store"},
                {"key": "data", "label": "Data"},
                {"key": "tipo", "label": "Tipo"},
                {"key": "fornitore", "label": "Fornitore"},
                {"key": "documento", "label": "Documento"},
                {"key": "importo", "label": "Importo"},
                {"key": "foto", "label": "Foto", "type": "photo"},
                {"key": "verificato", "label": "Verificato"},
                {"key": "nota", "label": "Nota"},
            ]

            for r0 in (res.get("rows") or []):
                r = dict(r0 or {})
                site = str(r.get("site") or "").strip()
                name = store_map.get(site) or ""
                foto_file = r.get("foto_file")
                foto_url = url_for("rendiconto.spese_photo_scoped", store_code=site, filename=foto_file) if foto_file else ""
                rows_out.append(
                    {
                        "record_key": _build_verifica_record_key(kind, r),
                        "site": site,
                        "store": site + ((" - " + name) if name else ""),
                        "data": r.get("data") or "",
                        "data_iso": r.get("data_iso") or "",
                        "tipo": r.get("tipo") or "",
                        "fornitore": r.get("fornitore") or "",
                        "documento": r.get("documento") or "",
                        "importo": _fmt_money(r.get("importo") or 0.0),
                        "importo_key": r.get("importo_key") or "",
                        "foto_url": foto_url,
                        "foto_file": str(foto_file or "") if foto_file else "",
                    }
                )
        else:
            res = search_versamenti_range_multi(store_codes, start=start_d, end=end_d) if store_codes else {"rows": [], "total": 0.0, "warnings": []}
            warnings = list(res.get("warnings") or [])
            total = float(res.get("total") or 0.0)

            columns = [
                {"key": "store", "label": "Store"},
                {"key": "data_versamento", "label": "Data versamento"},
                {"key": "dal", "label": "Dal"},
                {"key": "al", "label": "Al"},
                {"key": "nome", "label": "Nome"},
                {"key": "tipo", "label": "Tipo"},
                {"key": "tessera", "label": "Tessera"},
                {"key": "riferimento", "label": "Riferimento"},
                {"key": "valore", "label": "Valore"},
                {"key": "foto", "label": "Foto", "type": "photo"},
                {"key": "verificato", "label": "Verificato"},
                {"key": "nota", "label": "Nota"},
            ]

            for r0 in (res.get("rows") or []):
                r = dict(r0 or {})
                site = str(r.get("site") or "").strip()
                name = store_map.get(site) or ""
                foto_file = r.get("foto_file")
                foto_url = url_for("rendiconto.versamenti_photo_scoped", store_code=site, filename=foto_file) if foto_file else ""
                rows_out.append(
                    {
                        "record_key": _build_verifica_record_key(kind, r),
                        "site": site,
                        "store": site + ((" - " + name) if name else ""),
                        "data_versamento": r.get("data_versamento") or "",
                        "data_versamento_iso": r.get("data_versamento_iso") or "",
                        "dal": r.get("dal") or "",
                        "dal_iso": r.get("dal_iso") or "",
                        "al": r.get("al") or "",
                        "al_iso": r.get("al_iso") or "",
                        "nome": r.get("nome") or "",
                        "tipo": r.get("tipo") or "",
                        "tessera": r.get("tessera") or "",
                        "riferimento": r.get("riferimento") or "",
                        "valore": _fmt_money(r.get("valore") or 0.0),
                        "valore_key": r.get("valore_key") or "",
                        "foto_url": foto_url,
                        "foto_file": str(foto_file or "") if foto_file else "",
                    }
                )
    except Exception as ex:
        current_app.logger.exception("Errore api_verifica_rendiconto")
        return jsonify({"ok": False, "error": str(ex)}), 500

    return jsonify(
        {
            "ok": True,
            "kind": kind,
            "start": start_iso,
            "end": end_iso,
            "columns": columns,
            "rows": rows_out,
            "total": total,
            "total_display": _fmt_money(total),
            "warnings": warnings,
        }
    )


@estrazioni_bp.post("/api/verifica-rendiconto/save")
def api_verifica_rendiconto_save():
    r = _require_login()
    if r:
        return jsonify({"ok": False, "error": "Sessione scaduta."}), 401

    data = request.get_json(silent=True) or {}
    try:
        save_meta(
            record_key=str(data.get("record_key") or "").strip(),
            record_kind=str(data.get("kind") or "").strip().upper(),
            site=str(data.get("site") or "").strip(),
            verificato=bool(data.get("verificato")),
            nota=str(data.get("nota") or ""),
        )
        return jsonify({"ok": True})
    except Exception as ex:
        current_app.logger.exception("Errore salvataggio verifica rendiconto")
        return jsonify({"ok": False, "error": str(ex)}), 500


@estrazioni_bp.post("/api/verifica-rendiconto/export.xlsx")
def api_verifica_rendiconto_export_xlsx():
    r = _require_login()
    if r:
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    if Workbook is None:
        return jsonify({"ok": False, "error": "openpyxl non disponibile"}), 500

    data = request.get_json(silent=True) or {}
    kind = str(data.get("kind") or "verifica_rendiconto").strip().lower()
    start_iso = str(data.get("start") or "").strip()
    end_iso = str(data.get("end") or "").strip()
    columns = data.get("columns") or []
    rows = data.get("rows") or []

    cols = []
    for c in columns:
        k = str((c or {}).get("key") or "").strip()
        lbl = str((c or {}).get("label") or k).strip()
        typ = str((c or {}).get("type") or "").strip().lower()
        if k:
            cols.append({"key": k, "label": lbl, "type": typ})

    wb = Workbook()
    ws = wb.active
    ws.title = "Verifica"

    header = []
    for c in cols:
        if c.get("type") == "photo":
            header.append(c.get("label") + " URL")
            header.append(c.get("label") + " FILE")
        else:
            header.append(c.get("label"))
    ws.append(header)

    if Font and Alignment:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    for r in rows:
        out = []
        for c in cols:
            k = c.get("key")
            if c.get("type") == "photo":
                out.append(str((r or {}).get(k + "_url") or ""))
                out.append(str((r or {}).get(k + "_file") or ""))
            else:
                out.append(str((r or {}).get(k) or ""))
        ws.append(out)

    try:
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[chr(64 + col_idx)].width = min(max(10, max_len + 2), 60)
    except Exception:
        log_swallowed('estrazioni:1000')

    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = f"estrazioni_verifica_rendiconto_{kind}_{start_iso}_{end_iso}.xlsx".replace("__", "_")
    return send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@estrazioni_bp.route("/scheduling", methods=["GET", "POST"])
def scheduling():
    r = _require_login()
    if r:
        return r
    if not _is_scheduling_enabled_for_tenant():
        flash("Scheduling non attivo per questo tenant.", "warning")
        return redirect(url_for("dashboard"))

    stores = _available_stores_for_user()

    if request.method == "GET":
        # default: current ISO week
        today = date.today()
        w = today.isocalendar().week
        y = today.isocalendar().year
        week_value = f"{y:04d}-W{w:02d}"
        return render_template("estrazioni_scheduling.html", stores=stores, week_value=week_value)

    week_value = (request.form.get("week") or "").strip()
    selected_sites = request.form.getlist("sites")

    parsed = _parse_week(week_value)
    if not parsed:
        flash("Seleziona una settimana valida.", "warning")
        return redirect(url_for("estrazioni.scheduling"))
    week_start, week_end = parsed

    allowed_codes = {
        str((s.get("code") if isinstance(s, dict) else getattr(s, "code", "")) or "").strip()
        for s in (stores or [])
        if str((s.get("code") if isinstance(s, dict) else getattr(s, "code", "")) or "").strip()
    }
    names_by_code = {
        str((s.get("code") if isinstance(s, dict) else getattr(s, "code", "")) or "").strip():
        str((s.get("name") if isinstance(s, dict) else getattr(s, "name", "")) or "").strip()
        for s in (stores or [])
    }
    sites = []
    seen_sites = set()
    for raw_site in selected_sites:
        site = str(raw_site or "").strip()
        if not site or site not in allowed_codes or site in seen_sites:
            continue
        seen_sites.add(site)
        sites.append(site)
    if not sites:
        flash("Seleziona almeno uno store.", "warning")
        return redirect(url_for("estrazioni.scheduling"))

    try:
        rows = _build_csv_rows(week_start=week_start, week_end=week_end, progressivo="", sites=sites)
    except Exception as e:
        current_app.logger.exception("Errore generazione CSV scheduling")
        flash(f"Errore generazione CSV: {e}", "danger")
        return redirect(url_for("estrazioni.scheduling"))

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
    for r in rows:
        w.writerow(r)
    content = buf.getvalue().encode("utf-8")

    fname = f"Scheduling_{week_start.strftime('%Y%m%d')}_{week_end.strftime('%Y%m%d')}.csv"
    return Response(
        content,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@estrazioni_bp.route("/delivery-weekly", methods=["GET"])
def delivery_weekly():
    r = _require_login()
    if r:
        return r

    stores = _available_stores_for_user()
    allowed_codes = {
        str((s.get("code") if isinstance(s, dict) else getattr(s, "code", "")) or "").strip()
        for s in (stores or [])
        if str((s.get("code") if isinstance(s, dict) else getattr(s, "code", "")) or "").strip()
    }
    names_by_code = {
        str((s.get("code") if isinstance(s, dict) else getattr(s, "code", "")) or "").strip():
        str((s.get("name") if isinstance(s, dict) else getattr(s, "name", "")) or "").strip()
        for s in (stores or [])
    }

    today = date.today()
    cur_week = today.isocalendar().week
    cur_year = today.isocalendar().year
    default_week = f"{cur_year:04d}-W{cur_week:02d}"

    week_from_value = (request.args.get("week_from") or "").strip() or default_week
    week_to_value = (request.args.get("week_to") or "").strip() or week_from_value
    provider = (request.args.get("provider") or "ALL").strip().upper()
    selected_sites = [str(x or "").strip() for x in request.args.getlist("sites") if str(x or "").strip()]
    selected_sites = [s for s in selected_sites if s in allowed_codes]
    if not selected_sites:
        selected_sites = sorted(allowed_codes)

    rows = []
    summary = {
        "payment_online": 0.0,
        "payment_cash": 0.0,
        "payment_total": 0.0,
        "complaints_received": 0,
        "refund_value": 0.0,
        "complaints_contested": 0,
        "appeals_accepted": 0,
        "refunds_cancelled_value": 0.0,
        "stores_count": 0,
        "recovered_value_pct": None,
        "accepted_appeals_pct": None,
    }
    if selected_sites:
        from_parsed = _parse_week(week_from_value)
        to_parsed = _parse_week(week_to_value)
        if from_parsed and to_parsed:
            week_start_from, _ = from_parsed
            week_start_to, _ = to_parsed
            if week_start_to < week_start_from:
                week_start_from, week_start_to = week_start_to, week_start_from
            try:
                rows = export_weekly_rows(
                    store_codes=selected_sites,
                    week_start_from=week_start_from,
                    week_start_to=week_start_to,
                    platform=provider,
                )
                for row in (rows or []):
                    code = str((row or {}).get("store_code") or "").strip()
                    row["store_name"] = names_by_code.get(code, code)
                distinct_stores = set()
                for row in rows or []:
                    pay_online = float((row or {}).get("payment_online") or 0.0)
                    pay_cash = float((row or {}).get("payment_cash") or 0.0)
                    summary["payment_online"] += pay_online
                    summary["payment_cash"] += pay_cash
                    summary["payment_total"] += float((row or {}).get("payment_total") or (pay_online + pay_cash))
                    summary["complaints_received"] += int((row or {}).get("complaints_received") or 0)
                    summary["refund_value"] += float((row or {}).get("refund_value") or 0.0)
                    summary["complaints_contested"] += int((row or {}).get("complaints_contested") or 0)
                    summary["appeals_accepted"] += int((row or {}).get("appeals_accepted") or 0)
                    summary["refunds_cancelled_value"] += float((row or {}).get("refunds_cancelled_value") or 0.0)
                    store_code = str((row or {}).get("store_code") or "").strip()
                    if store_code:
                        distinct_stores.add(store_code)
                summary["stores_count"] = len(distinct_stores)
                refund_value = float(summary["refund_value"] or 0.0)
                complaints_received = int(summary["complaints_received"] or 0)
                if refund_value:
                    summary["recovered_value_pct"] = round((float(summary["refunds_cancelled_value"]) / refund_value) * 100.0, 2)
                if complaints_received:
                    summary["accepted_appeals_pct"] = round((int(summary["appeals_accepted"]) / complaints_received) * 100.0, 2)
            except Exception as e:
                current_app.logger.exception("Errore caricamento estrazione delivery")
                flash(f"Errore caricamento dati delivery: {e}", "danger")
        else:
            flash("Seleziona settimane valide.", "warning")

    return render_template(
        "estrazioni_delivery_weekly.html",
        stores=stores,
        week_from_value=week_from_value,
        week_to_value=week_to_value,
        provider=provider,
        selected_sites=selected_sites,
        rows=rows,
        summary=summary,
    )


@estrazioni_bp.get("/delivery-weekly.xlsx")
def delivery_weekly_xlsx():
    r = _require_login()
    if r:
        return r
    if Workbook is None:
        flash("Export Excel non disponibile: openpyxl mancante.", "warning")
        return redirect(url_for("estrazioni.delivery_weekly"))

    stores = _available_stores_for_user()
    allowed_codes = {
        str((s.get("code") if isinstance(s, dict) else getattr(s, "code", "")) or "").strip()
        for s in (stores or [])
        if str((s.get("code") if isinstance(s, dict) else getattr(s, "code", "")) or "").strip()
    }
    names_by_code = {
        str((s.get("code") if isinstance(s, dict) else getattr(s, "code", "")) or "").strip():
        str((s.get("name") if isinstance(s, dict) else getattr(s, "name", "")) or "").strip()
        for s in (stores or [])
    }

    today = date.today()
    cur_week = today.isocalendar().week
    cur_year = today.isocalendar().year
    default_week = f"{cur_year:04d}-W{cur_week:02d}"
    week_from_value = (request.args.get("week_from") or "").strip() or default_week
    week_to_value = (request.args.get("week_to") or "").strip() or week_from_value
    provider = (request.args.get("provider") or "ALL").strip().upper()
    selected_sites = [str(x or "").strip() for x in request.args.getlist("sites") if str(x or "").strip()]
    selected_sites = [s for s in selected_sites if s in allowed_codes]
    if not selected_sites:
        selected_sites = sorted(allowed_codes)

    from_parsed = _parse_week(week_from_value)
    to_parsed = _parse_week(week_to_value)
    if not (from_parsed and to_parsed and selected_sites):
        flash("Filtri non validi per export Excel.", "warning")
        return redirect(url_for("estrazioni.delivery_weekly"))

    week_start_from, _ = from_parsed
    week_start_to, _ = to_parsed
    if week_start_to < week_start_from:
        week_start_from, week_start_to = week_start_to, week_start_from

    rows = export_weekly_rows(
        store_codes=selected_sites,
        week_start_from=week_start_from,
        week_start_to=week_start_to,
        platform=provider,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Dati Delivery"
    headers = [
        "Settimana",
        "Store",
        "Nome store",
        "Provider",
        "Online",
        "Contanti",
        "Totale",
        "Ordini",
        "Cancellati",
        "Rimborsi",
        "Valore rimborsi",
        "Contestati",
        "Accettati",
        "Rimborsi annullati",
        "Apertura %",
        "Rating",
        "Unita rating",
    ]
    ws.append(headers)
    if Font is not None:
        for cell in ws[1]:
            cell.font = Font(bold=True)

    for row in rows:
        ws.append([
            str(row.get("week_start") or ""),
            str(row.get("store_code") or ""),
            names_by_code.get(str(row.get("store_code") or ""), ""),
            str(row.get("platform") or ""),
            float(row.get("payment_online") or 0),
            float(row.get("payment_cash") or 0),
            float(row.get("payment_total") or 0),
            int(row.get("orders") or 0),
            int(row.get("cancelled_orders") or 0),
            int(row.get("complaints_received") or 0),
            float(row.get("refund_value") or 0),
            int(row.get("complaints_contested") or 0),
            int(row.get("appeals_accepted") or 0),
            float(row.get("refunds_cancelled_value") or 0),
            "" if row.get("opening_pct") is None else float(row.get("opening_pct") or 0),
            "" if row.get("rating_value") is None else float(row.get("rating_value") or 0),
            str(row.get("rating_unit") or ""),
        ])

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[letter].width = min(max_len + 2, 24)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"dati_delivery_{week_start_from.isoformat()}_{week_start_to.isoformat()}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
    )


@estrazioni_bp.get("/finance")
def finance():
    r = _require_login()
    if r:
        return r

    stores = _available_stores_for_user()

    table_key = str(request.args.get("table") or "transazioni").strip().lower()
    if table_key not in FINANCE_TABLES:
        table_key = "transazioni"

    cfg = FINANCE_TABLES[table_key]
    has_store = bool(cfg.get("store_column"))

    date_field = str(request.args.get("date_field") or "DataContabile").strip()
    if date_field not in DATE_FIELDS:
        date_field = "DataContabile"

    start_iso = str(request.args.get("start") or "").strip()
    end_iso = str(request.args.get("end") or "").strip()
    banca = str(request.args.get("banca") or "").strip()
    societa = str(request.args.get("societa") or "").strip()
    type_value = str(request.args.get("type_value") or "").strip()
    tessera = str(request.args.get("tessera") or "").strip()
    atm = str(request.args.get("atm") or "").strip()
    search = str(request.args.get("q") or "").strip()
    try:
        page = int(request.args.get("page") or 1)
    except Exception:
        page = 1

    store_options = _finance_store_options_for_user(table_key, stores, str(session.get("role") or ""))
    allowed_store_values = {str((x or {}).get("value") or "").strip() for x in (store_options or [])}
    selected_store_values = [str(x or "").strip() for x in request.args.getlist("stores") if str(x or "").strip()]
    selected_store_values = [x for x in selected_store_values if x in allowed_store_values]

    include_blank_store = False
    if has_store:
        include_blank_store = str(request.args.get("include_blank_store") or "").strip().lower() in {"1", "true", "on", "yes"}
        # Nessun filtro store = mostra tutte le righe; il campo Store non è affidabile
        # come chiave anagrafica e non va usato come vincolo implicito.

    result = {
        "table": table_key,
        "title": cfg.get("title") or table_key.title(),
        "columns": [{"key": c, "label": lbl} for c, lbl in (cfg.get("columns") or [])],
        "type_label": cfg.get("type_label") or "Tipo",
        "has_store": has_store,
        "has_tessera": any(c == "NumeroTessera" for c, _ in (cfg.get("columns") or [])),
        "has_atm": any(c == "NumeroATM" for c, _ in (cfg.get("columns") or [])),
        "rows": [],
        "page": 1,
        "page_size": 200,
        "total_rows": 0,
        "total_importo": 0.0,
        "pages": 1,
    }

    try:
        result = list_finance_rows(
            table_key=table_key,
            date_field=date_field,
            start_iso=start_iso,
            end_iso=end_iso,
            store_values=selected_store_values if has_store else None,
            include_blank_store=include_blank_store if has_store else False,
            banca=banca,
            societa=societa,
            type_value=type_value,
            search=search,
            tessera=tessera,
            atm=atm,
            page=page,
            page_size=200,
        )
        result = _augment_finance_result_with_store_app(result, stores, table_key)
    except Exception as e:
        current_app.logger.exception("Errore caricamento dati finance")
        flash(f"Errore caricamento dati finance: {e}", "danger")

    summary = {
        "rows": int(result.get("total_rows") or 0),
        "total_importo": float(result.get("total_importo") or 0.0),
        "page": int(result.get("page") or 1),
        "pages": int(result.get("pages") or 1),
    }

    table_tabs = [
        {"key": key, "label": str(cfg2.get("title") or key.title())}
        for key, cfg2 in FINANCE_TABLES.items()
    ]

    return render_template(
        "estrazioni_finance.html",
        stores=stores,
        table_tabs=table_tabs,
        table_key=table_key,
        table_title=result.get("title") or cfg.get("title") or table_key.title(),
        date_field=date_field,
        date_field_options=[
            {"key": "DataContabile", "label": "Data contabile"},
            {"key": "DataValuta", "label": "Data valuta"},
            {"key": "DataInserimento", "label": "Data inserimento"},
        ],
        start_iso=start_iso,
        end_iso=end_iso,
        banca=banca,
        societa=societa,
        type_value=type_value,
        tessera=tessera,
        atm=atm,
        search=search,
        store_options=store_options,
        selected_store_values=selected_store_values,
        include_blank_store=include_blank_store,
        result=result,
        summary=summary,
        is_admin=str(session.get("role") or "").strip().lower() == "admin",
    )


@estrazioni_bp.get("/finance.xlsx")
def finance_xlsx():
    r = _require_login()
    if r:
        return r
    if Workbook is None:
        flash("Export Excel non disponibile: openpyxl mancante.", "warning")
        return redirect(url_for("estrazioni.finance"))

    table_key = str(request.args.get("table") or "transazioni").strip().lower()
    if table_key not in FINANCE_TABLES:
        table_key = "transazioni"

    cfg = FINANCE_TABLES[table_key]
    has_store = bool(cfg.get("store_column"))
    date_field = str(request.args.get("date_field") or "DataContabile").strip()
    if date_field not in DATE_FIELDS:
        date_field = "DataContabile"

    stores = _available_stores_for_user()
    store_options = _finance_store_options_for_user(table_key, stores, str(session.get("role") or ""))
    allowed_store_values = {str((x or {}).get("value") or "").strip() for x in (store_options or [])}
    selected_store_values = [str(x or "").strip() for x in request.args.getlist("stores") if str(x or "").strip()]
    selected_store_values = [x for x in selected_store_values if x in allowed_store_values]
    include_blank_store = has_store and str(request.args.get("include_blank_store") or "").strip().lower() in {"1", "true", "on", "yes"}

    result = list_finance_rows(
        table_key=table_key,
        date_field=date_field,
        start_iso=str(request.args.get("start") or "").strip(),
        end_iso=str(request.args.get("end") or "").strip(),
        store_values=selected_store_values if has_store else None,
        include_blank_store=include_blank_store if has_store else False,
        banca=str(request.args.get("banca") or "").strip(),
        societa=str(request.args.get("societa") or "").strip(),
        type_value=str(request.args.get("type_value") or "").strip(),
        search=str(request.args.get("q") or "").strip(),
        tessera=str(request.args.get("tessera") or "").strip(),
        atm=str(request.args.get("atm") or "").strip(),
        export_all=True,
    )
    result = _augment_finance_result_with_store_app(result, stores, table_key)

    wb = Workbook()
    ws = wb.active
    ws.title = str(result.get("title") or cfg.get("title") or table_key.title())[:31]
    headers = [str(c.get("label") or c.get("key") or "") for c in (result.get("columns") or [])]
    ws.append(headers)
    if Font is not None:
        for cell in ws[1]:
            cell.font = Font(bold=True)

    col_keys = [str(c.get("key") or "") for c in (result.get("columns") or [])]
    for row in (result.get("rows") or []):
        out = []
        for key in col_keys:
            value = (row or {}).get(key)
            if hasattr(value, "strftime"):
                try:
                    out.append(value.strftime("%d/%m/%Y %H:%M"))
                except Exception:
                    out.append(str(value))
            else:
                out.append(value)
        ws.append(out)

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[letter].width = min(max_len + 2, 36)

    meta = wb.create_sheet("Riepilogo")
    meta.append(["Tabella", str(result.get("title") or cfg.get("title") or table_key.title())])
    meta.append(["Data su", date_field])
    meta.append(["Dal", str(request.args.get("start") or "").strip()])
    meta.append(["Al", str(request.args.get("end") or "").strip()])
    meta.append(["Banca", str(request.args.get("banca") or "").strip()])
    meta.append(["Società", str(request.args.get("societa") or "").strip()])
    meta.append([str(cfg.get("type_label") or "Tipo"), str(request.args.get("type_value") or "").strip()])
    meta.append(["Ricerca", str(request.args.get("q") or "").strip()])
    meta.append(["Righe", int(result.get("total_rows") or 0)])
    meta.append(["Importo totale", float(result.get("total_importo") or 0.0)])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"finance_{table_key}_{date.today().isoformat()}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
    )




