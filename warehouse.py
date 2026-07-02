from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, current_app, Response, send_file
import time
import io

from datetime import datetime, date, timedelta
from decimal import Decimal

# Excel export (openpyxl)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except Exception:
    Workbook = None  # type: ignore
    Font = None  # type: ignore
    Alignment = None  # type: ignore

from access_db import test_connection
from app_db import get_connection, get_backend
from db_integration import get_warehouse_stores, get_user_warehouse_stores, upsert_warehouse_store
from delivery_repository import (
    get_recent_deliveries,
    get_delivery_table_name,
    get_suppliers_for_store,
    get_price_list_for_supplier,
    save_delivery_document,
    get_delivery_document_rows,
    delete_delivery_row,
    update_delivery_ddt_dates,
    _detect_delivery_layout,
)



from inventory_repository import (
    get_conversions_for_supplier,
    get_delivery_avg_prices_last_weeks,
    save_inventory_movement,
    get_inventory_document_rows,
    replace_inventory_movement,
    save_inventory_document,
    delete_inventory_row,
    update_inventory_movement_header,
)

from collect_repository import (
    get_collect_month_data,
    get_collect_day_breakdown,
    get_magazzino_month_summary,
)

from analysis_repository import (
    get_consumption_summary,
    get_consumption_summary_single_connection,
    get_consumption_summary_on_conn,
)

from consumi_repository import (
    get_product_consumption_table,
)

from warehouse_search_repository import (
    search_ddt_headers,
    search_inventory_headers,
)

from orders_repository import build_orders_suggestion
from supplier_orders_repository import (
    ensure_supplier_orders_schema,
    migrate_legacy_pricelists,
    list_fornitori as list_sql_fornitori,
    save_fornitore,
    delete_fornitore,
    list_fornitore_contacts,
    save_fornitore_contact,
    delete_fornitore_contact,
    list_listino_types,
    list_listino_groups,
    save_listino_type,
    save_listino_group,
    delete_listino_type,
    delete_listino_group,
    ensure_default_price_list_assignments,
    list_price_lists,
    save_price_list,
    set_default_price_list,
    delete_price_list,
    list_price_list_store_assignments,
    replace_price_list_store_assignments,
)
from supplier_order_flow_repository import (
    add_supplier_order_selected_product,
    ORDER_MODE_MAIL,
    ORDER_MODE_ONLINE,
    ORDER_STATUS_COMPLETED,
    ORDER_STATUS_INVITED,
    ORDER_STATUS_SENT,
    ORDER_STATUS_VIEWED,
    build_order_pdf_bytes,
    complete_supplier_order,
    create_supplier_order,
    delete_supplier_order,
    ensure_supplier_order_flow_schema,
    get_supplier_by_uuid,
    get_supplier_order_detail,
    get_supplier_user_supplier_ids,
    get_supplier_user_assignments,
    link_order_to_ddt,
    list_supplier_orders,
    load_supplier_pricelist_for_order,
    mark_order_viewed,
    reopen_supplier_order,
    replace_supplier_user_assignments,
    resolve_supplier_contact,
    resolve_supplier_user_supplier_ids,
    save_order_pdf,
    save_supplier_order_line,
    send_order_email,
    normalize_order_mode,
    update_order_pdf_and_email,
)

warehouse_bp = Blueprint("warehouse", __name__, url_prefix="/magazzino")
_AVAILABLE_STORES_CACHE_TTL_SECONDS = 60


def _ensure_admin_warehouse_access():
    role = str(session.get("role") or "").strip().lower()
    if role == "admin":
        return True
    if bool(session.get("is_master")) or role == "master":
        try:
            from tenant_config_repository import get_tenant

            tenant_key = str(session.get("master_admin_tenant_key") or "").strip()
            if not tenant_key:
                flash("Seleziona prima il tenant da amministrare.", "warning")
                return False
            if bool(get_tenant(str(tenant_key or "")).get("master_can_admin")):
                return True
        except Exception:
            pass
    if role != "admin":
        flash("Questa pagina è riservata agli admin.", "warning")
        return False
    return True


def _session_module_enabled(module_key: str) -> bool:
    role = str(session.get("role") or "").strip().lower()
    if role == "admin":
        return True
    mods = session.get("access_modules")
    if isinstance(mods, dict) and module_key in mods:
        return bool(mods.get(module_key))
    return True


def _ensure_supplier_orders_internal_access():
    role = str(session.get("role") or "").strip().lower()
    if role == "fornitore":
        flash("Questa pagina è riservata agli utenti interni.", "warning")
        return False
    if not _session_module_enabled("mod_supplier_orders"):
        flash("Non hai i permessi per accedere a questa sezione.", "warning")
        return False
    return True


def _ensure_supplier_orders_supplier_access():
    role = str(session.get("role") or "").strip().lower()
    if role not in {"fornitore", "admin"}:
        flash("Questa pagina è riservata ai profili fornitore.", "warning")
        return False
    if role != "admin" and not _session_module_enabled("mod_supplier_orders"):
        flash("Non hai i permessi per accedere a questa sezione.", "warning")
        return False
    return True


def _list_store_options_for_admin_context(active_only: bool = False):
    try:
        from tenant_config_repository import current_tenant_key, list_tenant_stores

        tenant_key = str(
            session.get("master_admin_tenant_key")
            or session.get("tenant_key")
            or ""
        ).strip()
        default_key = str(current_tenant_key() or "").strip() or "default"
        if tenant_key and tenant_key != default_key:
            rows = list_tenant_stores(tenant_key, active_only=active_only)
            stores = []
            for row in rows or []:
                code = str((row or {}).get("store_code") or "").strip()
                if not code:
                    continue
                stores.append(
                    {
                        "code": code,
                        "name": str((row or {}).get("store_name") or code).strip() or code,
                        "is_active": bool((row or {}).get("is_active", True)),
                    }
                )
            return stores
    except Exception:
        pass
    try:
        return get_warehouse_stores() or []
    except Exception:
        return []


def _norm(s: str) -> str:
    return (s or "").strip().lower()


def _find_col(cols, candidates):
    cand = {_norm(c) for c in (candidates or [])}
    for c in (cols or []):
        if _norm(c) in cand:
            return c
    return None


def _json_safe_value(v):
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, Decimal):
        try:
            return float(v)
        except Exception:
            return str(v)
    if isinstance(v, (datetime, date)):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    # fallback: try numeric then string
    try:
        return float(v)
    except Exception:
        return str(v)


def _json_safe_rows(rows):
    out = []
    for r in rows or []:
        if isinstance(r, dict):
            out.append({k: _json_safe_value(v) for k, v in r.items()})
        else:
            out.append(_json_safe_value(r))
    return out



def _json_safe_deep(v):
    """Recursively convert objects to JSON-serializable primitives."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, Decimal):
        try:
            return float(v)
        except Exception:
            return str(v)
    if isinstance(v, (datetime, date)):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    if isinstance(v, dict):
        return {str(k): _json_safe_deep(val) for k, val in v.items()}
    if isinstance(v, (list, tuple, set)):
        return [_json_safe_deep(x) for x in v]
    # fallback
    try:
        return float(v)
    except Exception:
        return str(v)


def _json_safe_rows_deep(rows):
    return [_json_safe_deep(r) for r in (rows or [])]


# -------------------------------------------------
#  Pagina Ricerca (DDT + Inventario/TX)
# -------------------------------------------------


@warehouse_bp.get("/ricerca")
def ricerca():
    """Ricerca movimentazioni (DDT e Inventario/TX) per intervallo date."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("warehouse.home"))

    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    kind = (request.args.get("kind") or "").strip().upper() or "ALL"

    results = []
    if start and end:
        if kind == "DDT":
            results = search_ddt_headers(store_code=str(store_code), start_iso=start, end_iso=end)
        elif kind in ("INV", "TXIN", "TXOUT"):
            results = search_inventory_headers(store_code=str(store_code), start_iso=start, end_iso=end, mov_type=kind)
        elif kind == "WASTE CRUDO":
            results = search_inventory_headers(store_code=str(store_code), start_iso=start, end_iso=end, mov_type=kind)
        elif kind in ("ALL", "TUTTO"):
            ddt = search_ddt_headers(store_code=str(store_code), start_iso=start, end_iso=end)
            inv = search_inventory_headers(store_code=str(store_code), start_iso=start, end_iso=end, mov_type=None)
            results = (ddt or []) + (inv or [])
        else:
            # fallback: prova inventario con mov_type specifico
            results = search_inventory_headers(store_code=str(store_code), start_iso=start, end_iso=end, mov_type=kind)

        # Mappa codice->nome fornitore (DDT usa il codice, in UI preferiamo il nome)
        try:
            suppliers_info = get_suppliers_for_store(store_code)
            suppliers = suppliers_info.get("suppliers") if suppliers_info and not suppliers_info.get("error") else []
            sup_map = {str(s.get("code")): (s.get("name") or str(s.get("code"))) for s in (suppliers or [])}
        except Exception:
            sup_map = {}

        for r in results or []:
            try:
                if (r.get("kind") or "").upper() == "DDT":
                    code = str(r.get("supplier") or "").strip()
                    r["supplier_code"] = code
                    r["supplier_label"] = sup_map.get(code, code)
                else:
                    # inventario: supplier è già il nome
                    r["supplier_label"] = r.get("supplier")
            except Exception:
                continue

        # ordinamento: data desc, poi tipo
        def _key(r):
            return (r.get("date") or "", r.get("kind") or "")

        results = sorted(results, key=_key, reverse=True)

    return render_template(
        "warehouse_search.html",
        store_code=store_code,
        store_name=store_name,
        start=start,
        end=end,
        kind=kind,
        results=_json_safe_rows_deep(results),
    )


# -------------------------------------------------
#  API - Raccolta dati (Dashboard mensile)
# -------------------------------------------------


@warehouse_bp.get("/api/collect/month")
def api_collect_month():
    """Ritorna i totali giornalieri del mese per movement+category."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    if not store_code:
        return jsonify({"error": "store_not_selected"}), 400

    try:
        year = int(request.args.get("year") or 0)
        month = int(request.args.get("month") or 0)
        movement = (request.args.get("movement") or "").strip()
        category = (request.args.get("category") or "FoodPaper").strip()
        if not (1 <= month <= 12) or year < 2000:
            return jsonify({"error": "invalid_year_month"}), 400

        data = get_collect_month_data(
            store_code=str(store_code),
            year=year,
            month=month,
            movement=movement,
            category=category,
        )
        try:
            summary = get_magazzino_month_summary(
                store_code=str(store_code),
                year=year,
                month=month,
                category=category,
                current_movement=str(data.get('movement') or movement).strip().upper(),
                current_total=float(data.get('total') or 0.0),
            )
            data['summary'] = summary
        except Exception:
            # non bloccare il caricamento del calendario se fallisce il riepilogo
            data['summary'] = None

        return jsonify(_json_safe_deep(data))
    except Exception as e:
        current_app.logger.exception("Errore api_collect_month")
        return jsonify({"error": str(e)}), 500


@warehouse_bp.get("/api/collect/breakdown")
def api_collect_breakdown():
    """Ritorna il breakdown per giorno (supplier x bucket) per un movement."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    if not store_code:
        return jsonify({"error": "store_not_selected"}), 400

    try:
        day_iso = (request.args.get("date") or "").strip()
        movement = (request.args.get("movement") or "").strip()
        if not day_iso:
            return jsonify({"error": "missing_date"}), 400

        data = get_collect_day_breakdown(
            store_code=str(store_code),
            day_iso=day_iso,
            movement=movement,
        )
        return jsonify(_json_safe_deep(data))
    except Exception as e:
        current_app.logger.exception("Errore api_collect_breakdown")
        return jsonify({"error": str(e)}), 500


# -------------------------------------------------
#  Pagina Analisi + API (Consumi)
# -------------------------------------------------


@warehouse_bp.get("/analisi")
def analysis():
    """Pagina Analisi (consumi)."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")

    suppliers_info = get_suppliers_for_store(store_code) if store_code else {}
    suppliers = suppliers_info.get("suppliers") if suppliers_info and not suppliers_info.get("error") else []
    suppliers_error = suppliers_info.get("error") if suppliers_info else None

    return render_template(
        "warehouse_analysis.html",
        store_code=store_code,
        store_name=store_name,
        suppliers=suppliers,
        suppliers_error=suppliers_error,
    )


@warehouse_bp.get("/api/analysis/summary")
def api_analysis_summary():
    """Ritorna i KPI consumo per periodo (start/end) ed eventuale fornitore."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    if not store_code:
        return jsonify({"error": "store_not_selected"}), 400

    try:
        start = (request.args.get("start") or "").strip()
        end = (request.args.get("end") or "").strip()
        supplier = (request.args.get("supplier") or "").strip() or None
        if not start or not end:
            return jsonify({"error": "missing_start_end"}), 400

        data = get_consumption_summary(
            store_code=str(store_code),
            start_iso=start,
            end_iso=end,
            supplier_name=supplier,
        )
        return jsonify(_json_safe_deep(data))
    except Exception as e:
        current_app.logger.exception("Errore api_analysis_summary")
        return jsonify({"error": str(e)}), 500


# -------------------------------------------------
#  Pagina Riepilogo (Mensile) + API
# -------------------------------------------------


def _parse_month_yyyy_mm(value: str) -> tuple[date, date]:
    """Ritorna (start, end) del mese indicato (YYYY-MM). Fallback: mese corrente."""
    value = (value or "").strip()
    today = date.today()
    y = today.year
    m = today.month
    try:
        if value and "-" in value:
            p = value.split("-")
            y2 = int(p[0])
            m2 = int(p[1])
            if 1 <= m2 <= 12 and y2 >= 2000:
                y, m = y2, m2
    except Exception:
        pass

    start = date(y, m, 1)
    # ultimo giorno del mese
    if m == 12:
        end = date(y + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(y, m + 1, 1) - timedelta(days=1)
    return start, end


@warehouse_bp.get("/riepilogo")
def riepilogo():
    """Pagina Riepilogo mensile multi-store (FoodPaper / Operating)."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")

    return render_template(
        "warehouse_riepilogo.html",
        store_code=store_code,
        store_name=store_name,
    )



@warehouse_bp.get("/riepilogo/export")
def riepilogo_export():
    """Esporta in Excel il riepilogo mensile (multi-store) mostrato nella pagina Riepilogo."""
    _ensure_session_keys()
    user_id = session.get("uid")
    if not user_id:
        return jsonify({"ok": False, "error": "not_logged_in"}), 401

    if Workbook is None:
        return jsonify({"ok": False, "error": "Dipendenza mancante: openpyxl"}), 500

    month = (request.args.get("month") or "").strip()  # YYYY-MM
    bucket = (request.args.get("bucket") or "FoodPaper").strip() or "FoodPaper"
    if bucket not in ("FoodPaper", "Operating"):
        bucket = "FoodPaper"

    start_d, end_d = _parse_month_yyyy_mm(month)
    start_iso = start_d.isoformat()
    end_iso = end_d.isoformat()
    month_norm = month or start_iso[:7]

    try:
        stores = _available_stores_for_user(str(user_id)) or []
    except Exception:
        stores = []


    rows = []
    warnings_all = []

    conn_shared = None
    if get_backend() == "sqlserver":
        try:
            conn_shared = get_connection(None, read_only=True)
        except Exception:
            conn_shared = None

    for s in stores:
        code = str((s or {}).get("code") or "").strip()
        name = str((s or {}).get("name") or code).strip()
        if not code:
            continue

        try:
            if conn_shared is not None and get_backend() == "sqlserver":

                try:

                    data = get_consumption_summary_on_conn(

                        conn_shared,

                        store_code=code,

                        start_iso=start_iso,

                        end_iso=end_iso,

                        supplier_name=None,

                        buckets=[bucket],

                    )

                except Exception:

                    data = get_consumption_summary_single_connection(

                        store_code=code,

                        start_iso=start_iso,

                        end_iso=end_iso,

                        supplier_name=None,

                        buckets=[bucket],

                    )

            else:

                data = get_consumption_summary_single_connection(

                    store_code=code,

                    start_iso=start_iso,

                    end_iso=end_iso,

                    supplier_name=None,

                    buckets=[bucket],

                )
            b = (data.get("buckets") or {}).get(bucket) or {}

            row = {
                "store_code": code,
                "store_name": name,
                "revenues_net": float(data.get("revenues_net") or 0.0),
                "inv_initial": float(b.get("inv_initial") or 0.0),
                "tx_in": float(b.get("tx_in") or 0.0),
                "tx_out": float(b.get("tx_out") or 0.0),
                "delivery": float(b.get("delivery") or 0.0),
                "inv_final": float(b.get("inv_final") or 0.0),
                "consumption": float(b.get("consumption") or 0.0),
                # in Excel usiamo frazione (0-1) così la formattazione percentuale è corretta
                "consumption_pct": float(b.get("consumption_pct") or 0.0) / 100.0,
                "waste": float(b.get("waste") or 0.0),
                "waste_pct": float(b.get("waste_pct") or 0.0) / 100.0,
            }

            w = data.get("warnings") or []
            if w:
                for msg in w:
                    msg = str(msg or "").strip()
                    if msg:
                        warnings_all.append(f"[{code}] {msg}")

            rows.append(row)
        except Exception as e:
            current_app.logger.exception("Errore export riepilogo store %s", code)
            rows.append({
                "store_code": code,
                "store_name": name,
                "error": str(e),
            })

    try:
        rows = sorted(rows, key=lambda r: str(r.get("store_code") or ""))
    except Exception:
        pass

    # Dedup warnings
    uniq_w = []
    seen = set()
    for w in warnings_all:
        if w not in seen:
            uniq_w.append(w)
            seen.add(w)

    wb = Workbook()
    ws = wb.active
    ws.title = "Riepilogo"

    # Intestazioni
    ws.append([f"Riepilogo Magazzino - {bucket}"])
    ws.append([f"Mese: {month_norm}", f"Periodo: {start_iso} → {end_iso}"])
    ws.append([])

    headers = [
        "STORE",
        "CODICE",
        "REV (net)",
        "INV INIZ",
        "TX IN",
        "TX OUT",
        "DELIVERY",
        "INV FIN",
        "CONS",
        "CONS %",
        "WASTE",
        "WASTE %",
    ]
    ws.append(headers)

    header_row_idx = ws.max_row
    if Font is not None:
        for c in ws[header_row_idx]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

    # Righe dati
    for r in rows:
        if r.get("error"):
            ws.append([
                r.get("store_name") or r.get("store_code") or "",
                r.get("store_code") or "",
                None, None, None, None, None, None, None, None, None, None
            ])
            # Scriviamo l'errore come nota in una colonna extra "nascosta" (colonna M)
            ws.cell(row=ws.max_row, column=13).value = str(r.get("error") or "")
            continue

        ws.append([
            r.get("store_name") or "",
            r.get("store_code") or "",
            r.get("revenues_net") or 0.0,
            r.get("inv_initial") or 0.0,
            r.get("tx_in") or 0.0,
            r.get("tx_out") or 0.0,
            r.get("delivery") or 0.0,
            r.get("inv_final") or 0.0,
            r.get("consumption") or 0.0,
            r.get("consumption_pct") or 0.0,
            r.get("waste") or 0.0,
            r.get("waste_pct") or 0.0,
        ])

    # Riga totale
    data_start = header_row_idx + 1
    data_end = ws.max_row
    ws.append([])
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1).value = "TOTALE"
    ws.cell(row=total_row, column=1).font = Font(bold=True) if Font else None
    # Somme colonne euro: C,D,E,F,G,H,I,K
    sum_cols = [3,4,5,6,7,8,9,11]
    for col in sum_cols:
        col_letter = ws.cell(row=header_row_idx, column=col).column_letter
        ws.cell(row=total_row, column=col).value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        if Font:
            ws.cell(row=total_row, column=col).font = Font(bold=True)

    # Formati
    eur_fmt = u"€ #,##0.00"
    pct_fmt = "0.00%"
    # applica formati per colonne dati (solo se ci sono righe)
    for row_idx in range(data_start, data_end + 1):
        # euro
        for col_idx in [3,4,5,6,7,8,9,11]:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = eur_fmt
                if Alignment:
                    cell.alignment = Alignment(horizontal="right")
        # percentuali
        for col_idx in [10,12]:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = pct_fmt
                if Alignment:
                    cell.alignment = Alignment(horizontal="right")

    # Formati anche sul totale
    for col_idx in [3,4,5,6,7,8,9,11]:
        cell = ws.cell(row=total_row, column=col_idx)
        cell.number_format = eur_fmt
        if Alignment:
            cell.alignment = Alignment(horizontal="right")

    # Larghezze colonne
    widths = [28, 10, 14, 14, 14, 14, 14, 14, 14, 10, 14, 10, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=header_row_idx, column=i).column_letter].width = w

    # Freeze header
    ws.freeze_panes = ws["A{}".format(header_row_idx + 1)]
    ws.auto_filter.ref = f"A{header_row_idx}:L{data_end}"

    # Foglio avvisi
    if uniq_w:
        ws2 = wb.create_sheet("Avvisi")
        ws2.append(["Avvisi / anomalie"])
        if Font:
            ws2["A1"].font = Font(bold=True)
        ws2.append([])
        for w in uniq_w[:500]:
            ws2.append([w])

        ws2.column_dimensions["A"].width = 80

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"magazzino_riepilogo_{bucket.lower()}_{month_norm}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@warehouse_bp.get("/api/riepilogo/mensile")
def api_riepilogo_mensile():
    """Ritorna il riepilogo mensile per tutti gli store disponibili all'utente."""
    _ensure_session_keys()
    user_id = session.get("uid")
    if not user_id:
        return jsonify({"ok": False, "error": "not_logged_in"}), 401

    month = (request.args.get("month") or "").strip()  # YYYY-MM
    bucket = (request.args.get("bucket") or "FoodPaper").strip() or "FoodPaper"
    if bucket not in ("FoodPaper", "Operating"):
        bucket = "FoodPaper"

    start_d, end_d = _parse_month_yyyy_mm(month)
    start_iso = start_d.isoformat()
    end_iso = end_d.isoformat()

    try:
        stores = _available_stores_for_user(str(user_id)) or []
    except Exception:
        stores = []


    rows = []
    warnings_all = []

    conn_shared = None
    if get_backend() == "sqlserver":
        try:
            conn_shared = get_connection(None, read_only=True)
        except Exception:
            conn_shared = None

    for s in stores:
        code = str((s or {}).get("code") or "").strip()
        name = str((s or {}).get("name") or code).strip()
        if not code:
            continue

        try:
            # Per evitare errori ODBC tipo "Too many client tasks" su multi-store,
            # riusiamo una sola connessione Access per store.
            if conn_shared is not None and get_backend() == "sqlserver":

                try:

                    data = get_consumption_summary_on_conn(

                        conn_shared,

                        store_code=code,

                        start_iso=start_iso,

                        end_iso=end_iso,

                        supplier_name=None,

                        buckets=[bucket],

                    )

                except Exception:

                    data = get_consumption_summary_single_connection(

                        store_code=code,

                        start_iso=start_iso,

                        end_iso=end_iso,

                        supplier_name=None,

                        buckets=[bucket],

                    )

            else:

                data = get_consumption_summary_single_connection(

                    store_code=code,

                    start_iso=start_iso,

                    end_iso=end_iso,

                    supplier_name=None,

                    buckets=[bucket],

                )
            b = (data.get("buckets") or {}).get(bucket) or {}

            rev = float(data.get("revenues_net") or 0.0)
            row = {
                "store_code": code,
                "store_name": name,
                "revenues_net": rev,
                "inv_initial": float(b.get("inv_initial") or 0.0),
                "tx_in": float(b.get("tx_in") or 0.0),
                "tx_out": float(b.get("tx_out") or 0.0),
                "delivery": float(b.get("delivery") or 0.0),
                "inv_final": float(b.get("inv_final") or 0.0),
                "consumption": float(b.get("consumption") or 0.0),
                "consumption_pct": float(b.get("consumption_pct") or 0.0),
                "waste": float(b.get("waste") or 0.0),
                "waste_pct": float(b.get("waste_pct") or 0.0),
            }

            w = data.get("warnings") or []
            if w:
                for msg in w:
                    msg = str(msg or "").strip()
                    if msg:
                        warnings_all.append(f"[{code}] {msg}")

            rows.append(row)
        except Exception as e:
            current_app.logger.exception("Errore riepilogo store %s", code)
            rows.append({
                "store_code": code,
                "store_name": name,
                "error": str(e),
            })


    # Chiudi connessione condivisa SQL Server (se usata)
    try:
        if conn_shared is not None:
            conn_shared.close()
    except Exception:
        pass
    # Ordina per store_code
    try:
        rows = sorted(rows, key=lambda r: str(r.get("store_code") or ""))
    except Exception:
        pass

    # Dedup warnings
    uniq_w = []
    seen = set()
    for w in warnings_all:
        if w not in seen:
            uniq_w.append(w)
            seen.add(w)
    out = {
        "ok": True,
        "bucket": bucket,
        "month": month or start_iso[:7],
        "period": {
            "start": start_iso,
            "end": end_iso,
        },
        "rows": rows,
        "warnings": uniq_w[:200],
    }

    return jsonify(_json_safe_deep(out))


# -------------------------------------------------
#  Pagina Consumi (pezzi) + API
# -------------------------------------------------


@warehouse_bp.get("/consumi")
def consumi():
    """Pagina Consumi: consumo in pezzi per prodotto."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")

    suppliers_info = get_suppliers_for_store(store_code) if store_code else {}
    suppliers = suppliers_info.get("suppliers") if suppliers_info and not suppliers_info.get("error") else []
    suppliers_error = suppliers_info.get("error") if suppliers_info else None

    return render_template(
        "warehouse_consumi.html",
        store_code=store_code,
        store_name=store_name,
        suppliers=suppliers,
        suppliers_error=suppliers_error,
    )


@warehouse_bp.get("/api/consumi/table")
def api_consumi_table():
    """Ritorna la tabella consumo in pezzi per prodotto (fornitore + listino)."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    if not store_code:
        return jsonify({"error": "store_not_selected"}), 400

    try:
        start = (request.args.get("start") or "").strip()
        end = (request.args.get("end") or "").strip()
        supplier = (request.args.get("supplier") or "").strip()
        listino = (request.args.get("listino") or "").strip() or "FoodPaper"
        if not start or not end:
            return jsonify({"error": "missing_start_end"}), 400
        if not supplier:
            return jsonify({"error": "missing_supplier"}), 400

        data = get_product_consumption_table(
            store_code=str(store_code),
            start_iso=start,
            end_iso=end,
            supplier_code=supplier,
            listino=listino,
        )
        return jsonify(_json_safe_deep(data))
    except Exception as e:
        current_app.logger.exception("Errore api_consumi_table")
        return jsonify({"error": str(e)}), 500




def _require_login():
    """Protegge le route del magazzino: richiede sessione Supabase."""
    if not session.get("uid") or not session.get("sb_token"):
        return redirect(url_for("login", next=request.full_path))
    return None


@warehouse_bp.before_request
def _warehouse_before_request():
    _ensure_session_keys()
    guard = _require_login()
    if guard:
        return guard

    # Route che NON richiedono store selezionato
    if request.endpoint in (
        "warehouse.select_store",
        "warehouse.stores_json",
        "warehouse.stores_json_all",
        "warehouse.home",
        "warehouse.fornitori",
        "warehouse.listini_anagrafica",
        "warehouse.listini_prezzi",
        "warehouse.riepilogo",
        "warehouse.api_riepilogo_mensile",
        "warehouse.riepilogo_export",
        "warehouse.supplier_orders_home",
        "warehouse.supplier_orders_received",
        "warehouse.supplier_orders_received_complete",
        "warehouse.supplier_orders_received_line",
        "warehouse.supplier_orders_received_reopen",
        "warehouse.supplier_orders_received_add_product",
        "warehouse.supplier_order_pdf",
    ):
        return None

    # Se non c'è store selezionato, riportiamo a /magazzino/ e facciamo scegliere via popup
    if not session.get("store_code"):
        flash("Seleziona uno store per continuare.", "warning")
        return redirect(url_for("warehouse.home"))

    return None


def _available_stores_for_user(user_id: str | None):
    """Ritorna la lista store visibile all'utente.
    - Admin (ruolo verificato): tutti gli store
    - User: solo store assegnati in warehouse_user_stores
    """
    cache_uid = str(user_id or "").strip()
    cache_role = str(session.get("tenant_role") or session.get("role") or "").strip().lower()
    cache_tenant = str(session.get("tenant_key") or session.get("master_admin_tenant_key") or "").strip().lower()
    try:
        cached_payload = session.get("available_stores_cache")
        cached_ts = float(session.get("available_stores_cache_ts") or 0)
        cached_uid = str(session.get("available_stores_cache_uid") or "").strip()
        cached_role = str(session.get("available_stores_cache_role") or "").strip().lower()
        cached_tenant = str(session.get("available_stores_cache_tenant") or "").strip().lower()
        if (
            isinstance(cached_payload, list)
            and (time.time() - cached_ts) < _AVAILABLE_STORES_CACHE_TTL_SECONDS
            and cached_uid == cache_uid
            and cached_role == cache_role
            and cached_tenant == cache_tenant
        ):
            return cached_payload
    except Exception:
        pass

    try:
        # Verifica ruolo con SERVICE_ROLE: l'elenco store deve riflettere subito
        # cambi ruolo fatti da Admin/Master, senza dipendere da cookie/sessioni vecchie.
        role = str(session.get("tenant_role") or session.get("role") or "").lower()
        try:
            if user_id and not session.get("tenant_role"):
                from db_integration import get_profile_role_by_id
                srv_role = get_profile_role_by_id(user_id)
                if srv_role:
                    role = str(srv_role).lower()
                    session["role"] = srv_role
                    session["role_verified_for"] = user_id
                    session["role_verified_at"] = int(time.time())
        except Exception:
            pass

        if role == "master":
            stores = []
        elif role == "admin":
            stores = get_warehouse_stores() or []
        else:
            assigned = []
            if user_id:
                try:
                    assigned = get_user_warehouse_stores(str(user_id))
                except Exception:
                    assigned = []

            all_stores = get_warehouse_stores() or []
            allowed_codes = {str(row.get("store_code")) for row in assigned if row.get("store_code")}

            if allowed_codes:
                stores = [s for s in all_stores if str((s or {}).get("code")) in allowed_codes]
            else:
                # Nessuna assegnazione: non mostrare lista completa
                cur_code = session.get("store_code")
                if cur_code:
                    stores = [s for s in all_stores if str((s or {}).get("code")) == str(cur_code)]
                else:
                    stores = []
    except Exception:
        stores = []

    # Ordina alfabeticamente per nome store (fallback su code)
    try:
        stores = sorted(
            stores or [],
            key=lambda s: (
                str((s or {}).get("name") or "").strip().lower(),
                str((s or {}).get("code") or "").strip().lower(),
            ),
        )
    except Exception:
        pass

    try:
        session["available_stores_cache"] = stores or []
        session["available_stores_cache_ts"] = time.time()
        session["available_stores_cache_uid"] = cache_uid
        session["available_stores_cache_role"] = cache_role
        session["available_stores_cache_tenant"] = cache_tenant
    except Exception:
        pass

    return stores

@warehouse_bp.get("/stores-json")
def stores_json():
    """Elenco store disponibili per l'utente loggato (per popup selezione store)."""
    _ensure_session_keys()
    user_id = session.get("uid")
    try:
        stores = _available_stores_for_user(str(user_id) if user_id else None)
        return jsonify({"stores": stores or []})
    except Exception as e:
        current_app.logger.exception("Errore caricamento store per utente")
        return jsonify({"stores": [], "error": str(e)}), 500


@warehouse_bp.get("/stores-json-all")
def stores_json_all():
    """Elenco store COMPLETO (non filtrato per utente).

    Usato per la selezione SITE2 nei trasferimenti TXIN/TXOUT.
    """
    _ensure_session_keys()
    try:
        stores = get_warehouse_stores() or []

        # Ordina alfabeticamente per nome store (fallback su code)
        try:
            stores = sorted(
                stores,
                key=lambda s: (
                    str((s or {}).get("name") or "").strip().lower(),
                    str((s or {}).get("code") or "").strip().lower(),
                ),
            )
        except Exception:
            pass

        return jsonify({"stores": stores})
    except Exception as e:
        current_app.logger.exception("Errore caricamento elenco store completo")
        return jsonify({"stores": [], "error": str(e)}), 500




def _ensure_session_keys() -> None:
    """Garantisce che le chiavi di sessione esistano sempre."""
    session.setdefault("store_code", None)
    session.setdefault("store_name", None)
    session.setdefault("supplier_order_ddt_prefill", None)


def _supplier_order_allowed_sites() -> list[dict]:
    uid = session.get("uid")
    return _available_stores_for_user(str(uid) if uid else None) or []


def _supplier_order_allowed_site_codes() -> list[str]:
    return [str(s.get("code")) for s in (_supplier_order_allowed_sites() or []) if s.get("code")]


def _supplier_order_store_name(store_code: str) -> str:
    code = str(store_code or "")
    for s in _supplier_order_allowed_sites():
        if str(s.get("code")) == code:
            return str(s.get("name") or code)
    return code


def _supplier_order_contact_summary(contact: dict | None) -> str:
    contact = contact or {}
    bits = []
    if contact.get("Referente"):
        bits.append(str(contact.get("Referente")))
    if contact.get("Email"):
        bits.append(str(contact.get("Email")))
    return " · ".join(bits)


def _supplier_order_mail_subject(store_name: str, requested_delivery_date: str) -> str:
    return f"Nuovo ordine - {store_name} - consegna {requested_delivery_date}"


def _supplier_order_mail_body(order_detail: dict) -> str:
    header = (order_detail or {}).get("header") or {}
    lines = [
        "Buongiorno,",
        "",
        f"in allegato il nuovo ordine {header.get('numero_ordine') or ''}.",
        f"Store: {header.get('site') or ''} {('- ' + str(header.get('store_name') or '')) if header.get('store_name') else ''}",
        f"Consegna richiesta: {header.get('requested_delivery_date') or ''}",
    ]
    note = str(header.get("note_ordine") or "").strip()
    if note:
        lines.extend(["", f"Note: {note}"])
    lines.extend(["", "Grazie.", "Store Hub 360"])
    return "\n".join(lines)


def _supplier_order_parse_form_rows(form) -> tuple[list[dict], Decimal]:
    rows: list[dict] = []
    total = Decimal("0")
    rows_count = int(form.get("rows_count") or 0)
    for i in range(rows_count):
        descr = str(form.get(f"row-{i}-descrizione") or "").strip()
        if not descr:
            continue
        qty_colli_s = str(form.get(f"row-{i}-qty_colli") or "").strip()
        qty_pezzi_s = str(form.get(f"row-{i}-qty_pezzi") or "").strip()
        qty_colli = Decimal(qty_colli_s.replace(",", ".")) if qty_colli_s else Decimal("0")
        qty_pezzi = Decimal(qty_pezzi_s.replace(",", ".")) if qty_pezzi_s else Decimal("0")
        if qty_colli < 0 or qty_pezzi < 0:
            raise ValueError("Le quantità non possono essere negative.")
        qta_car = form.get(f"row-{i}-qta_car")
        qta_car_dec = Decimal(str(qta_car)) if str(qta_car or "").strip() else Decimal("0")
        qty_total = qty_colli
        if qty_pezzi > 0:
            qty_total += (qty_pezzi / qta_car_dec) if qta_car_dec > 0 else qty_pezzi
        if qty_total <= 0:
            continue
        estimated_price = Decimal(str(form.get(f"row-{i}-estimated_price") or "0").replace(",", "."))
        subtotal = (qty_total * estimated_price).quantize(Decimal("0.01"))
        total += subtotal
        rows.append(
            {
                "sort_order": i,
                "tipo_listino": form.get(f"row-{i}-tipo_listino"),
                "product_code": form.get(f"row-{i}-product_code"),
                "descrizione": descr,
                "gruppo": form.get(f"row-{i}-gruppo"),
                "unita": form.get(f"row-{i}-unita"),
                "qta_car": form.get(f"row-{i}-qta_car"),
                "qta_int": form.get(f"row-{i}-qta_int"),
                "conv": form.get(f"row-{i}-conv"),
                "qty_colli": qty_colli,
                "qty_pezzi": qty_pezzi,
                "qty_ordered": qty_total,
                "estimated_price": estimated_price,
                "price_source": form.get(f"row-{i}-price_source"),
                "subtotal": subtotal,
            }
        )
    return rows, total


def _supplier_order_line_payload(form_data, line_id: str) -> dict:
    lid = str(line_id or "").strip()
    if not lid:
        return {"picked": False, "lotto": "", "scadenza": "", "note_fornitore": ""}
    picked_key = f"picked_{lid}"
    lotto_key = f"lotto_{lid}"
    scadenza_key = f"scadenza_{lid}"
    note_key = f"note_fornitore_{lid}"
    if any(k in form_data for k in (picked_key, lotto_key, scadenza_key, note_key)):
        return {
            "picked": form_data.get(picked_key) == "1",
            "lotto": form_data.get(lotto_key) or "",
            "scadenza": form_data.get(scadenza_key) or "",
            "note_fornitore": form_data.get(note_key) or "",
        }
    return {
        "picked": form_data.get("picked") == "1",
        "lotto": form_data.get("lotto") or "",
        "scadenza": form_data.get("scadenza") or "",
        "note_fornitore": form_data.get("note_fornitore") or "",
    }


def _supplier_order_catalog_options(detail: dict | None) -> list[dict]:
    if not detail:
        return []
    header = detail.get("header") or {}
    return load_supplier_pricelist_for_order(
        str(header.get("site") or "").strip(),
        str(header.get("supplier_name") or "").strip(),
    ) or []


def _supplier_order_effective_rows(detail: dict | None) -> list[dict]:
    rows = list((detail or {}).get("rows") or [])
    replacement_targets = {
        str(r.get("replacement_for_row_uuid") or "").strip()
        for r in rows
        if str(r.get("replacement_for_row_uuid") or "").strip()
    }
    effective = []
    for row in rows:
        row_id = str(row.get("row_uuid") or "").strip()
        is_supplier_added = bool(row.get("supplier_added"))
        if not is_supplier_added and row_id in replacement_targets:
            continue
        effective.append(row)
    return effective


def _supplier_order_access_supplier_code(store_code: str, supplier_name: str) -> str:
    info = get_suppliers_for_store(store_code) or {}
    suppliers = info.get("suppliers") or []
    target = str(supplier_name or "").strip().lower()
    for s in suppliers:
        name = str(s.get("name") or "").strip().lower()
        if name == target:
            return str(s.get("code") or "").strip()
    for s in suppliers:
        name = str(s.get("name") or "").strip().lower()
        if target and (target in name or name in target):
            return str(s.get("code") or "").strip()
    return ""


@warehouse_bp.route("/seleziona-store", methods=["GET", "POST"])
def select_store():
    _ensure_session_keys()
    user_id = session.get("uid")
    stores = _available_stores_for_user(str(user_id) if user_id else None)

    next_url = request.args.get("next") or request.form.get("next") or url_for("dashboard")

    if request.method == "POST":
        store_code = request.form.get("store_code")
        if not store_code:
            flash("Seleziona uno store prima di continuare.", "warning")
            return redirect(url_for("warehouse.select_store", next=next_url))

        store_name = None
        for s in stores or []:
            if str(s.get("code")) == str(store_code):
                store_name = s.get("name") or str(store_code)
                break

        session["store_code"] = store_code
        session["store_name"] = store_name
        flash(f"Store {store_code} selezionato.", "success")
        return redirect(next_url or url_for("dashboard"))

    return render_template("warehouse_select_store.html", stores=stores, next=next_url)


@warehouse_bp.route("/")
def home():
    _ensure_session_keys()
    return render_template("warehouse_home.html")


@warehouse_bp.route("/test-connessione")
def test_connessione():
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("warehouse.home"))

    info = test_connection(store_code)
    return render_template(
        "warehouse_test_connection.html",
        store_code=store_code,
        store_name=store_name,
        info=info,
    )


@warehouse_bp.route("/delivery")
def delivery():
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("warehouse.home"))

    info = get_recent_deliveries(store_code, limit=50)
    return render_template(
        "warehouse_delivery.html",
        store_code=store_code,
        store_name=store_name,
        delivery=info,
    )


@warehouse_bp.route("/delivery/nuovo", methods=["GET", "POST"])
def delivery_new():
    """Maschera di inserimento DDT basata sui listini del fornitore."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("warehouse.home"))

    # Header del documento
    header = {
        "data_doc": "",
        "data_rif": "",
        "numero_ddt": "",
        "causale": "",
        "supplier_code": "",
    }

    # Elenco fornitori da Access
    suppliers_info = get_suppliers_for_store(store_code)
    suppliers = suppliers_info.get("suppliers") if suppliers_info and not suppliers_info.get("error") else []
    suppliers_error = suppliers_info.get("error") if suppliers_info else None

    price_list = None
    order_prefill = session.get("supplier_order_ddt_prefill") if isinstance(session.get("supplier_order_ddt_prefill"), dict) else None

    def _apply_order_prefill():
        nonlocal header, price_list, order_prefill
        if not order_prefill:
            return
        if str(order_prefill.get("supplier_code") or "").strip():
            header["supplier_code"] = str(order_prefill.get("supplier_code") or "").strip()
        if str(order_prefill.get("data_doc") or "").strip():
            header["data_doc"] = str(order_prefill.get("data_doc") or "").strip()
        if str(order_prefill.get("data_rif") or "").strip():
            header["data_rif"] = str(order_prefill.get("data_rif") or "").strip()
        if header["supplier_code"]:
            price_list = get_price_list_for_supplier(store_code, header["supplier_code"])
            if price_list and price_list.get("rows"):
                wanted = {}
                for row in order_prefill.get("rows") or []:
                    key = str(row.get("descrizione") or "").strip().lower()
                    if key:
                        wanted[key] = row
                for pr in price_list.get("rows") or []:
                    key = str(pr.get("Descrizione") or "").strip().lower()
                    src = wanted.get(key)
                    if not src:
                        continue
                    pr["_prefill_colli"] = str(src.get("qty_colli") or "")
                    pr["_prefill_pezzi"] = str(src.get("qty_pezzi") or "")
                    pr["_prefill_prezzo_ddt"] = str(src.get("estimated_price") or "")
                    lot_scad = []
                    if src.get("lotto"):
                        lot_scad.append(f"Lotto {src.get('lotto')}")
                    if src.get("scadenza"):
                        lot_scad.append(f"Scadenza {src.get('scadenza')}")
                    if lot_scad:
                        pr["_supplier_prefill_note"] = " · ".join(lot_scad)
        order_prefill = None

    if request.method == "GET":
        _apply_order_prefill()

    if request.method == "POST":
        header["data_doc"] = request.form.get("data_doc") or ""
        header["data_rif"] = request.form.get("data_rif") or ""
        header["numero_ddt"] = request.form.get("numero_ddt") or ""
        header["causale"] = request.form.get("causale") or ""
        header["supplier_code"] = request.form.get("supplier_code") or ""

        submit_action = request.form.get("submit_action") or "load"

        if submit_action == "load":
            # Carica listino per il fornitore selezionato
            if not header["supplier_code"]:
                flash("Seleziona un fornitore prima di caricare il listino.", "warning")
            else:
                price_list = get_price_list_for_supplier(store_code, header["supplier_code"])
                if price_list.get("error"):
                    flash(price_list["error"], "danger")

        elif submit_action == "save":
            # Salvataggio del DDT in Access (tabella DatiDelivery)

            # Validazione campi data: devono essere presenti
            if not header["data_doc"] or not header["data_rif"]:
                flash("Compila 'Data documento' e 'Data consegna' prima di salvare il DDT.", "danger")
                if header["supplier_code"]:
                    price_list = get_price_list_for_supplier(store_code, header["supplier_code"])
                return render_template(
                    "warehouse_delivery_new.html",
                    store_code=store_code,
                    store_name=store_name,
                    header=header,
                    suppliers=suppliers,
                    suppliers_error=suppliers_error,
                    price_list=price_list,
                )
            rows_count = int(request.form.get("rows_count") or 0)
            cols_count = int(request.form.get("cols_count") or 0)
            unit_column = request.form.get("unit_column") or None
            code_column = request.form.get("code_column") or None

            # Ricostruisci elenco colonne di anagrafica
            cols = []
            for j in range(cols_count):
                col_name = request.form.get(f"col-{j}-name")
                if col_name is not None:
                    cols.append(col_name)

            # Ricostruisci le righe
            rows_data = []
            for i in range(rows_count):
                anag = {}
                for j in range(cols_count):
                    col_name = request.form.get(f"col-{j}-name")
                    if col_name is None:
                        continue
                    cell_key = f"row-{i}-col-{j}"
                    cell_val = request.form.get(cell_key)
                    anag[col_name] = cell_val

                row_struct = {
                    "anag": anag,
                    "colli": request.form.get(f"row-{i}-colli"),
                    "pezzi": request.form.get(f"row-{i}-pezzi"),
                    "prezzo_ddt": request.form.get(f"row-{i}-prezzo_ddt"),
                    "sconto": request.form.get(f"row-{i}-sconto"),
                }
                rows_data.append(row_struct)

            if not rows_data:
                flash("Nessuna riga da salvare.", "warning")
            else:
                try:
                    result = save_delivery_document(
                        store_code=store_code,
                        header=header,
                        cols=cols,
                        rows=rows_data,
                        unit_column=unit_column,
                        code_column=code_column,
                    )
                    if not result.get("success"):
                        flash(result.get("error") or "Errore durante il salvataggio del DDT.", "danger")
                    else:
                        inserted = result.get("inserted", 0)
                        skipped = result.get("skipped", 0)
                        prefill = session.get("supplier_order_ddt_prefill") if isinstance(session.get("supplier_order_ddt_prefill"), dict) else None
                        if prefill and str(prefill.get("order_row_uuid") or "").strip():
                            try:
                                link_order_to_ddt(
                                    str(prefill.get("order_row_uuid") or "").strip(),
                                    supplier_code=header["supplier_code"],
                                    data_doc=header["data_doc"],
                                    data_rif=header["data_rif"],
                                    numero_ddt=header.get("numero_ddt") or "",
                                )
                            except Exception:
                                current_app.logger.exception("Errore collegamento ordine->DDT")
                            try:
                                session["supplier_order_ddt_prefill"] = None
                            except Exception:
                                pass
                        flash(
                            f"DDT salvato correttamente. Righe inserite: {inserted}, righe ignorate: {skipped}.",
                            "success",
                        )
                        # Dopo il salvataggio ricarichiamo il listino per lo stesso fornitore
                        if header["supplier_code"]:
                            price_list = get_price_list_for_supplier(store_code, header["supplier_code"])
                except Exception as ex:
                    flash(f"Errore imprevisto durante il salvataggio del DDT: {ex}", "danger")

    return render_template(
        "warehouse_delivery_new.html",
        store_code=store_code,
        store_name=store_name,
        header=header,
        suppliers=suppliers,
        suppliers_error=suppliers_error,
        price_list=price_list,
    )


@warehouse_bp.route("/spesa", methods=["GET", "POST"])
def spesa_new():
    """Inserimento DDT manuale (Spesa) senza listino: anagrafica prodotti compilata a mano."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("warehouse.home"))

    # Header
    header = {
        "data_doc": "",
        "data_rif": "",
        "supplier_code": "",
    }

    # Layout tabella DDT (per mappare correttamente le colonne)
    layout = None
    try:
        conn = get_connection(store_code)
        try:
            layout = _detect_delivery_layout(conn, get_delivery_table_name())
        finally:
            try:
                conn.close()
            except Exception:
                pass
    except Exception:
        layout = None

    anag_cols = (layout or {}).get("anag_cols") or []
    col_desc = (layout or {}).get("desc_col") or _find_col(anag_cols, ["Descrizione", "Descr"]) or "Descrizione"
    col_code = (layout or {}).get("code_col") or _find_col(anag_cols, ["Codice", "Code", "SKU"]) or "Codice"
    col_group = _find_col(anag_cols, ["Gruppo", "Categoria"]) or "Gruppo"
    col_price = _find_col(anag_cols, ["Prezzo", "Price"]) or "Prezzo"
    col_qtacar = _find_col(
        anag_cols,
        [
            "QtaCar",
            "Qta Car",
            "Qta_Car",
            "Pezzi_per_collo",
            "Pezzi per collo",
            "Qta per collo",
            "QtaPerCollo",
        ],
    ) or "QtaCar"
    col_unita = _find_col(anag_cols, ["Unita", "Unità", "UM", "U.M."]) or "Unita"
    try:
        ensure_supplier_orders_schema()
        group_options = [str(r.get("gruppo") or "").strip() for r in list_listino_groups(False) if str(r.get("gruppo") or "").strip()]
    except Exception:
        group_options = ["FOOD", "PAPER", "OPERATING"]

    # Righe UI (default)
    rows_ui = []
    default_rows = 12
    for _ in range(default_rows):
        rows_ui.append({
            "code": "",
            "desc": "",
            "group": "",
            "price": "",
            "qtacar": "",
            "unita": "",
            "colli": "",
            "pezzi": "",
            "prezzo_ddt": "",
            "sconto": "",
        })

    if request.method == "POST":
        header["data_doc"] = (request.form.get("data_doc") or "").strip()
        header["data_rif"] = (request.form.get("data_rif") or "").strip()
        header["supplier_code"] = (request.form.get("supplier_code") or "").strip()

        # Ricostruisci righe UI per ripopolare la pagina in caso di errori
        rows_ui = []
        rows_count = int(request.form.get("rows_count") or 0)
        for i in range(rows_count):
            rows_ui.append({
                "code": request.form.get(f"row-{i}-code") or "",
                "desc": request.form.get(f"row-{i}-desc") or "",
                "group": request.form.get(f"row-{i}-group") or "",
                "price": request.form.get(f"row-{i}-price") or "",
                "qtacar": request.form.get(f"row-{i}-qtacar") or "",
                "unita": request.form.get(f"row-{i}-unita") or "",
                "colli": request.form.get(f"row-{i}-colli") or "",
                "pezzi": request.form.get(f"row-{i}-pezzi") or "",
                "prezzo_ddt": request.form.get(f"row-{i}-prezzo_ddt") or "",
                "sconto": request.form.get(f"row-{i}-sconto") or "",
            })

        submit_action = (request.form.get("submit_action") or "save").strip().lower()
        if submit_action == "save":
            if not header["data_doc"] or not header["data_rif"]:
                flash("Compila 'Data documento' e 'Data consegna' prima di salvare.", "danger")
            elif not header["supplier_code"]:
                flash("Compila il campo 'Fornitore' prima di salvare.", "danger")
            else:
                rows_data = []
                invalid_rows = []
                required_fields = ["code", "desc", "group", "price", "qtacar", "unita", "colli", "pezzi"]

                for row_idx, r in enumerate(rows_ui, start=1):
                    vals = {k: (r.get(k) or "").strip() for k in required_fields}

                    # riga totalmente vuota: la ignoriamo
                    if all(not v for v in vals.values()):
                        continue

                    # se l'utente ha iniziato a compilare la riga, allora TUTTI i campi devono esserci
                    missing = [k for k, v in vals.items() if not v]
                    if missing:
                        invalid_rows.append(row_idx)
                        continue

                    anag = {
                        col_desc: vals["desc"],
                        col_code: vals["code"],
                        col_group: vals["group"],
                        col_price: vals["price"],
                        col_qtacar: vals["qtacar"],
                        col_unita: vals["unita"],
                    }

                    # marca facoltativa (se presente la colonna)
                    tipolistino_col = _find_col(anag_cols, ["TipoListino", "Tipo Listino"])
                    if tipolistino_col:
                        anag[tipolistino_col] = "SPESA"

                    rows_data.append({
                        "anag": anag,
                        "colli": vals["colli"],
                        "pezzi": vals["pezzi"],
                        "prezzo_ddt": r.get("prezzo_ddt") or "",
                        "sconto": r.get("sconto") or "",
                    })
                if invalid_rows:
                    flash(
                        "Compila tutti i campi per le righe: " + ", ".join(str(x) for x in invalid_rows) + ".",
                        "danger",
                    )
                elif not rows_data:
                    flash("Nessuna riga valida da salvare.", "warning")
                else:
                    try:
                        result = save_delivery_document(
                            store_code=store_code,
                            header=header,
                            cols=[col_desc, col_group, col_price, col_qtacar, col_unita, col_code],
                            rows=rows_data,
                            unit_column=col_qtacar,
                            code_column=col_code,
                        )
                        if not result.get("success"):
                            flash(result.get("error") or "Errore durante il salvataggio della spesa.", "danger")
                        else:
                            flash(
                                f"Spesa salvata correttamente. Righe inserite: {result.get('inserted', 0)}, righe ignorate: {result.get('skipped', 0)}.",
                                "success",
                            )
                            # reset griglia mantenendo header
                            rows_ui = []
                            for _ in range(default_rows):
                                rows_ui.append({
                                    "code": "",
                                    "desc": "",
                                    "group": "",
                                    "price": "",
                                    "qtacar": "",
                                    "unita": "",
                                    "colli": "",
                                    "pezzi": "",
                                    "prezzo_ddt": "",
                                    "sconto": "",
                                })
                    except Exception as ex:
                        current_app.logger.exception("Errore salvataggio spesa")
                        flash(f"Errore imprevisto durante il salvataggio della spesa: {ex}", "danger")

    return render_template(
        "warehouse_spesa_new.html",
        store_code=store_code,
        store_name=store_name,
        header=header,
        rows_ui=rows_ui,
        col_desc=col_desc,
        col_code=col_code,
        col_group=col_group,
        col_price=col_price,
        col_qtacar=col_qtacar,
        col_unita=col_unita,
        group_options=group_options,
    )

@warehouse_bp.route("/delivery/modifica", methods=["GET", "POST"])
def delivery_edit():
    """Maschera per verificare/modificare un DDT già scritto in DatiDelivery."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("warehouse.home"))

    header = {
        "data_doc": "",
        "data_rif": "",
        "supplier_code": "",
    }

    suppliers_info = get_suppliers_for_store(store_code)
    suppliers = suppliers_info.get("suppliers") if suppliers_info and not suppliers_info.get("error") else []
    suppliers_error = suppliers_info.get("error") if suppliers_info else None

    doc = None

    # Supporto apertura da "Ricerca" via querystring (autoload su GET)
    if request.method == "GET":
        qs_supplier = (request.args.get("supplier_code") or "").strip()
        qs_data_doc = (request.args.get("data_doc") or "").strip()
        qs_data_rif = (request.args.get("data_rif") or "").strip()
        if qs_supplier and qs_data_doc and qs_data_rif:
            header["supplier_code"] = qs_supplier
            header["data_doc"] = qs_data_doc
            header["data_rif"] = qs_data_rif
            try:
                doc = get_delivery_document_rows(
                    store_code=store_code,
                    supplier_name=header["supplier_code"],
                    data_consegna=header["data_rif"],
                    data_documento=header["data_doc"],
                )
                if doc and doc.get("error"):
                    flash(doc.get("error"), "danger")
                    doc = None
            except Exception as ex:
                current_app.logger.exception("Errore caricamento DDT (GET)")
                flash(f"Errore caricamento DDT: {ex}", "danger")
                doc = None

    if request.method == "POST":
        header["data_doc"] = request.form.get("data_doc") or ""
        header["data_rif"] = request.form.get("data_rif") or ""
        header["supplier_code"] = request.form.get("supplier_code") or ""
        submit_action = request.form.get("submit_action") or "load"

        # campi obbligatori per qualunque azione
        if not header["supplier_code"] or not header["data_doc"] or not header["data_rif"]:
            flash("Seleziona fornitore e compila le due date.", "danger")
            return render_template(
                "warehouse_delivery_edit.html",
                store_code=store_code,
                store_name=store_name,
                header=header,
                suppliers=suppliers,
                suppliers_error=suppliers_error,
                doc=doc,
            )

        if submit_action == "load":
            doc = get_delivery_document_rows(
                store_code=store_code,
                supplier_name=header["supplier_code"],
                data_consegna=header["data_rif"],
                data_documento=header["data_doc"],
            )
            if doc.get("error"):
                flash(doc["error"], "danger")
            elif not doc.get("rows"):
                flash("Nessuna riga trovata per questo DDT.", "warning")

        elif submit_action == "save":
            # Ricostruisci doc corrente (serve per i metadati/colonne)
            doc = get_delivery_document_rows(
                store_code=store_code,
                supplier_name=header["supplier_code"],
                data_consegna=header["data_rif"],
                data_documento=header["data_doc"],
            )

            rows_count = int(request.form.get("rows_count") or 0)
            cols_count = int(request.form.get("cols_count") or 0)
            unit_column = request.form.get("unit_column") or None
            code_column = request.form.get("code_column") or None

            cols = []
            for j in range(cols_count):
                col_name = request.form.get(f"col-{j}-name")
                if col_name is not None:
                    cols.append(col_name)

            rows_data = []
            for i in range(rows_count):
                # controlla se la riga è stata cancellata dal DOM
                if request.form.get(f"row-{i}-deleted") == "1":
                    continue

                # valori originali (per capire se è cambiata)
                orig_colli = (request.form.get(f"row-{i}-orig-colli") or "").strip()
                orig_pezzi = (request.form.get(f"row-{i}-orig-pezzi") or "").strip()
                orig_prezzo = (request.form.get(f"row-{i}-orig-prezzo_ddt") or "").strip()

                new_colli = (request.form.get(f"row-{i}-colli") or "").strip()
                new_pezzi = (request.form.get(f"row-{i}-pezzi") or "").strip()
                new_prezzo = (request.form.get(f"row-{i}-prezzo_ddt") or "").strip()
                new_sconto = (request.form.get(f"row-{i}-sconto") or "").strip()

                changed = (new_colli != orig_colli) or (new_pezzi != orig_pezzi) or (new_prezzo != orig_prezzo) or (new_sconto and new_sconto != "0")

                if not changed:
                    continue

                anag = {}
                for j, col_name in enumerate(cols):
                    cell_key = f"row-{i}-col-{j}"
                    cell_val = request.form.get(cell_key)
                    anag[col_name] = cell_val

                row_struct = {
                    "anag": anag,
                    "colli": new_colli,
                    "pezzi": new_pezzi,
                    "prezzo_ddt": new_prezzo,
                    "sconto": new_sconto,
                }
                rows_data.append(row_struct)

            if not rows_data:
                flash("Nessuna modifica da salvare.", "warning")
            else:
                try:
                    result = save_delivery_document(
                        store_code=store_code,
                        header={
                            "data_doc": header["data_doc"],
                            "data_rif": header["data_rif"],
                            "supplier_code": header["supplier_code"],
                        },
                        cols=cols,
                        rows=rows_data,
                        unit_column=unit_column,
                        code_column=code_column,
                    )
                    if not result.get("success"):
                        flash(result.get("error") or "Errore durante il salvataggio delle modifiche.", "danger")
                    else:
                        flash(f"Modifiche salvate. Righe aggiornate: {result.get('inserted', 0)}.", "success")
                except Exception as ex:
                    flash(f"Errore imprevisto durante il salvataggio: {ex}", "danger")

            # ricarica doc aggiornato
            doc = get_delivery_document_rows(
                store_code=store_code,
                supplier_name=header["supplier_code"],
                data_consegna=header["data_rif"],
                data_documento=header["data_doc"],
            )

        elif submit_action == "change_dates":
            new_data_doc = request.form.get("new_data_doc") or ""
            new_data_rif = request.form.get("new_data_rif") or ""
            if not new_data_doc or not new_data_rif:
                flash("Compila le nuove date prima di applicare il cambio date.", "danger")
            else:
                res = update_delivery_ddt_dates(
                    store_code=store_code,
                    supplier_name=header["supplier_code"],
                    old_data_consegna=header["data_rif"],
                    old_data_documento=header["data_doc"],
                    new_data_consegna=new_data_rif,
                    new_data_documento=new_data_doc,
                )
                if not res.get("success"):
                    flash(res.get("error") or "Errore durante l'aggiornamento delle date.", "danger")
                else:
                    flash(f"Date aggiornate. Righe interessate: {res.get('updated', 0)}.", "success")
                    header["data_doc"] = new_data_doc
                    header["data_rif"] = new_data_rif

            doc = get_delivery_document_rows(
                store_code=store_code,
                supplier_name=header["supplier_code"],
                data_consegna=header["data_rif"],
                data_documento=header["data_doc"],
            )

    return render_template(
        "warehouse_delivery_edit.html",
        store_code=store_code,
        store_name=store_name,
        header=header,
        suppliers=suppliers,
        suppliers_error=suppliers_error,
        doc=doc,
    )


@warehouse_bp.post("/delivery/modifica/delete-row")
def delivery_delete_row():
    """Cancella in tempo reale una riga del DDT."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    if not store_code:
        return jsonify({"success": False, "error": "Store non selezionato."}), 400

    data = request.get_json(silent=True) or request.form or {}
    supplier = (data.get("supplier_code") or "").strip()
    data_doc = (data.get("data_doc") or "").strip()
    data_rif = (data.get("data_rif") or "").strip()
    descr = (data.get("descrizione") or "").strip()

    if not (supplier and data_doc and data_rif and descr):
        return jsonify({"success": False, "error": "Parametri mancanti."}), 400

    res = delete_delivery_row(
        store_code=store_code,
        supplier_name=supplier,
        data_consegna=data_rif,
        data_documento=data_doc,
        descrizione=descr,
    )
    status = 200 if res.get("success") else 500
    return jsonify(res), status
@warehouse_bp.route("/inventario/nuovo", methods=["GET", "POST"])
def inventory_new():
    """Maschera di inserimento Inventario / TX (solo caricamento listino + calcoli, scrittura in step successivo)."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("warehouse.home"))

    suppliers_info = get_suppliers_for_store(store_code)
    suppliers = suppliers_info.get("suppliers") if suppliers_info and not suppliers_info.get("error") else []
    suppliers_error = suppliers_info.get("error") if suppliers_info else None

    header = {
        "supplier_name": "",       # compat/backward
        "supplier_names": [],      # multi
        "totals_supplier": "",     # for dropdown totals
        "mov_type": "INV",
        "data_mov": datetime.now().strftime("%Y-%m-%d"),
    }

    price_list = None
    conv_map = {}
    avg_price_map = {}
    loaded_suppliers = []

    if request.method == "POST":
        # Multi-select fornitori
        supplier_names = request.form.getlist("supplier_name") or []
        supplier_names = [s.strip() for s in supplier_names if s and s.strip()]

        header["supplier_names"] = supplier_names
        header["supplier_name"] = supplier_names[0] if supplier_names else ""

        header["mov_type"] = (request.form.get("mov_type") or "INV").strip()
        header["data_mov"] = (request.form.get("data_mov") or "").strip()

        # vincoli minimi per far funzionare la maschera
        if not header["supplier_names"]:
            flash("Seleziona almeno un fornitore.", "warning")
        if not header["mov_type"]:
            flash("Seleziona un tipo movimentazione.", "warning")
        if not header["data_mov"]:
            flash("Inserisci la data movimentazione.", "warning")

        if header["supplier_names"] and header["mov_type"] and header["data_mov"]:
            merged_cols = []
            merged_rows = []

            for sup in header["supplier_names"]:
                pl = get_price_list_for_supplier(store_code, sup)
                if pl and pl.get("error"):
                    flash(f"{sup}: {pl['error']}", "danger")
                    continue

                pl_cols = (pl.get("columns") or []) if isinstance(pl, dict) else []
                pl_rows = (pl.get("rows") or []) if isinstance(pl, dict) else []

                # merge columns keeping stable order
                for c in pl_cols:
                    if c not in merged_cols:
                        merged_cols.append(c)

                if pl_rows:
                    loaded_suppliers.append(sup)

                for r in pl_rows:
                    if not isinstance(r, dict):
                        continue
                    rr = dict(r)
                    # ensure supplier available client-side (never shown as a column)
                    rr["__supplier__"] = sup
                    merged_rows.append(rr)

                # Conversioni KG->PEZ da listino e prezzi medi 4 settimane (da DatiDelivery) per prezzi=0
                try:
                    cm = get_conversions_for_supplier(store_code, sup)
                    for dk, v in (cm or {}).items():
                        key = f"{sup.strip().lower()}||{dk}"
                        conv_map[key] = v
                except Exception as e:
                    current_app.logger.exception("Errore lettura conversioni listino (multi)")
                    flash(f"{sup}: errore lettura conversioni listino: {e}", "danger")

                try:
                    am = get_delivery_avg_prices_last_weeks(store_code, sup, weeks=4)
                    for dk, v in (am or {}).items():
                        key = f"{sup.strip().lower()}||{dk}"
                        avg_price_map[key] = v
                except Exception:
                    current_app.logger.exception("Errore calcolo prezzo medio DDT (multi)")

            if merged_rows:
                # Rendi serializzabile per JS (evita problemi con Decimal/date)
                merged_rows = _json_safe_rows(merged_rows)
                price_list = {"columns": merged_cols, "rows": merged_rows}
                header["totals_supplier"] = (loaded_suppliers[0] if loaded_suppliers else header["supplier_name"]) or ""
            else:
                price_list = None

    return render_template(
        "warehouse_inventory_new.html",
        store_code=store_code,
        store_name=store_name,
        suppliers=suppliers,
        suppliers_error=suppliers_error,
        header=header,
        price_list=price_list,
        conv_map=conv_map,
        avg_price_map=avg_price_map,
        loaded_suppliers=loaded_suppliers,
    )


@warehouse_bp.post("/inventario/salva")
def inventory_save():
    """Salva movimenti inventario (datiinventario) e, per TXIN/TXOUT, anche DatiTX."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    if not store_code:
        return jsonify({"success": False, "error": "Store non selezionato."}), 400

    payload = request.get_json(silent=True) or {}
    header = payload.get("header") or {}
    rows = payload.get("rows") or []

    supplier_name = (header.get("supplier_name") or "").strip()
    supplier_names = header.get("supplier_names") or []
    if isinstance(supplier_names, str):
        supplier_names = [supplier_names]
    try:
        supplier_names = [str(s).strip() for s in supplier_names if str(s).strip()]
    except Exception:
        supplier_names = []

    mov_type = (header.get("mov_type") or "").strip()
    data_mov = (header.get("data_mov") or "").strip()
    site2 = (header.get("site2") or "").strip()

    if not supplier_name and not supplier_names:
        return jsonify({"success": False, "error": "Fornitore mancante."}), 400
    if not mov_type:
        return jsonify({"success": False, "error": "Tipo movimentazione mancante."}), 400
    if not data_mov:
        return jsonify({"success": False, "error": "Data movimentazione mancante."}), 400

    is_tx = mov_type.upper() in ("TXIN", "TXOUT")
    if is_tx and not site2:
        return jsonify({"success": False, "error": "Per TXIN/TXOUT è obbligatorio selezionare lo store (SITE2)."}), 400

    try:
        res = save_inventory_movement(
            store_code=str(store_code),
            data_mov=data_mov,
            mov_type=mov_type,
            rows=rows if isinstance(rows, list) else [],
            site2=site2 if site2 else None,
        )
        status = 200 if res.get("success") else 500
        return jsonify(res), status
    except Exception as e:
        current_app.logger.exception("Errore salvataggio inventario/TX")
        return jsonify({"success": False, "error": str(e)}), 500


@warehouse_bp.post("/inventario/pdf")
def inventory_pdf():
    """Genera PDF della tabella Inventario con soli campi editabili (CAR/INT/PEZ/KG)."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")
    if not store_code:
        return jsonify({"success": False, "error": "Store non selezionato."}), 400

    payload = request.get_json(silent=True) or {}
    header = payload.get("header") or {}
    rows = payload.get("rows") or []

    try:
        from inventory_pdf import build_inventory_pdf
    except ModuleNotFoundError:
        return render_template("inventory_pdf_missing.html")
    except ImportError:
        return render_template("inventory_pdf_missing.html")

    try:
        pdf_bytes = build_inventory_pdf(
            site=str(store_code),
            site_name=str(store_name or "").strip(),
            header=header if isinstance(header, dict) else {},
            rows=rows if isinstance(rows, list) else [],
        )
        data_mov = str((header or {}).get("data_mov") or "").strip()
        mov_type = str((header or {}).get("mov_type") or "").strip()
        safe_dt = data_mov.replace(":", "-") if data_mov else datetime.now().strftime("%Y-%m-%d")
        safe_mt = mov_type.replace(" ", "_") if mov_type else "INV"
        filename = f"Inventario_{store_code}_{safe_mt}_{safe_dt}.pdf"
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        current_app.logger.exception("Errore generazione PDF inventario")
        return jsonify({"success": False, "error": str(e)}), 500



@warehouse_bp.route("/inventario/modifica", methods=["GET", "POST"])
def inventory_edit():
    """Maschera per verificare/modificare una movimentazione Inventario/TX già scritta.

    Logica allineata a Modifica DDT:
      - Caricamento (LOAD) con i campi header.
      - Cancellazione riga realtime (endpoint dedicato).
      - Salvataggio massivo SOLO delle righe modificate (DELETE+INSERT per riga).
    """
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("warehouse.home"))

    suppliers_info = get_suppliers_for_store(store_code)
    suppliers = suppliers_info.get("suppliers") if suppliers_info and not suppliers_info.get("error") else []
    suppliers_error = suppliers_info.get("error") if suppliers_info else None

    header = {
        "supplier_name": "",
        "mov_type": "INV",
        "data_mov": datetime.now().strftime("%Y-%m-%d"),
        "site2": "",
    }

    doc = None
    conv_map = {}
    avg_price_map = {}

    def _parse_number_local(v):
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if not s:
            return 0.0
        s = s.replace("€", "").replace(" ", "")
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0.0

    # Supporto apertura da "Ricerca" via querystring (autoload su GET)
    if request.method == "GET":
        qs_supplier = (request.args.get("supplier_name") or "").strip()
        qs_mov_type = (request.args.get("mov_type") or "").strip() or "INV"
        qs_data_mov = (request.args.get("data_mov") or "").strip()
        qs_site2 = (request.args.get("site2") or "").strip()

        if qs_supplier and qs_mov_type and qs_data_mov:
            header["supplier_name"] = qs_supplier
            header["mov_type"] = qs_mov_type
            header["data_mov"] = qs_data_mov
            header["site2"] = qs_site2

            # calcoli (prezzi avg / conversioni) - servono in pagina per calcoli client
            try:
                conv_map = get_conversions_for_supplier(store_code, header["supplier_name"])
            except Exception as e:
                current_app.logger.exception("Errore lettura conversioni listino")
                flash(f"Errore lettura conversioni listino: {e}", "danger")
                conv_map = {}

            try:
                avg_price_map = get_delivery_avg_prices_last_weeks(store_code, header["supplier_name"], weeks=4)
            except Exception:
                current_app.logger.exception("Errore calcolo prezzi medi 4 settimane")
                avg_price_map = {}

            try:
                doc = get_inventory_document_rows(
                    store_code=store_code,
                    supplier_name=header["supplier_name"],
                    data_mov=header["data_mov"],
                    mov_type=header["mov_type"],
                    site2=header.get("site2") or None,
                )
                if doc and doc.get("error"):
                    flash(doc.get("error"), "danger")
                    doc = None
            except Exception as ex:
                current_app.logger.exception("Errore caricamento movimento (GET)")
                flash(f"Errore caricamento movimento: {ex}", "danger")
                doc = None

    if request.method == "POST":
        header["supplier_name"] = (request.form.get("supplier_name") or "").strip()
        header["mov_type"] = (request.form.get("mov_type") or "INV").strip()
        header["data_mov"] = (request.form.get("data_mov") or "").strip()
        header["site2"] = (request.form.get("site2") or "").strip()
        submit_action = (request.form.get("submit_action") or "load").strip()

        # campi obbligatori per qualunque azione
        if not header["supplier_name"] or not header["mov_type"] or not header["data_mov"]:
            flash("Compila data, tipo movimentazione e fornitore.", "danger")
        else:
            is_tx = header["mov_type"].upper() in ("TXIN", "TXOUT")
            if is_tx and not header["site2"]:
                flash("Per TXIN/TXOUT devi selezionare anche SITE2.", "danger")
            else:
                # calcoli (prezzi avg / conversioni) - servono in pagina per calcoli client
                try:
                    conv_map = get_conversions_for_supplier(store_code, header["supplier_name"])
                except Exception as e:
                    current_app.logger.exception("Errore lettura conversioni listino")
                    flash(f"Errore lettura conversioni listino: {e}", "danger")
                    conv_map = {}

                try:
                    avg_price_map = get_delivery_avg_prices_last_weeks(store_code, header["supplier_name"], weeks=4)
                except Exception:
                    current_app.logger.exception("Errore calcolo prezzi medi 4 settimane")
                    avg_price_map = {}

                if submit_action == "save":
                    rows_count = int(request.form.get("rows_count") or 0)
                    cols_count = int(request.form.get("cols_count") or 0)

                    cols = []
                    for j in range(cols_count):
                        cn = request.form.get(f"col-{j}-name")
                        if cn is not None:
                            cols.append(cn)

                    # detect price column name (if present)
                    price_col = None
                    for c in cols:
                        cl = (c or "").strip().lower()
                        if cl == "prezzo" or "prezzo" in cl:
                            price_col = c
                            break

                    rows_data = []
                    for i in range(rows_count):
                        # se riga rimossa dal DOM non avremo i campi: in quel caso la ignoriamo
                        new_car = request.form.get(f"row-{i}-car")
                        new_int = request.form.get(f"row-{i}-interno")
                        new_pez = request.form.get(f"row-{i}-pez")
                        new_kg = request.form.get(f"row-{i}-kg")
                        new_totpz = request.form.get(f"row-{i}-totpz")
                        new_toteuro = request.form.get(f"row-{i}-toteuro")

                        if new_car is None and new_int is None and new_pez is None and new_kg is None:
                            continue

                        orig_car = request.form.get(f"row-{i}-orig-car") or ""
                        orig_int = request.form.get(f"row-{i}-orig-interno") or ""
                        orig_pez = request.form.get(f"row-{i}-orig-pez") or ""
                        orig_totpz = request.form.get(f"row-{i}-orig-totpz") or ""
                        orig_toteuro = request.form.get(f"row-{i}-orig-toteuro") or ""

                        changed = (
                            abs(_parse_number_local(new_car) - _parse_number_local(orig_car)) > 1e-9 or
                            abs(_parse_number_local(new_int) - _parse_number_local(orig_int)) > 1e-9 or
                            abs(_parse_number_local(new_pez) - _parse_number_local(orig_pez)) > 1e-9 or
                            abs(_parse_number_local(new_totpz) - _parse_number_local(orig_totpz)) > 1e-9 or
                            abs(_parse_number_local(new_toteuro) - _parse_number_local(orig_toteuro)) > 1e-9
                        )

                        if not changed:
                            continue

                        anag = {}
                        for j, col_name in enumerate(cols):
                            cell_key = f"row-{i}-col-{j}"
                            anag[col_name] = request.form.get(cell_key)

                        # forza prezzo effettivo (se disponibile)
                        eff_price = request.form.get(f"row-{i}-effective-price")
                        if price_col and eff_price is not None:
                            anag[price_col] = _parse_number_local(eff_price)

                        row_struct = {
                            "anag": anag,
                            "car": new_car,
                            "interno": new_int,
                            "pez": new_pez,
                            "kg": new_kg,
                            "totpz": new_totpz,
                            "toteuro": new_toteuro,
                        }
                        rows_data.append(row_struct)

                    if not rows_data:
                        flash("Nessuna modifica da salvare.", "warning")
                    else:
                        try:
                            res = save_inventory_document(
                                store_code=str(store_code),
                                header=header,
                                cols=cols,
                                rows=rows_data,
                                site2=header["site2"] if is_tx else None,
                            )
                            if not res.get("success"):
                                flash(res.get("error") or "Errore durante il salvataggio delle modifiche.", "danger")
                            else:
                                msg = f"Modifiche salvate. Righe aggiornate (INV): {res.get('inserted_inventory', 0)}"
                                if is_tx:
                                    msg += f" | Righe aggiornate (TX): {res.get('inserted_tx', 0)}"
                                if res.get("skipped"):
                                    msg += f" | Ignorate: {res.get('skipped')}"
                                flash(msg, "success")
                        except Exception as ex:
                            flash(f"Errore imprevisto durante il salvataggio: {ex}", "danger")


                elif submit_action == "change_header":
                    # Cambio intestazione (data e, per TX, SITE2 e/o tipo)
                    orig_data_mov = (request.form.get("orig_data_mov") or header["data_mov"] or "").strip()
                    orig_mov_type = (request.form.get("orig_mov_type") or header["mov_type"] or "").strip()
                    orig_site2 = (request.form.get("orig_site2") or header.get("site2") or "").strip()

                    new_data_mov = (request.form.get("new_data_mov") or "").strip()
                    new_mov_type = (request.form.get("new_mov_type") or orig_mov_type or "").strip()
                    new_site2 = (request.form.get("new_site2") or orig_site2 or "").strip()

                    if not new_data_mov:
                        flash("Seleziona la nuova data prima di applicare il cambio intestazione.", "danger")
                    else:
                        is_tx_orig = (orig_mov_type or "").upper() in ("TXIN", "TXOUT")
                        if is_tx_orig and not orig_site2:
                            flash("SITE2 mancante per il trasferimento (chiave originale).", "danger")
                        elif is_tx_orig and not new_site2:
                            flash("Seleziona il nuovo SITE2 prima di applicare.", "danger")
                        else:
                            res_upd = update_inventory_movement_header(
                                store_code=str(store_code),
                                supplier_name=header["supplier_name"],
                                mov_type=orig_mov_type,
                                old_data_mov=orig_data_mov,
                                new_data_mov=new_data_mov,
                                old_site2=orig_site2 if is_tx_orig else None,
                                new_site2=new_site2 if is_tx_orig else None,
                                new_mov_type=new_mov_type or None,
                            )
                            if not res_upd.get("success"):
                                flash(res_upd.get("error") or "Errore durante il cambio intestazione.", "danger")
                            else:
                                # aggiorna header per ricaricare il documento con i nuovi valori
                                header["data_mov"] = new_data_mov
                                header["mov_type"] = (new_mov_type or orig_mov_type or header["mov_type"]).strip()
                                if is_tx_orig:
                                    header["site2"] = new_site2
                                flash(
                                    f"Intestazione aggiornata. Righe interessate: INV {res_upd.get('updated_inv', 0)}"
                                    + (f" | TX {res_upd.get('updated_tx', 0)}" if is_tx_orig else ""),
                                    "success",
                                )

                # carica/ricarica righe (dopo load o dopo save)
                res = get_inventory_document_rows(
                    store_code=str(store_code),
                    supplier_name=header["supplier_name"],
                    data_mov=header["data_mov"],
                    mov_type=header["mov_type"],
                    site2=header["site2"] if is_tx else None,
                )
                if not res.get("success"):
                    flash(res.get("error") or "Errore caricamento movimentazione.", "danger")
                    doc = None
                else:
                    rows = res.get("rows") or []
                    if not rows:
                        flash("Nessuna riga trovata per questa combinazione.", "warning")
                    doc = {
                        "cols": res.get("cols") or [],
                        "rows": _json_safe_rows_deep(rows),
                    }

    # Elenco store completo per selezione SITE2 nei trasferimenti (TXIN/TXOUT)
    # (evita dipendenze da fetch/redirect che possono lasciare la tendina vuota)
    stores_all = []
    try:
        stores_all = get_warehouse_stores() or []
    except Exception:
        stores_all = []

    return render_template(
        "warehouse_inventory_edit.html",
        store_code=store_code,
        store_name=store_name,
        header=header,
        suppliers=suppliers,
        suppliers_error=suppliers_error,
        doc=doc,
        conv_map=conv_map,
        avg_price_map=avg_price_map,
        stores_all=stores_all,
    )


@warehouse_bp.post("/inventario/modifica/delete-row")
def inventory_delete_row():
    """Cancella in tempo reale una riga del movimento Inventario/TX."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    if not store_code:
        return jsonify({"success": False, "error": "Store non selezionato."}), 400

    data = request.get_json(silent=True) or request.form or {}

    supplier_name = (data.get("supplier_name") or "").strip()
    mov_type = (data.get("mov_type") or "").strip()
    data_mov = (data.get("data_mov") or "").strip()
    site2 = (data.get("site2") or "").strip()
    descr = (data.get("descrizione") or "").strip()
    code = (data.get("codice") or "").strip()

    if not (supplier_name and mov_type and data_mov):
        return jsonify({"success": False, "error": "Parametri header mancanti."}), 400

    is_tx = mov_type.upper() in ("TXIN", "TXOUT")
    if is_tx and not site2:
        return jsonify({"success": False, "error": "SITE2 mancante per TXIN/TXOUT."}), 400

    if not (descr or code):
        return jsonify({"success": False, "error": "Parametri mancanti: descrizione o codice."}), 400

    res = delete_inventory_row(
        store_code=str(store_code),
        supplier_name=supplier_name,
        data_mov=data_mov,
        mov_type=mov_type,
        descrizione=descr if descr else None,
        codice=code if code else None,
        site2=site2 if is_tx else None,
    )
    status = 200 if res.get("success") else 500
    return jsonify(res), status


@warehouse_bp.post("/inventario/modifica/salva")
def inventory_edit_save():
    """Salva modifiche inventario/TX (delete + reinsert del set)."""
    _ensure_session_keys()
    store_code = session.get("store_code")
    if not store_code:
        return jsonify({"success": False, "error": "Store non selezionato."}), 400

    data = request.get_json(silent=True) or {}
    header = data.get("header") or {}
    rows = data.get("rows") or []

    supplier_name = (header.get("supplier_name") or "").strip()
    mov_type = (header.get("mov_type") or "").strip()
    data_mov = (header.get("data_mov") or "").strip()
    site2 = (header.get("site2") or "").strip()

    if not supplier_name or not mov_type or not data_mov:
        return jsonify({"success": False, "error": "Header incompleto (fornitore, tipo, data)."}), 400

    is_tx = mov_type.upper() in ("TXIN", "TXOUT")
    if is_tx and not site2:
        return jsonify({"success": False, "error": "SITE2 mancante per TXIN/TXOUT."}), 400

    try:
        res = replace_inventory_movement(
            store_code=str(store_code),
            supplier_name=supplier_name,
            data_mov=data_mov,
            mov_type=mov_type,
            rows=rows if isinstance(rows, list) else [],
            site2=site2 if is_tx else None,
        )
        status = 200 if res.get("success") else 500
        return jsonify(res), status
    except Exception as e:
        current_app.logger.exception("Errore salvataggio modifica inventario/TX")
        return jsonify({"success": False, "error": str(e)}), 500







@warehouse_bp.route("/ordini")
def ordini():
    store_code = session.get("store_code")
    store_name = session.get("store_name") or ""
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("warehouse.home"))

    # Normalizza fornitori:
    # get_suppliers_for_store può restituire:
    # - list[str]
    # - list[dict]/list[tuple]
    # - dict con chiave "suppliers" (tipico: {"suppliers":[{"code":..,"name":..},...] , "available_columns":[...], ...})
    try:
        ensure_supplier_orders_schema()
        raw = {"suppliers": [{"name": r.get("Fornitore"), "code": r.get("Fornitore")} for r in list_sql_fornitori()]}
    except Exception:
        try:
            raw = get_suppliers_for_store(store_code)
        except Exception:
            raw = []

    # Se è un dict "descrittivo", prendi la lista dentro "suppliers" (o "rows")
    if isinstance(raw, dict):
        if isinstance(raw.get("suppliers"), list):
            raw_list = raw.get("suppliers") or []
        elif isinstance(raw.get("rows"), list):
            raw_list = raw.get("rows") or []
        else:
            raw_list = []
    else:
        raw_list = raw or []

    suppliers = []
    for x in (raw_list or []):
        if x is None:
            continue

        # dict: preferisci nome fornitore
        if isinstance(x, dict):
            picked = None
            for k in (
                "name", "Nome", "fornitore", "Fornitore", "Supplier", "supplier",
                "descrizione", "Descrizione",
            ):
                if k in x and x.get(k):
                    picked = x.get(k)
                    break
            # fallback: se c'è 'code' ma non 'name', usa code
            if picked is None:
                for k in ("code", "Codice", "codice"):
                    if k in x and x.get(k):
                        picked = x.get(k)
                        break
            if picked:
                suppliers.append(str(picked).strip())
            continue

        # tuple/list -> primo elemento
        if isinstance(x, (list, tuple)) and len(x) > 0:
            suppliers.append(str(x[0]).strip())
            continue

        suppliers.append(str(x).strip())

    # pulizia finale: rimuovi elementi tecnici tipo 'available_columns', ecc.
    def _is_bad(s: str) -> bool:
        t = (s or "").strip().lower()
        if not t:
            return True
        if t in ("none", "null"):
            return True
        bad_tokens = (
            "available", "columns", "colonne", "column",
            "code_column", "name_column", "table", "error", "traceback",
        )
        return any(tok in t for tok in bad_tokens)

    suppliers = sorted({s for s in suppliers if not _is_bad(s)})
    try:
        ensure_supplier_orders_schema()
        listino_types = [str(r.get("tipo_listino") or "").strip() for r in list_listino_types(False) if str(r.get("tipo_listino") or "").strip()]
    except Exception:
        listino_types = ["FoodPaper", "Operating"]
    today = date.today()
    next_d = today + timedelta(days=7)

    return render_template(
        "warehouse_ordini.html",
        store_code=store_code,
        store_name=store_name,
        suppliers=suppliers,
        listino_types=listino_types,
        today_iso=today.isoformat(),
        next_iso=next_d.isoformat(),
    )


@warehouse_bp.route("/fornitori", methods=["GET", "POST"])
def fornitori():
    _ensure_session_keys()
    if not _ensure_admin_warehouse_access():
        return redirect(url_for("warehouse.home"))

    ensure_supplier_orders_schema()
    try:
        stores = get_warehouse_stores() or []
    except Exception:
        stores = []
    if request.method == "POST":
        action = (request.form.get("action") or "save").strip().lower()
        try:
            if action == "delete":
                delete_fornitore((request.form.get("row_uuid") or "").strip())
                flash("Fornitore eliminato.", "success")
            elif action == "save_contact":
                supplier_uuid = (request.form.get("supplier_row_uuid") or "").strip()
                save_fornitore_contact(
                    supplier_uuid,
                    (request.form.get("contact_row_uuid") or "").strip() or None,
                    {
                        "Referente": request.form.get("contact_referente"),
                        "Email": request.form.get("contact_email"),
                        "Telefono1": request.form.get("contact_telefono1"),
                        "Telefono2": request.form.get("contact_telefono2"),
                        "TipoOrdine": request.form.get("contact_tipo_ordine"),
                        "sort_order": request.form.get("contact_sort_order"),
                    },
                    request.form.getlist("contact_store_codes"),
                )
                flash("Cluster contatti salvato.", "success")
                return redirect(url_for("warehouse.fornitori", edit=supplier_uuid))
            elif action == "delete_contact":
                supplier_uuid = (request.form.get("supplier_row_uuid") or "").strip()
                delete_fornitore_contact(supplier_uuid, (request.form.get("contact_row_uuid") or "").strip())
                flash("Cluster contatti eliminato.", "success")
                return redirect(url_for("warehouse.fornitori", edit=supplier_uuid))
            else:
                save_fornitore(
                    (request.form.get("row_uuid") or "").strip() or None,
                    {
                        "Fornitore": request.form.get("fornitore"),
                        "Referente": request.form.get("referente"),
                        "Email": request.form.get("email"),
                        "Telefono1": request.form.get("telefono1"),
                        "Telefono2": request.form.get("telefono2"),
                        "TipoOrdine": request.form.get("tipo_ordine"),
                    },
                )
                flash("Fornitore salvato.", "success")
                if not (request.form.get("row_uuid") or "").strip():
                    rows = list_sql_fornitori()
                    supplier_name = str(request.form.get("fornitore") or "").strip()
                    match = next((r for r in rows if str(r.get("Fornitore") or "").strip() == supplier_name), None)
                    if match:
                        return redirect(url_for("warehouse.fornitori", edit=match.get("row_uuid")))
        except Exception as e:
            flash(f"Errore salvataggio fornitore: {e}", "danger")
        return redirect(url_for("warehouse.fornitori"))

    edit_id = (request.args.get("edit") or "").strip()
    edit_contact_id = (request.args.get("edit_contact") or "").strip()
    rows = list_sql_fornitori()
    edit_row = next((r for r in rows if str(r.get("row_uuid") or "") == edit_id), None)
    contact_rows = list_fornitore_contacts(edit_id) if edit_id else []
    edit_contact_row = next((r for r in contact_rows if str(r.get("row_uuid") or "") == edit_contact_id), None)
    return render_template(
        "warehouse_fornitori.html",
        rows=rows,
        edit_row=edit_row,
        stores=stores,
        contact_rows=contact_rows,
        edit_contact_row=edit_contact_row,
    )


@warehouse_bp.route("/anagrafica-listini-prezzi", methods=["GET", "POST"])
def listini_anagrafica():
    _ensure_session_keys()
    if not _ensure_admin_warehouse_access():
        return redirect(url_for("warehouse.home"))

    ensure_supplier_orders_schema()
    try:
        ensure_default_price_list_assignments()
    except Exception as e:
        current_app.logger.exception("Errore inizializzazione assegnazioni listini")
        flash(f"Listini multipli non ancora inizializzati: {e}", "warning")
    if request.method == "POST":
        entity = (request.form.get("entity") or "").strip().lower()
        action = (request.form.get("action") or "save").strip().lower()
        try:
            if entity == "tipo":
                if action == "delete":
                    delete_listino_type((request.form.get("row_uuid") or "").strip())
                    flash("Tipo listino eliminato.", "success")
                else:
                    save_listino_type(
                        (request.form.get("row_uuid") or "").strip() or None,
                        request.form.get("tipo_listino") or "",
                        int(request.form.get("sort_order") or 0),
                        (request.form.get("is_active") or "1") == "1",
                    )
                    flash("Tipo listino salvato.", "success")
            elif entity == "gruppo":
                if action == "delete":
                    delete_listino_group((request.form.get("row_uuid") or "").strip())
                    flash("Gruppo eliminato.", "success")
                else:
                    save_listino_group(
                        (request.form.get("row_uuid") or "").strip() or None,
                        request.form.get("gruppo") or "",
                        int(request.form.get("sort_order") or 0),
                        (request.form.get("is_active") or "1") == "1",
                    )
                    flash("Gruppo salvato.", "success")
            elif entity == "listino":
                row_uuid = (request.form.get("row_uuid") or "").strip() or None
                if action == "delete":
                    delete_price_list(row_uuid or "")
                    flash("Elenco prezzi eliminato.", "success")
                elif action == "set_default":
                    set_default_price_list(row_uuid or "")
                    flash("Listino impostato come default.", "success")
                else:
                    listino_id = save_price_list(
                        row_uuid,
                        request.form.get("nome") or "",
                        (request.form.get("is_active") or "1") == "1",
                    )
                    replace_price_list_store_assignments(listino_id, request.form.getlist("store_codes"))
                    flash("Elenco prezzi salvato e assegnazioni aggiornate.", "success")
        except Exception as e:
            flash(f"Errore salvataggio anagrafica listini: {e}", "danger")
        return redirect(url_for("warehouse.listini_anagrafica"))

    edit_tipo = (request.args.get("edit_tipo") or "").strip()
    edit_gruppo = (request.args.get("edit_gruppo") or "").strip()
    tipi = list_listino_types(include_inactive=True)
    gruppi = list_listino_groups(include_inactive=True)
    try:
        price_lists = list_price_lists(include_inactive=True)
    except Exception as e:
        current_app.logger.exception("Errore caricamento elenchi prezzi")
        flash(f"Errore caricamento elenchi prezzi: {e}", "warning")
        price_lists = [{"row_uuid": "", "nome": "Listino DOS", "is_default": True, "is_active": True, "store_count": 0, "row_count": 0}]
    stores = _list_store_options_for_admin_context(active_only=False)
    if not stores:
        flash("Nessuno store disponibile per il tenant corrente.", "warning")
    try:
        assignments = list_price_list_store_assignments()
    except Exception:
        assignments = {}
    edit_tipo_row = next((r for r in tipi if str(r.get("row_uuid") or "") == edit_tipo), None)
    edit_gruppo_row = next((r for r in gruppi if str(r.get("row_uuid") or "") == edit_gruppo), None)
    return render_template(
        "warehouse_listini_anagrafica.html",
        tipi=tipi,
        gruppi=gruppi,
        price_lists=price_lists,
        stores=stores,
        assignments=assignments,
        edit_tipo_row=edit_tipo_row,
        edit_gruppo_row=edit_gruppo_row,
    )


@warehouse_bp.route("/listini-prezzi")
def listini_prezzi():
    _ensure_session_keys()
    if not _ensure_admin_warehouse_access():
        return redirect(url_for("warehouse.home"))
    ensure_supplier_orders_schema()
    try:
        migrate_legacy_pricelists()
        ensure_default_price_list_assignments()
    except Exception as e:
        current_app.logger.exception("Errore inizializzazione listini prezzi")
        flash(f"Listini multipli non ancora inizializzati: {e}", "warning")
    tipo_options = [str(r.get("tipo_listino") or "").strip() for r in list_listino_types(False) if str(r.get("tipo_listino") or "").strip()]
    try:
        price_lists = list_price_lists(False)
    except Exception:
        price_lists = [{"row_uuid": "", "nome": "Listino DOS"}]
    stores = _list_store_options_for_admin_context(active_only=False)
    return render_template(
        "admin_listini.html",
        listino_types=tipo_options,
        price_lists=price_lists,
        stores=stores,
        page_title="Listini Prezzi",
        supplier_orders_context=True,
    )


@warehouse_bp.route("/ordini-fornitore")
def supplier_orders_home():
    _ensure_session_keys()
    role = str(session.get("role") or "").strip().lower()
    if role == "fornitore":
        return redirect(url_for("warehouse.supplier_orders_received"))
    if not _ensure_supplier_orders_internal_access():
        return redirect(url_for("warehouse.home"))
    return redirect(url_for("warehouse.supplier_orders_send"))


@warehouse_bp.route("/ordini-fornitore/invia", methods=["GET", "POST"])
def supplier_orders_send():
    _ensure_session_keys()
    if not _ensure_supplier_orders_internal_access():
        return redirect(url_for("warehouse.home"))
    ensure_supplier_order_flow_schema()

    store_code = str(session.get("store_code") or "").strip()
    store_name = str(session.get("store_name") or _supplier_order_store_name(store_code)).strip()
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("warehouse.home"))
    supplier_rows = list_sql_fornitori()
    supplier_options = sorted(supplier_rows, key=lambda x: str(x.get("Fornitore") or "").lower())

    selected_supplier_uuid = (request.values.get("supplier_row_uuid") or "").strip()
    requested_delivery_date = (request.values.get("requested_delivery_date") or "").strip()
    note_ordine = (request.values.get("note_ordine") or "").strip()
    selected_supplier = get_supplier_by_uuid(selected_supplier_uuid) if selected_supplier_uuid else None
    pricelist_rows = load_supplier_pricelist_for_order(store_code, str(selected_supplier.get("Fornitore") or "")) if selected_supplier else []
    selected_contact = None
    total_estimated = Decimal("0")

    if request.method == "POST":
        submit_action = (request.form.get("submit_action") or "").strip().lower()
        if not submit_action:
            submit_action = "send" if str(request.form.get("rows_count") or "").strip() else "load"
        if submit_action == "load":
            pass
        elif not selected_supplier:
            flash("Seleziona un fornitore.", "warning")
        elif not requested_delivery_date:
            flash("Inserisci la data consegna richiesta.", "warning")
        else:
            selected_order_mode = normalize_order_mode(selected_supplier.get("TipoOrdine") or ORDER_MODE_MAIL)
            selected_contact = resolve_supplier_contact(selected_supplier_uuid, store_code, selected_order_mode)
            try:
                parsed_rows, total_estimated = _supplier_order_parse_form_rows(request.form)
            except Exception as e:
                flash(str(e), "danger")
                parsed_rows = []
            if not parsed_rows:
                flash("Inserisci almeno una quantità valida.", "warning")
            elif selected_order_mode == ORDER_MODE_MAIL and not str(selected_contact.get("Email") or "").strip():
                flash("Il fornitore di tipo Mail non ha una mail associata per questo store.", "danger")
            else:
                actor = {
                    "uid": session.get("uid"),
                    "name": session.get("name"),
                    "email": session.get("email"),
                    "role": session.get("role"),
                }
                created = create_supplier_order(
                    site=store_code,
                    store_name=store_name,
                    supplier_row_uuid=selected_supplier_uuid,
                    supplier_name=str(selected_supplier.get("Fornitore") or ""),
                    supplier_contact=selected_contact,
                    requested_delivery_date=requested_delivery_date,
                    order_mode=selected_order_mode,
                    note_ordine=note_ordine,
                    rows=parsed_rows,
                    actor=actor,
                )
                try:
                    detail = get_supplier_order_detail(created["row_uuid"])
                    if not detail:
                        raise RuntimeError("Ordine creato ma non recuperabile.")
                    if selected_order_mode == ORDER_MODE_MAIL:
                        pdf_bytes = build_order_pdf_bytes(detail)
                        saved_pdf = save_order_pdf(detail, pdf_bytes)
                        update_order_pdf_and_email(created["row_uuid"], pdf_rel_path=saved_pdf["rel_path"], pdf_filename=saved_pdf["filename"], sent_email_at=False)
                        send_order_email(
                            to_email=str(selected_contact.get("Email") or "").strip(),
                            subject=_supplier_order_mail_subject(store_name or store_code, requested_delivery_date),
                            body=_supplier_order_mail_body(detail),
                            pdf_bytes=pdf_bytes,
                            pdf_filename=saved_pdf["filename"],
                        )
                        update_order_pdf_and_email(created["row_uuid"], pdf_rel_path=saved_pdf["rel_path"], pdf_filename=saved_pdf["filename"], sent_email_at=True)
                        flash("Ordine inviato via mail con PDF allegato.", "success")
                    else:
                        flash("Ordine online creato correttamente.", "success")
                    return redirect(url_for("warehouse.supplier_orders_sent", order_id=created["row_uuid"]))
                except Exception as e:
                    delete_supplier_order(created["row_uuid"])
                    flash(f"Errore durante invio ordine: {e}", "danger")

    if selected_supplier and not selected_contact:
        selected_contact = resolve_supplier_contact(selected_supplier_uuid, store_code, str(selected_supplier.get("TipoOrdine") or ORDER_MODE_MAIL))

    return render_template(
        "warehouse_supplier_order_send.html",
        store_code=store_code,
        store_name=store_name,
        supplier_options=supplier_options,
        selected_supplier=selected_supplier,
        selected_supplier_uuid=selected_supplier_uuid,
        selected_contact=selected_contact,
        requested_delivery_date=requested_delivery_date,
        note_ordine=note_ordine,
        pricelist_rows=pricelist_rows,
        total_estimated=total_estimated,
    )


@warehouse_bp.route("/ordini-fornitore/inviati")
def supplier_orders_sent():
    _ensure_session_keys()
    if not _ensure_supplier_orders_internal_access():
        return redirect(url_for("warehouse.home"))
    ensure_supplier_order_flow_schema()
    allowed_stores = _supplier_order_allowed_sites()
    allowed_codes = [str(s.get("code")) for s in allowed_stores if s.get("code")]
    filters = {
        "store": request.args.get("store"),
        "supplier": request.args.get("supplier"),
        "status": request.args.get("status"),
        "order_mode": request.args.get("order_mode"),
        "order_date_from": request.args.get("order_date_from"),
        "order_date_to": request.args.get("order_date_to"),
        "delivery_date_from": request.args.get("delivery_date_from"),
        "delivery_date_to": request.args.get("delivery_date_to"),
    }
    rows = list_supplier_orders(allowed_sites=allowed_codes, viewer_role="internal", filters=filters)
    selected_order_id = (request.args.get("order_id") or "").strip()
    detail = get_supplier_order_detail(selected_order_id) if selected_order_id else None
    if detail and allowed_codes and str(detail["header"].get("site") or "") not in allowed_codes and str(session.get("role") or "").strip().lower() != "admin":
        flash("Ordine non disponibile per questo utente.", "danger")
        return redirect(url_for("warehouse.supplier_orders_sent"))
    return render_template(
        "warehouse_supplier_orders_sent.html",
        rows=rows,
        detail=detail,
        selected_order_id=selected_order_id,
        filters=filters,
        store_options=allowed_stores,
        supplier_options=list_sql_fornitori(),
        status_options=[ORDER_STATUS_INVITED, ORDER_STATUS_VIEWED, ORDER_STATUS_COMPLETED, ORDER_STATUS_SENT],
    )


@warehouse_bp.post("/ordini-fornitore/inviati/<order_id>/carica-ddt")
def supplier_orders_sent_to_ddt(order_id: str):
    _ensure_session_keys()
    if not _ensure_supplier_orders_internal_access():
        return redirect(url_for("warehouse.home"))
    detail = get_supplier_order_detail(order_id)
    if not detail:
        flash("Ordine non trovato.", "warning")
        return redirect(url_for("warehouse.supplier_orders_sent"))
    allowed_codes = _supplier_order_allowed_site_codes()
    if allowed_codes and str(detail["header"].get("site") or "") not in allowed_codes and str(session.get("role") or "").strip().lower() != "admin":
        flash("Ordine non disponibile per questo utente.", "danger")
        return redirect(url_for("warehouse.supplier_orders_sent"))
    header = detail["header"]
    if header.get("ddt_data_doc") and header.get("ddt_data_rif") and header.get("ddt_supplier_code"):
        flash("Questo ordine è già collegato a un DDT.", "info")
        return redirect(
            url_for(
                "warehouse.delivery_edit",
                supplier_code=header.get("ddt_supplier_code"),
                data_doc=header.get("ddt_data_doc"),
                data_rif=header.get("ddt_data_rif"),
            )
        )
    supplier_code = _supplier_order_access_supplier_code(str(header.get("site") or ""), str(header.get("supplier_name") or ""))
    if not supplier_code:
        flash("Non riesco a collegare il fornitore ordine al fornitore DDT dello store.", "danger")
        return redirect(url_for("warehouse.supplier_orders_sent", order_id=order_id))
    session["store_code"] = str(header.get("site") or "")
    if header.get("store_name"):
        session["store_name"] = str(header.get("store_name"))
    session["supplier_order_ddt_prefill"] = {
        "order_row_uuid": str(header.get("row_uuid") or order_id),
        "supplier_code": supplier_code,
        "supplier_name": str(header.get("supplier_name") or ""),
        "data_doc": datetime.now().date().isoformat(),
        "data_rif": str(header.get("requested_delivery_date") or ""),
        "rows": _supplier_order_effective_rows(detail),
        "note_ordine": str(header.get("note_ordine") or ""),
    }
    return redirect(url_for("warehouse.delivery_new"))


@warehouse_bp.route("/ordini-fornitore/ricevuti", methods=["GET"])
def supplier_orders_received():
    _ensure_session_keys()
    if not _ensure_supplier_orders_supplier_access():
        return redirect(url_for("warehouse.home"))
    ensure_supplier_order_flow_schema()
    supplier_ids = resolve_supplier_user_supplier_ids(user_uid=str(session.get("uid") or ""), user_email=str(session.get("email") or ""))
    raw_status = (request.args.get("status") or "").strip()
    supplier_status_map = {
        "Ricevuto": ORDER_STATUS_INVITED,
        "Inviato": ORDER_STATUS_INVITED,
        "Invitato": ORDER_STATUS_INVITED,
        ORDER_STATUS_INVITED: ORDER_STATUS_INVITED,
        ORDER_STATUS_VIEWED: ORDER_STATUS_VIEWED,
        ORDER_STATUS_COMPLETED: ORDER_STATUS_COMPLETED,
    }
    status_filter = supplier_status_map.get(raw_status, raw_status)
    rows = list_supplier_orders(
        supplier_row_uuids=supplier_ids,
        viewer_role="supplier",
        filters={"status": status_filter},
    )
    selected_order_id = (request.args.get("order_id") or "").strip()
    detail = get_supplier_order_detail(selected_order_id) if selected_order_id else None
    if detail and str(detail["header"].get("supplier_row_uuid") or "") not in supplier_ids and str(session.get("role") or "").strip().lower() != "admin":
        flash("Ordine non disponibile per questo profilo fornitore.", "danger")
        return redirect(url_for("warehouse.supplier_orders_received"))
    if detail:
        mark_order_viewed(
            selected_order_id,
            {"uid": session.get("uid"), "name": session.get("name"), "email": session.get("email"), "role": session.get("role")},
        )
        detail = get_supplier_order_detail(selected_order_id)
    return render_template(
        "warehouse_supplier_orders_received.html",
        rows=rows,
        detail=detail,
        catalog_options=_supplier_order_catalog_options(detail),
        selected_order_id=selected_order_id,
        status_options=[
            {"value": "Ricevuto", "label": "Ricevuto"},
            {"value": ORDER_STATUS_VIEWED, "label": ORDER_STATUS_VIEWED},
            {"value": ORDER_STATUS_COMPLETED, "label": ORDER_STATUS_COMPLETED},
        ],
        selected_status=raw_status if raw_status else "",
    )


@warehouse_bp.post("/ordini-fornitore/ricevuti/<order_id>/linea/<line_id>")
def supplier_orders_received_line(order_id: str, line_id: str):
    _ensure_session_keys()
    if not _ensure_supplier_orders_supplier_access():
        return redirect(url_for("warehouse.home"))
    detail = get_supplier_order_detail(order_id)
    supplier_ids = resolve_supplier_user_supplier_ids(user_uid=str(session.get("uid") or ""), user_email=str(session.get("email") or ""))
    if not detail or (str(detail["header"].get("supplier_row_uuid") or "") not in supplier_ids and str(session.get("role") or "").strip().lower() != "admin"):
        flash("Ordine non disponibile per questo profilo fornitore.", "danger")
        return redirect(url_for("warehouse.supplier_orders_received"))
    if str(detail["header"].get("status") or "") == ORDER_STATUS_COMPLETED:
        flash("Ordine già completato: ulteriori modifiche bloccate.", "warning")
        return redirect(url_for("warehouse.supplier_orders_received", order_id=order_id))
    try:
        payload = _supplier_order_line_payload(request.form, line_id)
        save_supplier_order_line(
            order_id,
            line_id,
            picked=bool(payload.get("picked")),
            lotto=str(payload.get("lotto") or ""),
            scadenza=str(payload.get("scadenza") or ""),
            note_fornitore=str(payload.get("note_fornitore") or ""),
        )
        flash("Riga aggiornata.", "success")
    except Exception as e:
        flash(f"Errore salvataggio riga: {e}", "danger")
    return redirect(url_for("warehouse.supplier_orders_received", order_id=order_id))


@warehouse_bp.post("/ordini-fornitore/ricevuti/<order_id>/completa")
def supplier_orders_received_complete(order_id: str):
    _ensure_session_keys()
    if not _ensure_supplier_orders_supplier_access():
        return redirect(url_for("warehouse.home"))
    detail = get_supplier_order_detail(order_id)
    supplier_ids = resolve_supplier_user_supplier_ids(user_uid=str(session.get("uid") or ""), user_email=str(session.get("email") or ""))
    if not detail or (str(detail["header"].get("supplier_row_uuid") or "") not in supplier_ids and str(session.get("role") or "").strip().lower() != "admin"):
        flash("Ordine non disponibile per questo profilo fornitore.", "danger")
        return redirect(url_for("warehouse.supplier_orders_received"))
    if str(detail["header"].get("status") or "") == ORDER_STATUS_COMPLETED:
        flash("Ordine già completato.", "info")
        return redirect(url_for("warehouse.supplier_orders_received", order_id=order_id))
    try:
        for row in detail.get("rows") or []:
            line_id = str(row.get("row_uuid") or "").strip()
            if not line_id:
                continue
            payload = _supplier_order_line_payload(request.form, line_id)
            save_supplier_order_line(
                order_id,
                line_id,
                picked=bool(payload.get("picked")),
                lotto=str(payload.get("lotto") or ""),
                scadenza=str(payload.get("scadenza") or ""),
                note_fornitore=str(payload.get("note_fornitore") or ""),
            )
        complete_supplier_order(
            order_id,
            {"uid": session.get("uid"), "name": session.get("name"), "email": session.get("email"), "role": session.get("role")},
        )
        flash("Ordine completato.", "success")
    except Exception as e:
        flash(str(e), "danger")
    return redirect(url_for("warehouse.supplier_orders_received", order_id=order_id))


@warehouse_bp.post("/ordini-fornitore/ricevuti/<order_id>/riapri")
def supplier_orders_received_reopen(order_id: str):
    _ensure_session_keys()
    if not _ensure_supplier_orders_supplier_access():
        return redirect(url_for("warehouse.home"))
    detail = get_supplier_order_detail(order_id)
    supplier_ids = resolve_supplier_user_supplier_ids(user_uid=str(session.get("uid") or ""), user_email=str(session.get("email") or ""))
    if not detail or (str(detail["header"].get("supplier_row_uuid") or "") not in supplier_ids and str(session.get("role") or "").strip().lower() != "admin"):
        flash("Ordine non disponibile per questo profilo fornitore.", "danger")
        return redirect(url_for("warehouse.supplier_orders_received"))
    try:
        reopen_supplier_order(
            order_id,
            {"uid": session.get("uid"), "name": session.get("name"), "email": session.get("email"), "role": session.get("role")},
        )
        flash("Ordine riaperto.", "success")
    except Exception as e:
        flash(str(e), "danger")
    return redirect(url_for("warehouse.supplier_orders_received", order_id=order_id))


@warehouse_bp.post("/ordini-fornitore/ricevuti/<order_id>/aggiungi-prodotto")
def supplier_orders_received_add_product(order_id: str):
    _ensure_session_keys()
    if not _ensure_supplier_orders_supplier_access():
        return redirect(url_for("warehouse.home"))
    detail = get_supplier_order_detail(order_id)
    supplier_ids = resolve_supplier_user_supplier_ids(user_uid=str(session.get("uid") or ""), user_email=str(session.get("email") or ""))
    if not detail or (str(detail["header"].get("supplier_row_uuid") or "") not in supplier_ids and str(session.get("role") or "").strip().lower() != "admin"):
        flash("Ordine non disponibile per questo profilo fornitore.", "danger")
        return redirect(url_for("warehouse.supplier_orders_received"))
    if str(detail["header"].get("status") or "") == ORDER_STATUS_COMPLETED and detail["header"].get("archived_at"):
        flash("Ordine archiviato: riaprilo prima di modificare le righe.", "warning")
        return redirect(url_for("warehouse.supplier_orders_received", order_id=order_id))
    catalog = _supplier_order_catalog_options(detail)
    selected_code = str(request.form.get("selected_product_code") or "").strip().lower()
    selected_descr = str(request.form.get("selected_product_descrizione") or "").strip().lower()
    selected_product = next(
        (r for r in catalog if str(r.get("product_code") or "").strip().lower() == selected_code and selected_code),
        None,
    ) or next(
        (r for r in catalog if str(r.get("descrizione") or "").strip().lower() == selected_descr and selected_descr),
        None,
    )
    try:
        if not selected_product:
            raise ValueError("Seleziona un prodotto valido dal listino.")
        add_supplier_order_selected_product(
            order_id,
            selected_product=selected_product,
            qty_ordered=request.form.get("qty_added") or "",
            replacement_for_row_uuid=request.form.get("replacement_for_row_uuid") or "",
            picked=(request.form.get("picked_added") == "1"),
            lotto=str(request.form.get("lotto_added") or ""),
            scadenza=str(request.form.get("scadenza_added") or ""),
            note_fornitore=str(request.form.get("note_added") or ""),
        )
        flash("Prodotto aggiunto all'ordine.", "success")
    except Exception as e:
        flash(str(e), "danger")
    return redirect(url_for("warehouse.supplier_orders_received", order_id=order_id))


@warehouse_bp.get("/ordini-fornitore/pdf/<order_id>")
def supplier_order_pdf(order_id: str):
    _ensure_session_keys()
    detail = get_supplier_order_detail(order_id)
    if not detail:
        flash("Ordine non trovato.", "warning")
        return redirect(url_for("warehouse.supplier_orders_home"))
    role = str(session.get("role") or "").strip().lower()
    if role == "fornitore":
        supplier_ids = get_supplier_user_supplier_ids(str(session.get("email") or ""))
        if str(detail["header"].get("supplier_row_uuid") or "") not in supplier_ids and role != "admin":
            flash("Ordine non disponibile per questo profilo fornitore.", "danger")
            return redirect(url_for("warehouse.supplier_orders_received"))
    else:
        allowed_codes = _supplier_order_allowed_site_codes()
        if allowed_codes and str(detail["header"].get("site") or "") not in allowed_codes and role != "admin":
            flash("Ordine non disponibile per questo utente.", "danger")
            return redirect(url_for("warehouse.supplier_orders_sent"))
    pdf_bytes = build_order_pdf_bytes(detail)
    return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf", download_name=f"{detail['header'].get('numero_ordine') or 'ordine'}.pdf", as_attachment=False, max_age=300)


@warehouse_bp.route("/api/ordini")
def api_ordini():
    store_code = session.get("store_code")
    if not store_code:
        return jsonify({"error": "Store non selezionato"}), 400

    supplier = (request.args.get("supplier") or "").strip()
    listino = (request.args.get("listino") or "FoodPaper").strip()
    order_date = (request.args.get("order_date") or "").strip()
    next_delivery = (request.args.get("next_delivery") or "").strip()

    if not supplier:
        return jsonify({"error": "Seleziona un fornitore"}), 400
    if not order_date or not next_delivery:
        return jsonify({"error": "Inserisci le date di consegna"}), 400

    try:
        order_day = datetime.fromisoformat(order_date).date()
        next_day = datetime.fromisoformat(next_delivery).date()
    except Exception:
        return jsonify({"error": "Formato data non valido"}), 400

    try:
        data = build_orders_suggestion(
            store_code=str(store_code),
            supplier_name=str(supplier),
            listino=str(listino),
            order_day=order_day,
            next_delivery_day=next_day,
        )
        return jsonify(data)
    except Exception as e:
        current_app.logger.exception("Errore api_ordini")
        return jsonify({"error": str(e)}), 500
