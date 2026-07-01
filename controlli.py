from __future__ import annotations

import io
import datetime
import re
from typing import Any, Dict, List, Optional

from flask import Blueprint, jsonify, render_template, request, session, send_file

from controlli_repository import get_pnl, get_andamento

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except Exception:
    Workbook = None  # type: ignore


controlli_bp = Blueprint("controlli", __name__, url_prefix="/controlli")


def _get_int(name: str, default: int) -> int:
    try:
        return int(request.args.get(name, default))
    except Exception:
        return default


def _get_store_codes() -> str:
    # Priorità: querystring 'stores' (comma-separated) -> session['store_code']
    stores = (request.args.get("stores") or request.args.get("store") or "").strip()
    if stores:
        return stores
    return (session.get("store_code") or "").strip()


@controlli_bp.get("/pnl")
def pnl():
    today = datetime.date.today()
    store_code = (session.get("store_code") or "").strip()
    store_name = (session.get("store_name") or "").strip()
    return render_template(
        "controlli_pnl.html",
        store_code=store_code,
        store_name=store_name,
        default_year=today.year,
        default_month=today.month,
    )



@controlli_bp.get("/kpi")
def kpi():
    today = datetime.date.today()
    store_code = (session.get("store_code") or "").strip()
    store_name = (session.get("store_name") or "").strip()
    return render_template(
        "controlli_kpi.html",
        store_code=store_code,
        store_name=store_name,
        default_year=today.year,
        default_month=today.month,
    )

@controlli_bp.get("/andamento")
def andamento():
    today = datetime.date.today()
    store_code = (session.get("store_code") or "").strip()
    store_name = (session.get("store_name") or "").strip()
    return render_template(
        "controlli_andamento.html",
        store_code=store_code,
        store_name=store_name,
        default_year=today.year,
        default_month=today.month,
    )


@controlli_bp.get("/api/pnl")
def api_pnl():
    try:
        stores = _get_store_codes()
        year = _get_int("year", datetime.date.today().year)
        month_from = _get_int("month_from", datetime.date.today().month)
        month_to = _get_int("month_to", month_from)

        data = get_pnl(stores, year, month_from, month_to)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 400


def _build_kpi_items_from_pnl(pnl: Dict[str, Any]) -> List[Dict[str, Any]]:
    def _norm(s: str) -> str:
        return re.sub(r"\s+", "", (s or "").strip().lower())

    wanted = ["REVENUES", "COGS", "LABOUR COST", "EBITDA"]
    wanted_norm = {_norm(v): v for v in wanted}

    rows_by_voice = {}
    for r in (pnl.get("rows") or []):
        v = str(r.get("voice") or "").strip()
        if not v:
            continue
        rows_by_voice[_norm(v)] = r

    out_items = []
    for key_norm, label in wanted_norm.items():
        r = rows_by_voice.get(key_norm) or rows_by_voice.get(_norm(label))
        if not r and key_norm == _norm("REVENUES"):
            r = rows_by_voice.get(_norm("Revenues"))

        item = {"kpi": label}
        if not r:
            item.update({"missing": True, "status": "na"})
            out_items.append(item)
            continue

        b = float(r.get("budget") or 0.0)
        a = float(r.get("actual") or 0.0)
        b_pct = float(r.get("budget_pct") or 0.0)
        a_pct = float(r.get("actual_pct") or 0.0)
        diff = float(r.get("diff") or (a - b))
        diff_pct = r.get("diff_pct", None)

        status = "na"
        if label == "REVENUES":
            if diff_pct is None:
                status = "na"
            else:
                dp = float(diff_pct)
                if dp >= 0:
                    status = "green"
                elif dp >= -0.005:
                    status = "yellow"
                else:
                    status = "red"
            item.update(
                {
                    "budget": b,
                    "actual": a,
                    "diff": diff,
                    "diff_pct": diff_pct,
                }
            )
        else:
            delta_pp = a_pct - b_pct
            if label in ("COGS", "LABOUR COST"):
                if delta_pp <= 0:
                    status = "green"
                elif delta_pp <= 0.005:
                    status = "yellow"
                else:
                    status = "red"
            elif label == "EBITDA":
                if delta_pp >= 0:
                    status = "green"
                elif delta_pp >= -0.005:
                    status = "yellow"
                else:
                    status = "red"
            item.update(
                {
                    "budget_pct": b_pct,
                    "actual_pct": a_pct,
                    "delta_pp": delta_pp,
                }
            )

        item["status"] = status
        out_items.append(item)

    return out_items


@controlli_bp.get("/api/kpi")
def api_kpi():
    """KPI sintetici: REVENUES / COGS / LABOUR COST / EBITDA vs Budget."""
    try:
        stores = _get_store_codes()
        year = _get_int("year", datetime.date.today().year)
        month_from = _get_int("month_from", datetime.date.today().month)
        month_to = _get_int("month_to", month_from)

        pnl = get_pnl(stores, year, month_from, month_to)
        selected_stores = [s.strip() for s in str(stores or "").split(",") if s.strip()]
        selected_stores = list(dict.fromkeys(selected_stores))

        store_breakdown = []
        if len(selected_stores) > 1:
            for code in selected_stores:
                store_pnl = get_pnl(code, year, month_from, month_to)
                store_breakdown.append(
                    {
                        "store_code": code,
                        "items": _build_kpi_items_from_pnl(store_pnl),
                    }
                )

        return jsonify(
            {
                "store_code": pnl.get("store_code"),
                "stores": pnl.get("stores"),
                "year": pnl.get("year"),
                "month_from": pnl.get("month_from"),
                "month_to": pnl.get("month_to"),
                "months_label": pnl.get("months_label"),
                "items": _build_kpi_items_from_pnl(pnl),
                "store_breakdown": store_breakdown,
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@controlli_bp.get("/api/andamento")
def api_andamento():
    try:
        stores = _get_store_codes()
        year = _get_int("year", datetime.date.today().year)
        month_from = _get_int("month_from", datetime.date.today().month)
        month_to = _get_int("month_to", month_from)

        data = get_andamento(stores, year, month_from, month_to)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@controlli_bp.get("/api/pnl/export")
def api_pnl_export():
    # Export Excel del P&L (stesso contenuto della tabella)
    if Workbook is None:
        return jsonify({"error": "openpyxl non disponibile"}), 500

    try:
        stores = _get_store_codes()
        year = _get_int("year", datetime.date.today().year)
        month_from = _get_int("month_from", datetime.date.today().month)
        month_to = _get_int("month_to", month_from)

        data = get_pnl(stores, year, month_from, month_to)
        rows = data.get("rows") or []

        wb = Workbook()
        ws = wb.active
        ws.title = "P&L"

        headers = [
            "Voce",
            "Budget",
            "Budget %",
            "Actual",
            "Actual %",
            "Δ Actual vs Budget",
            "Δ% Actual vs Budget",
            "Anno precedente",
            "Anno precedente %",
            "Δ Actual vs Anno precedente",
            "Δ% Actual vs Anno precedente",
        ]
        ws.append(headers)

        bold = Font(bold=True)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = bold
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _nf(v: Any) -> float:
            try:
                return float(v)
            except Exception:
                return 0.0

        for r in rows:
            ws.append(
                [
                    r.get("voice", ""),
                    _nf(r.get("budget")),
                    _nf(r.get("budget_pct")),
                    _nf(r.get("actual")),
                    _nf(r.get("actual_pct")),
                    _nf(r.get("diff")),
                    _nf(r.get("diff_pct")),
                    _nf(r.get("last_year")),
                    _nf(r.get("last_year_pct")),
                    _nf(r.get("diff_last_year")),
                    _nf(r.get("diff_last_year_pct")),
                ]
            )

        # format
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws.max_row):
            for cell in row:
                cell.number_format = "#,##0.00"
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, max_row=ws.max_row):
            for cell in row:
                cell.number_format = "#,##0.00"
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=8, max_row=ws.max_row):
            for cell in row:
                cell.number_format = "#,##0.00"

        pct_cols = [3, 5, 7, 9, 11]
        for col in pct_cols:
            for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col, max_row=ws.max_row):
                for cc in cell:
                    cc.number_format = "0.00%"

        ws.freeze_panes = "B2"
        # column widths
        ws.column_dimensions["A"].width = 34
        for col in "BCDEFGHIJK":
            ws.column_dimensions[col].width = 18

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"PnL_{year}_{month_from:02d}-{month_to:02d}.xlsx"
        return send_file(
            bio,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 400
