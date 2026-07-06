from __future__ import annotations

from app_logging import log_swallowed
import io
from datetime import datetime

from flask import Blueprint, flash, redirect, render_template, request, send_file, session, url_for

from mbo_repository import (
    build_custom_survey_dashboard,
    build_audit_address_mapping_matrix,
    build_audit_month_matrix,
    build_mbo_google_scores,
    build_mbo_glovo_scores,
    build_mbo_deliveroo_scores,
    build_mbo_pnl_scores,
    build_mbo_report_dashboard,
    build_mbo_store_month_matrix,
    build_soft_skill_dashboard,
    load_mbo_reward_settings,
    create_custom_survey,
    delete_audit_manual_value,
    delete_custom_survey_submission,
    delete_soft_skill_submission,
    delete_google_manual_value,
    delete_glovo_manual_value,
    delete_deliveroo_manual_value,
    delete_area_manager,
    get_custom_survey,
    get_mbo_report_definition,
    ignore_audit_row,
    ensure_mbo_schema,
    import_audit_file,
    list_area_managers,
    list_mbo_report_definitions,
    parse_custom_survey_questions_file,
    restore_audit_row,
    recalculate_soft_skill_submissions,
    save_audit_address_mapping,
    save_audit_manual_value,
    save_audit_row_assignment,
    save_area_manager,
    save_custom_survey_submission,
    save_google_manual_value,
    save_glovo_manual_value,
    save_deliveroo_manual_value,
    save_mbo_report_definition,
    save_mbo_report_multipliers,
    get_soft_skill_period,
    get_soft_skill_period_by_token,
    get_soft_skill_questions,
    get_soft_skill_submission,
    save_soft_skill_period,
    save_soft_skill_submission,
    save_store_month_assignments,
    save_area_manager_monthly_settings,
    save_reward_weights,
    save_role_monthly_rewards,
    set_custom_survey_active,
    survey_store_area_manager_map,
    update_custom_survey_names,
    propagate_area_manager_monthly_setting,
    propagate_role_monthly_reward,
    sync_area_managers_from_ilp,
)


mbo_bp = Blueprint("mbo", __name__, url_prefix="/mbo")


def _safe_xlsx_name(value: str, fallback: str = "export") -> str:
    name = "".join(ch if ch.isalnum() else "_" for ch in str(value or "")).strip("_")
    return name or fallback


def _xlsx_response(workbook, filename: str):
    bio = io.BytesIO()
    workbook.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name=filename,
        as_attachment=True,
    )


def _style_header(ws):
    try:
        from openpyxl.styles import Font, PatternFill

        fill = PatternFill("solid", fgColor="E9EEF6")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
    except Exception:
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)


def _auto_width(ws, max_width: int = 55):
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        width = 10
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def _append_rows_sheet(wb, title: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_header(ws)
    _auto_width(ws)
    ws.freeze_panes = "A2"
    return ws


def _pct_value(value):
    if value is None:
        return None
    try:
        return round(float(value), 4)
    except Exception:
        return None


def _session_module_enabled(module_key: str) -> bool:
    role = str(session.get("role") or "").strip().lower()
    if role == "admin":
        return True
    mods = session.get("access_modules")
    if isinstance(mods, dict) and module_key in mods:
        return bool(mods.get(module_key))
    return True


def _ensure_mbo_access() -> bool:
    if not session.get("uid"):
        return False
    role = str(session.get("role") or "").strip().lower()
    if role == "admin":
        return True
    return _session_module_enabled("mod_mbo")


def _ensure_admin_access() -> bool:
    return str(session.get("role") or "").strip().lower() == "admin"


@mbo_bp.before_request
def _mbo_before_request():
    if request.endpoint in {"mbo.soft_skills_fill", "mbo.custom_survey_fill"}:
        ensure_mbo_schema()
        return None
    if not session.get("uid"):
        return redirect(url_for("login", next=request.full_path if request.query_string else request.path))
    if not _ensure_mbo_access():
        flash("Non hai i permessi per accedere alla sezione MBO.", "warning")
        return redirect(url_for("dashboard"))
    ensure_mbo_schema()
    return None


@mbo_bp.route("", methods=["GET", "POST"])
@mbo_bp.route("/", methods=["GET", "POST"])
def home():
    year = int((request.values.get("year") or datetime.today().year))
    selected_report_uuid = str(request.values.get("report_uuid") or "").strip()

    if request.method == "POST":
        action = str(request.form.get("action") or "").strip().lower()
        try:
            if action == "save_report":
                selected_report_uuid = save_mbo_report_definition(
                    report_uuid=str(request.form.get("report_uuid") or "").strip() or None,
                    report_name=str(request.form.get("report_name") or "").strip(),
                    start_year_month=str(request.form.get("start_year_month") or "").strip(),
                    end_year_month=str(request.form.get("end_year_month") or "").strip(),
                    created_by=str(session.get("email") or session.get("uid") or ""),
                )
                flash("Report MBO salvato.", "success")
            elif action == "save_staff_multipliers":
                save_mbo_report_multipliers(
                    report_uuid=str(request.form.get("report_uuid") or "").strip(),
                    report_type="staff",
                    form_data=request.form,
                )
                selected_report_uuid = str(request.form.get("report_uuid") or "").strip()
                flash("Parametri Multi salvati.", "success")
            elif action == "save_area_manager_multipliers":
                save_mbo_report_multipliers(
                    report_uuid=str(request.form.get("report_uuid") or "").strip(),
                    report_type="area_manager",
                    form_data=request.form,
                )
                selected_report_uuid = str(request.form.get("report_uuid") or "").strip()
                flash("Parametri Ricalcolo salvati.", "success")
            else:
                flash("Azione riepilogo MBO non riconosciuta.", "warning")
        except Exception as e:
            flash(f"Errore riepilogo MBO: {e}", "danger")
        return redirect(url_for("mbo.home", year=year, report_uuid=selected_report_uuid) if selected_report_uuid else url_for("mbo.home", year=year))

    try:
        matrix = build_mbo_store_month_matrix(year)
    except Exception as e:
        flash(f"Errore caricamento riepilogo MBO: {e}", "danger")
        matrix = {"store_count": 0, "default_count": 0, "override_count": 0, "unmatched_count": 0}
    try:
        reports = list_mbo_report_definitions()
        if not selected_report_uuid and reports:
            selected_report_uuid = str(reports[0].get("report_uuid") or "")
        selected_report = get_mbo_report_definition(selected_report_uuid) if selected_report_uuid else None
        report_dashboard = build_mbo_report_dashboard(selected_report_uuid) if selected_report_uuid else None
    except Exception as e:
        flash(f"Errore caricamento report MBO: {e}", "danger")
        reports = []
        selected_report = None
        report_dashboard = None
    return render_template(
        "mbo_home.html",
        mbo_year=year,
        mbo_year_options=list(range(datetime.today().year - 1, datetime.today().year + 3)),
        mbo_reports=reports,
        selected_report=selected_report,
        report_dashboard=report_dashboard,
        mbo_summary={
            "store_count": matrix.get("store_count", 0),
            "default_count": matrix.get("default_count", 0),
            "override_count": matrix.get("override_count", 0),
            "unmatched_count": matrix.get("unmatched_count", 0),
        },
    )


def _mbo_metric_row(metric: dict | None) -> list:
    metric = metric or {}
    return [
        _pct_value(metric.get("pct")),
        round(float(metric.get("points_taken") or 0), 4),
        round(float(metric.get("points_theoretical") or 0), 4),
    ]


def _mbo_category_row(category: dict | None) -> list:
    category = category or {}
    return [
        _pct_value(category.get("pct")),
        _pct_value(category.get("weight_pct")),
        _pct_value(category.get("weighted_pct")),
    ]


def _am_export_filter(request_key: str = "area_manager_uuid") -> str:
    return str(request.args.get(request_key) or "").strip()


@mbo_bp.route("/report/<report_uuid>/export.xlsx")
def mbo_report_export_xlsx(report_uuid: str):
    if not _ensure_mbo_access():
        return redirect(url_for("login", next=request.full_path if request.query_string else request.path))
    section = str(request.args.get("section") or "all").strip().lower()
    if section not in {"all", "store", "staff", "area_manager"}:
        section = "all"
    try:
        from openpyxl import Workbook
    except Exception as e:
        flash(f"Export Excel non disponibile: {e}", "danger")
        return redirect(url_for("mbo.home", report_uuid=report_uuid))

    try:
        dashboard = build_mbo_report_dashboard(report_uuid)
    except Exception as e:
        flash(f"Errore export riepilogo MBO: {e}", "danger")
        return redirect(url_for("mbo.home", report_uuid=report_uuid))

    report = dashboard.get("report") or {}
    selected_am_uuid = _am_export_filter()
    try:
        report_store_am_map = survey_store_area_manager_map(
            str(report.get("start_year_month") or ""),
            str(report.get("end_year_month") or ""),
        )
    except Exception:
        report_store_am_map = {}

    def _report_am_for_store(store_code: str) -> tuple[str, str]:
        am = report_store_am_map.get(str(store_code or "").strip()) or {}
        return (
            str(am.get("area_manager_name") or "Senza Area Manager"),
            str(am.get("area_manager_uuid") or ""),
        )

    wb = Workbook()
    wb.remove(wb.active)

    if section == "all":
        _append_rows_sheet(
            wb,
            "Report",
            ["Voce", "Valore"],
            [
                ["Nome", report.get("report_name") or ""],
                ["Periodo", f"{report.get('start_year_month') or ''} / {report.get('end_year_month') or ''}"],
                ["Mesi", ", ".join(str(x) for x in dashboard.get("months") or [])],
            ],
        )

    if section in {"all", "store"}:
        top_rows = []
        top10_rows = []
        for row in (dashboard.get("store") or {}).get("top10") or []:
            _am_name, am_uuid = _report_am_for_store(str(row.get("store_code") or ""))
            if selected_am_uuid and am_uuid != selected_am_uuid:
                continue
            top10_rows.append(row)
        for idx, row in enumerate(top10_rows, start=1):
            metrics = row.get("metrics") or {}
            top_score = row.get("top_score") or {}
            top_rows.append(
                [
                    idx,
                    row.get("store_code") or "",
                    row.get("store_name") or "",
                    *_report_am_for_store(str(row.get("store_code") or "")),
                    _pct_value(top_score.get("pct")),
                    round(float(top_score.get("points_taken") or 0), 4),
                    round(float(top_score.get("points_theoretical") or 0), 4),
                    *_mbo_metric_row(metrics.get("pnl")),
                    *_mbo_metric_row(metrics.get("rating")),
                    *_mbo_metric_row(metrics.get("google")),
                ]
            )
        _append_rows_sheet(
            wb,
            "TOP 10 Store",
            [
                "Rank", "Store code", "Store", "Area Manager", "AM UUID",
                "Totale %", "Totale punti", "Totale teorici",
                "P&L %", "P&L punti", "P&L teorici",
                "Rating %", "Rating punti", "Rating teorici",
                "Google %", "Google punti", "Google teorici",
            ],
            top_rows,
        )
        rows = []
        for row in (dashboard.get("store") or {}).get("rows") or []:
            _am_name, am_uuid = _report_am_for_store(str(row.get("store_code") or ""))
            if selected_am_uuid and am_uuid != selected_am_uuid:
                continue
            metrics = row.get("metrics") or {}
            rows.append(
                [
                    row.get("store_code") or "",
                    row.get("store_name") or "",
                    *_report_am_for_store(str(row.get("store_code") or "")),
                    *_mbo_metric_row(metrics.get("pnl")),
                    *_mbo_metric_row(metrics.get("rating")),
                    *_mbo_metric_row(metrics.get("google")),
                ]
            )
        _append_rows_sheet(
            wb,
            "Store",
            [
                "Store code", "Store", "Area Manager", "AM UUID",
                "P&L %", "P&L punti", "P&L teorici",
                "Rating %", "Rating punti", "Rating teorici",
                "Google %", "Google punti", "Google teorici",
            ],
            rows,
        )

    if section in {"all", "staff"}:
        rows = []
        for row in (dashboard.get("staff") or {}).get("rows") or []:
            categories = row.get("categories") or {}
            am_name, am_uuid = _report_am_for_store(str(row.get("store_code") or ""))
            if selected_am_uuid and am_uuid != selected_am_uuid:
                continue
            rows.append(
                [
                    row.get("full_name") or "",
                    row.get("role") or "",
                    row.get("store_code") or "",
                    row.get("store_name") or "",
                    am_name,
                    am_uuid,
                    *_mbo_category_row(categories.get("pnl")),
                    *_mbo_category_row(categories.get("rating")),
                    *_mbo_category_row(categories.get("google")),
                    *_mbo_category_row(categories.get("survey_soft")),
                    _pct_value(row.get("soft_pct")),
                    _pct_value(row.get("survey_pct")),
                    int(row.get("survey_count") or 0),
                    _pct_value(row.get("total_pct")),
                    _pct_value(row.get("multi_pct")),
                    round(float(row.get("reward_total") or 0), 2),
                    round(float(row.get("reward_due") or 0), 2),
                ]
            )
        _append_rows_sheet(
            wb,
            "Staff",
            [
                "Persona", "Ruolo", "Store code", "Store", "Area Manager", "AM UUID",
                "P&L %", "P&L peso", "P&L riparam.",
                "Rating %", "Rating peso", "Rating riparam.",
                "Google %", "Google peso", "Google riparam.",
                "Survey+Soft %", "Survey+Soft peso", "Survey+Soft riparam.",
                "Soft %", "Survey %", "N. survey",
                "Totale %", "Multi %", "Premio potenziale", "Premio",
            ],
            rows,
        )

    if section in {"all", "area_manager"}:
        rows = []
        store_detail_rows = []
        for row in (dashboard.get("area_manager") or {}).get("rows") or []:
            if selected_am_uuid and str(row.get("subject_key") or "") != selected_am_uuid:
                continue
            categories = row.get("categories") or {}
            rows.append(
                [
                    row.get("manager_name") or "",
                    int(row.get("store_count") or 0),
                    *_mbo_category_row(categories.get("pnl")),
                    *_mbo_category_row(categories.get("rating")),
                    *_mbo_category_row(categories.get("google")),
                    *_mbo_category_row(categories.get("survey")),
                    int((categories.get("survey") or {}).get("count") or 0),
                    _pct_value(row.get("total_pct")),
                    _pct_value(row.get("ricalcolo_pct")),
                    round(float(row.get("reward_total") or 0), 2),
                    round(float(row.get("reward_due") or 0), 2),
                ]
            )
            for store in row.get("store_rows") or []:
                store_detail_rows.append(
                    [
                        row.get("manager_name") or "",
                        store.get("store_code") or "",
                        store.get("store_name") or "",
                        ", ".join(str(x) for x in store.get("months") or []),
                    ]
                )
        _append_rows_sheet(
            wb,
            "Area Manager",
            [
                "Area Manager", "Store",
                "P&L %", "P&L peso", "P&L riparam.",
                "Rating %", "Rating peso", "Rating riparam.",
                "Google %", "Google peso", "Google riparam.",
                "Survey %", "Survey peso", "Survey riparam.", "N. survey",
                "Totale %", "Ricalcolo %", "Premio potenziale", "Premio",
            ],
            rows,
        )
        _append_rows_sheet(wb, "AM store dettaglio", ["Area Manager", "Store code", "Store", "Mesi"], store_detail_rows)

    safe_name = _safe_xlsx_name(str(report.get("report_name") or "mbo_report"), "mbo_report")
    suffix = "completo" if section == "all" else section
    return _xlsx_response(wb, f"{safe_name}_{suffix}.xlsx")


@mbo_bp.route("/audit", methods=["GET", "POST"])
def audit():
    year = int((request.values.get("year") or datetime.today().year))

    if request.method == "POST":
        action = str(request.form.get("action") or "").strip().lower()
        try:
            if action == "upload_audit":
                upload = request.files.get("audit_file")
                if not upload or not upload.filename:
                    raise ValueError("Seleziona un file CSV o ZIP.")
                content = upload.read()
                if not content:
                    raise ValueError("Il file caricato e vuoto.")
                stats = import_audit_file(
                    upload.filename,
                    content,
                    imported_by=str(session.get("email") or session.get("uid") or ""),
                )
                flash(f"Audit importati. Righe lette: {stats.get('parsed', 0)} | salvate: {stats.get('imported', 0)}.", "success")
            elif action == "save_address_mapping":
                save_audit_address_mapping(
                    address_norm=str(request.form.get("address_norm") or "").strip(),
                    address_sample=str(request.form.get("address_sample") or "").strip(),
                    store_code=str(request.form.get("store_code") or "").strip(),
                )
                flash("Associazione indirizzo-store salvata.", "success")
            elif action == "save_audit_assignment":
                save_audit_row_assignment(
                    row_uuid=str(request.form.get("row_uuid") or "").strip(),
                    store_code=str(request.form.get("store_code") or "").strip(),
                    year_month=str(request.form.get("year_month") or "").strip(),
                )
                flash("Audit riassegnato.", "success")
            elif action == "ignore_audit_row":
                ignore_audit_row(str(request.form.get("row_uuid") or "").strip())
                flash("Audit escluso dal calcolo.", "success")
            elif action == "restore_audit_row":
                restore_audit_row(str(request.form.get("row_uuid") or "").strip())
                flash("Audit ripristinato nel calcolo.", "success")
            elif action == "save_manual_value":
                save_audit_manual_value(
                    store_code=str(request.form.get("store_code") or "").strip(),
                    year_month=str(request.form.get("year_month") or "").strip(),
                    score=str(request.form.get("score") or "").strip(),
                    note=str(request.form.get("note") or "").strip() or None,
                )
                flash("Valore manuale salvato.", "success")
            elif action == "reset_manual_value":
                delete_audit_manual_value(
                    store_code=str(request.form.get("store_code") or "").strip(),
                    year_month=str(request.form.get("year_month") or "").strip(),
                )
                flash("Valore ripristinato dal caricamento.", "success")
            else:
                flash("Azione audit non riconosciuta.", "warning")
        except Exception as e:
            flash(f"Errore dati audit: {e}", "danger")
        return redirect(url_for("mbo.audit", year=year))

    try:
        audit_matrix = build_audit_month_matrix(year)
    except Exception as e:
        flash(f"Errore caricamento dati audit: {e}", "danger")
        audit_matrix = {
            "year": year,
            "month_labels": [{"month_num": i, "month_label": lbl, "year_month": f"{int(year):04d}-{int(i):02d}"} for i, lbl in enumerate(["", "Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]) if i],
            "stores": [],
            "rows": [],
            "unmatched_addresses": [],
            "missing_cells": [],
            "duplicate_cells": [],
            "duplicate_audits": [],
            "excess_audits": [],
            "ignored_audits": [],
            "store_count": 0,
            "audit_count": 0,
            "unmatched_count": 0,
            "missing_count": 0,
            "duplicate_count": 0,
            "excess_count": 0,
        }
    return render_template(
        "mbo_audit.html",
        mbo_year=year,
        mbo_year_options=list(range(datetime.today().year - 1, datetime.today().year + 3)),
        audit_matrix=audit_matrix,
    )


@mbo_bp.route("/audit/indirizzi", methods=["GET", "POST"])
def audit_address_mapping():
    year = int((request.values.get("year") or datetime.today().year))

    if request.method == "POST":
        action = str(request.form.get("action") or "").strip().lower()
        try:
            if action == "save_address_mapping":
                save_audit_address_mapping(
                    address_norm=str(request.form.get("address_norm") or "").strip(),
                    address_sample=str(request.form.get("address_sample") or "").strip(),
                    store_code=str(request.form.get("store_code") or "").strip(),
                )
                flash("Associazione address-store aggiornata.", "success")
            else:
                flash("Azione address-store non riconosciuta.", "warning")
        except Exception as e:
            flash(f"Errore associazione address-store: {e}", "danger")
        return redirect(url_for("mbo.audit_address_mapping", year=year))

    try:
        mapping_matrix = build_audit_address_mapping_matrix(year)
    except Exception as e:
        flash(f"Errore caricamento associazioni address-store: {e}", "danger")
        mapping_matrix = {
            "year": year,
            "stores": [],
            "rows": [],
            "total_count": 0,
            "manual_count": 0,
            "automatic_count": 0,
            "unmatched_count": 0,
        }
    return render_template(
        "mbo_audit_address_mapping.html",
        mbo_year=year,
        mbo_year_options=list(range(datetime.today().year - 1, datetime.today().year + 3)),
        mapping_matrix=mapping_matrix,
    )


@mbo_bp.route("/calcoli-deliveroo", methods=["GET", "POST"])
def deliveroo_scores():
    year = int((request.values.get("year") or datetime.today().year))

    if request.method == "POST":
        action = str(request.form.get("action") or "").strip().lower()
        try:
            if action == "save_deliveroo_value":
                save_deliveroo_manual_value(
                    store_code=str(request.form.get("store_code") or "").strip(),
                    year_month=str(request.form.get("year_month") or "").strip(),
                    rating=str(request.form.get("rating") or "").strip(),
                    note=str(request.form.get("note") or "").strip() or None,
                )
                flash("Valore Deliveroo salvato.", "success")
            elif action == "reset_deliveroo_value":
                delete_deliveroo_manual_value(
                    store_code=str(request.form.get("store_code") or "").strip(),
                    year_month=str(request.form.get("year_month") or "").strip(),
                )
                flash("Valore Deliveroo ripristinato.", "success")
            else:
                flash("Azione Deliveroo non riconosciuta.", "warning")
        except Exception as e:
            flash(f"Errore calcoli Deliveroo: {e}", "danger")
        return redirect(url_for("mbo.deliveroo_scores", year=year))

    try:
        deliveroo_matrix = build_mbo_deliveroo_scores(year)
    except Exception as e:
        flash(f"Errore calcoli Deliveroo MBO: {e}", "danger")
        deliveroo_matrix = {
            "year": year,
            "month_labels": [{"month_num": i, "month_label": lbl, "year_month": f"{int(year):04d}-{int(i):02d}"} for i, lbl in enumerate(["", "Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]) if i],
            "rows": [],
            "store_count": 0,
            "active_months": 0,
            "points_taken": 0.0,
            "points_theoretical": 0.0,
            "pct": 0.0,
        }
    return render_template(
        "mbo_deliveroo_scores.html",
        mbo_year=year,
        mbo_year_options=list(range(datetime.today().year - 1, datetime.today().year + 3)),
        deliveroo_matrix=deliveroo_matrix,
    )


@mbo_bp.route("/soft-skills", methods=["GET", "POST"])
def soft_skills():
    if request.method == "POST":
        action = str(request.form.get("action") or "").strip().lower()
        selected_period_uuid = str(request.form.get("period_uuid") or request.args.get("period_uuid") or "").strip()
        try:
            if action == "save_period":
                selected_period_uuid = save_soft_skill_period(
                    period_uuid=str(request.form.get("edit_period_uuid") or "").strip() or None,
                    period_label=str(request.form.get("period_label") or "").strip(),
                    start_year_month=str(request.form.get("start_year_month") or "").strip(),
                    end_year_month=str(request.form.get("end_year_month") or "").strip(),
                    is_active=("is_active" in request.form),
                    created_by=str(session.get("email") or session.get("uid") or ""),
                )
                flash("Link Soft skills salvato.", "success")
            elif action == "delete_submission":
                delete_soft_skill_submission(str(request.form.get("submission_uuid") or "").strip())
                flash("Compilazione eliminata.", "success")
            elif action == "recalculate_submissions":
                result = recalculate_soft_skill_submissions(
                    selected_period_uuid,
                    user_id=str(session.get("email") or session.get("uid") or ""),
                )
                msg = f"Ricalcolo Soft skills completato: {int(result.get('updated') or 0)} compilazioni aggiornate."
                if int(result.get("missing_pnl") or 0):
                    msg += f" {int(result.get('missing_pnl') or 0)} compilazioni hanno ancora P&L senza punti teorici."
                    flash(msg, "warning")
                else:
                    flash(msg, "success")
            elif action == "save_submission":
                submission_uuid = str(request.form.get("submission_uuid") or "").strip()
                current = get_soft_skill_submission(submission_uuid)
                if not current:
                    raise ValueError("Compilazione non trovata.")
                save_soft_skill_submission(
                    token=str(current.get("token") or "").strip(),
                    submission_uuid=submission_uuid,
                    full_name=str(request.form.get("full_name") or "").strip(),
                    role=str(request.form.get("role") or "").strip(),
                    store_code=str(request.form.get("store_code") or "").strip(),
                    store_codes=request.form.getlist("store_codes"),
                    answers_source=request.form,
                    user_id=str(session.get("email") or session.get("uid") or ""),
                )
                selected_period_uuid = str(current.get("period_uuid") or selected_period_uuid)
                flash("Compilazione aggiornata.", "success")
            else:
                flash("Azione Soft skills non riconosciuta.", "warning")
        except Exception as e:
            flash(f"Errore Soft skills: {e}", "danger")
        return redirect(url_for("mbo.soft_skills", period_uuid=selected_period_uuid))

    selected_period_uuid = str(request.args.get("period_uuid") or "").strip()
    edit_submission_uuid = str(request.args.get("edit") or "").strip()
    new_period_mode = str(request.args.get("new") or "").strip() == "1"
    try:
        dashboard = build_soft_skill_dashboard(selected_period_uuid)
        if new_period_mode:
            dashboard["selected_period_uuid"] = ""
        edit_submission = get_soft_skill_submission(edit_submission_uuid) if edit_submission_uuid else None
        if edit_submission:
            dashboard["selected_period_uuid"] = str(edit_submission.get("period_uuid") or dashboard.get("selected_period_uuid") or "")
    except Exception as e:
        flash(f"Errore caricamento Soft skills: {e}", "danger")
        dashboard = {"periods": [], "selected_period_uuid": "", "submissions": [], "stores": [], "roles": [], "questions": get_soft_skill_questions(), "submission_count": 0, "avg_score": None}
        edit_submission = None
    return render_template("mbo_soft_skills.html", soft_skills=dashboard, edit_submission=edit_submission, new_period_mode=new_period_mode)


@mbo_bp.route("/soft-skills/compila/<token>", methods=["GET", "POST"])
def soft_skills_fill(token: str):
    period = get_soft_skill_period_by_token(token, require_active=False)
    if not period:
        flash("Link Soft skills non valido.", "danger")
        return redirect(url_for("mbo.home"))
    if not period.get("is_active"):
        flash("Questo link Soft skills non e attivo.", "warning")
        return redirect(url_for("mbo.home"))

    if request.method == "POST":
        try:
            save_soft_skill_submission(
                token=str(token or "").strip(),
                submission_uuid=None,
                full_name=str(request.form.get("full_name") or "").strip(),
                role=str(request.form.get("role") or "").strip(),
                store_code=str(request.form.get("store_code") or "").strip(),
                store_codes=request.form.getlist("store_codes"),
                answers_source=request.form,
                user_id=str(session.get("email") or session.get("uid") or ""),
            )
            flash("Compilazione Soft skills salvata.", "success")
            return redirect(url_for("mbo.soft_skills_fill", token=token))
        except Exception as e:
            flash(f"Errore compilazione Soft skills: {e}", "danger")

    dashboard = build_soft_skill_dashboard(str(period.get("period_uuid") or ""))
    return render_template(
        "mbo_soft_skills_fill.html",
        period=period,
        stores=dashboard.get("stores") or [],
        roles=dashboard.get("roles") or [],
        questions=dashboard.get("questions") or get_soft_skill_questions(),
        form_values=request.form if request.method == "POST" else {},
        selected_extra_store_codes=request.form.getlist("store_codes") if request.method == "POST" else [],
    )


@mbo_bp.route("/survey", methods=["GET", "POST"])
def custom_surveys():
    selected_survey_uuid = str(request.values.get("survey_uuid") or "").strip()
    if request.method == "POST":
        action = str(request.form.get("action") or "").strip().lower()
        try:
            if action == "create_survey":
                upload = request.files.get("questions_file")
                if not upload or not upload.filename:
                    raise ValueError("Carica un file Excel o CSV con le domande.")
                questions = parse_custom_survey_questions_file(upload.filename, upload.read() or b"")
                selected_survey_uuid = create_custom_survey(
                    survey_name=str(request.form.get("survey_name") or "").strip(),
                    display_name=str(request.form.get("display_name") or "").strip(),
                    target_type=str(request.form.get("target_type") or "").strip(),
                    start_year_month=str(request.form.get("start_year_month") or "").strip(),
                    end_year_month=str(request.form.get("end_year_month") or "").strip(),
                    is_active=("is_active" in request.form),
                    questions=questions,
                    created_by=str(session.get("email") or session.get("uid") or ""),
                )
                flash("Survey creata e link generati.", "success")
            elif action == "toggle_survey":
                selected_survey_uuid = str(request.form.get("survey_uuid") or "").strip()
                set_custom_survey_active(selected_survey_uuid, "is_active" in request.form)
                flash("Stato survey aggiornato.", "success")
            elif action == "update_survey_names":
                selected_survey_uuid = str(request.form.get("survey_uuid") or "").strip()
                update_custom_survey_names(
                    selected_survey_uuid,
                    survey_name=str(request.form.get("survey_name") or "").strip(),
                    display_name=str(request.form.get("display_name") or "").strip(),
                )
                flash("Nomi survey aggiornati.", "success")
            elif action == "delete_submission":
                selected_survey_uuid = str(request.form.get("survey_uuid") or "").strip()
                delete_custom_survey_submission(str(request.form.get("submission_uuid") or "").strip())
                flash("Compilazione eliminata.", "success")
            else:
                flash("Azione survey non riconosciuta.", "warning")
        except Exception as e:
            flash(f"Errore survey: {e}", "danger")
        return redirect(url_for("mbo.custom_surveys", survey_uuid=selected_survey_uuid))

    try:
        dashboard = build_custom_survey_dashboard(selected_survey_uuid)
    except Exception as e:
        flash(f"Errore caricamento survey: {e}", "danger")
        dashboard = {"surveys": [], "links": [], "submissions": [], "aggregates": [], "category_rows": [], "submission_count": 0, "avg_pct": None}
    return render_template("mbo_custom_surveys.html", survey=dashboard)


@mbo_bp.route("/survey/link/<token>", methods=["GET", "POST"])
def custom_survey_fill(token: str):
    survey = get_custom_survey(token=token)
    if not survey:
        flash("Link survey non valido.", "danger")
        return render_template("mbo_custom_survey_fill.html", survey=None, form_values={}), 404
    if not survey.get("is_active") or not survey.get("link_active"):
        flash("Questa survey non e attiva.", "warning")
        return render_template("mbo_custom_survey_fill.html", survey=survey, form_values={}), 403
    if request.method == "POST":
        try:
            save_custom_survey_submission(
                token=token,
                respondent_name="",
                answers_source=request.form,
            )
            flash("Survey inviata correttamente.", "success")
            return redirect(url_for("mbo.custom_survey_fill", token=token))
        except Exception as e:
            flash(f"Errore salvataggio survey: {e}", "danger")
    return render_template("mbo_custom_survey_fill.html", survey=survey, form_values=request.form if request.method == "POST" else {})


@mbo_bp.route("/survey/<survey_uuid>/links.xlsx")
def custom_survey_links_xlsx(survey_uuid: str):
    if not _ensure_mbo_access():
        return redirect(url_for("login", next=request.full_path if request.query_string else request.path))
    dashboard = build_custom_survey_dashboard(survey_uuid)
    selected = dashboard.get("selected_survey") or {}
    selected_am_uuid = _am_export_filter()
    target_type = str(selected.get("target_type") or "")
    try:
        store_am_map = survey_store_area_manager_map(
            str(selected.get("start_year_month") or ""),
            str(selected.get("end_year_month") or ""),
        ) if target_type == "store_manager" else {}
    except Exception:
        store_am_map = {}

    def _link_matches_am(link: dict) -> bool:
        if not selected_am_uuid:
            return True
        if target_type == "area_manager":
            return str(link.get("target_code") or "") == selected_am_uuid
        am = store_am_map.get(str(link.get("target_code") or "").strip()) or {}
        return str(am.get("area_manager_uuid") or "") == selected_am_uuid

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except Exception as e:
        flash(f"Export Excel non disponibile: {e}", "danger")
        return redirect(url_for("mbo.custom_surveys", survey_uuid=survey_uuid))

    wb = Workbook()
    ws = wb.active
    ws.title = "Link survey"
    target_label = "Area manager" if str(selected.get("target_type") or "") == "area_manager" else "Store"
    ws.append([target_label, "Link"])
    ws[1][0].font = Font(bold=True)
    ws[1][1].font = Font(bold=True)
    for link in dashboard.get("links") or []:
        if not _link_matches_am(link):
            continue
        ws.append([
            str(link.get("target_name") or link.get("target_code") or ""),
            url_for("mbo.custom_survey_fill", token=link.get("token"), _external=True, _scheme="https"),
        ])
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 110
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    safe_name = "".join(ch if ch.isalnum() else "_" for ch in str(selected.get("survey_name") or "survey")).strip("_") or "survey"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name=f"{safe_name}_link.xlsx",
        as_attachment=True,
    )


@mbo_bp.route("/survey/<survey_uuid>/export.xlsx")
def custom_survey_export_xlsx(survey_uuid: str):
    if not _ensure_mbo_access():
        return redirect(url_for("login", next=request.full_path if request.query_string else request.path))
    try:
        from openpyxl import Workbook
    except Exception as e:
        flash(f"Export Excel non disponibile: {e}", "danger")
        return redirect(url_for("mbo.custom_surveys", survey_uuid=survey_uuid))

    dashboard = build_custom_survey_dashboard(survey_uuid)
    selected = dashboard.get("selected_survey") or {}
    target_type = str(selected.get("target_type") or "")
    selected_am_uuid = _am_export_filter()
    try:
        store_am_map = survey_store_area_manager_map(
            str(selected.get("start_year_month") or ""),
            str(selected.get("end_year_month") or ""),
        ) if target_type == "store_manager" else {}
    except Exception:
        store_am_map = {}

    def _survey_am_for_store(store_code: str) -> tuple[str, str]:
        am = store_am_map.get(str(store_code or "").strip()) or {}
        return (
            str(am.get("area_manager_name") or "Senza Area Manager"),
            str(am.get("area_manager_uuid") or ""),
        )

    def _survey_row_matches_am(code: str) -> bool:
        if not selected_am_uuid:
            return True
        if target_type == "area_manager":
            return str(code or "") == selected_am_uuid
        return _survey_am_for_store(code)[1] == selected_am_uuid

    wb = Workbook()
    wb.remove(wb.active)

    _append_rows_sheet(
        wb,
        "Panoramica",
        ["Voce", "Valore"],
        [
            ["Survey", selected.get("survey_name") or ""],
            ["Titolo compilazione", selected.get("display_name") or ""],
            ["Target", "Area Manager" if str(selected.get("target_type") or "") == "area_manager" else "Store Manager"],
            ["Periodo", f"{selected.get('start_year_month') or ''} / {selected.get('end_year_month') or ''}"],
            ["Attiva", "Si" if selected.get("is_active") else "No"],
            ["Compilazioni", dashboard.get("submission_count") or 0],
            ["Media %", _pct_value(dashboard.get("avg_pct"))],
            ["Link generati", len(dashboard.get("links") or [])],
        ],
    )

    _append_rows_sheet(
        wb,
        "Link",
        ["Target", "Codice", "Area Manager", "AM UUID", "Tipo", "Attivo", "Link"],
        [
            [
                link.get("target_name") or "",
                link.get("target_code") or "",
                _survey_am_for_store(str(link.get("target_code") or ""))[0] if target_type == "store_manager" else "",
                _survey_am_for_store(str(link.get("target_code") or ""))[1] if target_type == "store_manager" else "",
                link.get("target_type") or "",
                "Si" if link.get("is_active") else "No",
                url_for("mbo.custom_survey_fill", token=link.get("token"), _external=True, _scheme="https"),
            ]
            for link in (dashboard.get("links") or [])
            if _survey_row_matches_am(str(link.get("target_code") or ""))
        ],
    )

    _append_rows_sheet(
        wb,
        "Aggregati",
        ["Target", "Codice", "Area Manager", "AM UUID", "Compilazioni", "Media %"],
        [
            [
                row.get("target_name") or "",
                row.get("target_code") or "",
                _survey_am_for_store(str(row.get("target_code") or ""))[0] if target_type == "store_manager" else "",
                _survey_am_for_store(str(row.get("target_code") or ""))[1] if target_type == "store_manager" else "",
                row.get("count") or 0,
                _pct_value(row.get("avg_pct")),
            ]
            for row in (dashboard.get("aggregates") or [])
            if _survey_row_matches_am(str(row.get("target_code") or ""))
        ],
    )

    _append_rows_sheet(
        wb,
        "Categorie",
        ["Categoria", "Risposte", "Media %"],
        [
            [row.get("category") or "", row.get("count") or 0, _pct_value(row.get("avg_pct"))]
            for row in dashboard.get("category_rows") or []
        ],
    )

    submission_rows = []
    answer_rows = []
    for row in dashboard.get("aggregates") or []:
        if not _survey_row_matches_am(str(row.get("target_code") or "")):
            continue
        am_name, am_uuid = _survey_am_for_store(str(row.get("target_code") or "")) if target_type == "store_manager" else ("", "")
        for sub in row.get("submissions") or []:
            submission_rows.append(
                [
                    row.get("target_name") or sub.get("target_name") or "",
                    row.get("target_code") or sub.get("target_code") or "",
                    am_name,
                    am_uuid,
                    sub.get("created_at") or "",
                    sub.get("respondent_name") or "",
                    _pct_value(sub.get("score_pct")),
                ]
            )
            for ans in sub.get("answer_rows") or []:
                answer_rows.append(
                    [
                        row.get("target_name") or sub.get("target_name") or "",
                        row.get("target_code") or sub.get("target_code") or "",
                        am_name,
                        am_uuid,
                        sub.get("created_at") or "",
                        ans.get("category") or "",
                        ans.get("question") or "",
                        ans.get("answer") or "",
                    ]
                )
    _append_rows_sheet(wb, "Compilazioni", ["Target", "Codice", "Area Manager", "AM UUID", "Data", "Compilatore", "Score %"], submission_rows)
    _append_rows_sheet(wb, "Risposte", ["Target", "Codice", "Area Manager", "AM UUID", "Data", "Categoria", "Domanda", "Risposta"], answer_rows)

    overview = dashboard.get("overview") or {}
    overview_rows = []
    for section_key, section_label in [
        ("area_manager_results", "Area Manager"),
        ("store_manager_results", "Store Manager"),
        ("store_manager_by_area_manager", "Store Manager per Area Manager"),
    ]:
        for row in overview.get(section_key) or []:
            row_code = str(row.get("target_code") or row.get("area_manager_uuid") or "")
            if section_key == "store_manager_by_area_manager":
                row_code = str(row.get("area_manager_uuid") or "")
            if selected_am_uuid and row_code != selected_am_uuid:
                continue
            overview_rows.append(
                [
                    section_label,
                    row.get("target_name") or row.get("area_manager_name") or "",
                    row.get("count") or 0,
                    _pct_value(row.get("avg_pct")),
                    "; ".join(f"{x.get('category')} {_pct_value(x.get('avg_pct'))}%" for x in (row.get("strengths") or [])[:3]),
                    "; ".join(f"{x.get('category')} {_pct_value(x.get('avg_pct'))}%" for x in (row.get("improvements") or [])[:3]),
                    len(row.get("text_answers") or []),
                ]
            )
    _append_rows_sheet(
        wb,
        "Panoramica risultati",
        ["Sezione", "Nome", "Compilazioni", "Media %", "Punti di forza", "Aree miglioramento", "Commenti"],
        overview_rows,
    )

    if target_type == "store_manager":
        _append_rows_sheet(
            wb,
            "Store Manager per AM",
            ["Area Manager", "AM UUID", "Store", "Media %", "Store collegati", "Commenti"],
            [
                [
                    row.get("area_manager_name") or "",
                    row.get("area_manager_uuid") or "",
                    row.get("count") or 0,
                    _pct_value(row.get("avg_pct")),
                    "; ".join(str(store.get("target_name") or "") for store in (row.get("stores") or [])),
                    len(row.get("text_answers") or []),
                ]
                for row in (overview.get("store_manager_by_area_manager") or [])
                if (not selected_am_uuid or str(row.get("area_manager_uuid") or "") == selected_am_uuid)
            ],
        )

    safe_name = _safe_xlsx_name(str(selected.get("survey_name") or "survey"), "survey")
    return _xlsx_response(wb, f"{safe_name}_risultati.xlsx")


@mbo_bp.route("/calcoli-glovo", methods=["GET", "POST"])
def glovo_scores():
    year = int((request.values.get("year") or datetime.today().year))

    if request.method == "POST":
        action = str(request.form.get("action") or "").strip().lower()
        try:
            if action == "save_glovo_value":
                save_glovo_manual_value(
                    store_code=str(request.form.get("store_code") or "").strip(),
                    year_month=str(request.form.get("year_month") or "").strip(),
                    value_pct=str(request.form.get("value_pct") or "").strip(),
                    note=str(request.form.get("note") or "").strip() or None,
                )
                flash("Valore Glovo salvato.", "success")
            elif action == "reset_glovo_value":
                delete_glovo_manual_value(
                    store_code=str(request.form.get("store_code") or "").strip(),
                    year_month=str(request.form.get("year_month") or "").strip(),
                )
                flash("Valore Glovo ripristinato.", "success")
            else:
                flash("Azione Glovo non riconosciuta.", "warning")
        except Exception as e:
            flash(f"Errore calcoli Glovo: {e}", "danger")
        return redirect(url_for("mbo.glovo_scores", year=year))

    try:
        glovo_matrix = build_mbo_glovo_scores(year)
    except Exception as e:
        flash(f"Errore calcoli Glovo MBO: {e}", "danger")
        glovo_matrix = {
            "year": year,
            "month_labels": [{"month_num": i, "month_label": lbl, "year_month": f"{int(year):04d}-{int(i):02d}"} for i, lbl in enumerate(["", "Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]) if i],
            "rows": [],
            "store_count": 0,
            "active_months": 0,
            "points_taken": 0.0,
            "points_theoretical": 0.0,
            "pct": 0.0,
        }
    return render_template(
        "mbo_glovo_scores.html",
        mbo_year=year,
        mbo_year_options=list(range(datetime.today().year - 1, datetime.today().year + 3)),
        glovo_matrix=glovo_matrix,
    )


@mbo_bp.route("/calcoli-google", methods=["GET", "POST"])
def google_scores():
    year = int((request.values.get("year") or datetime.today().year))

    if request.method == "POST":
        action = str(request.form.get("action") or "").strip().lower()
        try:
            if action == "save_google_value":
                save_google_manual_value(
                    store_code=str(request.form.get("store_code") or "").strip(),
                    year_month=str(request.form.get("year_month") or "").strip(),
                    rating=str(request.form.get("rating") or "").strip(),
                    note=str(request.form.get("note") or "").strip() or None,
                )
                flash("Valore Google salvato.", "success")
            elif action == "reset_google_value":
                delete_google_manual_value(
                    store_code=str(request.form.get("store_code") or "").strip(),
                    year_month=str(request.form.get("year_month") or "").strip(),
                )
                flash("Valore Google ripristinato.", "success")
            else:
                flash("Azione Google non riconosciuta.", "warning")
        except Exception as e:
            flash(f"Errore calcoli Google: {e}", "danger")
        return redirect(url_for("mbo.google_scores", year=year))

    try:
        google_matrix = build_mbo_google_scores(year)
    except Exception as e:
        flash(f"Errore calcoli Google MBO: {e}", "danger")
        google_matrix = {
            "year": year,
            "month_labels": [{"month_num": i, "month_label": lbl, "year_month": f"{int(year):04d}-{int(i):02d}"} for i, lbl in enumerate(["", "Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]) if i],
            "rows": [],
            "store_count": 0,
            "active_months": 0,
            "points_taken": 0.0,
            "points_theoretical": 0.0,
            "pct": 0.0,
            "default_count": 0,
        }
    return render_template(
        "mbo_google_scores.html",
        mbo_year=year,
        mbo_year_options=list(range(datetime.today().year - 1, datetime.today().year + 3)),
        google_matrix=google_matrix,
    )


@mbo_bp.route("/calcoli-pl")
def pnl_scores():
    year = int((request.args.get("year") or datetime.today().year))
    try:
        pnl_matrix = build_mbo_pnl_scores(year)
    except Exception as e:
        flash(f"Errore calcoli P&L MBO: {e}", "danger")
        pnl_matrix = {
            "year": year,
            "month_labels": [{"month_num": i, "month_label": lbl, "year_month": f"{int(year):04d}-{int(i):02d}"} for i, lbl in enumerate(["", "Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]) if i],
            "rows": [],
            "store_count": 0,
            "active_months": 0,
            "points_taken": 0.0,
            "points_theoretical": 0.0,
            "pct": 0.0,
            "voices": ["REVENUES", "COGS", "LABOUR COST", "EBITDA"],
        }
    return render_template(
        "mbo_pnl_scores.html",
        mbo_year=year,
        mbo_year_options=list(range(datetime.today().year - 1, datetime.today().year + 3)),
        pnl_matrix=pnl_matrix,
    )


@mbo_bp.route("/impostazioni", methods=["GET", "POST"])
def settings():
    if not _ensure_admin_access():
        flash("La pagina impostazioni MBO Ã¨ riservata agli admin.", "warning")
        return redirect(url_for("mbo.home"))

    year = int((request.values.get("year") or datetime.today().year))

    if request.method == "POST":
        action = str(request.form.get("action") or "save_matrix").strip().lower()
        try:
            if action == "sync_managers":
                created = sync_area_managers_from_ilp()
                flash(f"Sincronizzazione completata. Nuovi area manager importati: {created}.", "success")
            elif action == "save_manager":
                save_area_manager(
                    row_uuid=(request.form.get("row_uuid") or "").strip() or None,
                    manager_name=(request.form.get("manager_name") or "").strip(),
                    ilp_am_value=(request.form.get("ilp_am_value") or "").strip() or None,
                    sort_order=int((request.form.get("sort_order") or "0").strip() or "0"),
                    is_active=("is_active" in request.form),
                )
                flash("Area manager salvato.", "success")
            elif action == "delete_manager":
                delete_area_manager((request.form.get("row_uuid") or "").strip())
                flash("Area manager eliminato.", "success")
            elif action == "save_matrix":
                assignments = {}
                for key, value in request.form.items():
                    if not key.startswith("assign__"):
                        continue
                    _, store_code, year_month = key.split("__", 2)
                    assignments[(store_code, year_month)] = str(value or "").strip()
                stats = save_store_month_assignments(year, assignments)
                flash(
                    f"Associazioni salvate. Aggiornate: {stats.get('updated', 0)} | eliminate override: {stats.get('deleted', 0)}.",
                    "success",
                )
            elif action == "save_am_monthly":
                stats = save_area_manager_monthly_settings(year, request.form)
                flash(f"Configurazione mensile Area Manager salvata: {stats.get('saved', 0)} mesi.", "success")
            elif action == "propagate_am_monthly":
                stats = propagate_area_manager_monthly_setting(
                    area_manager_uuid=str(request.form.get("area_manager_row_uuid") or "").strip(),
                    from_year_month=str(request.form.get("year_month") or "").strip(),
                    is_active=("is_active" in request.form),
                    reward_eur=str(request.form.get("reward_eur") or "").strip(),
                )
                flash(f"Valore Area Manager propagato sui mesi successivi: {stats.get('saved', 0)}.", "success")
            elif action == "save_role_rewards":
                stats = save_role_monthly_rewards(year, request.form)
                flash(f"Premi mensili ruoli salvati: {stats.get('saved', 0)} mesi.", "success")
            elif action == "propagate_role_reward":
                stats = propagate_role_monthly_reward(
                    role_name=str(request.form.get("role_name") or "").strip(),
                    from_year_month=str(request.form.get("year_month") or "").strip(),
                    reward_eur=str(request.form.get("reward_eur") or "").strip(),
                )
                flash(f"Premio ruolo propagato sui mesi successivi: {stats.get('saved', 0)}.", "success")
            elif action == "save_reward_weights":
                stats = save_reward_weights(request.form)
                flash(f"Pesi premio salvati: {stats.get('saved', 0)} valori.", "success")
            else:
                flash("Azione non riconosciuta.", "warning")
        except Exception as e:
            flash(f"Errore impostazioni MBO: {e}", "danger")
        return redirect(url_for("mbo.settings", year=year))

    try:
        sync_area_managers_from_ilp()
    except Exception:
        log_swallowed('mbo:1280')

    try:
        matrix = build_mbo_store_month_matrix(year)
    except Exception as e:
        flash(f"Errore caricamento matrice MBO: {e}", "danger")
        matrix = {
            "store_count": 0,
            "default_count": 0,
            "override_count": 0,
            "unmatched_count": 0,
            "month_labels": [{"month_num": i, "month_label": lbl, "year_month": f"{int(year):04d}-{int(i):02d}"} for i, lbl in enumerate(["", "Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]) if i],
            "rows": [],
        }
    managers = list_area_managers(include_inactive=True)
    try:
        reward_settings = load_mbo_reward_settings(year)
    except Exception as e:
        flash(f"Errore caricamento impostazioni premi: {e}", "danger")
        reward_settings = {"month_labels": matrix.get("month_labels", []), "area_managers": [], "roles": [], "weights": []}
    return render_template(
        "mbo_settings.html",
        mbo_year=year,
        mbo_year_options=list(range(datetime.today().year - 1, datetime.today().year + 3)),
        mbo_matrix=matrix,
        area_managers=managers,
        reward_settings=reward_settings,
    )

