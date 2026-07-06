from __future__ import annotations

from app_logging import log_swallowed
import csv
import hashlib
import io
import json
import os
import uuid
import zipfile
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from calendar import monthrange
from datetime import date
from datetime import datetime
from datetime import timedelta
from typing import Any, Dict, List

from app_db import get_connection_sqlserver_database, get_storehub_database_name
from app_db import get_connection_database_new, get_connection_ilp
from controlli_repository import _detect_pl_layout
from db_integration import get_warehouse_stores
from store_registry_repository import list_store_registry


def _conn(database: str | None = None, read_only: bool = False):
    return get_connection_sqlserver_database(database or get_storehub_database_name(), read_only=read_only)


def _norm(v: Any) -> str:
    return str(v or "").strip().lower()


def _norm_key(v: Any) -> str:
    return " ".join(str(v or "").strip().lower().split())


def _norm_address(v: Any) -> str:
    return _norm_key(v)[:450]


def _month_key(dt: date | datetime | None) -> str:
    if not dt:
        return ""
    return f"{int(dt.year):04d}-{int(dt.month):02d}"


def _month_end(year_month: str) -> date | None:
    try:
        year_s, month_s = str(year_month or "").split("-", 1)
        year_i = int(year_s)
        month_i = int(month_s)
        return date(year_i, month_i, monthrange(year_i, month_i)[1])
    except Exception:
        return None


def _parse_date(value: Any) -> date | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    if raw.endswith("Z"):
        raw = raw[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(raw).date()
    except Exception:
        log_swallowed('mbo_repository:66')
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw[:10], fmt).date()
        except Exception:
            continue
    return None


def _parse_decimal(value: Any) -> Decimal | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    raw = raw.replace("%", "").replace(" ", "")
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    else:
        raw = raw.replace(",", ".")
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def ensure_mbo_schema() -> None:
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
IF OBJECT_ID('dbo.MboAreaManagers','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboAreaManagers (
    row_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    manager_name NVARCHAR(255) NOT NULL,
    ilp_am_value NVARCHAR(255) NULL,
    sort_order INT NOT NULL DEFAULT 0,
    is_active BIT NOT NULL DEFAULT 1,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboAreaManagers_manager_name ON dbo.MboAreaManagers(manager_name);
  CREATE UNIQUE INDEX UX_MboAreaManagers_ilp_am_value ON dbo.MboAreaManagers(ilp_am_value) WHERE ilp_am_value IS NOT NULL;
END

IF OBJECT_ID('dbo.MboStoreAreaManagerMonthly','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboStoreAreaManagerMonthly (
    row_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    store_code NVARCHAR(50) NOT NULL,
    year_month CHAR(7) NOT NULL,
    area_manager_row_uuid UNIQUEIDENTIFIER NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboStoreAreaManagerMonthly_store_month
    ON dbo.MboStoreAreaManagerMonthly(store_code, year_month);
  CREATE INDEX IX_MboStoreAreaManagerMonthly_month
    ON dbo.MboStoreAreaManagerMonthly(year_month, store_code);
END

IF OBJECT_ID('dbo.MboAuditImports','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboAuditImports (
    import_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    source_filename NVARCHAR(255) NULL,
    imported_by NVARCHAR(255) NULL,
    rows_total INT NOT NULL DEFAULT 0,
    rows_imported INT NOT NULL DEFAULT 0,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
END

IF OBJECT_ID('dbo.MboAuditRows','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboAuditRows (
    row_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    import_uuid UNIQUEIDENTIFIER NULL,
    source_key NVARCHAR(80) NOT NULL,
    mission_data_id NVARCHAR(80) NULL,
    mission_id NVARCHAR(80) NULL,
    mission_title NVARCHAR(255) NULL,
    address NVARCHAR(1000) NOT NULL,
    address_norm NVARCHAR(450) NOT NULL,
    audit_date DATE NOT NULL,
    audit_year_month CHAR(7) NOT NULL,
    score DECIMAL(9,4) NULL,
    matched_store_code NVARCHAR(50) NULL,
    assigned_store_code NVARCHAR(50) NULL,
    assigned_year_month CHAR(7) NULL,
    ignored BIT NOT NULL DEFAULT 0,
    username NVARCHAR(255) NULL,
    user_full_name NVARCHAR(255) NULL,
    raw_json NVARCHAR(MAX) NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboAuditRows_source_key ON dbo.MboAuditRows(source_key);
  CREATE INDEX IX_MboAuditRows_store_month ON dbo.MboAuditRows(assigned_store_code, assigned_year_month);
  CREATE INDEX IX_MboAuditRows_audit_month ON dbo.MboAuditRows(audit_year_month);
END

IF OBJECT_ID('dbo.MboAuditRows','U') IS NOT NULL
   AND COL_LENGTH('dbo.MboAuditRows', 'ignored') IS NULL
BEGIN
  ALTER TABLE dbo.MboAuditRows ADD ignored BIT NOT NULL CONSTRAINT DF_MboAuditRows_ignored DEFAULT 0;
END

IF OBJECT_ID('dbo.MboAuditAddressStoreMap','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboAuditAddressStoreMap (
    address_norm NVARCHAR(450) NOT NULL PRIMARY KEY,
    address_sample NVARCHAR(1000) NOT NULL,
    store_code NVARCHAR(50) NOT NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE INDEX IX_MboAuditAddressStoreMap_store ON dbo.MboAuditAddressStoreMap(store_code);
END

IF OBJECT_ID('dbo.MboAuditManualMonthlyValues','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboAuditManualMonthlyValues (
    row_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    store_code NVARCHAR(50) NOT NULL,
    year_month CHAR(7) NOT NULL,
    score DECIMAL(9,4) NOT NULL DEFAULT 0,
    note NVARCHAR(500) NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboAuditManualMonthlyValues_store_month
    ON dbo.MboAuditManualMonthlyValues(store_code, year_month);
END

IF OBJECT_ID('dbo.MboGoogleManualMonthlyValues','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboGoogleManualMonthlyValues (
    row_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    store_code NVARCHAR(50) NOT NULL,
    year_month CHAR(7) NOT NULL,
    rating DECIMAL(9,4) NULL,
    note NVARCHAR(500) NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboGoogleManualMonthlyValues_store_month
    ON dbo.MboGoogleManualMonthlyValues(store_code, year_month);
END

IF OBJECT_ID('dbo.MboGlovoManualMonthlyValues','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboGlovoManualMonthlyValues (
    row_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    store_code NVARCHAR(50) NOT NULL,
    year_month CHAR(7) NOT NULL,
    value_pct DECIMAL(9,4) NULL,
    note NVARCHAR(500) NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboGlovoManualMonthlyValues_store_month
    ON dbo.MboGlovoManualMonthlyValues(store_code, year_month);
END

IF OBJECT_ID('dbo.MboDeliverooManualMonthlyValues','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboDeliverooManualMonthlyValues (
    row_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    store_code NVARCHAR(50) NOT NULL,
    year_month CHAR(7) NOT NULL,
    rating DECIMAL(9,4) NULL,
    note NVARCHAR(500) NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboDeliverooManualMonthlyValues_store_month
    ON dbo.MboDeliverooManualMonthlyValues(store_code, year_month);
END

IF OBJECT_ID('dbo.MboSoftSkillPeriods','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboSoftSkillPeriods (
    period_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    token NVARCHAR(80) NOT NULL,
    period_label NVARCHAR(255) NOT NULL,
    start_year_month CHAR(7) NOT NULL,
    end_year_month CHAR(7) NOT NULL,
    is_active BIT NOT NULL DEFAULT 1,
    created_by NVARCHAR(255) NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboSoftSkillPeriods_token ON dbo.MboSoftSkillPeriods(token);
  CREATE INDEX IX_MboSoftSkillPeriods_months ON dbo.MboSoftSkillPeriods(start_year_month, end_year_month);
END

IF OBJECT_ID('dbo.MboSoftSkillSubmissions','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboSoftSkillSubmissions (
    submission_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    group_uuid UNIQUEIDENTIFIER NULL,
    period_uuid UNIQUEIDENTIFIER NOT NULL,
    full_name NVARCHAR(255) NOT NULL,
    role NVARCHAR(50) NOT NULL,
    store_code NVARCHAR(50) NOT NULL,
    answers_json NVARCHAR(MAX) NOT NULL,
    computed_json NVARCHAR(MAX) NULL,
    final_score DECIMAL(9,4) NULL,
    created_by NVARCHAR(255) NULL,
    updated_by NVARCHAR(255) NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    deleted_at DATETIME2 NULL
  );
  CREATE INDEX IX_MboSoftSkillSubmissions_period ON dbo.MboSoftSkillSubmissions(period_uuid, deleted_at);
  CREATE INDEX IX_MboSoftSkillSubmissions_store ON dbo.MboSoftSkillSubmissions(store_code);
  CREATE INDEX IX_MboSoftSkillSubmissions_group ON dbo.MboSoftSkillSubmissions(group_uuid);
END

IF OBJECT_ID('dbo.MboSoftSkillSubmissions','U') IS NOT NULL
   AND COL_LENGTH('dbo.MboSoftSkillSubmissions', 'group_uuid') IS NULL
BEGIN
  ALTER TABLE dbo.MboSoftSkillSubmissions ADD group_uuid UNIQUEIDENTIFIER NULL;
  CREATE INDEX IX_MboSoftSkillSubmissions_group ON dbo.MboSoftSkillSubmissions(group_uuid);
END
"""
        )
        cur.execute(
            """
IF OBJECT_ID('dbo.MboSurveys','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboSurveys (
    survey_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    survey_name NVARCHAR(255) NOT NULL,
    display_name NVARCHAR(255) NULL,
    target_type NVARCHAR(30) NOT NULL,
    start_year_month CHAR(7) NOT NULL,
    end_year_month CHAR(7) NOT NULL,
    is_active BIT NOT NULL DEFAULT 1,
    created_by NVARCHAR(255) NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE INDEX IX_MboSurveys_active ON dbo.MboSurveys(is_active, created_at);
END

IF OBJECT_ID('dbo.MboSurveys','U') IS NOT NULL
   AND COL_LENGTH('dbo.MboSurveys', 'display_name') IS NULL
BEGIN
  ALTER TABLE dbo.MboSurveys ADD display_name NVARCHAR(255) NULL;
END

IF OBJECT_ID('dbo.MboSurveyQuestions','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboSurveyQuestions (
    question_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    survey_uuid UNIQUEIDENTIFIER NOT NULL,
    question_text NVARCHAR(1000) NOT NULL,
    question_type NVARCHAR(30) NOT NULL,
    category NVARCHAR(120) NULL,
    max_score DECIMAL(9,4) NOT NULL DEFAULT 0,
    sort_order INT NOT NULL DEFAULT 0,
    is_required BIT NOT NULL DEFAULT 1,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE INDEX IX_MboSurveyQuestions_survey ON dbo.MboSurveyQuestions(survey_uuid, sort_order);
END

IF OBJECT_ID('dbo.MboSurveyQuestionOptions','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboSurveyQuestionOptions (
    option_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    question_uuid UNIQUEIDENTIFIER NOT NULL,
    option_text NVARCHAR(500) NOT NULL,
    score DECIMAL(9,4) NOT NULL DEFAULT 0,
    sort_order INT NOT NULL DEFAULT 0
  );
  CREATE INDEX IX_MboSurveyQuestionOptions_question ON dbo.MboSurveyQuestionOptions(question_uuid, sort_order);
END

IF OBJECT_ID('dbo.MboSurveyLinks','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboSurveyLinks (
    link_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    survey_uuid UNIQUEIDENTIFIER NOT NULL,
    token NVARCHAR(80) NOT NULL,
    target_type NVARCHAR(30) NOT NULL,
    target_code NVARCHAR(80) NOT NULL,
    target_name NVARCHAR(255) NOT NULL,
    is_active BIT NOT NULL DEFAULT 1,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboSurveyLinks_token ON dbo.MboSurveyLinks(token);
  CREATE UNIQUE INDEX UX_MboSurveyLinks_target ON dbo.MboSurveyLinks(survey_uuid, target_type, target_code);
END

IF OBJECT_ID('dbo.MboSurveySubmissions','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboSurveySubmissions (
    submission_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    survey_uuid UNIQUEIDENTIFIER NOT NULL,
    link_uuid UNIQUEIDENTIFIER NOT NULL,
    target_type NVARCHAR(30) NOT NULL,
    target_code NVARCHAR(80) NOT NULL,
    target_name NVARCHAR(255) NOT NULL,
    respondent_name NVARCHAR(255) NULL,
    answers_json NVARCHAR(MAX) NOT NULL,
    computed_json NVARCHAR(MAX) NULL,
    score_pct DECIMAL(9,4) NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    deleted_at DATETIME2 NULL
  );
  CREATE INDEX IX_MboSurveySubmissions_survey ON dbo.MboSurveySubmissions(survey_uuid, deleted_at);
  CREATE INDEX IX_MboSurveySubmissions_target ON dbo.MboSurveySubmissions(survey_uuid, target_type, target_code);
END
IF OBJECT_ID('dbo.MboAreaManagerMonthlySettings','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboAreaManagerMonthlySettings (
    row_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    area_manager_row_uuid UNIQUEIDENTIFIER NOT NULL,
    year_month CHAR(7) NOT NULL,
    is_active BIT NOT NULL DEFAULT 1,
    reward_eur DECIMAL(12,2) NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboAreaManagerMonthlySettings_am_month
    ON dbo.MboAreaManagerMonthlySettings(area_manager_row_uuid, year_month);
END

IF OBJECT_ID('dbo.MboRoleMonthlyRewards','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboRoleMonthlyRewards (
    row_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    role_name NVARCHAR(50) NOT NULL,
    year_month CHAR(7) NOT NULL,
    reward_eur DECIMAL(12,2) NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboRoleMonthlyRewards_role_month
    ON dbo.MboRoleMonthlyRewards(role_name, year_month);
END

IF OBJECT_ID('dbo.MboRewardWeights','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboRewardWeights (
    row_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    target_profile NVARCHAR(50) NOT NULL,
    category_key NVARCHAR(50) NOT NULL,
    weight_pct DECIMAL(9,4) NOT NULL DEFAULT 0,
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboRewardWeights_profile_category
    ON dbo.MboRewardWeights(target_profile, category_key);
END

IF OBJECT_ID('dbo.MboReportDefinitions','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboReportDefinitions (
    report_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    report_name NVARCHAR(255) NOT NULL,
    start_year_month CHAR(7) NOT NULL,
    end_year_month CHAR(7) NOT NULL,
    created_by NVARCHAR(255) NULL,
    created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE INDEX IX_MboReportDefinitions_months
    ON dbo.MboReportDefinitions(start_year_month, end_year_month, created_at);
END

IF OBJECT_ID('dbo.MboReportMultipliers','U') IS NULL
BEGIN
  CREATE TABLE dbo.MboReportMultipliers (
    row_uuid UNIQUEIDENTIFIER NOT NULL DEFAULT NEWID() PRIMARY KEY,
    report_uuid UNIQUEIDENTIFIER NOT NULL,
    report_type NVARCHAR(30) NOT NULL,
    subject_key NVARCHAR(120) NOT NULL,
    store_code NVARCHAR(50) NOT NULL DEFAULT '',
    multiplier_pct DECIMAL(9,4) NOT NULL DEFAULT 100,
    updated_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME()
  );
  CREATE UNIQUE INDEX UX_MboReportMultipliers_row
    ON dbo.MboReportMultipliers(report_uuid, report_type, subject_key, store_code);
END
"""
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:463')


def _find_column(columns: List[str], exact: List[str] | None = None, contains: List[str] | None = None, exclude: List[str] | None = None) -> str:
    exact = [str(x).strip().lower() for x in (exact or [])]
    contains = [str(x).strip().lower() for x in (contains or [])]
    exclude = [str(x).strip().lower() for x in (exclude or [])]

    for col in columns:
        c = _norm(col)
        if exact and c in exact and c not in exclude:
            return col
    for col in columns:
        c = _norm(col)
        if c in exclude:
            continue
        for token in contains:
            if token and token in c:
                return col
    return ""


def load_ilp_store_area_managers() -> List[Dict[str, Any]]:
    """
    Lettura best-effort dell'anagrafica store tenant.
    Serve solo a derivare l'associazione di default store -> area manager.
    """
    try:
        rows = []
        for row in list_store_registry(include_inactive=True):
            rows.append(
                {
                    "store_code": str(row.get("store_code") or "").strip(),
                    "store_name": str(row.get("store_name") or row.get("store_code") or "").strip(),
                    "am_value": str(row.get("area_manager") or "").strip(),
                }
            )
        if rows:
            return rows
    except Exception:
        log_swallowed('mbo_repository:503')

    conn = _conn(database="ILP", read_only=True)
    try:
        cur = conn.cursor()
        try:
            cur.execute("SELECT TOP 3000 * FROM dbo.[STORE]")
        except Exception:
            cur.execute("SELECT TOP 3000 * FROM [STORE]")

        columns = [d[0] for d in (cur.description or [])]
        code_col = _find_column(
            columns,
            exact=["codice", "storecode", "site", "pdv", "idpdv"],
            contains=["store code", "storecode", "codice", "site", "pdv"],
            exclude=["am"],
        )
        name_col = _find_column(
            columns,
            exact=["store", "storename", "name", "descrizione", "negozio", "puntovendita"],
            contains=["store", "descr", "negozio", "punto vendita", "ragione"],
            exclude=["am"],
        )
        am_col = _find_column(
            columns,
            exact=["am"],
            contains=["area manager", "am"],
        )

        if not am_col:
            raise RuntimeError(f"Colonna AM non trovata nella sorgente legacy ILP STORE. Colonne lette: {columns}")

        rows: List[Dict[str, Any]] = []
        for raw in cur.fetchall() or []:
            row = dict(zip(columns, raw))
            rows.append(
                {
                    "store_code": str(row.get(code_col) or "").strip() if code_col else "",
                    "store_name": str(row.get(name_col) or row.get(code_col) or "").strip() if (name_col or code_col) else "",
                    "am_value": str(row.get(am_col) or "").strip(),
                }
            )
        return rows
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:550')


def load_ilp_store_audit_refs() -> List[Dict[str, Any]]:
    """
    Lettura best-effort dell'anagrafica store tenant per match Yoobic e chiusura.
    Il risultato viene poi filtrato sul perimetro Supabase.
    """
    try:
        rows = []
        for row in list_store_registry(include_inactive=True):
            rows.append(
                {
                    "store_code": str(row.get("store_code") or "").strip(),
                    "store_name": str(row.get("store_name") or row.get("store_code") or "").strip(),
                    "yoobic": str(row.get("yoobic") or row.get("yoobic_address") or "").strip(),
                    "zucchetti": str(row.get("zucchetti") or row.get("ipratico_api_key") or "").strip(),
                    "google": str(row.get("google") or row.get("google_location_id") or "").strip(),
                    "glovo": str(row.get("glovo") or row.get("glovo_store_id") or "").strip(),
                    "deliveroo": str(row.get("deliveroo") or row.get("deliveroo_store_id") or "").strip(),
                    "closure_date": row.get("closure_date"),
                }
            )
        if rows:
            return rows
    except Exception:
        log_swallowed('mbo_repository:576')

    conn = _conn(database="ILP", read_only=True)
    try:
        cur = conn.cursor()
        try:
            cur.execute("SELECT TOP 5000 * FROM dbo.[STORE]")
        except Exception:
            cur.execute("SELECT TOP 5000 * FROM [STORE]")

        columns = [d[0] for d in (cur.description or [])]
        code_col = _find_column(
            columns,
            exact=["codice", "storecode", "site", "pdv", "idpdv"],
            contains=["store code", "storecode", "codice", "site", "pdv"],
            exclude=["am"],
        )
        name_col = _find_column(
            columns,
            exact=["store", "storename", "name", "descrizione", "negozio", "puntovendita"],
            contains=["store", "descr", "negozio", "punto vendita", "ragione"],
            exclude=["am"],
        )
        yoobic_col = _find_column(columns, exact=["yoobic"], contains=["yoobic"])
        zucchetti_col = _find_column(columns, exact=["zucchetti"], contains=["zucchetti"])
        google_col = _find_column(columns, exact=["google"], contains=["google"])
        glovo_col = _find_column(columns, exact=["glovo"], contains=["glovo"])
        deliveroo_col = _find_column(columns, exact=["deliveroo"], contains=["deliveroo"])
        closure_col = _find_column(columns, exact=["chiusura"], contains=["chiusura", "closing"])

        rows: List[Dict[str, Any]] = []
        for raw in cur.fetchall() or []:
            row = dict(zip(columns, raw))
            closure_raw = row.get(closure_col) if closure_col else None
            closure_date = None
            if isinstance(closure_raw, datetime):
                closure_date = closure_raw.date()
            elif isinstance(closure_raw, date):
                closure_date = closure_raw
            elif closure_raw:
                closure_date = _parse_date(closure_raw)
            rows.append(
                {
                    "store_code": str(row.get(code_col) or "").strip() if code_col else "",
                    "store_name": str(row.get(name_col) or row.get(code_col) or "").strip() if (name_col or code_col) else "",
                    "yoobic": str(row.get(yoobic_col) or "").strip() if yoobic_col else "",
                    "zucchetti": str(row.get(zucchetti_col) or "").strip() if zucchetti_col else "",
                    "google": str(row.get(google_col) or "").strip() if google_col else "",
                    "glovo": str(row.get(glovo_col) or "").strip() if glovo_col else "",
                    "deliveroo": str(row.get(deliveroo_col) or "").strip() if deliveroo_col else "",
                    "closure_date": closure_date,
                }
            )
        return rows
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:634')


def _load_valid_store_refs() -> List[Dict[str, Any]]:
    stores = get_warehouse_stores() or []
    try:
        registry_rows = list_store_registry(include_inactive=True)
    except Exception:
        registry_rows = []
    if not registry_rows:
        registry_rows = load_ilp_store_audit_refs()
    registry_by_code: Dict[str, Dict[str, Any]] = {}
    registry_by_name: Dict[str, Dict[str, Any]] = {}
    for row in registry_rows:
        code_key = _norm_key(row.get("store_code"))
        name_key = _norm_key(row.get("store_name"))
        if code_key and code_key not in registry_by_code:
            registry_by_code[code_key] = row
        if name_key and name_key not in registry_by_name:
            registry_by_name[name_key] = row

    refs: List[Dict[str, Any]] = []
    for store in stores:
        code = str((store or {}).get("code") or "").strip()
        name = str((store or {}).get("name") or code).strip()
        store_row = registry_by_code.get(_norm_key(code)) or registry_by_name.get(_norm_key(name)) or {}
        refs.append(
            {
                "store_code": code,
                "store_name": name,
                "yoobic": str(store_row.get("yoobic") or store_row.get("yoobic_address") or "").strip(),
                "yoobic_norm": _norm_address(store_row.get("yoobic") or store_row.get("yoobic_address")),
                "zucchetti": str(store_row.get("zucchetti") or store_row.get("ipratico_api_key") or "").strip(),
                "zucchetti_norm": _norm_key(store_row.get("zucchetti") or store_row.get("ipratico_api_key")),
                "google": str(store_row.get("google") or store_row.get("google_location_id") or "").strip(),
                "google_norm": _norm_key(store_row.get("google") or store_row.get("google_location_id")),
                "glovo": str(store_row.get("glovo") or store_row.get("glovo_store_id") or "").strip(),
                "glovo_norm": _norm_key(store_row.get("glovo") or store_row.get("glovo_store_id")),
                "deliveroo": str(store_row.get("deliveroo") or store_row.get("deliveroo_store_id") or "").strip(),
                "deliveroo_norm": _norm_key(store_row.get("deliveroo") or store_row.get("deliveroo_store_id")),
                "closure_date": store_row.get("closure_date"),
            }
        )
    return refs


def _active_months_for_store(year: int, closure_date: date | None = None) -> List[str]:
    months: List[str] = []
    for month in range(1, 13):
        ym = f"{int(year):04d}-{month:02d}"
        end_dt = _month_end(ym)
        if closure_date and end_dt and end_dt > closure_date and (end_dt.year, end_dt.month) > (closure_date.year, closure_date.month):
            continue
        months.append(ym)
    return months


def _read_audit_upload_rows(filename: str, content: bytes) -> List[Dict[str, Any]]:
    name = str(filename or "").lower()
    payloads: List[tuple[str, bytes]] = []
    if name.endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for info in zf.infolist():
                if not info.is_dir() and info.filename.lower().endswith(".csv"):
                    payloads.append((info.filename, zf.read(info)))
    else:
        payloads.append((filename, content))

    rows: List[Dict[str, Any]] = []
    for inner_name, raw in payloads:
        text = None
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            text = raw.decode("utf-8", errors="replace")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ";"
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        for idx, row in enumerate(reader, start=1):
            normalized = {str(k or "").strip().lower(): v for k, v in (row or {}).items()}
            address = str(normalized.get("address") or "").strip()
            audit_date = _parse_date(normalized.get("date"))
            if not address or not audit_date:
                continue
            score = _parse_decimal(normalized.get("score"))
            mission_data_id = str(normalized.get("mission_data_id") or "").strip()
            mission_id = str(normalized.get("mission_id") or "").strip()
            source_seed = mission_data_id or f"{inner_name}|{idx}|{address}|{normalized.get('date') or ''}|{normalized.get('score') or ''}"
            rows.append(
                {
                    "source_file": inner_name,
                    "source_key": hashlib.sha1(source_seed.encode("utf-8", errors="ignore")).hexdigest(),
                    "mission_data_id": mission_data_id,
                    "mission_id": mission_id,
                    "mission_title": str(normalized.get("mission_title") or "").strip(),
                    "address": address,
                    "address_norm": _norm_address(address),
                    "audit_date": audit_date,
                    "audit_year_month": _month_key(audit_date),
                    "score": score,
                    "username": str(normalized.get("username") or "").strip(),
                    "user_full_name": str(normalized.get("user_full_name") or "").strip(),
                    "raw_json": json.dumps(row, ensure_ascii=False, default=str),
                }
            )
    return rows


def _load_audit_manual_maps() -> Dict[str, str]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute("SELECT address_norm, store_code FROM dbo.MboAuditAddressStoreMap")
        return {str(r[0] or ""): str(r[1] or "").strip() for r in cur.fetchall() or []}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:761')


def import_audit_file(filename: str, content: bytes, imported_by: str | None = None) -> Dict[str, int]:
    ensure_mbo_schema()
    parsed_rows = _read_audit_upload_rows(filename, content)
    store_refs = _load_valid_store_refs()
    valid_codes = {str(s.get("store_code") or "").strip() for s in store_refs if str(s.get("store_code") or "").strip()}
    by_yoobic = {str(s.get("yoobic_norm") or ""): str(s.get("store_code") or "").strip() for s in store_refs if str(s.get("yoobic_norm") or "").strip()}
    manual_maps = _load_audit_manual_maps()

    conn = _conn(read_only=False)
    imported = 0
    try:
        cur = conn.cursor()
        cur.execute(
            """
INSERT INTO dbo.MboAuditImports (source_filename, imported_by, rows_total, rows_imported)
OUTPUT inserted.import_uuid
VALUES (?, ?, ?, 0)
""",
            (str(filename or "")[:255], str(imported_by or "")[:255] or None, len(parsed_rows)),
        )
        import_uuid = str(cur.fetchone()[0])

        for row in parsed_rows:
            matched_code = manual_maps.get(row["address_norm"]) or by_yoobic.get(row["address_norm"]) or ""
            if matched_code not in valid_codes:
                matched_code = ""
            assigned_store = matched_code or None
            assigned_month = row["audit_year_month"] if assigned_store else None
            cur.execute(
                """
IF EXISTS (SELECT 1 FROM dbo.MboAuditRows WHERE source_key = ?)
BEGIN
  UPDATE dbo.MboAuditRows
  SET import_uuid = ?, mission_data_id = ?, mission_id = ?, mission_title = ?,
      address = ?, address_norm = ?, audit_date = ?, audit_year_month = ?, score = ?,
      matched_store_code = ?, assigned_store_code = COALESCE(assigned_store_code, ?),
      assigned_year_month = COALESCE(assigned_year_month, ?),
      username = ?, user_full_name = ?, raw_json = ?, updated_at = SYSUTCDATETIME()
  WHERE source_key = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboAuditRows (
    import_uuid, source_key, mission_data_id, mission_id, mission_title, address, address_norm,
    audit_date, audit_year_month, score, matched_store_code, assigned_store_code, assigned_year_month,
    username, user_full_name, raw_json
  )
  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
END
""",
                (
                    row["source_key"],
                    import_uuid,
                    row["mission_data_id"] or None,
                    row["mission_id"] or None,
                    row["mission_title"] or None,
                    row["address"],
                    row["address_norm"],
                    row["audit_date"],
                    row["audit_year_month"],
                    float(row["score"]) if row["score"] is not None else None,
                    matched_code or None,
                    assigned_store,
                    assigned_month,
                    row["username"] or None,
                    row["user_full_name"] or None,
                    row["raw_json"],
                    row["source_key"],
                    import_uuid,
                    row["source_key"],
                    row["mission_data_id"] or None,
                    row["mission_id"] or None,
                    row["mission_title"] or None,
                    row["address"],
                    row["address_norm"],
                    row["audit_date"],
                    row["audit_year_month"],
                    float(row["score"]) if row["score"] is not None else None,
                    matched_code or None,
                    assigned_store,
                    assigned_month,
                    row["username"] or None,
                    row["user_full_name"] or None,
                    row["raw_json"],
                ),
            )
            imported += 1

        cur.execute("UPDATE dbo.MboAuditImports SET rows_imported = ? WHERE import_uuid = ?", (imported, import_uuid))
        conn.commit()
        return {"parsed": len(parsed_rows), "imported": imported}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:859')


def save_audit_address_mapping(address_norm: str, address_sample: str, store_code: str) -> None:
    ensure_mbo_schema()
    store_code = str(store_code or "").strip()
    if not address_norm or not store_code:
        raise ValueError("Indirizzo e store sono obbligatori.")
    valid_codes = {str(s.get("store_code") or "").strip() for s in _load_valid_store_refs()}
    if store_code not in valid_codes:
        raise ValueError("Store non valido per il perimetro Supabase.")
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
IF EXISTS (SELECT 1 FROM dbo.MboAuditAddressStoreMap WHERE address_norm = ?)
BEGIN
  UPDATE dbo.MboAuditAddressStoreMap
  SET address_sample = ?, store_code = ?, updated_at = SYSUTCDATETIME()
  WHERE address_norm = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboAuditAddressStoreMap (address_norm, address_sample, store_code)
  VALUES (?, ?, ?);
END

UPDATE dbo.MboAuditRows
SET assigned_store_code = ?, matched_store_code = ?, assigned_year_month = COALESCE(assigned_year_month, audit_year_month),
    updated_at = SYSUTCDATETIME()
WHERE address_norm = ?;
""",
            (address_norm, address_sample, store_code, address_norm, address_norm, address_sample, store_code, store_code, store_code, address_norm),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:899')


def save_audit_row_assignment(row_uuid: str, store_code: str, year_month: str) -> None:
    ensure_mbo_schema()
    row_uuid = str(row_uuid or "").strip()
    store_code = str(store_code or "").strip()
    year_month = str(year_month or "").strip()
    if not row_uuid or not store_code or not year_month:
        raise ValueError("Riga audit, store e mese sono obbligatori.")
    store_refs = _load_valid_store_refs()
    valid_codes = {str(s.get("store_code") or "").strip() for s in store_refs}
    if store_code not in valid_codes:
        raise ValueError("Store non valido per il perimetro Supabase.")
    store_ref = next((s for s in store_refs if str(s.get("store_code") or "").strip() == store_code), {})
    if year_month not in _active_months_for_store(int(str(year_month)[:4]), store_ref.get("closure_date")):
        raise ValueError("Mese non valido: lo store risulta chiuso per quel periodo.")
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
UPDATE dbo.MboAuditRows
SET assigned_store_code = ?, assigned_year_month = ?, ignored = 0, updated_at = SYSUTCDATETIME()
WHERE row_uuid = ?
""",
            (store_code, year_month, row_uuid),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:932')


def ignore_audit_row(row_uuid: str) -> None:
    ensure_mbo_schema()
    row_uuid = str(row_uuid or "").strip()
    if not row_uuid:
        raise ValueError("Riga audit obbligatoria.")
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
UPDATE dbo.MboAuditRows
SET ignored = 1, updated_at = SYSUTCDATETIME()
WHERE row_uuid = ?
""",
            (row_uuid,),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:956')


def restore_audit_row(row_uuid: str) -> None:
    ensure_mbo_schema()
    row_uuid = str(row_uuid or "").strip()
    if not row_uuid:
        raise ValueError("Riga audit obbligatoria.")
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
UPDATE dbo.MboAuditRows
SET ignored = 0, updated_at = SYSUTCDATETIME()
WHERE row_uuid = ?
""",
            (row_uuid,),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:980')


def save_audit_manual_value(store_code: str, year_month: str, score: Any, note: str | None = None) -> None:
    ensure_mbo_schema()
    score_dec = _parse_decimal(score)
    if score_dec is None:
        raise ValueError("Valore audit non valido.")
    store_refs = _load_valid_store_refs()
    store_ref = next((s for s in store_refs if str(s.get("store_code") or "").strip() == str(store_code or "").strip()), None)
    if not store_ref:
        raise ValueError("Store non valido per il perimetro Supabase.")
    if year_month not in _active_months_for_store(int(str(year_month)[:4]), store_ref.get("closure_date")):
        raise ValueError("Mese non valido: lo store risulta chiuso per quel periodo.")
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
IF EXISTS (SELECT 1 FROM dbo.MboAuditManualMonthlyValues WHERE store_code = ? AND year_month = ?)
BEGIN
  UPDATE dbo.MboAuditManualMonthlyValues
  SET score = ?, note = ?, updated_at = SYSUTCDATETIME()
  WHERE store_code = ? AND year_month = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboAuditManualMonthlyValues (store_code, year_month, score, note)
  VALUES (?, ?, ?, ?);
END
""",
            (store_code, year_month, float(score_dec), note, store_code, year_month, store_code, year_month, float(score_dec), note),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:1018')


def delete_audit_manual_value(store_code: str, year_month: str) -> None:
    ensure_mbo_schema()
    store_code = str(store_code or "").strip()
    year_month = str(year_month or "").strip()
    if not store_code or not year_month:
        raise ValueError("Store e mese sono obbligatori.")
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            "DELETE FROM dbo.MboAuditManualMonthlyValues WHERE store_code = ? AND year_month = ?",
            (store_code, year_month),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:1039')


def _load_audit_rows_for_year(year: int) -> List[Dict[str, Any]]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute(
            """
SELECT row_uuid, source_key, mission_data_id, address, address_norm, audit_date, audit_year_month,
       score, matched_store_code, assigned_store_code, assigned_year_month, username, user_full_name, created_at, ignored
FROM dbo.MboAuditRows
WHERE COALESCE(ignored, 0) = 0
  AND ((audit_year_month >= ? AND audit_year_month <= ?)
   OR (assigned_year_month >= ? AND assigned_year_month <= ?))
ORDER BY COALESCE(assigned_year_month, audit_year_month) DESC, audit_date DESC, address ASC
""",
            (f"{int(year):04d}-01", f"{int(year):04d}-12", f"{int(year):04d}-01", f"{int(year):04d}-12"),
        )
        out: List[Dict[str, Any]] = []
        for r in cur.fetchall() or []:
            out.append(
                {
                    "row_uuid": str(r[0]),
                    "source_key": str(r[1] or ""),
                    "mission_data_id": str(r[2] or ""),
                    "address": str(r[3] or ""),
                    "address_norm": str(r[4] or ""),
                    "audit_date": r[5],
                    "audit_year_month": str(r[6] or ""),
                    "score": float(r[7]) if r[7] is not None else None,
                    "matched_store_code": str(r[8] or ""),
                    "assigned_store_code": str(r[9] or ""),
                    "assigned_year_month": str(r[10] or ""),
                    "username": str(r[11] or ""),
                    "user_full_name": str(r[12] or ""),
                    "created_at": r[13],
                    "ignored": bool(r[14]),
                }
            )
        return out
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:1085')


def list_ignored_audit_rows(year: int) -> List[Dict[str, Any]]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute(
            """
SELECT row_uuid, source_key, mission_data_id, address, address_norm, audit_date, audit_year_month,
       score, matched_store_code, assigned_store_code, assigned_year_month, username, user_full_name, created_at, ignored
FROM dbo.MboAuditRows
WHERE COALESCE(ignored, 0) = 1
  AND ((audit_year_month >= ? AND audit_year_month <= ?)
   OR (assigned_year_month >= ? AND assigned_year_month <= ?))
ORDER BY COALESCE(assigned_year_month, audit_year_month) DESC, audit_date DESC, address ASC
""",
            (f"{int(year):04d}-01", f"{int(year):04d}-12", f"{int(year):04d}-01", f"{int(year):04d}-12"),
        )
        out: List[Dict[str, Any]] = []
        for r in cur.fetchall() or []:
            out.append(
                {
                    "row_uuid": str(r[0]),
                    "source_key": str(r[1] or ""),
                    "mission_data_id": str(r[2] or ""),
                    "address": str(r[3] or ""),
                    "address_norm": str(r[4] or ""),
                    "audit_date": r[5],
                    "audit_year_month": str(r[6] or ""),
                    "score": float(r[7]) if r[7] is not None else None,
                    "matched_store_code": str(r[8] or ""),
                    "assigned_store_code": str(r[9] or ""),
                    "assigned_year_month": str(r[10] or ""),
                    "username": str(r[11] or ""),
                    "user_full_name": str(r[12] or ""),
                    "created_at": r[13],
                    "ignored": bool(r[14]),
                }
            )
        return out
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:1131')


def build_audit_address_mapping_matrix(year: int) -> Dict[str, Any]:
    ensure_mbo_schema()
    store_refs = _load_valid_store_refs()
    stores_by_code = {str(s.get("store_code") or "").strip(): s for s in store_refs}
    by_yoobic = {str(s.get("yoobic_norm") or ""): str(s.get("store_code") or "").strip() for s in store_refs if str(s.get("yoobic_norm") or "").strip()}
    manual_maps = _load_audit_manual_maps()

    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute(
            """
SELECT address_norm, MIN(address) AS address_sample, COUNT(*) AS audit_count, MAX(audit_date) AS latest_audit_date
FROM dbo.MboAuditRows
WHERE audit_year_month >= ? AND audit_year_month <= ?
GROUP BY address_norm
ORDER BY MIN(address)
""",
            (f"{int(year):04d}-01", f"{int(year):04d}-12"),
        )
        rows: List[Dict[str, Any]] = []
        for r in cur.fetchall() or []:
            address_norm = str(r[0] or "")
            address_sample = str(r[1] or "")
            manual_code = str(manual_maps.get(address_norm) or "").strip()
            automatic_code = str(by_yoobic.get(address_norm) or "").strip()
            effective_code = manual_code or automatic_code
            source = "manuale" if manual_code else ("automatico" if automatic_code else "senza match")
            store = stores_by_code.get(effective_code, {})
            rows.append(
                {
                    "address_norm": address_norm,
                    "address": address_sample,
                    "audit_count": int(r[2] or 0),
                    "latest_audit_date": r[3],
                    "manual_store_code": manual_code,
                    "automatic_store_code": automatic_code,
                    "effective_store_code": effective_code,
                    "effective_store_name": str(store.get("store_name") or store.get("name") or effective_code),
                    "source": source,
                }
            )
        return {
            "year": int(year),
            "stores": store_refs,
            "rows": rows,
            "total_count": len(rows),
            "manual_count": sum(1 for r in rows if r.get("manual_store_code")),
            "automatic_count": sum(1 for r in rows if (not r.get("manual_store_code")) and r.get("automatic_store_code")),
            "unmatched_count": sum(1 for r in rows if not r.get("effective_store_code")),
        }
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:1189')


def _load_audit_manual_values_for_year(year: int) -> Dict[tuple[str, str], Dict[str, Any]]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute(
            """
SELECT store_code, year_month, score, note
FROM dbo.MboAuditManualMonthlyValues
WHERE year_month >= ? AND year_month <= ?
""",
            (f"{int(year):04d}-01", f"{int(year):04d}-12"),
        )
        out: Dict[tuple[str, str], Dict[str, Any]] = {}
        for r in cur.fetchall() or []:
            out[(str(r[0] or ""), str(r[1] or ""))] = {
                "score": float(r[2]) if r[2] is not None else 0.0,
                "note": str(r[3] or ""),
            }
        return out
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:1216')


def build_audit_month_matrix(year: int) -> Dict[str, Any]:
    ensure_mbo_schema()
    store_refs = _load_valid_store_refs()
    rows = _load_audit_rows_for_year(year)
    manual_values = _load_audit_manual_values_for_year(year)

    store_by_code = {str(s.get("store_code") or ""): s for s in store_refs}
    month_labels = [
        {"month_num": i, "month_label": lbl, "year_month": f"{int(year):04d}-{i:02d}"}
        for i, lbl in enumerate(["", "Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"])
        if i
    ]

    assigned: Dict[tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    unmatched: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        code = str(row.get("assigned_store_code") or "").strip()
        ym = str(row.get("assigned_year_month") or row.get("audit_year_month") or "").strip()
        if code and ym:
            assigned[(code, ym)].append(row)
        else:
            key = str(row.get("address_norm") or "")
            if key and key not in unmatched:
                unmatched[key] = {
                    "address_norm": key,
                    "address": row.get("address") or "",
                    "count": 0,
                    "latest_date": row.get("audit_date"),
                }
            if key:
                unmatched[key]["count"] += 1

    matrix_rows: List[Dict[str, Any]] = []
    missing_cells: List[Dict[str, Any]] = []
    duplicate_cells: List[Dict[str, Any]] = []
    excess_audits: List[Dict[str, Any]] = []
    duplicate_audits: List[Dict[str, Any]] = []

    for store in sorted(store_refs, key=lambda s: (_norm_key(s.get("store_name")), _norm_key(s.get("store_code")))):
        code = str(store.get("store_code") or "").strip()
        active_months = set(_active_months_for_store(year, store.get("closure_date")))
        cells: List[Dict[str, Any]] = []
        for month in month_labels:
            ym = month["year_month"]
            active = ym in active_months
            audits = assigned.get((code, ym), []) if active else []
            manual = manual_values.get((code, ym))
            status = "closed" if not active else "ok"
            value = None
            imported_value = None
            if active:
                if len(audits) == 1:
                    status = "ok"
                    imported_value = audits[0].get("score")
                    value = imported_value
                elif len(audits) > 1:
                    status = "duplicate"
                    imported_value = audits[0].get("score")
                    value = imported_value
                    duplicate_cells.append({"store": store, "year_month": ym, "audits": audits})
                    available_months = [m for m in month_labels if m["year_month"] in active_months]
                    for idx, audit in enumerate(audits, start=1):
                        duplicate_item = (
                            {
                                **audit,
                                "store_name": store.get("store_name") or code,
                                "store_code": code,
                                "duplicate_index": idx,
                                "duplicate_total": len(audits),
                                "available_months": available_months,
                            }
                        )
                        duplicate_audits.append(duplicate_item)
                        if idx > 1:
                            excess_audits.append(duplicate_item)
                else:
                    status = "missing"
                    value = 0.0
                    missing = {"store": store, "year_month": ym, "score": (manual or {}).get("score", value), "note": (manual or {}).get("note", "")}
                    missing_cells.append(missing)
                if manual:
                    status = "manual"
                    value = manual.get("score", value if value is not None else 0.0)
            cells.append(
                {
                    "year_month": ym,
                    "month_label": month["month_label"],
                    "active": active,
                    "status": status,
                    "score": value,
                    "imported_score": imported_value,
                    "audits": audits,
                    "manual": manual,
                }
            )
        matrix_rows.append({**store, "months": cells, "active_months": len(active_months)})

    return {
        "year": int(year),
        "month_labels": month_labels,
        "stores": store_refs,
        "store_by_code": store_by_code,
        "rows": matrix_rows,
        "unmatched_addresses": list(unmatched.values()),
        "missing_cells": missing_cells,
        "duplicate_cells": duplicate_cells,
        "duplicate_audits": duplicate_audits,
        "excess_audits": excess_audits,
        "all_audits": rows,
        "ignored_audits": list_ignored_audit_rows(year),
        "store_count": len(store_refs),
        "audit_count": len(rows),
        "unmatched_count": len(unmatched),
        "missing_count": len(missing_cells),
        "duplicate_count": len(duplicate_cells),
        "excess_count": len(duplicate_audits),
        "generated_at": datetime.utcnow(),
    }


def _to_float(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, Decimal):
        try:
            return float(v)
        except Exception:
            return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def _voice_get_ci(d: Dict[str, float], voice: str) -> float:
    if not d:
        return 0.0
    wanted = _norm_key(voice)
    for key, value in d.items():
        if _norm_key(key) == wanted:
            return _to_float(value)
    return 0.0


_INVENTORY_OPENING_ALIASES = (
    "Magazzino Iniziale",
    "Inventario Iniziale",
    "Initial Inventory",
    "Opening Inventory",
)
_INVENTORY_CLOSING_ALIASES = (
    "Magazzino Finale",
    "Inventario Finale",
    "Final Inventory",
    "Closing Inventory",
)


def _voice_get_any_ci(d: Dict[str, float], voices: tuple[str, ...]) -> float:
    return sum(_voice_get_ci(d, voice) for voice in voices)


def _fetch_store_month_voice_matrix(
    conn,
    table_or_view: str,
    store_codes: List[str],
    year: int,
    month_from: int = 1,
    month_to: int = 12,
) -> Dict[tuple[str, int], Dict[str, float]]:
    layout = _detect_pl_layout(conn, table_or_view)
    stores = [str(s or "").strip() for s in (store_codes or []) if str(s or "").strip()]
    if not stores:
        return {}

    site_expr = f"LTRIM(RTRIM(CAST([{layout.site_col}] AS NVARCHAR(255))))"
    month_expr = f"TRY_CONVERT(int, [{layout.month_col}])"
    year_expr = f"TRY_CONVERT(int, [{layout.year_col}])"
    voice_expr = f"LTRIM(RTRIM(CAST([{layout.voice_col}] AS NVARCHAR(255))))"
    placeholders = ",".join(["?"] * len(stores))
    sql = (
        f"SELECT {site_expr} AS site, {month_expr} AS month_num, {voice_expr} AS voice, "
        f"SUM(TRY_CONVERT(decimal(18,4), [{layout.value_col}])) AS value "
        f"FROM {layout.table_ref} "
        f"WHERE {site_expr} IN ({placeholders}) "
        f"AND {year_expr} = ? "
        f"AND {month_expr} >= ? AND {month_expr} <= ? "
        f"GROUP BY {site_expr}, {month_expr}, {voice_expr}"
    )
    params: List[Any] = [*stores, int(year), int(month_from), int(month_to)]
    cur = conn.cursor()
    cur.execute(sql, params)
    out: Dict[tuple[str, int], Dict[str, float]] = {}
    for site, month_num, voice, value in cur.fetchall() or []:
        site_s = str(site or "").strip()
        month_i = int(month_num or 0)
        voice_s = str(voice or "").strip()
        if not site_s or month_i < 1 or month_i > 12 or not voice_s:
            continue
        out.setdefault((site_s, month_i), {})
        out[(site_s, month_i)][voice_s] = out[(site_s, month_i)].get(voice_s, 0.0) + _to_float(value)
    return out


def _compute_pnl_totals(src: Dict[str, float], *, waste_to_cogs: bool = False) -> Dict[str, float]:
    out = dict(src or {})
    revenues = _voice_get_ci(out, "REVENUES")

    other_ga_extra = [
        "Casse e HiTec",
        "Altri servizi esterni",
        "Commissioni Ticket",
        "Piccole attrezzaure - Cancelleria",
        "Costi assicurativi",
        "Affitto attrezzature",
        "Altro",
    ]
    labour_items = ["Labour fixed", "Stage", "External Labour", "Trasferimento", "Costo formazione", "Other cost"]
    delivery_items = ["Variable fees", "Other delivery fees"]
    ga_items = [
        "Rent",
        "Spese Condominiali",
        "Utilities",
        "Cleaning+Security",
        "Marketing",
        "Maintenance",
        "Spese Trasporto",
        "Other G&A",
    ]
    ebitda_other = ["Other personnel cost", "Bank commissions", "Consultancies", "Other taxes", "Other revenues"]

    other_ga_base = _voice_get_ci(out, "Other G&A")
    other_ga_sum = sum(_voice_get_ci(out, v) for v in other_ga_extra)
    out["Other G&A"] = other_ga_base + other_ga_sum

    cogs = (
        _voice_get_any_ci(out, _INVENTORY_OPENING_ALIASES)
        + _voice_get_ci(out, "Acquistato")
        + _voice_get_ci(out, "Trasferimenti")
        - _voice_get_any_ci(out, _INVENTORY_CLOSING_ALIASES)
    )
    if waste_to_cogs:
        cogs += _voice_get_ci(out, "Waste")
    out["COGS"] = cogs
    out["MARGINE DI CONTRIBUZIONE"] = revenues - cogs
    out["LABOUR COST"] = sum(_voice_get_ci(out, v) for v in labour_items)
    out["DELIVERY FEES"] = sum(_voice_get_ci(out, v) for v in delivery_items)
    out["G&A STORE"] = sum(_voice_get_ci(out, v) for v in ga_items)
    out["TOTALE COSTI CONTROLLABILI"] = out["COGS"] + out["LABOUR COST"] + out["DELIVERY FEES"] + out["G&A STORE"]
    out["STORE EBITDA"] = revenues - out["TOTALE COSTI CONTROLLABILI"]
    out["EBITDA"] = out["STORE EBITDA"] - sum(_voice_get_ci(out, v) for v in ebitda_other)
    return out


def _incidence(value: float, revenues: float) -> float | None:
    if revenues == 0:
        return None
    return float(value) / float(revenues)


def _score_positive_delta(actual: float, budget: float) -> Dict[str, Any]:
    actual_f = _to_float(actual)
    budget_f = _to_float(budget)
    diff = actual_f - budget_f
    if diff >= 0:
        points = 1.0
    elif budget_f and (diff / budget_f) >= -0.005:
        points = 0.5
    else:
        points = 0.0
    return {
        "points": points,
        "actual": actual_f,
        "budget": budget_f,
        "delta": diff,
        "delta_pct": (diff / budget_f) if budget_f else None,
    }


def _score_cost_incidence(actual_value: float, actual_revenues: float, budget_value: float, budget_revenues: float) -> Dict[str, Any]:
    actual_pct = _incidence(actual_value, actual_revenues)
    budget_pct = _incidence(budget_value, budget_revenues)
    if actual_pct is None or budget_pct is None:
        points = 0.0
        delta_pct = None
    else:
        delta_pct = actual_pct - budget_pct
        if delta_pct <= 0:
            points = 1.0
        elif delta_pct <= 0.005:
            points = 0.5
        else:
            points = 0.0
    return {
        "points": points,
        "actual": _to_float(actual_value),
        "budget": _to_float(budget_value),
        "actual_pct": actual_pct,
        "budget_pct": budget_pct,
        "delta_pct": delta_pct,
    }


def build_mbo_pnl_scores(year: int) -> Dict[str, Any]:
    ensure_mbo_schema()
    store_refs = _load_valid_store_refs()
    store_codes = [str(s.get("store_code") or "").strip() for s in store_refs if str(s.get("store_code") or "").strip()]
    audit_matrix = build_audit_month_matrix(year)
    audit_by_pair: Dict[tuple[str, str], Dict[str, Any]] = {}
    for row in audit_matrix.get("rows") or []:
        code = str(row.get("store_code") or "").strip()
        for cell in row.get("months") or []:
            audit_by_pair[(code, str(cell.get("year_month") or ""))] = cell

    month_labels = [
        {"month_num": i, "month_label": lbl, "year_month": f"{int(year):04d}-{i:02d}"}
        for i, lbl in enumerate(["", "Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"])
        if i
    ]

    if not store_codes:
        return {
            "year": int(year),
            "month_labels": month_labels,
            "rows": [],
            "store_count": 0,
            "points_taken": 0.0,
            "points_theoretical": 0.0,
            "pct": 0.0,
        }

    conn_budget = get_connection_ilp(read_only=True)
    conn_actual = get_connection_database_new(read_only=True)
    try:
        budget_matrix = _fetch_store_month_voice_matrix(conn_budget, "BudgetPL", store_codes, year, 1, 12)
        actual_matrix = _fetch_store_month_voice_matrix(conn_actual, "vw_DATIPL", store_codes, year, 1, 12)
    finally:
        for conn in (conn_budget, conn_actual):
            try:
                conn.close()
            except Exception:
                log_swallowed('mbo_repository:1560')

    rows: List[Dict[str, Any]] = []
    points_taken_total = 0.0
    points_theoretical_total = 0.0
    active_months_total = 0

    for store in sorted(store_refs, key=lambda s: (_norm_key(s.get("store_name")), _norm_key(s.get("store_code")))):
        code = str(store.get("store_code") or "").strip()
        active_months = set(_active_months_for_store(year, store.get("closure_date")))
        months: List[Dict[str, Any]] = []
        store_points = 0.0
        store_theoretical = 0.0
        for month in month_labels:
            month_num = int(month["month_num"])
            ym = str(month["year_month"])
            active = ym in active_months
            if not active:
                months.append({"year_month": ym, "month_label": month["month_label"], "active": False})
                continue

            budget = _compute_pnl_totals(budget_matrix.get((code, month_num), {}), waste_to_cogs=True)
            actual = _compute_pnl_totals(actual_matrix.get((code, month_num), {}), waste_to_cogs=False)
            rev_budget = _voice_get_ci(budget, "REVENUES")
            rev_actual = _voice_get_ci(actual, "REVENUES")

            voice_scores = {
                "REVENUES": _score_positive_delta(_voice_get_ci(actual, "REVENUES"), _voice_get_ci(budget, "REVENUES")),
                "COGS": _score_cost_incidence(_voice_get_ci(actual, "COGS"), rev_actual, _voice_get_ci(budget, "COGS"), rev_budget),
                "LABOUR COST": _score_cost_incidence(_voice_get_ci(actual, "LABOUR COST"), rev_actual, _voice_get_ci(budget, "LABOUR COST"), rev_budget),
                "EBITDA": _score_positive_delta(_voice_get_ci(actual, "EBITDA"), _voice_get_ci(budget, "EBITDA")),
            }
            theoretical = 4.0
            raw_points = sum(float(v.get("points") or 0.0) for v in voice_scores.values())
            audit_cell = audit_by_pair.get((code, ym), {})
            audit_score = audit_cell.get("score")
            audit_ok = audit_score is not None and float(audit_score or 0.0) >= 90.0
            final_points = raw_points if audit_ok else 0.0

            points_taken_total += final_points
            points_theoretical_total += theoretical
            active_months_total += 1
            store_points += final_points
            store_theoretical += theoretical

            months.append(
                {
                    "year_month": ym,
                    "month_label": month["month_label"],
                    "active": True,
                    "voice_scores": voice_scores,
                    "raw_points": raw_points,
                    "final_points": final_points,
                    "theoretical_points": theoretical,
                    "audit_score": audit_score,
                    "audit_ok": audit_ok,
                    "audit_status": audit_cell.get("status", "missing"),
                }
            )
        rows.append(
            {
                **store,
                "months": months,
                "points_taken": store_points,
                "points_theoretical": store_theoretical,
                "pct": (store_points / store_theoretical * 100.0) if store_theoretical else 0.0,
            }
        )

    pct = (points_taken_total / points_theoretical_total * 100.0) if points_theoretical_total else 0.0
    return {
        "year": int(year),
        "month_labels": month_labels,
        "rows": rows,
        "store_count": len(rows),
        "active_months": active_months_total,
        "points_taken": points_taken_total,
        "points_theoretical": points_theoretical_total,
        "pct": pct,
        "voices": ["REVENUES", "COGS", "LABOUR COST", "EBITDA"],
    }


def _detect_google_reviews_layout(conn) -> Dict[str, str]:
    table_ref = "[dbo].[RecensioniGoogle]"
    cur = conn.cursor()
    cur.execute(f"SELECT TOP 0 * FROM {table_ref}")
    columns = [d[0] for d in (cur.description or [])]
    id_col = _find_column(columns, exact=["id_location", "idlocation"], contains=["id_location", "location"])
    date_col = _find_column(columns, exact=["data_recensione"], contains=["data_recensione", "recensione", "date", "data"])
    rating_col = _find_column(
        columns,
        exact=["rating", "valutazione", "punteggio", "stelle", "score", "voto"],
        contains=["rating", "valut", "punteggio", "stelle", "score", "voto"],
    )
    if not id_col or not date_col or not rating_col:
        raise RuntimeError(f"Colonne RecensioniGoogle non trovate. Colonne lette: {columns}")
    return {"table_ref": table_ref, "id_col": id_col, "date_col": date_col, "rating_col": rating_col}


def _load_google_review_averages_for_year(year: int) -> Dict[tuple[str, str], Dict[str, Any]]:
    conn = get_connection_ilp(read_only=True)
    try:
        layout = _detect_google_reviews_layout(conn)
        id_expr = f"LTRIM(RTRIM(CAST([{layout['id_col']}] AS NVARCHAR(255))))"
        date_expr = f"TRY_CONVERT(date, [{layout['date_col']}])"
        rating_expr = f"TRY_CONVERT(decimal(9,4), [{layout['rating_col']}])"
        sql = (
            f"SELECT {id_expr} AS id_location, "
            f"FORMAT({date_expr}, 'yyyy-MM') AS year_month, "
            f"AVG({rating_expr}) AS avg_rating, COUNT(*) AS review_count "
            f"FROM {layout['table_ref']} "
            f"WHERE {date_expr} >= ? AND {date_expr} < ? AND {rating_expr} IS NOT NULL "
            f"GROUP BY {id_expr}, FORMAT({date_expr}, 'yyyy-MM')"
        )
        cur = conn.cursor()
        cur.execute(sql, (f"{int(year):04d}-01-01", f"{int(year) + 1:04d}-01-01"))
        out: Dict[tuple[str, str], Dict[str, Any]] = {}
        for r in cur.fetchall() or []:
            id_location = str(r[0] or "").strip()
            ym = str(r[1] or "").strip()
            if not id_location or not ym:
                continue
            out[(_norm_key(id_location), ym)] = {
                "rating": _to_float(r[2]),
                "review_count": int(r[3] or 0),
                "id_location": id_location,
            }
        return out
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:1693')


def _load_google_manual_values_for_year(year: int) -> Dict[tuple[str, str], Dict[str, Any]]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute(
            """
SELECT store_code, year_month, rating, note
FROM dbo.MboGoogleManualMonthlyValues
WHERE year_month >= ? AND year_month <= ?
""",
            (f"{int(year):04d}-01", f"{int(year):04d}-12"),
        )
        out: Dict[tuple[str, str], Dict[str, Any]] = {}
        for r in cur.fetchall() or []:
            out[(str(r[0] or ""), str(r[1] or ""))] = {
                "rating": float(r[2]) if r[2] is not None else None,
                "note": str(r[3] or ""),
            }
        return out
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:1720')


def save_google_manual_value(store_code: str, year_month: str, rating: Any, note: str | None = None) -> None:
    ensure_mbo_schema()
    rating_dec = _parse_decimal(rating)
    if rating_dec is None:
        raise ValueError("Valore Google non valido.")
    if rating_dec < 0 or rating_dec > 5:
        raise ValueError("Il valore Google deve essere compreso tra 0 e 5.")
    store_refs = _load_valid_store_refs()
    store_ref = next((s for s in store_refs if str(s.get("store_code") or "").strip() == str(store_code or "").strip()), None)
    if not store_ref:
        raise ValueError("Store non valido per il perimetro Supabase.")
    if year_month not in _active_months_for_store(int(str(year_month)[:4]), store_ref.get("closure_date")):
        raise ValueError("Mese non valido: lo store risulta chiuso per quel periodo.")
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
IF EXISTS (SELECT 1 FROM dbo.MboGoogleManualMonthlyValues WHERE store_code = ? AND year_month = ?)
BEGIN
  UPDATE dbo.MboGoogleManualMonthlyValues
  SET rating = ?, note = ?, updated_at = SYSUTCDATETIME()
  WHERE store_code = ? AND year_month = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboGoogleManualMonthlyValues (store_code, year_month, rating, note)
  VALUES (?, ?, ?, ?);
END
""",
            (store_code, year_month, float(rating_dec), note, store_code, year_month, store_code, year_month, float(rating_dec), note),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:1760')


def delete_google_manual_value(store_code: str, year_month: str) -> None:
    ensure_mbo_schema()
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            "DELETE FROM dbo.MboGoogleManualMonthlyValues WHERE store_code = ? AND year_month = ?",
            (str(store_code or "").strip(), str(year_month or "").strip()),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:1777')


def _score_google_rating(rating: float | None) -> float | None:
    if rating is None:
        return None
    rating_f = float(rating)
    if rating_f >= 4.4:
        return 1.0
    if rating_f >= 4.0:
        return 0.5
    return -1.0


def build_mbo_google_scores(year: int) -> Dict[str, Any]:
    ensure_mbo_schema()
    store_refs = _load_valid_store_refs()
    review_avgs = _load_google_review_averages_for_year(year)
    manual_values = _load_google_manual_values_for_year(year)
    audit_matrix = build_audit_month_matrix(year)
    audit_by_pair: Dict[tuple[str, str], Dict[str, Any]] = {}
    for row in audit_matrix.get("rows") or []:
        code = str(row.get("store_code") or "").strip()
        for cell in row.get("months") or []:
            audit_by_pair[(code, str(cell.get("year_month") or ""))] = cell

    month_labels = [
        {"month_num": i, "month_label": lbl, "year_month": f"{int(year):04d}-{i:02d}"}
        for i, lbl in enumerate(["", "Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"])
        if i
    ]

    rows: List[Dict[str, Any]] = []
    points_taken_total = 0.0
    points_theoretical_total = 0.0
    active_months_total = 0
    default_count = 0

    for store in sorted(store_refs, key=lambda s: (_norm_key(s.get("store_name")), _norm_key(s.get("store_code")))):
        code = str(store.get("store_code") or "").strip()
        google_raw = str(store.get("google") or "").strip()
        has_google_id = bool(google_raw) and _norm_key(google_raw) != "vuoto"
        active_months = set(_active_months_for_store(year, store.get("closure_date")))
        months: List[Dict[str, Any]] = []
        store_points = 0.0
        store_theoretical = 0.0
        for month in month_labels:
            ym = str(month["year_month"])
            active = ym in active_months
            if not active:
                months.append({"year_month": ym, "month_label": month["month_label"], "active": False})
                continue

            manual = manual_values.get((code, ym))
            source = "vuoto"
            base_rating = None
            review_count = 0
            review = review_avgs.get((str(store.get("google_norm") or ""), ym))
            if review:
                base_rating = review.get("rating")
                review_count = int(review.get("review_count") or 0)
                source = "importato"
            elif has_google_id:
                base_rating = 5.0
                default_count += 1
                source = "default"

            rating = base_rating
            if manual:
                rating = manual.get("rating")
                source = "manuale"

            raw_points = _score_google_rating(rating)
            theoretical = 1.0 if rating is not None else 0.0
            audit_cell = audit_by_pair.get((code, ym), {})
            audit_score = audit_cell.get("score")
            audit_ok = audit_score is not None and float(audit_score or 0.0) >= 90.0
            final_points = (raw_points or 0.0) if audit_ok and raw_points is not None else 0.0

            points_taken_total += final_points
            points_theoretical_total += theoretical
            active_months_total += 1
            store_points += final_points
            store_theoretical += theoretical

            months.append(
                {
                    "year_month": ym,
                    "month_label": month["month_label"],
                    "active": True,
                    "rating": rating,
                    "base_rating": base_rating,
                    "raw_points": raw_points,
                    "final_points": final_points,
                    "theoretical_points": theoretical,
                    "source": source,
                    "review_count": review_count,
                    "manual": manual,
                    "audit_score": audit_score,
                    "audit_ok": audit_ok,
                }
            )
        rows.append(
            {
                **store,
                "months": months,
                "points_taken": store_points,
                "points_theoretical": store_theoretical,
                "pct": (store_points / store_theoretical * 100.0) if store_theoretical else 0.0,
            }
        )

    pct = (points_taken_total / points_theoretical_total * 100.0) if points_theoretical_total else 0.0
    return {
        "year": int(year),
        "month_labels": month_labels,
        "rows": rows,
        "store_count": len(rows),
        "active_months": active_months_total,
        "points_taken": points_taken_total,
        "points_theoretical": points_theoretical_total,
        "pct": pct,
        "default_count": default_count,
    }


def _detect_glovo_layout(conn) -> Dict[str, str]:
    table_ref = "[dbo].[GLOVO]"
    cur = conn.cursor()
    cur.execute(f"SELECT TOP 0 * FROM {table_ref}")
    columns = [d[0] for d in (cur.description or [])]
    store_col = _find_column(columns, exact=["storeid", "store_id"], contains=["storeid", "store id"])
    year_col = _find_column(columns, exact=["year", "anno"], contains=["year", "anno"])
    week_col = _find_column(columns, exact=["week", "settimana"], contains=["week", "settimana"])
    value_col = _find_column(
        columns,
        exact=["opening_pct", "chiusura", "pct_chiusura", "percentuale", "pct", "score", "rating", "valore", "value"],
        contains=["opening", "chiusura", "percent", "pct", "score", "rating", "valore", "value"],
        exclude=[store_col, year_col, week_col],
    )
    if not store_col or not year_col or not week_col or not value_col:
        raise RuntimeError(f"Colonne GLOVO non trovate. Colonne lette: {columns}")
    return {"table_ref": table_ref, "store_col": store_col, "year_col": year_col, "week_col": week_col, "value_col": value_col}


def _normalize_pct_value(value: Any) -> float | None:
    if value is None:
        return None
    val = _to_float(value)
    if val <= 1.5:
        val *= 100.0
    return val


def _week_months(year: int, week: int) -> List[str]:
    try:
        start = date.fromisocalendar(int(year), int(week), 1)
    except Exception:
        return []
    end = start + timedelta(days=6)
    months = {f"{start.year:04d}-{start.month:02d}", f"{end.year:04d}-{end.month:02d}"}
    return sorted(m for m in months if m.startswith(f"{int(year):04d}-"))


def _load_glovo_month_averages_for_year(year: int) -> Dict[tuple[str, str], Dict[str, Any]]:
    conn = get_connection_ilp(read_only=True)
    try:
        layout = _detect_glovo_layout(conn)
        store_expr = f"LTRIM(RTRIM(CAST([{layout['store_col']}] AS NVARCHAR(255))))"
        year_expr = f"TRY_CONVERT(int, [{layout['year_col']}])"
        week_expr = f"TRY_CONVERT(int, [{layout['week_col']}])"
        value_expr = f"TRY_CONVERT(decimal(9,4), [{layout['value_col']}])"
        sql = (
            f"SELECT {store_expr} AS store_id, {year_expr} AS year_num, {week_expr} AS week_num, {value_expr} AS value_pct "
            f"FROM {layout['table_ref']} "
            f"WHERE {year_expr} IN (?, ?) AND {value_expr} IS NOT NULL"
        )
        cur = conn.cursor()
        cur.execute(sql, (int(year) - 1, int(year)))
        buckets: Dict[tuple[str, str], List[float]] = {}
        weeks_by_pair: Dict[tuple[str, str], int] = {}
        for r in cur.fetchall() or []:
            store_id = str(r[0] or "").strip()
            year_num = int(r[1] or 0)
            week_num = int(r[2] or 0)
            value = _normalize_pct_value(r[3])
            if not store_id or not year_num or not week_num or value is None:
                continue
            for ym in _week_months(year_num, week_num):
                if not ym.startswith(f"{int(year):04d}-"):
                    continue
                key = (_norm_key(store_id), ym)
                buckets.setdefault(key, []).append(value)
                weeks_by_pair[key] = weeks_by_pair.get(key, 0) + 1
        out: Dict[tuple[str, str], Dict[str, Any]] = {}
        for key, values in buckets.items():
            out[key] = {
                "value_pct": (sum(values) / len(values)) if values else None,
                "week_count": weeks_by_pair.get(key, 0),
            }
        return out
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:1982')


def _load_glovo_manual_values_for_year(year: int) -> Dict[tuple[str, str], Dict[str, Any]]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute(
            """
SELECT store_code, year_month, value_pct, note
FROM dbo.MboGlovoManualMonthlyValues
WHERE year_month >= ? AND year_month <= ?
""",
            (f"{int(year):04d}-01", f"{int(year):04d}-12"),
        )
        out: Dict[tuple[str, str], Dict[str, Any]] = {}
        for r in cur.fetchall() or []:
            out[(str(r[0] or ""), str(r[1] or ""))] = {
                "value_pct": float(r[2]) if r[2] is not None else None,
                "note": str(r[3] or ""),
            }
        return out
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2009')


def save_glovo_manual_value(store_code: str, year_month: str, value_pct: Any, note: str | None = None) -> None:
    ensure_mbo_schema()
    pct = _normalize_pct_value(_parse_decimal(value_pct))
    if pct is None:
        raise ValueError("Valore Glovo non valido.")
    if pct < 0 or pct > 100:
        raise ValueError("Il valore Glovo deve essere compreso tra 0 e 100.")
    store_refs = _load_valid_store_refs()
    store_ref = next((s for s in store_refs if str(s.get("store_code") or "").strip() == str(store_code or "").strip()), None)
    if not store_ref:
        raise ValueError("Store non valido per il perimetro Supabase.")
    if year_month not in _active_months_for_store(int(str(year_month)[:4]), store_ref.get("closure_date")):
        raise ValueError("Mese non valido: lo store risulta chiuso per quel periodo.")
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
IF EXISTS (SELECT 1 FROM dbo.MboGlovoManualMonthlyValues WHERE store_code = ? AND year_month = ?)
BEGIN
  UPDATE dbo.MboGlovoManualMonthlyValues
  SET value_pct = ?, note = ?, updated_at = SYSUTCDATETIME()
  WHERE store_code = ? AND year_month = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboGlovoManualMonthlyValues (store_code, year_month, value_pct, note)
  VALUES (?, ?, ?, ?);
END
""",
            (store_code, year_month, float(pct), note, store_code, year_month, store_code, year_month, float(pct), note),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2049')


def delete_glovo_manual_value(store_code: str, year_month: str) -> None:
    ensure_mbo_schema()
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            "DELETE FROM dbo.MboGlovoManualMonthlyValues WHERE store_code = ? AND year_month = ?",
            (str(store_code or "").strip(), str(year_month or "").strip()),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2066')


def _score_glovo_pct(value_pct: float | None) -> float | None:
    if value_pct is None:
        return None
    val = float(value_pct)
    if val >= 97.0:
        return 1.0
    if val >= 90.0:
        return 0.5
    return -1.0


def build_mbo_glovo_scores(year: int) -> Dict[str, Any]:
    ensure_mbo_schema()
    store_refs = _load_valid_store_refs()
    glovo_avgs = _load_glovo_month_averages_for_year(year)
    manual_values = _load_glovo_manual_values_for_year(year)
    audit_matrix = build_audit_month_matrix(year)
    audit_by_pair: Dict[tuple[str, str], Dict[str, Any]] = {}
    for row in audit_matrix.get("rows") or []:
        code = str(row.get("store_code") or "").strip()
        for cell in row.get("months") or []:
            audit_by_pair[(code, str(cell.get("year_month") or ""))] = cell

    month_labels = [
        {"month_num": i, "month_label": lbl, "year_month": f"{int(year):04d}-{i:02d}"}
        for i, lbl in enumerate(["", "Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"])
        if i
    ]

    rows: List[Dict[str, Any]] = []
    points_taken_total = 0.0
    points_theoretical_total = 0.0
    active_months_total = 0

    for store in sorted(store_refs, key=lambda s: (_norm_key(s.get("store_name")), _norm_key(s.get("store_code")))):
        code = str(store.get("store_code") or "").strip()
        active_months = set(_active_months_for_store(year, store.get("closure_date")))
        months: List[Dict[str, Any]] = []
        store_points = 0.0
        store_theoretical = 0.0
        for month in month_labels:
            ym = str(month["year_month"])
            active = ym in active_months
            if not active:
                months.append({"year_month": ym, "month_label": month["month_label"], "active": False})
                continue

            manual = manual_values.get((code, ym))
            source = "vuoto"
            base_value = None
            week_count = 0
            imported = glovo_avgs.get((str(store.get("glovo_norm") or ""), ym))
            if imported:
                base_value = imported.get("value_pct")
                week_count = int(imported.get("week_count") or 0)
                source = "importato"
            value = base_value
            if manual:
                value = manual.get("value_pct")
                source = "manuale"

            raw_points = _score_glovo_pct(value)
            theoretical = 1.0 if value is not None else 0.0
            audit_cell = audit_by_pair.get((code, ym), {})
            audit_score = audit_cell.get("score")
            audit_ok = audit_score is not None and float(audit_score or 0.0) >= 90.0
            final_points = (raw_points or 0.0) if audit_ok and raw_points is not None else 0.0

            points_taken_total += final_points
            points_theoretical_total += theoretical
            active_months_total += 1
            store_points += final_points
            store_theoretical += theoretical

            months.append(
                {
                    "year_month": ym,
                    "month_label": month["month_label"],
                    "active": True,
                    "value_pct": value,
                    "base_value_pct": base_value,
                    "raw_points": raw_points,
                    "final_points": final_points,
                    "theoretical_points": theoretical,
                    "source": source,
                    "week_count": week_count,
                    "manual": manual,
                    "audit_score": audit_score,
                    "audit_ok": audit_ok,
                }
            )
        rows.append(
            {
                **store,
                "months": months,
                "points_taken": store_points,
                "points_theoretical": store_theoretical,
                "pct": (store_points / store_theoretical * 100.0) if store_theoretical else 0.0,
            }
        )

    pct = (points_taken_total / points_theoretical_total * 100.0) if points_theoretical_total else 0.0
    return {
        "year": int(year),
        "month_labels": month_labels,
        "rows": rows,
        "store_count": len(rows),
        "active_months": active_months_total,
        "points_taken": points_taken_total,
        "points_theoretical": points_theoretical_total,
        "pct": pct,
    }


def _detect_deliveroo_layout(conn) -> Dict[str, str]:
    table_ref = "[dbo].[DELIVEROO_RATING]"
    cur = conn.cursor()
    cur.execute(f"SELECT TOP 0 * FROM {table_ref}")
    columns = [d[0] for d in (cur.description or [])]
    store_col = _find_column(columns, exact=["storeid", "store_id"], contains=["storeid", "store id"])
    date_col = _find_column(columns, exact=["data", "date"], contains=["data", "date"])
    rating_col = _find_column(columns, exact=["rating"], contains=["rating"])
    count_col = _find_column(columns, exact=["conteggio", "count", "cnt"], contains=["conteggio", "count"])
    if not store_col or not date_col or not rating_col or not count_col:
        raise RuntimeError(f"Colonne DELIVEROO_RATING non trovate. Colonne lette: {columns}")
    return {"table_ref": table_ref, "store_col": store_col, "date_col": date_col, "rating_col": rating_col, "count_col": count_col}


def _date_months_for_week(day_value: Any, target_year: int) -> List[str]:
    d = day_value
    if isinstance(d, datetime):
        d = d.date()
    elif not isinstance(d, date):
        d = _parse_date(d)
    if not d:
        return []
    start = d - timedelta(days=d.weekday())
    end = start + timedelta(days=6)
    months = {f"{start.year:04d}-{start.month:02d}", f"{end.year:04d}-{end.month:02d}"}
    return sorted(m for m in months if m.startswith(f"{int(target_year):04d}-"))


def _load_deliveroo_month_averages_for_year(year: int) -> Dict[tuple[str, str], Dict[str, Any]]:
    conn = get_connection_ilp(read_only=True)
    try:
        layout = _detect_deliveroo_layout(conn)
        store_expr = f"LTRIM(RTRIM(CAST([{layout['store_col']}] AS NVARCHAR(255))))"
        date_expr = f"TRY_CONVERT(date, [{layout['date_col']}])"
        rating_expr = f"TRY_CONVERT(decimal(9,4), [{layout['rating_col']}])"
        count_expr = f"TRY_CONVERT(decimal(18,4), [{layout['count_col']}])"
        sql = (
            f"SELECT {store_expr} AS store_id, {date_expr} AS week_date, {rating_expr} AS rating, {count_expr} AS count_value "
            f"FROM {layout['table_ref']} "
            f"WHERE {date_expr} >= ? AND {date_expr} < ? AND {rating_expr} IS NOT NULL AND {count_expr} IS NOT NULL"
        )
        cur = conn.cursor()
        cur.execute(sql, (f"{int(year) - 1:04d}-12-20", f"{int(year) + 1:04d}-01-10"))
        totals: Dict[tuple[str, str], Dict[str, float]] = {}
        week_counts: Dict[tuple[str, str], int] = {}
        for r in cur.fetchall() or []:
            store_id = str(r[0] or "").strip()
            rating = _to_float(r[2])
            count_value = _to_float(r[3])
            if not store_id or count_value <= 0:
                continue
            for ym in _date_months_for_week(r[1], year):
                key = (_norm_key(store_id), ym)
                bucket = totals.setdefault(key, {"weighted": 0.0, "count": 0.0})
                bucket["weighted"] += rating * count_value
                bucket["count"] += count_value
                week_counts[key] = week_counts.get(key, 0) + 1
        out: Dict[tuple[str, str], Dict[str, Any]] = {}
        for key, bucket in totals.items():
            cnt = float(bucket.get("count") or 0.0)
            out[key] = {
                "rating": (float(bucket.get("weighted") or 0.0) / cnt) if cnt else None,
                "rating_count": cnt,
                "week_count": week_counts.get(key, 0),
            }
        return out
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2253')


def _load_deliveroo_manual_values_for_year(year: int) -> Dict[tuple[str, str], Dict[str, Any]]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute(
            """
SELECT store_code, year_month, rating, note
FROM dbo.MboDeliverooManualMonthlyValues
WHERE year_month >= ? AND year_month <= ?
""",
            (f"{int(year):04d}-01", f"{int(year):04d}-12"),
        )
        out: Dict[tuple[str, str], Dict[str, Any]] = {}
        for r in cur.fetchall() or []:
            out[(str(r[0] or ""), str(r[1] or ""))] = {
                "rating": float(r[2]) if r[2] is not None else None,
                "note": str(r[3] or ""),
            }
        return out
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2280')


def save_deliveroo_manual_value(store_code: str, year_month: str, rating: Any, note: str | None = None) -> None:
    ensure_mbo_schema()
    rating_dec = _parse_decimal(rating)
    if rating_dec is None:
        raise ValueError("Valore Deliveroo non valido.")
    if rating_dec < 0 or rating_dec > 5:
        raise ValueError("Il valore Deliveroo deve essere compreso tra 0 e 5.")
    store_refs = _load_valid_store_refs()
    store_ref = next((s for s in store_refs if str(s.get("store_code") or "").strip() == str(store_code or "").strip()), None)
    if not store_ref:
        raise ValueError("Store non valido per il perimetro Supabase.")
    if year_month not in _active_months_for_store(int(str(year_month)[:4]), store_ref.get("closure_date")):
        raise ValueError("Mese non valido: lo store risulta chiuso per quel periodo.")
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
IF EXISTS (SELECT 1 FROM dbo.MboDeliverooManualMonthlyValues WHERE store_code = ? AND year_month = ?)
BEGIN
  UPDATE dbo.MboDeliverooManualMonthlyValues
  SET rating = ?, note = ?, updated_at = SYSUTCDATETIME()
  WHERE store_code = ? AND year_month = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboDeliverooManualMonthlyValues (store_code, year_month, rating, note)
  VALUES (?, ?, ?, ?);
END
""",
            (store_code, year_month, float(rating_dec), note, store_code, year_month, store_code, year_month, float(rating_dec), note),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2320')


def delete_deliveroo_manual_value(store_code: str, year_month: str) -> None:
    ensure_mbo_schema()
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            "DELETE FROM dbo.MboDeliverooManualMonthlyValues WHERE store_code = ? AND year_month = ?",
            (str(store_code or "").strip(), str(year_month or "").strip()),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2337')


def _score_deliveroo_rating(rating: float | None) -> float | None:
    if rating is None:
        return None
    val = float(rating)
    if val >= 4.0:
        return 1.0
    if val >= 3.5:
        return 0.5
    return -1.0


def build_mbo_deliveroo_scores(year: int) -> Dict[str, Any]:
    ensure_mbo_schema()
    store_refs = _load_valid_store_refs()
    deliveroo_avgs = _load_deliveroo_month_averages_for_year(year)
    manual_values = _load_deliveroo_manual_values_for_year(year)
    audit_matrix = build_audit_month_matrix(year)
    audit_by_pair: Dict[tuple[str, str], Dict[str, Any]] = {}
    for row in audit_matrix.get("rows") or []:
        code = str(row.get("store_code") or "").strip()
        for cell in row.get("months") or []:
            audit_by_pair[(code, str(cell.get("year_month") or ""))] = cell

    month_labels = [
        {"month_num": i, "month_label": lbl, "year_month": f"{int(year):04d}-{i:02d}"}
        for i, lbl in enumerate(["", "Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"])
        if i
    ]

    rows: List[Dict[str, Any]] = []
    points_taken_total = 0.0
    points_theoretical_total = 0.0
    active_months_total = 0

    for store in sorted(store_refs, key=lambda s: (_norm_key(s.get("store_name")), _norm_key(s.get("store_code")))):
        code = str(store.get("store_code") or "").strip()
        active_months = set(_active_months_for_store(year, store.get("closure_date")))
        months: List[Dict[str, Any]] = []
        store_points = 0.0
        store_theoretical = 0.0
        for month in month_labels:
            ym = str(month["year_month"])
            active = ym in active_months
            if not active:
                months.append({"year_month": ym, "month_label": month["month_label"], "active": False})
                continue

            manual = manual_values.get((code, ym))
            source = "vuoto"
            base_rating = None
            week_count = 0
            rating_count = 0.0
            imported = deliveroo_avgs.get((str(store.get("deliveroo_norm") or ""), ym))
            if imported:
                base_rating = imported.get("rating")
                week_count = int(imported.get("week_count") or 0)
                rating_count = float(imported.get("rating_count") or 0.0)
                source = "importato"
            rating = base_rating
            if manual:
                rating = manual.get("rating")
                source = "manuale"

            raw_points = _score_deliveroo_rating(rating)
            theoretical = 1.0 if rating is not None else 0.0
            audit_cell = audit_by_pair.get((code, ym), {})
            audit_score = audit_cell.get("score")
            audit_ok = audit_score is not None and float(audit_score or 0.0) >= 90.0
            final_points = (raw_points or 0.0) if audit_ok and raw_points is not None else 0.0

            points_taken_total += final_points
            points_theoretical_total += theoretical
            active_months_total += 1
            store_points += final_points
            store_theoretical += theoretical

            months.append(
                {
                    "year_month": ym,
                    "month_label": month["month_label"],
                    "active": True,
                    "rating": rating,
                    "base_rating": base_rating,
                    "raw_points": raw_points,
                    "final_points": final_points,
                    "theoretical_points": theoretical,
                    "source": source,
                    "week_count": week_count,
                    "rating_count": rating_count,
                    "manual": manual,
                    "audit_score": audit_score,
                    "audit_ok": audit_ok,
                }
            )
        rows.append(
            {
                **store,
                "months": months,
                "points_taken": store_points,
                "points_theoretical": store_theoretical,
                "pct": (store_points / store_theoretical * 100.0) if store_theoretical else 0.0,
            }
        )

    pct = (points_taken_total / points_theoretical_total * 100.0) if points_theoretical_total else 0.0
    return {
        "year": int(year),
        "month_labels": month_labels,
        "rows": rows,
        "store_count": len(rows),
        "active_months": active_months_total,
        "points_taken": points_taken_total,
        "points_theoretical": points_theoretical_total,
        "pct": pct,
    }


SOFT_SKILL_ROLES = ["STORE MANAGER", "ASSISTANT", "MULTISTORE"]

SOFT_SKILL_QUESTIONS: List[Dict[str, str]] = [
    {"id": "cogs_01", "area": "controlli", "group": "cogs", "text": "LIVELLI DI PRODUZIONE ADEGUATI"},
    {"id": "cogs_02", "area": "controlli", "group": "cogs", "text": "LIVELLI DI DECONGELO ADEGUATI"},
    {"id": "cogs_03", "area": "controlli", "group": "cogs", "text": "GESTIONE DELLO SCARTO"},
    {"id": "cogs_04", "area": "controlli", "group": "cogs", "text": "VERIFICA DELLE RICETTE DI PRODUZIONE"},
    {"id": "cogs_05", "area": "controlli", "group": "cogs", "text": "UTILIZZO DEGLI SPALLINATORI CORRETTI"},
    {"id": "cogs_06", "area": "controlli", "group": "cogs", "text": "CORRETTO UTILIZZO DEGLI SPALLINATORI"},
    {"id": "cogs_07", "area": "controlli", "group": "cogs", "text": "CONTROLLO SCONTI E CANCELLAZIONI"},
    {"id": "cogs_08", "area": "controlli", "group": "cogs", "text": "CONTROLLO CONSUMI"},
    {"id": "cogs_09", "area": "controlli", "group": "cogs", "text": "ORDINI ADEGUATI"},
    {"id": "labour_01", "area": "controlli", "group": "labour", "text": "CORRETTEZZA DEGLI ORARI"},
    {"id": "labour_02", "area": "controlli", "group": "labour", "text": "PIANIFICAZIONE SECONDO I VOLUMI ATTESI"},
    {"id": "labour_03", "area": "controlli", "group": "labour", "text": "FORMAZIONE DELLO STAFF"},
    {"id": "labour_04", "area": "controlli", "group": "labour", "text": "PRESENZA SUL PIANO"},
    {"id": "labour_05", "area": "controlli", "group": "labour", "text": "ANALISI DELLA PRODUTTIVITA"},
    {"id": "labour_06", "area": "controlli", "group": "labour", "text": "GESTIONE DINAMICA DELLE ORE"},
    {"id": "labour_07", "area": "controlli", "group": "labour", "text": "GESTIONE PIANO FERIE"},
    {"id": "business_01", "area": "business", "group": "business", "text": "FIDELIZZAZIONE DEL CLIENTE"},
    {"id": "business_02", "area": "business", "group": "business", "text": "FORMAZIONE DELLO STAFF SULLE TECNICHE DI VENDITA"},
    {"id": "business_03", "area": "business", "group": "business", "text": "PULIZIA DELLO STORE"},
    {"id": "business_04", "area": "business", "group": "business", "text": "DISPONIBILITA PRODOTTI"},
    {"id": "business_05", "area": "business", "group": "business", "text": "QUALITA PRODOTTI"},
    {"id": "business_06", "area": "business", "group": "business", "text": "GESTIONE COMUNICAZIONE MARKETING IN STORE"},
    {"id": "business_07", "area": "business", "group": "business", "text": "CORRETTA GESTIONE DEI DELIVERY"},
    {"id": "business_08", "area": "business", "group": "business", "text": "INIZIATIVE SUL TERRITORIO"},
    {"id": "people_01", "area": "people", "group": "people", "text": "PIANIFICAZIONE DELLA FORMAZIONE DEI NUOVI ASSUNTI"},
    {"id": "people_02", "area": "people", "group": "people", "text": "GESTIONE DELLE PROBLEMATICHE"},
    {"id": "people_03", "area": "people", "group": "people", "text": "CORRECTIVE COACHING"},
    {"id": "people_04", "area": "people", "group": "people", "text": "ASCOLTO VERSO IL PERSONALE"},
    {"id": "people_05", "area": "people", "group": "people", "text": "MONITORAGGIO PERIODO DI PROVA NUOVI ASSUNTI"},
    {"id": "people_06", "area": "people", "group": "people", "text": "ORARI EQUI E BILANCIATI"},
]


def get_soft_skill_questions() -> List[Dict[str, str]]:
    return [dict(q) for q in SOFT_SKILL_QUESTIONS]


def _validate_year_month(value: str) -> str:
    raw = str(value or "").strip()
    try:
        year_s, month_s = raw.split("-", 1)
        y = int(year_s)
        m = int(month_s)
        if 1 <= m <= 12:
            return f"{y:04d}-{m:02d}"
    except Exception:
        log_swallowed('mbo_repository:2506')
    raise ValueError("Mese periodo non valido.")


def _period_months(start_year_month: str, end_year_month: str) -> List[str]:
    start = _validate_year_month(start_year_month)
    end = _validate_year_month(end_year_month)
    sy, sm = [int(x) for x in start.split("-", 1)]
    ey, em = [int(x) for x in end.split("-", 1)]
    if (ey, em) < (sy, sm):
        raise ValueError("Il mese finale deve essere successivo o uguale al mese iniziale.")
    months: List[str] = []
    y, m = sy, sm
    while (y, m) <= (ey, em):
        months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            y += 1
            m = 1
    return months


def _parse_answers(form_values: Dict[str, Any]) -> Dict[str, int]:
    answers: Dict[str, int] = {}
    for q in SOFT_SKILL_QUESTIONS:
        qid = str(q["id"])
        raw = str(form_values.get(qid) or "").strip()
        if raw == "":
            raise ValueError("Tutte le domande soft skills sono obbligatorie.")
        try:
            val = int(raw)
        except Exception as exc:
            raise ValueError("Le risposte soft skills devono essere da 1 a 4.") from exc
        if val < 1 or val > 4:
            raise ValueError("Le risposte soft skills devono essere da 1 a 4.")
        answers[qid] = val
    return answers


def _average(values: List[float]) -> float | None:
    clean = [float(v) for v in values if v is not None]
    return (sum(clean) / len(clean)) if clean else None


def _score_from_points(points_taken: float, points_theoretical: float) -> float | None:
    if not points_theoretical:
        return None
    pct = (float(points_taken or 0.0) / float(points_theoretical or 0.0)) * 100.0
    pct = max(0.0, min(100.0, pct))
    return 1.0 + (pct / 100.0 * 3.0)


def _weighted_score(items: List[tuple[float | None, float]]) -> float | None:
    total_weight = sum(float(w) for score, w in items if score is not None)
    if not total_weight:
        return None
    return sum(float(score) * float(weight) for score, weight in items if score is not None) / total_weight


def _find_matrix_row(matrix: Dict[str, Any], store_code: str) -> Dict[str, Any] | None:
    target = str(store_code or "").strip()
    for row in matrix.get("rows") or []:
        if str(row.get("store_code") or "").strip() == target:
            return row
    return None


def _aggregate_month_cells(row: Dict[str, Any] | None, months: List[str], metric: str) -> Dict[str, Any]:
    points = 0.0
    theoretical = 0.0
    month_details: List[Dict[str, Any]] = []
    if not row:
        return {"points_taken": 0.0, "points_theoretical": 0.0, "pct": None, "score": None, "months": []}

    by_month = {str(cell.get("year_month") or ""): cell for cell in row.get("months") or []}
    for ym in months:
        cell = by_month.get(ym) or {}
        if not cell.get("active", True):
            continue
        taken = 0.0
        theoretical_cell = 0.0
        if metric in {"REVENUES", "COGS", "LABOUR COST"}:
            voice = (cell.get("voice_scores") or {}).get(metric) or {}
            theoretical_cell = 1.0 if voice else 0.0
            taken = float(voice.get("points") or 0.0) if cell.get("audit_ok") else 0.0
        else:
            theoretical_cell = float(cell.get("theoretical_points") or 0.0)
            taken = float(cell.get("final_points") or 0.0)
        points += taken
        theoretical += theoretical_cell
        month_details.append({"year_month": ym, "points_taken": taken, "points_theoretical": theoretical_cell})

    score = _score_from_points(points, theoretical)
    return {
        "points_taken": points,
        "points_theoretical": theoretical,
        "pct": (points / theoretical * 100.0) if theoretical else None,
        "score": score,
        "months": month_details,
    }


def _soft_skill_auto_metrics(store_code: str, start_year_month: str, end_year_month: str) -> Dict[str, Any]:
    months = _period_months(start_year_month, end_year_month)
    years = sorted({int(ym[:4]) for ym in months})
    metrics: Dict[str, Dict[str, Any]] = {
        "cogs": {"points_taken": 0.0, "points_theoretical": 0.0, "months": []},
        "labour": {"points_taken": 0.0, "points_theoretical": 0.0, "months": []},
        "revenues": {"points_taken": 0.0, "points_theoretical": 0.0, "months": []},
        "google": {"points_taken": 0.0, "points_theoretical": 0.0, "months": []},
        "rating_totali": {"points_taken": 0.0, "points_theoretical": 0.0, "months": []},
    }

    for year in years:
        year_months = [ym for ym in months if int(ym[:4]) == year]
        pnl_row = _find_matrix_row(build_mbo_pnl_scores(year), store_code)
        google_row = _find_matrix_row(build_mbo_google_scores(year), store_code)
        glovo_row = _find_matrix_row(build_mbo_glovo_scores(year), store_code)
        deliveroo_row = _find_matrix_row(build_mbo_deliveroo_scores(year), store_code)

        for key, row, metric in (
            ("cogs", pnl_row, "COGS"),
            ("labour", pnl_row, "LABOUR COST"),
            ("revenues", pnl_row, "REVENUES"),
            ("google", google_row, "monthly"),
        ):
            part = _aggregate_month_cells(row, year_months, metric)
            metrics[key]["points_taken"] += part["points_taken"]
            metrics[key]["points_theoretical"] += part["points_theoretical"]
            metrics[key]["months"].extend(part["months"])

        for row in (glovo_row, deliveroo_row):
            part = _aggregate_month_cells(row, year_months, "monthly")
            metrics["rating_totali"]["points_taken"] += part["points_taken"]
            metrics["rating_totali"]["points_theoretical"] += part["points_theoretical"]
            metrics["rating_totali"]["months"].extend(part["months"])

    for key, metric in metrics.items():
        theoretical = float(metric.get("points_theoretical") or 0.0)
        taken = float(metric.get("points_taken") or 0.0)
        metric["pct"] = (taken / theoretical * 100.0) if theoretical else None
        metric["score"] = _score_from_points(taken, theoretical)
    return metrics


def _compute_soft_skill_result(answers: Dict[str, int], auto_metrics: Dict[str, Any]) -> Dict[str, Any]:
    groups: Dict[str, List[float]] = defaultdict(list)
    for q in SOFT_SKILL_QUESTIONS:
        groups[str(q["group"])].append(float(answers.get(str(q["id"]), 0)))

    cogs_manual = _average(groups["cogs"])
    labour_manual = _average(groups["labour"])
    business_manual = _average(groups["business"])
    people_score = _average(groups["people"])

    cogs_score = _weighted_score([(cogs_manual, 50.0), ((auto_metrics.get("cogs") or {}).get("score"), 50.0)])
    labour_score = _weighted_score([(labour_manual, 50.0), ((auto_metrics.get("labour") or {}).get("score"), 50.0)])
    controls_score = _weighted_score([(cogs_score, 50.0), (labour_score, 50.0)])
    business_score = _weighted_score(
        [
            (business_manual, 50.5),
            ((auto_metrics.get("google") or {}).get("score"), 16.5),
            ((auto_metrics.get("rating_totali") or {}).get("score"), 16.5),
            ((auto_metrics.get("revenues") or {}).get("score"), 16.5),
        ]
    )
    final_score = _weighted_score([(controls_score, 40.0), (business_score, 30.0), (people_score, 30.0)])
    return {
        "groups": {
            "cogs_manual": cogs_manual,
            "labour_manual": labour_manual,
            "business_manual": business_manual,
            "people": people_score,
            "cogs": cogs_score,
            "labour": labour_score,
            "controlli": controls_score,
            "business": business_score,
        },
        "auto_metrics": auto_metrics,
        "final_score": final_score,
    }


def list_soft_skill_periods(include_inactive: bool = True) -> List[Dict[str, Any]]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        sql = """
SELECT period_uuid, token, period_label, start_year_month, end_year_month, is_active, created_by, created_at, updated_at
FROM dbo.MboSoftSkillPeriods
"""
        if not include_inactive:
            sql += " WHERE is_active = 1"
        sql += " ORDER BY start_year_month DESC, created_at DESC"
        cur.execute(sql)
        rows = []
        for r in cur.fetchall() or []:
            rows.append(
                {
                    "period_uuid": str(r[0]),
                    "token": str(r[1] or ""),
                    "period_label": str(r[2] or ""),
                    "start_year_month": str(r[3] or ""),
                    "end_year_month": str(r[4] or ""),
                    "is_active": bool(r[5]),
                    "created_by": str(r[6] or ""),
                    "created_at": r[7],
                    "updated_at": r[8],
                }
            )
        return rows
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2722')


def get_soft_skill_period_by_token(token: str, require_active: bool = False) -> Dict[str, Any] | None:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        sql = """
SELECT period_uuid, token, period_label, start_year_month, end_year_month, is_active, created_by, created_at, updated_at
FROM dbo.MboSoftSkillPeriods
WHERE token = ?
"""
        params: list[Any] = [str(token or "").strip()]
        if require_active:
            sql += " AND is_active = 1"
        cur.execute(sql, params)
        r = cur.fetchone()
        if not r:
            return None
        return {
            "period_uuid": str(r[0]),
            "token": str(r[1] or ""),
            "period_label": str(r[2] or ""),
            "start_year_month": str(r[3] or ""),
            "end_year_month": str(r[4] or ""),
            "is_active": bool(r[5]),
            "created_by": str(r[6] or ""),
            "created_at": r[7],
            "updated_at": r[8],
        }
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2757')


def get_soft_skill_period(period_uuid: str) -> Dict[str, Any] | None:
    token = ""
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute("SELECT token FROM dbo.MboSoftSkillPeriods WHERE period_uuid = ?", (str(period_uuid or "").strip(),))
        r = cur.fetchone()
        token = str(r[0] or "") if r else ""
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2772')
    return get_soft_skill_period_by_token(token) if token else None


def save_soft_skill_period(*, period_uuid: str | None, period_label: str, start_year_month: str, end_year_month: str, is_active: bool, created_by: str | None = None) -> str:
    ensure_mbo_schema()
    label = str(period_label or "").strip()
    if not label:
        raise ValueError("Nome periodo obbligatorio.")
    start = _validate_year_month(start_year_month)
    end = _validate_year_month(end_year_month)
    _period_months(start, end)
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        if period_uuid:
            cur.execute(
                """
UPDATE dbo.MboSoftSkillPeriods
SET period_label = ?, start_year_month = ?, end_year_month = ?, is_active = ?, updated_at = SYSUTCDATETIME()
WHERE period_uuid = ?
""",
                (label, start, end, 1 if is_active else 0, str(period_uuid)),
            )
            conn.commit()
            return str(period_uuid)
        token = uuid.uuid4().hex
        cur.execute(
            """
INSERT INTO dbo.MboSoftSkillPeriods (token, period_label, start_year_month, end_year_month, is_active, created_by)
OUTPUT inserted.period_uuid
VALUES (?, ?, ?, ?, ?, ?)
""",
            (token, label, start, end, 1 if is_active else 0, str(created_by or "")),
        )
        new_id = str(cur.fetchone()[0])
        conn.commit()
        return new_id
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2814')


def _valid_store_codes() -> Dict[str, Dict[str, Any]]:
    return {str(s.get("store_code") or "").strip(): s for s in _load_valid_store_refs() if str(s.get("store_code") or "").strip()}


def _clean_store_code_list(values: List[Any], valid_codes: Dict[str, Dict[str, Any]]) -> List[str]:
    out: List[str] = []
    seen = set()
    for raw in values or []:
        code = str(raw or "").strip()
        if not code or code in seen:
            continue
        if code not in valid_codes:
            raise ValueError(f"Store non valido: {code}.")
        seen.add(code)
        out.append(code)
    return out


def list_soft_skill_submissions(period_uuid: str | None = None, include_deleted: bool = False) -> List[Dict[str, Any]]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        sql = """
SELECT s.submission_uuid, s.group_uuid, s.period_uuid, p.token, p.period_label, p.start_year_month, p.end_year_month,
       s.full_name, s.role, s.store_code, s.answers_json, s.computed_json, s.final_score,
       s.created_by, s.updated_by, s.created_at, s.updated_at, s.deleted_at
FROM dbo.MboSoftSkillSubmissions s
JOIN dbo.MboSoftSkillPeriods p ON p.period_uuid = s.period_uuid
WHERE 1 = 1
"""
        params: list[Any] = []
        if period_uuid:
            sql += " AND s.period_uuid = ?"
            params.append(str(period_uuid))
        if not include_deleted:
            sql += " AND s.deleted_at IS NULL"
        sql += " ORDER BY s.created_at DESC"
        cur.execute(sql, params)
        rows = []
        stores = _valid_store_codes()
        for r in cur.fetchall() or []:
            code = str(r[9] or "").strip()
            computed = {}
            try:
                computed = json.loads(str(r[11] or "{}"))
            except Exception:
                computed = {}
            rows.append(
                {
                    "submission_uuid": str(r[0]),
                    "group_uuid": str(r[1] or ""),
                    "period_uuid": str(r[2]),
                    "token": str(r[3] or ""),
                    "period_label": str(r[4] or ""),
                    "start_year_month": str(r[5] or ""),
                    "end_year_month": str(r[6] or ""),
                    "full_name": str(r[7] or ""),
                    "role": str(r[8] or ""),
                    "store_code": code,
                    "store_name": str((stores.get(code) or {}).get("store_name") or code),
                    "answers": json.loads(str(r[10] or "{}")),
                    "computed": computed,
                    "final_score": float(r[12]) if r[12] is not None else None,
                    "created_by": str(r[13] or ""),
                    "updated_by": str(r[14] or ""),
                    "created_at": r[15],
                    "updated_at": r[16],
                    "deleted_at": r[17],
                }
            )
        return rows
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2893')


def get_soft_skill_submission(submission_uuid: str) -> Dict[str, Any] | None:
    rows = list_soft_skill_submissions(include_deleted=False)
    target = str(submission_uuid or "").strip()
    return next((r for r in rows if str(r.get("submission_uuid") or "") == target), None)


def save_soft_skill_submission(
    *,
    token: str,
    submission_uuid: str | None,
    full_name: str,
    role: str,
    store_code: str,
    answers_source: Dict[str, Any],
    user_id: str | None = None,
    store_codes: List[Any] | None = None,
) -> str:
    ensure_mbo_schema()
    period = get_soft_skill_period_by_token(token, require_active=False)
    if not period:
        raise ValueError("Link soft skills non valido.")
    if not period.get("is_active") and not submission_uuid:
        raise ValueError("Link soft skills non attivo.")
    name = str(full_name or "").strip()
    if not name:
        raise ValueError("Nome e cognome obbligatorio.")
    role_clean = str(role or "").strip().upper()
    if role_clean not in SOFT_SKILL_ROLES:
        raise ValueError("Ruolo non valido.")
    stores = _valid_store_codes()
    requested_codes = [store_code]
    if role_clean == "MULTISTORE":
        requested_codes.extend(store_codes or [])
    codes = _clean_store_code_list(requested_codes, stores)
    if not codes:
        raise ValueError("Store non valido.")
    if role_clean != "MULTISTORE":
        codes = codes[:1]
    answers = _parse_answers(answers_source)
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        if submission_uuid:
            code = codes[0]
            auto_metrics = _soft_skill_auto_metrics(code, str(period["start_year_month"]), str(period["end_year_month"]))
            computed = _compute_soft_skill_result(answers, auto_metrics)
            final_score = computed.get("final_score")
            cur.execute(
                """
UPDATE dbo.MboSoftSkillSubmissions
SET full_name = ?, role = ?, store_code = ?, answers_json = ?, computed_json = ?,
    final_score = ?, updated_by = ?, updated_at = SYSUTCDATETIME()
WHERE submission_uuid = ? AND deleted_at IS NULL
""",
                (
                    name,
                    role_clean,
                    code,
                    json.dumps(answers, ensure_ascii=False),
                    json.dumps(computed, ensure_ascii=False, default=str),
                    float(final_score) if final_score is not None else None,
                    str(user_id or ""),
                    str(submission_uuid),
                ),
            )
            conn.commit()
            return str(submission_uuid)

        group_uuid = str(uuid.uuid4())
        new_ids: List[str] = []
        for code in codes:
            auto_metrics = _soft_skill_auto_metrics(code, str(period["start_year_month"]), str(period["end_year_month"]))
            computed = _compute_soft_skill_result(answers, auto_metrics)
            final_score = computed.get("final_score")
            cur.execute(
                """
INSERT INTO dbo.MboSoftSkillSubmissions
  (group_uuid, period_uuid, full_name, role, store_code, answers_json, computed_json, final_score, created_by, updated_by)
OUTPUT inserted.submission_uuid
VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
""",
                (
                    group_uuid,
                    str(period["period_uuid"]),
                    name,
                    role_clean,
                    code,
                    json.dumps(answers, ensure_ascii=False),
                    json.dumps(computed, ensure_ascii=False, default=str),
                    float(final_score) if final_score is not None else None,
                    str(user_id or ""),
                    str(user_id or ""),
                ),
            )
            new_ids.append(str(cur.fetchone()[0]))
        conn.commit()
        return new_ids[0] if new_ids else ""
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:2997')


def recalculate_soft_skill_submissions(period_uuid: str | None = None, user_id: str | None = None) -> Dict[str, Any]:
    ensure_mbo_schema()
    rows = list_soft_skill_submissions(period_uuid=period_uuid, include_deleted=False)
    updated = 0
    missing_pnl = 0
    details: List[Dict[str, Any]] = []
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        for row in rows:
            answers = row.get("answers") or {}
            code = str(row.get("store_code") or "").strip()
            start_ym = str(row.get("start_year_month") or "").strip()
            end_ym = str(row.get("end_year_month") or "").strip()
            if not code or not start_ym or not end_ym:
                continue
            auto_metrics = _soft_skill_auto_metrics(code, start_ym, end_ym)
            computed = _compute_soft_skill_result(answers, auto_metrics)
            final_score = computed.get("final_score")
            pnl_missing_keys = [
                key
                for key in ("cogs", "labour", "revenues")
                if not float((auto_metrics.get(key) or {}).get("points_theoretical") or 0.0)
            ]
            if pnl_missing_keys:
                missing_pnl += 1
                details.append(
                    {
                        "submission_uuid": row.get("submission_uuid"),
                        "full_name": row.get("full_name"),
                        "store_code": code,
                        "store_name": row.get("store_name"),
                        "missing": pnl_missing_keys,
                    }
                )
            cur.execute(
                """
UPDATE dbo.MboSoftSkillSubmissions
SET computed_json = ?, final_score = ?, updated_by = ?, updated_at = SYSUTCDATETIME()
WHERE submission_uuid = ? AND deleted_at IS NULL
""",
                (
                    json.dumps(computed, ensure_ascii=False, default=str),
                    float(final_score) if final_score is not None else None,
                    str(user_id or ""),
                    str(row.get("submission_uuid") or ""),
                ),
            )
            updated += int(cur.rowcount or 0)
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3054')
    return {"updated": updated, "missing_pnl": missing_pnl, "details": details}


def delete_soft_skill_submission(submission_uuid: str) -> None:
    ensure_mbo_schema()
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            "UPDATE dbo.MboSoftSkillSubmissions SET deleted_at = SYSUTCDATETIME(), updated_at = SYSUTCDATETIME() WHERE submission_uuid = ?",
            (str(submission_uuid or "").strip(),),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3072')


def build_soft_skill_dashboard(period_uuid: str | None = None) -> Dict[str, Any]:
    periods = list_soft_skill_periods(include_inactive=True)
    selected_uuid = str(period_uuid or "").strip() or (str(periods[0].get("period_uuid")) if periods else "")
    submissions = list_soft_skill_submissions(selected_uuid) if selected_uuid else []
    return {
        "periods": periods,
        "selected_period_uuid": selected_uuid,
        "submissions": submissions,
        "submission_count": len(submissions),
        "avg_score": _average([float(s.get("final_score")) for s in submissions if s.get("final_score") is not None]),
        "stores": sorted(_valid_store_codes().values(), key=lambda s: (_norm_key(s.get("store_name")), _norm_key(s.get("store_code")))),
        "roles": SOFT_SKILL_ROLES,
        "questions": get_soft_skill_questions(),
    }


def _survey_token() -> str:
    return uuid.uuid4().hex + uuid.uuid4().hex[:8]


def _survey_target_rows(target_type: str) -> List[Dict[str, str]]:
    target_type = str(target_type or "").strip()
    if target_type == "area_manager":
        return [
            {"target_type": "area_manager", "target_code": str(r.get("row_uuid") or ""), "target_name": str(r.get("manager_name") or "")}
            for r in list_area_managers(include_inactive=False)
            if str(r.get("row_uuid") or "").strip() and str(r.get("manager_name") or "").strip()
        ]
    if target_type == "store_manager":
        return [
            {"target_type": "store_manager", "target_code": str(s.get("store_code") or ""), "target_name": str(s.get("store_name") or s.get("store_code") or "")}
            for s in _load_valid_store_refs()
            if str(s.get("store_code") or "").strip()
        ]
    raise ValueError("Target survey non valido.")


def parse_custom_survey_questions(form: Any) -> List[Dict[str, Any]]:
    questions: List[Dict[str, Any]] = []
    texts = list(form.getlist("question_text"))
    types = list(form.getlist("question_type"))
    categories = list(form.getlist("question_category"))
    max_scores = list(form.getlist("question_max_score"))
    options_list = list(form.getlist("question_options"))
    for idx, text in enumerate(texts):
        q_text = str(text or "").strip()
        if not q_text:
            continue
        q_type = str(types[idx] if idx < len(types) else "score_1_4").strip() or "score_1_4"
        if q_type not in {"score_1_4", "single_choice", "text", "number"}:
            q_type = "score_1_4"
        category = str(categories[idx] if idx < len(categories) else "").strip() or "Generale"
        max_score_dec = _parse_decimal(max_scores[idx] if idx < len(max_scores) else "")
        max_score = float(max_score_dec) if max_score_dec is not None else (4.0 if q_type == "score_1_4" else 0.0)
        if q_type in {"text", "number"}:
            max_score = 0.0
        options = []
        if q_type == "single_choice":
            for opt_idx, line in enumerate(str(options_list[idx] if idx < len(options_list) else "").splitlines()):
                raw = str(line or "").strip()
                if not raw:
                    continue
                label, sep, score_raw = raw.partition("=")
                if not sep:
                    label, sep, score_raw = raw.partition(";")
                score_dec = _parse_decimal(score_raw) if sep else Decimal("0")
                options.append({"text": label.strip(), "score": float(score_dec) if score_dec is not None else 0.0, "sort_order": opt_idx})
            if not options:
                raise ValueError(f"Inserisci almeno una opzione per la domanda: {q_text}")
            max_score = max([float(o.get("score") or 0.0) for o in options] + [max_score, 0.0])
        questions.append({"text": q_text, "type": q_type, "category": category, "max_score": max_score, "sort_order": len(questions), "required": True, "options": options})
    if not questions:
        raise ValueError("Inserisci almeno una domanda.")
    return questions


def _hnorm(value: Any) -> str:
    raw = str(value or "").strip().lower()
    return "".join(ch for ch in raw if ch.isalnum())


def _decode_survey_csv(content: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return (content or b"").decode(enc)
        except Exception:
            continue
    return (content or b"").decode("utf-8", errors="replace")


def parse_custom_survey_questions_file(filename: str, content: bytes) -> List[Dict[str, Any]]:
    filename_l = str(filename or "").lower()
    rows: List[Dict[str, Any]] = []
    if filename_l.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise ValueError(f"Lettura Excel non disponibile: {e}")
        wb = load_workbook(io.BytesIO(content or b""), read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            raise ValueError("File domande vuoto.")
        headers = [str(x or "").strip() for x in values[0]]
        for vals in values[1:]:
            rows.append({headers[i]: vals[i] if i < len(vals) else "" for i in range(len(headers))})
    else:
        text = _decode_survey_csv(content or b"")
        if not text.strip():
            raise ValueError("File domande vuoto.")
        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|,")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        rows = list(csv.DictReader(io.StringIO(text), dialect=dialect))

    if not rows:
        raise ValueError("Nessuna domanda trovata nel file.")

    def pick(row: Dict[str, Any], *names: str) -> Any:
        by_norm = {_hnorm(k): v for k, v in (row or {}).items()}
        for name in names:
            if _hnorm(name) in by_norm:
                return by_norm[_hnorm(name)]
        return ""

    questions: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows):
        text = str(pick(row, "domanda", "question", "testo", "question_text") or "").strip()
        if not text:
            continue
        q_type = str(pick(row, "tipo", "type", "question_type", "tipo risposta") or "score_1_4").strip().lower()
        type_map = {
            "punteggio": "score_1_4",
            "score": "score_1_4",
            "score_1_4": "score_1_4",
            "1-4": "score_1_4",
            "scelta": "single_choice",
            "single_choice": "single_choice",
            "choice": "single_choice",
            "testo": "text",
            "text": "text",
            "numero": "number",
            "number": "number",
        }
        q_type = type_map.get(q_type, q_type)
        if q_type not in {"score_1_4", "single_choice", "text", "number"}:
            q_type = "score_1_4"
        category = str(pick(row, "categoria", "category", "area") or "Generale").strip() or "Generale"
        max_score_dec = _parse_decimal(pick(row, "punteggio massimo", "max_score", "max", "peso"))
        max_score = float(max_score_dec) if max_score_dec is not None else (4.0 if q_type == "score_1_4" else 0.0)
        options = []
        if q_type == "single_choice":
            raw_options = str(pick(row, "opzioni", "options", "risposte") or "").strip()
            for opt_idx, chunk in enumerate(raw_options.replace("\r", "\n").replace("|", "\n").split("\n")):
                raw = str(chunk or "").strip()
                if not raw:
                    continue
                label, sep, score_raw = raw.partition("=")
                if not sep:
                    label, sep, score_raw = raw.partition(";")
                score_dec = _parse_decimal(score_raw) if sep else Decimal("0")
                options.append({"text": label.strip(), "score": float(score_dec) if score_dec is not None else 0.0, "sort_order": opt_idx})
            if not options:
                raise ValueError(f"Riga {idx + 2}: per una scelta personalizzata servono opzioni nel formato testo=punteggio.")
            max_score = max([float(o.get("score") or 0.0) for o in options] + [max_score, 0.0])
        if q_type in {"text", "number"}:
            max_score = 0.0
        questions.append({"text": text, "type": q_type, "category": category, "max_score": max_score, "sort_order": len(questions), "required": True, "options": options})
    if not questions:
        raise ValueError("Nessuna domanda valida trovata nel file.")
    return questions


def create_custom_survey(*, survey_name: str, display_name: str | None = None, target_type: str, start_year_month: str, end_year_month: str, is_active: bool, questions: List[Dict[str, Any]], created_by: str | None = None) -> str:
    ensure_mbo_schema()
    survey_name = str(survey_name or "").strip()
    display_name = str(display_name or "").strip() or None
    target_type = str(target_type or "").strip()
    start_year_month = str(start_year_month or "").strip()
    end_year_month = str(end_year_month or "").strip()
    if not survey_name:
        raise ValueError("Nome survey obbligatorio.")
    if target_type not in {"store_manager", "area_manager"}:
        raise ValueError("Associazione survey non valida.")
    if not start_year_month or not end_year_month or start_year_month > end_year_month:
        raise ValueError("Periodo survey non valido.")
    targets = _survey_target_rows(target_type)
    if not targets:
        raise ValueError("Nessun target disponibile per generare i link.")
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
INSERT INTO dbo.MboSurveys (survey_name, display_name, target_type, start_year_month, end_year_month, is_active, created_by)
OUTPUT inserted.survey_uuid
VALUES (?, ?, ?, ?, ?, ?, ?)
""",
            (survey_name, display_name, target_type, start_year_month, end_year_month, 1 if is_active else 0, created_by),
        )
        survey_uuid = str(cur.fetchone()[0])
        for q in questions:
            cur.execute(
                """
INSERT INTO dbo.MboSurveyQuestions (survey_uuid, question_text, question_type, category, max_score, sort_order, is_required)
OUTPUT inserted.question_uuid
VALUES (?, ?, ?, ?, ?, ?, ?)
""",
                (survey_uuid, q["text"], q["type"], q["category"], float(q.get("max_score") or 0.0), int(q.get("sort_order") or 0), 1),
            )
            question_uuid = str(cur.fetchone()[0])
            for opt in q.get("options") or []:
                cur.execute(
                    "INSERT INTO dbo.MboSurveyQuestionOptions (question_uuid, option_text, score, sort_order) VALUES (?, ?, ?, ?)",
                    (question_uuid, str(opt.get("text") or "").strip(), float(opt.get("score") or 0.0), int(opt.get("sort_order") or 0)),
                )
        for target in targets:
            cur.execute(
                "INSERT INTO dbo.MboSurveyLinks (survey_uuid, token, target_type, target_code, target_name, is_active) VALUES (?, ?, ?, ?, ?, ?)",
                (survey_uuid, _survey_token(), target["target_type"], target["target_code"], target["target_name"], 1 if is_active else 0),
            )
        conn.commit()
        return survey_uuid
    except Exception:
        try:
            conn.rollback()
        except Exception:
            log_swallowed('mbo_repository:3305')
        raise
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3311')


def set_custom_survey_active(survey_uuid: str, is_active: bool) -> None:
    ensure_mbo_schema()
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
UPDATE dbo.MboSurveys SET is_active = ?, updated_at = SYSUTCDATETIME() WHERE survey_uuid = ?;
UPDATE dbo.MboSurveyLinks SET is_active = ?, updated_at = SYSUTCDATETIME() WHERE survey_uuid = ?;
""",
            (1 if is_active else 0, str(survey_uuid or "").strip(), 1 if is_active else 0, str(survey_uuid or "").strip()),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3331')


def update_custom_survey_names(survey_uuid: str, survey_name: str, display_name: str | None = None) -> None:
    ensure_mbo_schema()
    sid = str(survey_uuid or "").strip()
    name = str(survey_name or "").strip()
    display = str(display_name or "").strip() or None
    if not sid:
        raise ValueError("Survey non valida.")
    if not name:
        raise ValueError("Nome survey interno obbligatorio.")
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
UPDATE dbo.MboSurveys
SET survey_name = ?, display_name = ?, updated_at = SYSUTCDATETIME()
WHERE survey_uuid = ?;
""",
            (name, display, sid),
        )
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3359')


def get_custom_survey(token: str | None = None, survey_uuid: str | None = None) -> Dict[str, Any] | None:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        if token:
            cur.execute(
                """
SELECT s.survey_uuid, s.survey_name, s.display_name, s.target_type, s.start_year_month, s.end_year_month, s.is_active,
       l.link_uuid, l.token, l.target_code, l.target_name, l.is_active
FROM dbo.MboSurveyLinks l
JOIN dbo.MboSurveys s ON s.survey_uuid = l.survey_uuid
WHERE l.token = ?
""",
                (str(token or "").strip(),),
            )
        else:
            cur.execute(
                """
SELECT s.survey_uuid, s.survey_name, s.display_name, s.target_type, s.start_year_month, s.end_year_month, s.is_active,
       CAST(NULL AS UNIQUEIDENTIFIER), CAST(NULL AS NVARCHAR(80)), CAST(NULL AS NVARCHAR(80)), CAST(NULL AS NVARCHAR(255)), CAST(1 AS BIT)
FROM dbo.MboSurveys s
WHERE s.survey_uuid = ?
""",
                (str(survey_uuid or "").strip(),),
            )
        r = cur.fetchone()
        if not r:
            return None
        survey = {
            "survey_uuid": str(r[0]), "survey_name": str(r[1] or ""), "display_name": str(r[2] or "") or str(r[1] or ""), "target_type": str(r[3] or ""),
            "start_year_month": str(r[4] or ""), "end_year_month": str(r[5] or ""), "is_active": bool(r[6]),
            "link_uuid": str(r[7]) if r[7] else "", "token": str(r[8] or ""), "target_code": str(r[9] or ""),
            "target_name": str(r[10] or ""), "link_active": bool(r[11]), "questions": [],
        }
        cur.execute(
            """
SELECT question_uuid, question_text, question_type, category, max_score, sort_order, is_required
FROM dbo.MboSurveyQuestions
WHERE survey_uuid = ?
ORDER BY sort_order, created_at
""",
            (survey["survey_uuid"],),
        )
        for q in cur.fetchall() or []:
            qd = {"question_uuid": str(q[0]), "text": str(q[1] or ""), "type": str(q[2] or ""), "category": str(q[3] or "Generale"), "max_score": float(q[4] or 0.0), "sort_order": int(q[5] or 0), "required": bool(q[6]), "options": []}
            cur2 = conn.cursor()
            cur2.execute("SELECT option_uuid, option_text, score, sort_order FROM dbo.MboSurveyQuestionOptions WHERE question_uuid = ? ORDER BY sort_order, option_text", (qd["question_uuid"],))
            qd["options"] = [{"option_uuid": str(o[0]), "text": str(o[1] or ""), "score": float(o[2] or 0.0), "sort_order": int(o[3] or 0)} for o in cur2.fetchall() or []]
            try:
                cur2.close()
            except Exception:
                log_swallowed('mbo_repository:3414')
            survey["questions"].append(qd)
        return survey
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3421')


def _compute_custom_survey_score(survey: Dict[str, Any], answers: Dict[str, Any]) -> Dict[str, Any]:
    total = 0.0
    theoretical = 0.0
    categories: Dict[str, Dict[str, float]] = {}
    for q in survey.get("questions") or []:
        qid = str(q.get("question_uuid") or "")
        q_type = str(q.get("type") or "")
        category = str(q.get("category") or "Generale")
        raw = answers.get(qid)
        max_score = float(q.get("max_score") or 0.0)
        score = 0.0
        if q_type == "score_1_4":
            theoretical += max_score
            try:
                score = max(0.0, min(4.0, float(raw or 0.0))) / 4.0 * max_score
            except Exception:
                score = 0.0
        elif q_type == "single_choice":
            if max_score <= 0:
                continue
            theoretical += max_score
            match = next((o for o in q.get("options") or [] if str(o.get("option_uuid") or "") == str(raw or "")), None)
            score = float((match or {}).get("score") or 0.0)
        else:
            continue
        total += score
        bucket = categories.setdefault(category, {"score": 0.0, "theoretical": 0.0, "pct": 0.0})
        bucket["score"] += score
        bucket["theoretical"] += max_score
    for bucket in categories.values():
        bucket["pct"] = (bucket["score"] / bucket["theoretical"] * 100.0) if bucket["theoretical"] else 0.0
    return {"score": total, "theoretical": theoretical, "pct": (total / theoretical * 100.0) if theoretical else None, "categories": categories}


def save_custom_survey_submission(*, token: str, respondent_name: str | None, answers_source: Any) -> str:
    survey = get_custom_survey(token=token)
    if not survey:
        raise ValueError("Link survey non valido.")
    if not survey.get("is_active") or not survey.get("link_active"):
        raise ValueError("Questa survey non e attiva.")
    answers: Dict[str, Any] = {}
    for q in survey.get("questions") or []:
        qid = str(q.get("question_uuid") or "")
        value = str(answers_source.get(qid) or "").strip()
        if q.get("required") and not value:
            raise ValueError(f"Risposta obbligatoria mancante: {q.get('text')}")
        answers[qid] = value
    computed = _compute_custom_survey_score(survey, answers)
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute(
            """
INSERT INTO dbo.MboSurveySubmissions (survey_uuid, link_uuid, target_type, target_code, target_name, respondent_name, answers_json, computed_json, score_pct)
OUTPUT inserted.submission_uuid
VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
""",
            (survey["survey_uuid"], survey["link_uuid"], survey["target_type"], survey["target_code"], survey["target_name"], str(respondent_name or "").strip() or None, json.dumps(answers, ensure_ascii=False), json.dumps(computed, ensure_ascii=False), computed.get("pct")),
        )
        submission_uuid = str(cur.fetchone()[0])
        conn.commit()
        return submission_uuid
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3490')


def delete_custom_survey_submission(submission_uuid: str) -> None:
    ensure_mbo_schema()
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute("UPDATE dbo.MboSurveySubmissions SET deleted_at = SYSUTCDATETIME(), updated_at = SYSUTCDATETIME() WHERE submission_uuid = ?", (str(submission_uuid or "").strip(),))
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3504')


def _survey_pct_label(value: float | None) -> str:
    return "-" if value is None else f"{float(value):.1f}%"


def _survey_category_summary(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    buckets: Dict[str, List[float]] = defaultdict(list)
    for sub in rows or []:
        for cat, data in ((sub.get("computed") or {}).get("categories") or {}).items():
            if data and float(data.get("theoretical") or 0.0) > 0.0 and data.get("pct") is not None:
                buckets[str(cat)].append(float(data.get("pct") or 0.0))
    categories = [
        {"category": cat, "avg_pct": (sum(vals) / len(vals)) if vals else None, "count": len(vals)}
        for cat, vals in sorted(buckets.items(), key=lambda x: _norm_key(x[0]))
    ]
    strengths = sorted([c for c in categories if c.get("avg_pct") is not None], key=lambda c: float(c.get("avg_pct") or 0.0), reverse=True)[:3]
    improvements = sorted([c for c in categories if c.get("avg_pct") is not None], key=lambda c: float(c.get("avg_pct") or 0.0))[:3]
    return {"categories": categories, "strengths": strengths, "improvements": improvements}


def _survey_text_answers_for_submission(questions: List[Dict[str, Any]], answers: Dict[str, Any]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    for q in questions or []:
        qid = str(q.get("question_uuid") or "")
        q_type = str(q.get("type") or "")
        if q_type not in {"text", "number"}:
            continue
        value = str((answers or {}).get(qid) or "").strip()
        if not value:
            continue
        out.append({"question": str(q.get("text") or ""), "category": str(q.get("category") or "Generale"), "answer": value})
    return out


def _survey_answer_rows_for_submission(questions: List[Dict[str, Any]], answers: Dict[str, Any]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    for q in questions or []:
        qid = str(q.get("question_uuid") or "")
        q_type = str(q.get("type") or "")
        raw = str((answers or {}).get(qid) or "").strip()
        if not raw:
            continue
        value = raw
        if q_type == "single_choice":
            opt = next((o for o in q.get("options") or [] if str(o.get("option_uuid") or "") == raw), None)
            value = str((opt or {}).get("text") or raw)
        out.append({"question": str(q.get("text") or ""), "category": str(q.get("category") or "Generale"), "type": q_type, "answer": value})
    return out


def _survey_store_am_map(start_year_month: str, end_year_month: str) -> Dict[str, Dict[str, str]]:
    months = _period_months(start_year_month, end_year_month)
    managers_by_year: Dict[int, Dict[str, str]] = {}
    votes: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    labels: Dict[str, str] = {}
    for year in sorted({int(ym[:4]) for ym in months}):
        matrix = build_mbo_store_month_matrix(year)
        managers_by_year[year] = {str(m.get("row_uuid") or ""): str(m.get("manager_name") or "") for m in matrix.get("area_managers") or []}
        for row in matrix.get("rows") or []:
            store_code = str(row.get("store_code") or "").strip()
            if not store_code:
                continue
            for cell in row.get("months") or []:
                ym = str(cell.get("year_month") or "")
                if ym not in months:
                    continue
                am_uuid = str(cell.get("selected_row_uuid") or "").strip()
                if not am_uuid:
                    continue
                votes[store_code][am_uuid] += 1
                labels[am_uuid] = managers_by_year.get(year, {}).get(am_uuid, am_uuid)
    out: Dict[str, Dict[str, str]] = {}
    for store_code, am_votes in votes.items():
        am_uuid = sorted(am_votes.items(), key=lambda x: (-x[1], _norm_key(labels.get(x[0], x[0]))))[0][0]
        out[store_code] = {"area_manager_uuid": am_uuid, "area_manager_name": labels.get(am_uuid, am_uuid)}
    return out


def survey_store_area_manager_map(start_year_month: str, end_year_month: str) -> Dict[str, Dict[str, str]]:
    return _survey_store_am_map(start_year_month, end_year_month)


def build_custom_survey_dashboard(survey_uuid: str | None = None) -> Dict[str, Any]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute("SELECT survey_uuid, survey_name, display_name, target_type, start_year_month, end_year_month, is_active, created_at FROM dbo.MboSurveys ORDER BY created_at DESC")
        surveys = [{"survey_uuid": str(r[0]), "survey_name": str(r[1] or ""), "display_name": str(r[2] or "") or str(r[1] or ""), "target_type": str(r[3] or ""), "start_year_month": str(r[4] or ""), "end_year_month": str(r[5] or ""), "is_active": bool(r[6]), "created_at": str(r[7] or "")} for r in cur.fetchall() or []]
        selected = str(survey_uuid or "").strip() or (surveys[0]["survey_uuid"] if surveys else "")
        selected_survey = next((s for s in surveys if s["survey_uuid"] == selected), None)
        links: List[Dict[str, Any]] = []
        submissions: List[Dict[str, Any]] = []
        aggregates: List[Dict[str, Any]] = []
        category_rows: List[Dict[str, Any]] = []
        selected_full = get_custom_survey(survey_uuid=selected) if selected else None
        questions = (selected_full or {}).get("questions") or []
        if selected:
            cur.execute("SELECT link_uuid, token, target_type, target_code, target_name, is_active FROM dbo.MboSurveyLinks WHERE survey_uuid = ? ORDER BY target_name", (selected,))
            links = [{"link_uuid": str(r[0]), "token": str(r[1] or ""), "target_type": str(r[2] or ""), "target_code": str(r[3] or ""), "target_name": str(r[4] or ""), "is_active": bool(r[5])} for r in cur.fetchall() or []]
            cur.execute("SELECT submission_uuid, target_type, target_code, target_name, respondent_name, answers_json, computed_json, score_pct, created_at FROM dbo.MboSurveySubmissions WHERE survey_uuid = ? AND deleted_at IS NULL ORDER BY target_name, created_at DESC", (selected,))
            for r in cur.fetchall() or []:
                try:
                    answers = json.loads(r[5] or "{}")
                except Exception:
                    answers = {}
                try:
                    computed = json.loads(r[6] or "{}")
                except Exception:
                    computed = {}
                answer_rows = _survey_answer_rows_for_submission(questions, answers)
                text_answers = _survey_text_answers_for_submission(questions, answers)
                submissions.append({"submission_uuid": str(r[0]), "target_type": str(r[1] or ""), "target_code": str(r[2] or ""), "target_name": str(r[3] or ""), "respondent_name": str(r[4] or ""), "answers": answers, "answer_rows": answer_rows, "text_answers": text_answers, "text_answer_count": len(text_answers), "computed": computed, "score_pct": float(r[7]) if r[7] is not None else None, "created_at": str(r[8] or "")})
            by_target: Dict[tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
            for sub in submissions:
                by_target[(sub["target_code"], sub["target_name"])].append(sub)
            for (code, name), rows in sorted(by_target.items(), key=lambda x: _norm_key(x[0][1])):
                vals = [float(r.get("score_pct")) for r in rows if r.get("score_pct") is not None]
                summary = _survey_category_summary(rows)
                text_answers = [a for sub in rows for a in (sub.get("text_answers") or [])]
                aggregates.append({"target_code": code, "target_name": name, "count": len(rows), "avg_pct": (sum(vals) / len(vals)) if vals else None, "submissions": rows, "text_answers": text_answers, "text_answer_count": len(text_answers), **summary})
            cat_bucket: Dict[str, List[float]] = defaultdict(list)
            for sub in submissions:
                for cat, data in ((sub.get("computed") or {}).get("categories") or {}).items():
                    if data and float(data.get("theoretical") or 0.0) > 0.0 and data.get("pct") is not None:
                        cat_bucket[str(cat)].append(float(data.get("pct") or 0.0))
            category_rows = [{"category": cat, "avg_pct": (sum(vals) / len(vals)) if vals else None, "count": len(vals)} for cat, vals in sorted(cat_bucket.items(), key=lambda x: _norm_key(x[0]))]
        vals_all = [float(s.get("score_pct")) for s in submissions if s.get("score_pct") is not None]
        overview = {
            "overall": {**_survey_category_summary(submissions), "avg_pct": (sum(vals_all) / len(vals_all)) if vals_all else None, "text_answers": [a for sub in submissions for a in (sub.get("text_answers") or [])]},
            "area_manager_results": [],
            "store_manager_results": [],
            "store_manager_by_area_manager": [],
        }
        target_type = str((selected_survey or {}).get("target_type") or "")
        if target_type == "area_manager":
            overview["area_manager_results"] = aggregates
        elif target_type == "store_manager":
            overview["store_manager_results"] = aggregates
            try:
                store_am = _survey_store_am_map(str((selected_survey or {}).get("start_year_month") or ""), str((selected_survey or {}).get("end_year_month") or ""))
            except Exception:
                store_am = {}
            by_am: Dict[tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
            for agg in aggregates:
                am = store_am.get(str(agg.get("target_code") or "")) or {"area_manager_uuid": "", "area_manager_name": "Senza Area Manager"}
                by_am[(str(am.get("area_manager_uuid") or ""), str(am.get("area_manager_name") or "Senza Area Manager"))].append(agg)
            for (am_uuid, am_name), rows in sorted(by_am.items(), key=lambda x: _norm_key(x[0][1])):
                vals = [float(r.get("avg_pct")) for r in rows if r.get("avg_pct") is not None]
                all_subs = [sub for r in rows for sub in (r.get("submissions") or [])]
                overview["store_manager_by_area_manager"].append(
                    {
                        "area_manager_uuid": am_uuid,
                        "area_manager_name": am_name,
                        "count": len(rows),
                        "avg_pct": (sum(vals) / len(vals)) if vals else None,
                        "stores": rows,
                        "text_answers": [a for r in rows for a in (r.get("text_answers") or [])],
                        **_survey_category_summary(all_subs),
                    }
                )
        return {"surveys": surveys, "selected_survey_uuid": selected, "selected_survey": selected_survey, "selected_survey_full": selected_full, "links": links, "submissions": submissions, "aggregates": aggregates, "category_rows": category_rows, "overview": overview, "submission_count": len(submissions), "avg_pct": (sum(vals_all) / len(vals_all)) if vals_all else None}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3672')


def list_area_managers(include_inactive: bool = True) -> List[Dict[str, Any]]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        sql = """
SELECT row_uuid, manager_name, ilp_am_value, sort_order, is_active, created_at, updated_at
FROM dbo.MboAreaManagers
"""
        params: list[Any] = []
        if not include_inactive:
            sql += " WHERE is_active = 1"
        sql += " ORDER BY sort_order ASC, manager_name ASC"
        cur.execute(sql, params)
        out: List[Dict[str, Any]] = []
        for r in cur.fetchall() or []:
            out.append(
                {
                    "row_uuid": str(r[0]),
                    "manager_name": str(r[1] or "").strip(),
                    "ilp_am_value": str(r[2] or "").strip(),
                    "sort_order": int(r[3] or 0),
                    "is_active": bool(r[4]),
                    "created_at": r[5],
                    "updated_at": r[6],
                }
            )
        return out
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3707')


def save_area_manager(*, row_uuid: str | None, manager_name: str, ilp_am_value: str | None, sort_order: int = 0, is_active: bool = True) -> str:
    ensure_mbo_schema()
    name = str(manager_name or "").strip()
    if not name:
        raise ValueError("Nome area manager obbligatorio.")
    ilp_val = str(ilp_am_value or "").strip() or None

    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        if row_uuid:
            cur.execute(
                """
UPDATE dbo.MboAreaManagers
SET manager_name = ?, ilp_am_value = ?, sort_order = ?, is_active = ?, updated_at = SYSUTCDATETIME()
WHERE row_uuid = ?
""",
                (name, ilp_val, int(sort_order or 0), 1 if is_active else 0, str(row_uuid)),
            )
            conn.commit()
            return str(row_uuid)

        cur.execute(
            """
INSERT INTO dbo.MboAreaManagers (manager_name, ilp_am_value, sort_order, is_active)
OUTPUT inserted.row_uuid
VALUES (?, ?, ?, ?)
""",
            (name, ilp_val, int(sort_order or 0), 1 if is_active else 0),
        )
        new_id = str(cur.fetchone()[0])
        conn.commit()
        return new_id
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3747')


def delete_area_manager(row_uuid: str) -> None:
    ensure_mbo_schema()
    rid = str(row_uuid or "").strip()
    if not rid:
        return
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM dbo.MboStoreAreaManagerMonthly WHERE area_manager_row_uuid = ?", (rid,))
        cur.execute("DELETE FROM dbo.MboAreaManagers WHERE row_uuid = ?", (rid,))
        conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3765')


def sync_area_managers_from_ilp() -> int:
    ensure_mbo_schema()
    defaults = load_ilp_store_area_managers()
    distinct_vals = []
    seen = set()
    for row in defaults:
        raw = str(row.get("am_value") or "").strip()
        key = _norm(raw)
        if not raw or key in seen:
            continue
        seen.add(key)
        distinct_vals.append(raw)

    conn = _conn(read_only=False)
    created = 0
    try:
        cur = conn.cursor()
        for idx, raw in enumerate(distinct_vals, start=1):
            cur.execute(
                """
IF NOT EXISTS (SELECT 1 FROM dbo.MboAreaManagers WHERE LTRIM(RTRIM(COALESCE(ilp_am_value, ''))) = LTRIM(RTRIM(?)))
BEGIN
  INSERT INTO dbo.MboAreaManagers (manager_name, ilp_am_value, sort_order, is_active)
  VALUES (?, ?, ?, 1)
END
""",
                (raw, raw, raw, idx * 10),
            )
            try:
                created += int(cur.rowcount or 0)
            except Exception:
                log_swallowed('mbo_repository:3799')
        conn.commit()
        return created
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3806')


def _match_default_manager_uuid(raw_am_value: str, managers: List[Dict[str, Any]]) -> str:
    raw_norm = _norm(raw_am_value)
    if not raw_norm:
        return ""
    for m in managers:
        if _norm(m.get("ilp_am_value")) == raw_norm:
            return str(m.get("row_uuid") or "")
    for m in managers:
        if _norm(m.get("manager_name")) == raw_norm:
            return str(m.get("row_uuid") or "")
    return ""


def _load_monthly_overrides_for_year(year: int) -> Dict[tuple[str, str], str]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute(
            """
SELECT store_code, year_month, area_manager_row_uuid
FROM dbo.MboStoreAreaManagerMonthly
WHERE year_month >= ? AND year_month <= ?
""",
            (f"{int(year):04d}-01", f"{int(year):04d}-12"),
        )
        out: Dict[tuple[str, str], str] = {}
        for r in cur.fetchall() or []:
            out[(str(r[0] or "").strip(), str(r[1] or "").strip())] = str(r[2] or "").strip()
        return out
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3843')


def build_mbo_store_month_matrix(year: int) -> Dict[str, Any]:
    ensure_mbo_schema()
    stores = get_warehouse_stores() or []
    managers = list_area_managers(include_inactive=True)
    ilp_rows = load_ilp_store_area_managers()
    overrides = _load_monthly_overrides_for_year(year)

    ilp_by_code: Dict[str, Dict[str, Any]] = {}
    ilp_by_name: Dict[str, Dict[str, Any]] = {}
    for row in ilp_rows:
        code_key = _norm(row.get("store_code"))
        name_key = _norm(row.get("store_name"))
        if code_key and code_key not in ilp_by_code:
            ilp_by_code[code_key] = row
        if name_key and name_key not in ilp_by_name:
            ilp_by_name[name_key] = row

    month_labels = [
        (1, "Gen"),
        (2, "Feb"),
        (3, "Mar"),
        (4, "Apr"),
        (5, "Mag"),
        (6, "Giu"),
        (7, "Lug"),
        (8, "Ago"),
        (9, "Set"),
        (10, "Ott"),
        (11, "Nov"),
        (12, "Dic"),
    ]

    matrix_rows: List[Dict[str, Any]] = []
    for store in sorted(stores, key=lambda s: (str((s or {}).get("name") or "").lower(), str((s or {}).get("code") or "").lower())):
        code = str((store or {}).get("code") or "").strip()
        name = str((store or {}).get("name") or code).strip()
        ilp_row = ilp_by_code.get(_norm(code)) or ilp_by_name.get(_norm(name)) or {}
        raw_am = str(ilp_row.get("am_value") or "").strip()
        default_manager_uuid = _match_default_manager_uuid(raw_am, managers)
        months: List[Dict[str, Any]] = []
        for m_num, m_label in month_labels:
            ym = f"{int(year):04d}-{int(m_num):02d}"
            override_uuid = overrides.get((code, ym), "")
            selected_uuid = override_uuid or default_manager_uuid or ""
            months.append(
                {
                    "month_num": m_num,
                    "month_label": m_label,
                    "year_month": ym,
                    "selected_row_uuid": selected_uuid,
                    "override_row_uuid": override_uuid,
                    "default_row_uuid": default_manager_uuid,
                    "default_label": next((str(m.get("manager_name") or "") for m in managers if str(m.get("row_uuid") or "") == default_manager_uuid), "") or raw_am,
                    "is_override": bool(override_uuid),
                }
            )
        matrix_rows.append(
            {
                "store_code": code,
                "store_name": name,
                "ilp_am_value": raw_am,
                "default_manager_row_uuid": default_manager_uuid,
                "months": months,
            }
        )

    return {
        "year": int(year),
        "month_labels": [{"month_num": m, "month_label": label, "year_month": f"{int(year):04d}-{int(m):02d}"} for m, label in month_labels],
        "area_managers": managers,
        "rows": matrix_rows,
        "store_count": len(matrix_rows),
        "default_count": sum(1 for r in matrix_rows if r.get("default_manager_row_uuid")),
        "override_count": sum(1 for r in matrix_rows for c in (r.get("months") or []) if c.get("is_override")),
        "unmatched_count": sum(1 for r in matrix_rows if not str(r.get("default_manager_row_uuid") or "").strip()),
        "generated_at": datetime.utcnow(),
    }


def save_store_month_assignments(year: int, assignments: Dict[tuple[str, str], str]) -> Dict[str, int]:
    ensure_mbo_schema()
    matrix = build_mbo_store_month_matrix(year)
    defaults_by_pair: Dict[tuple[str, str], str] = {}
    for row in matrix.get("rows") or []:
        for month in row.get("months") or []:
            defaults_by_pair[(str(row.get("store_code") or ""), str(month.get("year_month") or ""))] = str(month.get("default_manager_row_uuid") or "")

    conn = _conn(read_only=False)
    inserted = updated = deleted = 0
    try:
        cur = conn.cursor()
        for (store_code, year_month), selected_uuid in (assignments or {}).items():
            store_code = str(store_code or "").strip()
            year_month = str(year_month or "").strip()
            selected_uuid = str(selected_uuid or "").strip()
            if not store_code or not year_month:
                continue
            default_uuid = str(defaults_by_pair.get((store_code, year_month)) or "").strip()
            should_delete = (not selected_uuid) or (selected_uuid == default_uuid)
            if should_delete:
                cur.execute(
                    "DELETE FROM dbo.MboStoreAreaManagerMonthly WHERE store_code = ? AND year_month = ?",
                    (store_code, year_month),
                )
                try:
                    deleted += int(cur.rowcount or 0)
                except Exception:
                    log_swallowed('mbo_repository:3953')
                continue

            cur.execute(
                """
IF EXISTS (SELECT 1 FROM dbo.MboStoreAreaManagerMonthly WHERE store_code = ? AND year_month = ?)
BEGIN
  UPDATE dbo.MboStoreAreaManagerMonthly
  SET area_manager_row_uuid = ?, updated_at = SYSUTCDATETIME()
  WHERE store_code = ? AND year_month = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboStoreAreaManagerMonthly (store_code, year_month, area_manager_row_uuid)
  VALUES (?, ?, ?);
END
""",
                (store_code, year_month, selected_uuid, store_code, year_month, store_code, year_month, selected_uuid),
            )
            # rowcount on IF/ELSE is unreliable: recount best-effort
            if (store_code, year_month) in defaults_by_pair:
                updated += 1
            else:
                inserted += 1
        conn.commit()
        return {"inserted": inserted, "updated": updated, "deleted": deleted}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:3983')


MBO_REWARD_ROLE_PROFILES = [
    {"key": "STORE MANAGER", "label": "Store Manager"},
    {"key": "ASSISTANT", "label": "Assistant"},
    {"key": "MULTISTORE", "label": "Multistore"},
]

MBO_REWARD_WEIGHT_PROFILES = [
    {"key": "STORE MANAGER", "label": "Store Manager", "categories": [("pnl", "P&L"), ("rating", "Rating"), ("google", "Google"), ("survey_soft", "Survey + Soft skills")]},
    {"key": "ASSISTANT", "label": "Assistant", "categories": [("pnl", "P&L"), ("rating", "Rating"), ("google", "Google"), ("survey_soft", "Survey + Soft skills")]},
    {"key": "MULTISTORE", "label": "Multistore", "categories": [("pnl", "P&L"), ("rating", "Rating"), ("google", "Google"), ("survey_soft", "Survey + Soft skills")]},
    {"key": "AREA_MANAGER", "label": "Area Manager", "categories": [("pnl", "P&L"), ("rating", "Rating"), ("google", "Google"), ("survey", "Survey")]},
]


def _mbo_month_labels(year: int) -> List[Dict[str, Any]]:
    return [
        {"month_num": i, "month_label": lbl, "year_month": f"{int(year):04d}-{i:02d}"}
        for i, lbl in enumerate(["", "Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"])
        if i
    ]


def load_mbo_reward_settings(year: int) -> Dict[str, Any]:
    ensure_mbo_schema()
    month_labels = _mbo_month_labels(year)
    managers = list_area_managers(include_inactive=True)
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute(
            """
SELECT area_manager_row_uuid, year_month, is_active, reward_eur
FROM dbo.MboAreaManagerMonthlySettings
WHERE year_month >= ? AND year_month <= ?
""",
            (f"{int(year):04d}-01", f"{int(year):04d}-12"),
        )
        am_settings = {(str(r[0]), str(r[1])): {"is_active": bool(r[2]), "reward_eur": float(r[3]) if r[3] is not None else None} for r in cur.fetchall() or []}

        cur.execute(
            """
SELECT role_name, year_month, reward_eur
FROM dbo.MboRoleMonthlyRewards
WHERE year_month >= ? AND year_month <= ?
""",
            (f"{int(year):04d}-01", f"{int(year):04d}-12"),
        )
        role_rewards = {(str(r[0]), str(r[1])): float(r[2]) if r[2] is not None else None for r in cur.fetchall() or []}

        cur.execute("SELECT target_profile, category_key, weight_pct FROM dbo.MboRewardWeights")
        weights = {(str(r[0]), str(r[1])): float(r[2] or 0.0) for r in cur.fetchall() or []}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4041')

    manager_rows = []
    for am in managers:
        rid = str(am.get("row_uuid") or "")
        months = []
        for m in month_labels:
            ym = str(m["year_month"])
            setting = am_settings.get((rid, ym), {})
            months.append(
                {
                    **m,
                    "is_active": bool(setting.get("is_active", am.get("is_active", True))),
                    "reward_eur": setting.get("reward_eur"),
                }
            )
        manager_rows.append({**am, "months": months})

    role_rows = []
    for role in MBO_REWARD_ROLE_PROFILES:
        months = []
        for m in month_labels:
            ym = str(m["year_month"])
            months.append({**m, "reward_eur": role_rewards.get((role["key"], ym))})
        role_rows.append({**role, "months": months})

    weight_rows = []
    for profile in MBO_REWARD_WEIGHT_PROFILES:
        cats = []
        for key, label in profile["categories"]:
            cats.append({"key": key, "label": label, "weight_pct": weights.get((profile["key"], key), 0.0)})
        weight_rows.append({"key": profile["key"], "label": profile["label"], "categories": cats})

    return {"month_labels": month_labels, "area_managers": manager_rows, "roles": role_rows, "weights": weight_rows}


def _parse_money_or_none(value: Any) -> float | None:
    dec = _parse_decimal(value)
    if dec is None:
        return None
    return float(dec)


def save_area_manager_monthly_settings(year: int, form_data: Any) -> Dict[str, int]:
    ensure_mbo_schema()
    managers = list_area_managers(include_inactive=True)
    months = _mbo_month_labels(year)
    conn = _conn(read_only=False)
    saved = 0
    try:
        cur = conn.cursor()
        for am in managers:
            rid = str(am.get("row_uuid") or "")
            for month in months:
                ym = str(month["year_month"])
                active = f"am_active__{rid}__{ym}" in form_data
                reward = _parse_money_or_none(form_data.get(f"am_reward__{rid}__{ym}"))
                cur.execute(
                    """
IF EXISTS (SELECT 1 FROM dbo.MboAreaManagerMonthlySettings WHERE area_manager_row_uuid = ? AND year_month = ?)
BEGIN
  UPDATE dbo.MboAreaManagerMonthlySettings
  SET is_active = ?, reward_eur = ?, updated_at = SYSUTCDATETIME()
  WHERE area_manager_row_uuid = ? AND year_month = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboAreaManagerMonthlySettings (area_manager_row_uuid, year_month, is_active, reward_eur)
  VALUES (?, ?, ?, ?);
END
""",
                    (rid, ym, 1 if active else 0, reward, rid, ym, rid, ym, 1 if active else 0, reward),
                )
                saved += 1
        conn.commit()
        return {"saved": saved}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4121')


def propagate_area_manager_monthly_setting(area_manager_uuid: str, from_year_month: str, is_active: bool, reward_eur: Any) -> Dict[str, int]:
    ensure_mbo_schema()
    rid = str(area_manager_uuid or "").strip()
    ym_from = str(from_year_month or "").strip()
    if not rid or len(ym_from) != 7:
        raise ValueError("Area manager o mese non valido.")
    reward = _parse_money_or_none(reward_eur)
    year = int(ym_from[:4])
    months = [m for m in _mbo_month_labels(year) if str(m["year_month"]) >= ym_from]
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        for m in months:
            ym = str(m["year_month"])
            cur.execute(
                """
IF EXISTS (SELECT 1 FROM dbo.MboAreaManagerMonthlySettings WHERE area_manager_row_uuid = ? AND year_month = ?)
BEGIN
  UPDATE dbo.MboAreaManagerMonthlySettings
  SET is_active = ?, reward_eur = ?, updated_at = SYSUTCDATETIME()
  WHERE area_manager_row_uuid = ? AND year_month = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboAreaManagerMonthlySettings (area_manager_row_uuid, year_month, is_active, reward_eur)
  VALUES (?, ?, ?, ?);
END
""",
                (rid, ym, 1 if is_active else 0, reward, rid, ym, rid, ym, 1 if is_active else 0, reward),
            )
        conn.commit()
        return {"saved": len(months)}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4160')


def save_role_monthly_rewards(year: int, form_data: Any) -> Dict[str, int]:
    ensure_mbo_schema()
    months = _mbo_month_labels(year)
    conn = _conn(read_only=False)
    saved = 0
    try:
        cur = conn.cursor()
        for role in MBO_REWARD_ROLE_PROFILES:
            role_key = role["key"]
            for month in months:
                ym = str(month["year_month"])
                reward = _parse_money_or_none(form_data.get(f"role_reward__{role_key}__{ym}"))
                cur.execute(
                    """
IF EXISTS (SELECT 1 FROM dbo.MboRoleMonthlyRewards WHERE role_name = ? AND year_month = ?)
BEGIN
  UPDATE dbo.MboRoleMonthlyRewards
  SET reward_eur = ?, updated_at = SYSUTCDATETIME()
  WHERE role_name = ? AND year_month = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboRoleMonthlyRewards (role_name, year_month, reward_eur)
  VALUES (?, ?, ?);
END
""",
                    (role_key, ym, reward, role_key, ym, role_key, ym, reward),
                )
                saved += 1
        conn.commit()
        return {"saved": saved}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4198')


def propagate_role_monthly_reward(role_name: str, from_year_month: str, reward_eur: Any) -> Dict[str, int]:
    ensure_mbo_schema()
    role_key = str(role_name or "").strip()
    ym_from = str(from_year_month or "").strip()
    if role_key not in {r["key"] for r in MBO_REWARD_ROLE_PROFILES} or len(ym_from) != 7:
        raise ValueError("Ruolo o mese non valido.")
    reward = _parse_money_or_none(reward_eur)
    year = int(ym_from[:4])
    months = [m for m in _mbo_month_labels(year) if str(m["year_month"]) >= ym_from]
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        for m in months:
            ym = str(m["year_month"])
            cur.execute(
                """
IF EXISTS (SELECT 1 FROM dbo.MboRoleMonthlyRewards WHERE role_name = ? AND year_month = ?)
BEGIN
  UPDATE dbo.MboRoleMonthlyRewards
  SET reward_eur = ?, updated_at = SYSUTCDATETIME()
  WHERE role_name = ? AND year_month = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboRoleMonthlyRewards (role_name, year_month, reward_eur)
  VALUES (?, ?, ?);
END
""",
                (role_key, ym, reward, role_key, ym, role_key, ym, reward),
            )
        conn.commit()
        return {"saved": len(months)}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4237')


def save_reward_weights(form_data: Any) -> Dict[str, int]:
    ensure_mbo_schema()
    conn = _conn(read_only=False)
    saved = 0
    try:
        cur = conn.cursor()
        for profile in MBO_REWARD_WEIGHT_PROFILES:
            profile_key = profile["key"]
            for category_key, _label in profile["categories"]:
                weight = _parse_decimal(form_data.get(f"weight__{profile_key}__{category_key}"))
                weight_f = float(weight) if weight is not None else 0.0
                if weight_f < 0 or weight_f > 100:
                    raise ValueError("I pesi devono essere compresi tra 0 e 100.")
                cur.execute(
                    """
IF EXISTS (SELECT 1 FROM dbo.MboRewardWeights WHERE target_profile = ? AND category_key = ?)
BEGIN
  UPDATE dbo.MboRewardWeights
  SET weight_pct = ?, updated_at = SYSUTCDATETIME()
  WHERE target_profile = ? AND category_key = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboRewardWeights (target_profile, category_key, weight_pct)
  VALUES (?, ?, ?);
END
""",
                    (profile_key, category_key, weight_f, profile_key, category_key, profile_key, category_key, weight_f),
                )
                saved += 1
        conn.commit()
        return {"saved": saved}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4276')


def save_mbo_report_definition(*, report_uuid: str | None, report_name: str, start_year_month: str, end_year_month: str, created_by: str | None = None) -> str:
    ensure_mbo_schema()
    name = str(report_name or "").strip()
    if not name:
        raise ValueError("Nome report obbligatorio.")
    start = _validate_year_month(start_year_month)
    end = _validate_year_month(end_year_month)
    _period_months(start, end)
    conn = _conn(read_only=False)
    try:
        cur = conn.cursor()
        if report_uuid:
            cur.execute(
                """
UPDATE dbo.MboReportDefinitions
SET report_name = ?, start_year_month = ?, end_year_month = ?, updated_at = SYSUTCDATETIME()
WHERE report_uuid = ?
""",
                (name, start, end, str(report_uuid)),
            )
            conn.commit()
            return str(report_uuid)
        cur.execute(
            """
INSERT INTO dbo.MboReportDefinitions (report_name, start_year_month, end_year_month, created_by)
OUTPUT inserted.report_uuid
VALUES (?, ?, ?, ?)
""",
            (name, start, end, str(created_by or "")),
        )
        new_id = str(cur.fetchone()[0])
        conn.commit()
        return new_id
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4316')


def list_mbo_report_definitions() -> List[Dict[str, Any]]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute(
            """
SELECT report_uuid, report_name, start_year_month, end_year_month, created_by, created_at, updated_at
FROM dbo.MboReportDefinitions
ORDER BY created_at DESC
"""
        )
        return [
            {
                "report_uuid": str(r[0]),
                "report_name": str(r[1] or ""),
                "start_year_month": str(r[2] or ""),
                "end_year_month": str(r[3] or ""),
                "created_by": str(r[4] or ""),
                "created_at": r[5],
                "updated_at": r[6],
            }
            for r in cur.fetchall() or []
        ]
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4347')


def get_mbo_report_definition(report_uuid: str) -> Dict[str, Any] | None:
    target = str(report_uuid or "").strip()
    if not target:
        return None
    return next((r for r in list_mbo_report_definitions() if str(r.get("report_uuid") or "") == target), None)


def _load_report_multipliers(report_uuid: str, report_type: str) -> Dict[tuple[str, str], float]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute(
            """
SELECT subject_key, store_code, multiplier_pct
FROM dbo.MboReportMultipliers
WHERE report_uuid = ? AND report_type = ?
""",
            (str(report_uuid or "").strip(), str(report_type or "").strip()),
        )
        return {(str(r[0] or ""), str(r[1] or "")): float(r[2] if r[2] is not None else 100.0) for r in cur.fetchall() or []}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4375')


def save_mbo_report_multipliers(report_uuid: str, report_type: str, form_data: Any) -> Dict[str, int]:
    ensure_mbo_schema()
    rid = str(report_uuid or "").strip()
    rtype = str(report_type or "").strip()
    if not rid or rtype not in {"staff", "area_manager"}:
        raise ValueError("Report o tipo riepilogo non valido.")
    prefix = "multi__" if rtype == "staff" else "ricalcolo__"
    conn = _conn(read_only=False)
    saved = 0
    try:
        cur = conn.cursor()
        for key in form_data:
            key_s = str(key or "")
            if not key_s.startswith(prefix):
                continue
            parts = key_s.split("__", 2)
            if len(parts) < 2:
                continue
            subject_key = parts[1].strip()
            store_code = parts[2].strip() if len(parts) > 2 else ""
            if not subject_key:
                continue
            pct = _parse_decimal(form_data.get(key_s))
            pct_f = float(pct) if pct is not None else 100.0
            if rtype == "staff" and (pct_f < 0 or pct_f > 100):
                raise ValueError("Il valore Multi deve essere compreso tra 0% e 100%.")
            if rtype == "area_manager" and pct_f < 0:
                raise ValueError("Il valore Ricalcolo non puo essere negativo.")
            cur.execute(
                """
IF EXISTS (SELECT 1 FROM dbo.MboReportMultipliers WHERE report_uuid = ? AND report_type = ? AND subject_key = ? AND store_code = ?)
BEGIN
  UPDATE dbo.MboReportMultipliers
  SET multiplier_pct = ?, updated_at = SYSUTCDATETIME()
  WHERE report_uuid = ? AND report_type = ? AND subject_key = ? AND store_code = ?;
END
ELSE
BEGIN
  INSERT INTO dbo.MboReportMultipliers (report_uuid, report_type, subject_key, store_code, multiplier_pct)
  VALUES (?, ?, ?, ?, ?);
END
""",
                (rid, rtype, subject_key, store_code, pct_f, rid, rtype, subject_key, store_code, rid, rtype, subject_key, store_code, pct_f),
            )
            saved += 1
        conn.commit()
        return {"saved": saved}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4429')


def _pct(points_taken: float, points_theoretical: float) -> float | None:
    return (float(points_taken or 0.0) / float(points_theoretical or 0.0) * 100.0) if points_theoretical else None


def _reward_pct(value: float | None) -> float:
    return max(0.0, min(100.0, float(value or 0.0)))


def _sum_metric(target: Dict[str, float], metric: Dict[str, Any]) -> None:
    target["points_taken"] += float(metric.get("points_taken") or 0.0)
    target["points_theoretical"] += float(metric.get("points_theoretical") or 0.0)


def _aggregate_pnl_total(row: Dict[str, Any] | None, months: List[str], *, apply_audit: bool) -> Dict[str, Any]:
    if not row:
        return {"points_taken": 0.0, "points_theoretical": 0.0, "pct": None}
    by_month = {str(cell.get("year_month") or ""): cell for cell in row.get("months") or []}
    points = 0.0
    theoretical = 0.0
    for ym in months:
        cell = by_month.get(ym) or {}
        if not cell.get("active", True):
            continue
        theoretical += float(cell.get("theoretical_points") or 0.0)
        points += float((cell.get("final_points") if apply_audit else cell.get("raw_points")) or 0.0)
    return {"points_taken": points, "points_theoretical": theoretical, "pct": _pct(points, theoretical)}


def _aggregate_monthly_total(row: Dict[str, Any] | None, months: List[str], *, apply_audit: bool) -> Dict[str, Any]:
    if not row:
        return {"points_taken": 0.0, "points_theoretical": 0.0, "pct": None}
    by_month = {str(cell.get("year_month") or ""): cell for cell in row.get("months") or []}
    points = 0.0
    theoretical = 0.0
    for ym in months:
        cell = by_month.get(ym) or {}
        if not cell.get("active", True):
            continue
        theoretical += float(cell.get("theoretical_points") or 0.0)
        points += float((cell.get("final_points") if apply_audit else cell.get("raw_points")) or 0.0)
    return {"points_taken": points, "points_theoretical": theoretical, "pct": _pct(points, theoretical)}


def _load_report_score_matrices(months: List[str]) -> Dict[int, Dict[str, Any]]:
    matrices: Dict[int, Dict[str, Any]] = {}
    for year in sorted({int(str(ym)[:4]) for ym in months}):
        matrices[year] = {
            "pnl": build_mbo_pnl_scores(year),
            "google": build_mbo_google_scores(year),
            "glovo": build_mbo_glovo_scores(year),
            "deliveroo": build_mbo_deliveroo_scores(year),
            "am": build_mbo_store_month_matrix(year),
        }
    return matrices


def _store_metric_for_months(store_code: str, months: List[str], matrices: Dict[int, Dict[str, Any]], *, apply_audit: bool) -> Dict[str, Any]:
    totals = {
        "pnl": {"points_taken": 0.0, "points_theoretical": 0.0},
        "rating": {"points_taken": 0.0, "points_theoretical": 0.0},
        "google": {"points_taken": 0.0, "points_theoretical": 0.0},
    }
    for year in sorted({int(str(ym)[:4]) for ym in months}):
        year_months = [ym for ym in months if int(str(ym)[:4]) == year]
        bundle = matrices.get(year) or {}
        _sum_metric(totals["pnl"], _aggregate_pnl_total(_find_matrix_row(bundle.get("pnl") or {}, store_code), year_months, apply_audit=apply_audit))
        _sum_metric(totals["google"], _aggregate_monthly_total(_find_matrix_row(bundle.get("google") or {}, store_code), year_months, apply_audit=apply_audit))
        _sum_metric(totals["rating"], _aggregate_monthly_total(_find_matrix_row(bundle.get("glovo") or {}, store_code), year_months, apply_audit=apply_audit))
        _sum_metric(totals["rating"], _aggregate_monthly_total(_find_matrix_row(bundle.get("deliveroo") or {}, store_code), year_months, apply_audit=apply_audit))
    for metric in totals.values():
        metric["pct"] = _pct(metric["points_taken"], metric["points_theoretical"])
    return totals


def _load_role_rewards_for_months(months: List[str]) -> Dict[str, float]:
    if not months:
        return {}
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        placeholders = ",".join("?" for _ in months)
        cur.execute(f"SELECT role_name, SUM(COALESCE(reward_eur, 0)) FROM dbo.MboRoleMonthlyRewards WHERE year_month IN ({placeholders}) GROUP BY role_name", tuple(months))
        return {str(r[0] or ""): float(r[1] or 0.0) for r in cur.fetchall() or []}
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4519')


def _load_am_settings_for_months(months: List[str]) -> Dict[tuple[str, str], Dict[str, Any]]:
    if not months:
        return {}
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        placeholders = ",".join("?" for _ in months)
        cur.execute(f"SELECT area_manager_row_uuid, year_month, is_active, reward_eur FROM dbo.MboAreaManagerMonthlySettings WHERE year_month IN ({placeholders})", tuple(months))
        return {
            (str(r[0] or ""), str(r[1] or "")): {"is_active": bool(r[2]), "reward_eur": float(r[3] or 0.0)}
            for r in cur.fetchall() or []
        }
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4538')


def _load_reward_weight_map() -> Dict[str, Dict[str, float]]:
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute("SELECT target_profile, category_key, weight_pct FROM dbo.MboRewardWeights")
        out: Dict[str, Dict[str, float]] = {}
        for r in cur.fetchall() or []:
            out.setdefault(str(r[0] or ""), {})[str(r[1] or "")] = float(r[2] or 0.0)
        return out
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4554')


def _load_custom_survey_averages(start_year_month: str, end_year_month: str) -> Dict[tuple[str, str], Dict[str, Any]]:
    ensure_mbo_schema()
    conn = _conn(read_only=True)
    try:
        cur = conn.cursor()
        cur.execute(
            """
SELECT sub.target_type, sub.target_code, AVG(sub.score_pct), COUNT(*)
FROM dbo.MboSurveySubmissions sub
JOIN dbo.MboSurveys s ON s.survey_uuid = sub.survey_uuid
WHERE sub.deleted_at IS NULL
  AND s.start_year_month <= ?
  AND s.end_year_month >= ?
GROUP BY sub.target_type, sub.target_code
""",
            (str(end_year_month or ""), str(start_year_month or "")),
        )
        return {
            (str(r[0] or ""), str(r[1] or "")): {"pct": float(r[2]) if r[2] is not None else None, "count": int(r[3] or 0)}
            for r in cur.fetchall() or []
        }
    finally:
        try:
            conn.close()
        except Exception:
            log_swallowed('mbo_repository:4582')


def _soft_skill_pct(final_score: float | None) -> float | None:
    if final_score is None:
        return None
    score = float(final_score or 0.0)
    if score < 3.0:
        return 0.0
    return max(0.0, min(100.0, score / 4.0 * 100.0))


def _average_optional(values: List[float | None]) -> float | None:
    clean = [float(v) for v in values if v is not None]
    return (sum(clean) / len(clean)) if clean else None


def _build_staff_report(report: Dict[str, Any], months: List[str], matrices: Dict[int, Dict[str, Any]]) -> Dict[str, Any]:
    start_ym = str(report.get("start_year_month") or "")
    end_ym = str(report.get("end_year_month") or "")
    survey_avgs = _load_custom_survey_averages(start_ym, end_ym)
    weights = _load_reward_weight_map()
    role_rewards = _load_role_rewards_for_months(months)
    multipliers = _load_report_multipliers(str(report.get("report_uuid") or ""), "staff")
    stores = _valid_store_codes()

    submissions = [
        r
        for r in list_soft_skill_submissions(include_deleted=False)
        if str(r.get("start_year_month") or "") <= end_ym and str(r.get("end_year_month") or "") >= start_ym
    ]
    grouped: Dict[tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for row in submissions:
        role = str(row.get("role") or "").strip().upper()
        if role not in {r["key"] for r in MBO_REWARD_ROLE_PROFILES}:
            continue
        name = str(row.get("full_name") or "").strip()
        code = str(row.get("store_code") or "").strip()
        if not name or not code:
            continue
        grouped[(name, role, code)].append(row)

    rows: List[Dict[str, Any]] = []
    for (name, role, store_code), group_rows in sorted(grouped.items(), key=lambda x: (_norm_key(x[0][1]), _norm_key(x[0][0]), _norm_key(x[0][2]))):
        subject_key = hashlib.sha1(f"{role}|{_norm_key(name)}".encode("utf-8")).hexdigest()
        store_metrics = _store_metric_for_months(store_code, months, matrices, apply_audit=True)
        soft_pct = _average_optional([_soft_skill_pct(r.get("final_score")) for r in group_rows])
        survey = survey_avgs.get(("store_manager", store_code), {})
        survey_pct = None if role == "ASSISTANT" else survey.get("pct")
        survey_soft_pct = _average_optional([soft_pct, survey_pct])
        role_weights = weights.get(role) or {}
        categories = {
            "pnl": {**store_metrics["pnl"], "weight_pct": role_weights.get("pnl", 0.0)},
            "rating": {**store_metrics["rating"], "weight_pct": role_weights.get("rating", 0.0)},
            "google": {**store_metrics["google"], "weight_pct": role_weights.get("google", 0.0)},
            "survey_soft": {"pct": survey_soft_pct, "soft_pct": soft_pct, "survey_pct": survey_pct, "weight_pct": role_weights.get("survey_soft", 0.0)},
        }
        for cat in categories.values():
            cat["weighted_pct"] = _reward_pct(cat.get("pct")) * float(cat.get("weight_pct") or 0.0) / 100.0
        total_pct = sum(float(c.get("weighted_pct") or 0.0) for c in categories.values())
        reward_base = float(role_rewards.get(role) or 0.0)
        multi_pct = multipliers.get((subject_key, store_code), 100.0)
        reward_due = reward_base * 0.93 * (total_pct / 100.0) * (float(multi_pct or 0.0) / 100.0)
        rows.append(
            {
                "subject_key": subject_key,
                "full_name": name,
                "role": role,
                "store_code": store_code,
                "store_name": str((stores.get(store_code) or {}).get("store_name") or store_code),
                "soft_pct": soft_pct,
                "survey_pct": survey_pct,
                "survey_count": int(survey.get("count") or 0),
                "categories": categories,
                "total_pct": total_pct,
                "reward_base": reward_base,
                "multi_pct": multi_pct,
                "reward_due": reward_due,
            }
        )
    return {"rows": rows, "row_count": len(rows), "reward_total": sum(float(r.get("reward_due") or 0.0) for r in rows)}


def _stores_by_area_manager(months: List[str], matrices: Dict[int, Dict[str, Any]], managers: List[Dict[str, Any]], settings: Dict[tuple[str, str], Dict[str, Any]]) -> Dict[str, Dict[str, List[str]]]:
    manager_active_default = {str(m.get("row_uuid") or ""): bool(m.get("is_active", True)) for m in managers}
    out: Dict[str, Dict[str, List[str]]] = defaultdict(lambda: defaultdict(list))
    for year in sorted({int(str(ym)[:4]) for ym in months}):
        matrix = (matrices.get(year) or {}).get("am") or {}
        for row in matrix.get("rows") or []:
            code = str(row.get("store_code") or "").strip()
            for cell in row.get("months") or []:
                ym = str(cell.get("year_month") or "")
                if ym not in months:
                    continue
                am_uuid = str(cell.get("selected_row_uuid") or "").strip()
                if not am_uuid:
                    continue
                setting = settings.get((am_uuid, ym), {})
                active = bool(setting.get("is_active", manager_active_default.get(am_uuid, True)))
                if active:
                    out[am_uuid][code].append(ym)
    return out


def _build_area_manager_report(report: Dict[str, Any], months: List[str], matrices: Dict[int, Dict[str, Any]]) -> Dict[str, Any]:
    managers = list_area_managers(include_inactive=True)
    settings = _load_am_settings_for_months(months)
    survey_avgs = _load_custom_survey_averages(str(report.get("start_year_month") or ""), str(report.get("end_year_month") or ""))
    weights = _load_reward_weight_map().get("AREA_MANAGER") or {}
    multipliers = _load_report_multipliers(str(report.get("report_uuid") or ""), "area_manager")
    stores = _valid_store_codes()
    stores_by_am = _stores_by_area_manager(months, matrices, managers, settings)
    manager_by_id = {str(m.get("row_uuid") or ""): m for m in managers}
    manager_active_default = {str(m.get("row_uuid") or ""): bool(m.get("is_active", True)) for m in managers}
    rows: List[Dict[str, Any]] = []

    for am_uuid, store_months in sorted(stores_by_am.items(), key=lambda x: _norm_key((manager_by_id.get(x[0]) or {}).get("manager_name"))):
        totals = {
            "pnl": {"points_taken": 0.0, "points_theoretical": 0.0},
            "rating": {"points_taken": 0.0, "points_theoretical": 0.0},
            "google": {"points_taken": 0.0, "points_theoretical": 0.0},
        }
        store_rows = []
        active_months = set()
        for store_code, store_month_list in sorted(store_months.items(), key=lambda x: _norm_key((stores.get(x[0]) or {}).get("store_name") or x[0])):
            active_months.update(store_month_list)
            metrics = _store_metric_for_months(store_code, store_month_list, matrices, apply_audit=False)
            for key in totals:
                _sum_metric(totals[key], metrics[key])
            store_rows.append({"store_code": store_code, "store_name": str((stores.get(store_code) or {}).get("store_name") or store_code), "months": store_month_list, "metrics": metrics})

        categories = {}
        for key in ("pnl", "rating", "google"):
            totals[key]["pct"] = _pct(totals[key]["points_taken"], totals[key]["points_theoretical"])
            categories[key] = {**totals[key], "weight_pct": weights.get(key, 0.0)}
        survey = survey_avgs.get(("area_manager", am_uuid), {})
        categories["survey"] = {"pct": survey.get("pct"), "count": int(survey.get("count") or 0), "weight_pct": weights.get("survey", 0.0)}
        for cat in categories.values():
            cat["weighted_pct"] = _reward_pct(cat.get("pct")) * float(cat.get("weight_pct") or 0.0) / 100.0
        total_pct = sum(float(c.get("weighted_pct") or 0.0) for c in categories.values())
        reward_base = 0.0
        for ym in active_months:
            setting = settings.get((am_uuid, ym), {})
            active = bool(setting.get("is_active", manager_active_default.get(am_uuid, True)))
            if active:
                reward_base += float(setting.get("reward_eur") or 0.0)
        ricalcolo_pct = multipliers.get((am_uuid, ""), 100.0)
        reward_due = reward_base * (total_pct / 100.0) * (float(ricalcolo_pct or 0.0) / 100.0)
        manager = manager_by_id.get(am_uuid) or {}
        rows.append(
            {
                "subject_key": am_uuid,
                "manager_name": str(manager.get("manager_name") or am_uuid),
                "store_rows": store_rows,
                "store_count": len(store_rows),
                "categories": categories,
                "total_pct": total_pct,
                "reward_base": reward_base,
                "ricalcolo_pct": ricalcolo_pct,
                "reward_due": reward_due,
            }
        )
    return {"rows": rows, "row_count": len(rows), "reward_total": sum(float(r.get("reward_due") or 0.0) for r in rows)}


def build_mbo_report_dashboard(report_uuid: str) -> Dict[str, Any]:
    ensure_mbo_schema()
    report = get_mbo_report_definition(report_uuid)
    if not report:
        raise ValueError("Report MBO non trovato.")
    months = _period_months(str(report.get("start_year_month") or ""), str(report.get("end_year_month") or ""))
    matrices = _load_report_score_matrices(months)
    stores = sorted(_valid_store_codes().values(), key=lambda s: (_norm_key(s.get("store_name")), _norm_key(s.get("store_code"))))

    store_rows = []
    store_totals = {
        "pnl": {"points_taken": 0.0, "points_theoretical": 0.0},
        "rating": {"points_taken": 0.0, "points_theoretical": 0.0},
        "google": {"points_taken": 0.0, "points_theoretical": 0.0},
    }
    for store in stores:
        code = str(store.get("store_code") or "").strip()
        metrics = _store_metric_for_months(code, months, matrices, apply_audit=True)
        for key in store_totals:
            _sum_metric(store_totals[key], metrics[key])
        top_score = {
            "points_taken": sum(float((metrics.get(key) or {}).get("points_taken") or 0.0) for key in ("pnl", "rating", "google")),
            "points_theoretical": sum(float((metrics.get(key) or {}).get("points_theoretical") or 0.0) for key in ("pnl", "rating", "google")),
        }
        top_score["pct"] = _pct(top_score["points_taken"], top_score["points_theoretical"])
        store_rows.append({"store_code": code, "store_name": str(store.get("store_name") or code), "metrics": metrics, "top_score": top_score})
    for metric in store_totals.values():
        metric["pct"] = _pct(metric["points_taken"], metric["points_theoretical"])
    store_top10 = [
        row
        for row in sorted(
            store_rows,
            key=lambda r: (
                -float((r.get("top_score") or {}).get("points_taken") or 0.0),
                -float((r.get("top_score") or {}).get("pct") or 0.0),
                _norm_key(r.get("store_name") or ""),
            ),
        )
        if float((row.get("top_score") or {}).get("points_theoretical") or 0.0) > 0
    ][:10]

    return {
        "report": report,
        "months": months,
        "store": {"rows": store_rows, "totals": store_totals, "top10": store_top10, "row_count": len(store_rows)},
        "staff": _build_staff_report(report, months, matrices),
        "area_manager": _build_area_manager_report(report, months, matrices),
    }
