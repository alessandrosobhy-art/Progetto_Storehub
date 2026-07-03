from __future__ import annotations

from datetime import date as _date, datetime, timedelta
import io
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Dict, List

import time
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app, jsonify, send_file, abort

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore
    Font = None  # type: ignore
    Alignment = None  # type: ignore

from app_db import get_backend

from db_integration import get_warehouse_stores, get_user_warehouse_stores

import json

from spese_repository import (
    insert_spesa,
    list_spese_month,
    delete_spesa,
    update_spesa,
    get_spesa_photo_file,
    sum_spese_day,
    sum_spese_month_by_day,
    sum_spese_month_total_net,
    sum_spese_month_total_net_multi,
    search_spese_range_multi,
)

from sharepoint_photos_repository import (
    SharePointTestError,
    upload_spesa_photo,
    download_spesa_photo,
    delete_spesa_photo,
    upload_versamento_photo,
    download_versamento_photo,
    delete_versamento_photo,
    upload_azzeramento_photo,
    download_azzeramento_photo,
    delete_azzeramento_photo,
)
from primanota_repository import (
    get_elenchi_options,
    load_primanota_day,
    replace_primanota_day,
    delete_primanota_day,
    load_primanota_month_agg,
    load_primanota_month_agg_totals,
    load_primanota_month_agg_totals_multi,
    sum_categoria_period,
    sum_categoria_by_day_range,
    sum_delivery_voce_range,
)

from dati_database_repository import upsert_datidatabase_from_distinta, delete_datidatabase_day
from daily_sales_repository import (
    upsert_daily_sales_from_distinta,
    delete_daily_sales_day,
    get_daily_sales_day,
    list_daily_sales_range,
)
from cash_statement_config_repository import list_cash_statement_config, list_cash_statement_dashboard_customizations

from versamenti_repository import (
    insert_versamento,
    list_versamenti_month,
    list_versamenti_periods_overlapping,
    sum_versamenti_month_total,
    sum_versamenti_month_total_multi,
    delete_versamento,
    update_versamento,
    get_versamento_photo_file,
    search_versamenti_range_multi,
)

from delivery_repository import (
    week_monday,
    list_delivery_providers,
    get_weekly as get_delivery_weekly,
    get_prev_rating as get_delivery_prev_rating,
    upsert_weekly as upsert_delivery_weekly,
    list_refunds_agg as list_delivery_refunds_agg,
)

from ipratico_repository import (
    import_distinta_day as import_ipratico_distinta_day,
    import_distinta_day_detailed as import_ipratico_distinta_day_detailed,
)
from ipratico_config_repository import get_ipratico_config, list_ipratico_payment_mappings
from distinta_cassa_photo_repository import (
    get_distinta_cassa_photo_file,
    upsert_distinta_cassa_photo_file,
    delete_distinta_cassa_photo_assoc,
    list_distinta_cassa_photo_days,
)
from distinta_cassa_ipratico_repository import (
    get_distinta_cassa_ipratico_snapshot,
    sum_distinta_cassa_ipratico_contanti_month_multi,
    upsert_distinta_cassa_ipratico_snapshot,
)
from rendiconto_convalide_repository import (
    delete_convalida_by_id,
    insert_convalida_periodo,
    list_convalida_days_month,
    list_convalide_overlapping,
    list_convalide_store,
)
from rendiconto_legacy_schema_repository import ensure_rendiconto_legacy_schema

rendiconto_bp = Blueprint("rendiconto", __name__, url_prefix="/rendiconto")
_REN_SCHEMA_ENSURED: set[str] = set()
_RENDICONTO_TTL_CACHE: dict[str, dict[str, Any]] = {}


def _rendiconto_ttl_cached(key: str, ttl_seconds: int, loader):
    now = time.time()
    cached = _RENDICONTO_TTL_CACHE.get(key)
    if isinstance(cached, dict) and (now - float(cached.get("ts") or 0)) < ttl_seconds:
        return cached.get("value")
    value = loader()
    _RENDICONTO_TTL_CACHE[key] = {"ts": now, "value": value}
    return value


def _session_tenant_key() -> str:
    return str(session.get("tenant_key") or "default").strip() or "default"


def _session_ui_language() -> str:
    return str(session.get("ui_language") or "it").strip().lower() or "it"


@rendiconto_bp.before_request
def _ensure_rendiconto_operational_schema():
    tenant_db = str(session.get("tenant_database") or "").strip() or "default"
    if tenant_db in _REN_SCHEMA_ENSURED:
        return None
    try:
        ensure_rendiconto_legacy_schema()
        _REN_SCHEMA_ENSURED.add(tenant_db)
    except Exception:
        current_app.logger.exception("Errore inizializzazione schema operativo Rendiconto")
    return None


def _ensure_session_keys() -> None:
    session.setdefault("store_code", None)
    session.setdefault("store_name", None)


def _require_login():
    if not session.get("uid"):
        nxt = request.full_path if request.query_string else request.path
        return redirect(url_for("login", next=nxt))
    return None


def _parse_ym(value: str):
    value = (value or "").strip()
    if not value:
        today = _date.today()
        return today.year, today.month, f"{today.year:04d}-{today.month:02d}"
    try:
        y, m = value.split("-", 1)
        y = int(y); m = int(m)
        if not (1 <= m <= 12) or y < 2000:
            raise ValueError
        return y, m, f"{y:04d}-{m:02d}"
    except Exception:
        today = _date.today()
        return today.year, today.month, f"{today.year:04d}-{today.month:02d}"


def _parse_date_iso(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return _date.today().isoformat()
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return value
    except Exception:
        return _date.today().isoformat()


def _is_valid_iso_date(value: str) -> bool:
    value = (value or "").strip()
    if not value:
        return False
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except Exception:
        return False


def _same_month_iso(dal_iso: str, al_iso: str) -> bool:
    dal_iso = (dal_iso or "").strip()
    al_iso = (al_iso or "").strip()
    if not (_is_valid_iso_date(dal_iso) and _is_valid_iso_date(al_iso)):
        return False
    return dal_iso[:7] == al_iso[:7]


def _parse_float(v: str) -> float:
    s = (v or "").strip().replace(".", "").replace(",", ".")
    # gestione input come 1.234,56 -> 1234.56
    # se l'utente inserisce 1234.56 resta ok dopo replace sopra ("." tolto),
    # quindi forziamo: se la stringa originale contiene punto come decimale,
    # l'utente puÃ² scrivere 1234.56 senza separatori migliaia.
    if v and ("." in v) and ("," not in v):
        s = (v or "").strip().replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _parse_money_decimal_value(v: str) -> Decimal:
    return Decimal(str(_parse_float(v or "0"))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _friendly_photo_upload_error(exc: Exception, *, subject: str = "foto") -> str:
    msg = str(exc or "").strip()
    low = msg.lower()
    if (
        "aadsts7000222" in low
        or "invalid_client" in low
        or "client secret" in low
        or "refresh token microsoft fallito" in low
    ):
        return f"Errore upload {subject}: servizio Microsoft temporaneamente non disponibile. Le credenziali dell'app risultano scadute, contatta l'amministratore."
    return f"Errore upload {subject}: {msg or 'errore non specificato'}"


def _iter_days_iso(start_iso: str, end_iso: str):
    """Yield giorni ISO inclusivi."""
    try:
        d0 = datetime.strptime(start_iso, "%Y-%m-%d").date()
        d1 = datetime.strptime(end_iso, "%Y-%m-%d").date()
    except Exception:
        return
    if d1 < d0:
        return
    cur = d0
    while cur <= d1:
        yield cur.isoformat()
        cur += timedelta(days=1)


def _overlap_days_with_existing_versamenti(
    *,
    store_code: str,
    dal_iso: str,
    al_iso: str,
    exclude_id: str | None = None,
    exclude_signature: dict | None = None,
):
    """Ritorna (giorni_overlap, summaries_versamenti).

    Se exclude_id Ã¨ valorizzato, il versamento con quell'id viene ignorato.
    Se exclude_signature Ã¨ valorizzato e le righe non hanno id, prova a escludere la riga matching.
    """
    exclude_id = (exclude_id or "").strip()
    exclude_signature = exclude_signature or None

    vres = list_versamenti_periods_overlapping(store_code=store_code, start_iso=dal_iso, end_iso=al_iso)
    rows = (vres or {}).get("rows") or []
    if not rows:
        return [], []

    req_start = datetime.strptime(dal_iso, "%Y-%m-%d").date()
    req_end = datetime.strptime(al_iso, "%Y-%m-%d").date()

    covered = set()
    summaries = []

    for r in rows:
        rid = str(r.get("id") or "").strip()
        if exclude_id and rid and rid == exclude_id:
            continue

        if not rid and exclude_signature:
            if (
                str(r.get("dal_iso") or "").strip() == str(exclude_signature.get("dal_iso") or "").strip()
                and str(r.get("al_iso") or "").strip() == str(exclude_signature.get("al_iso") or "").strip()
                and str(r.get("data_versamento_iso") or "").strip()
                == str(exclude_signature.get("data_versamento_iso") or "").strip()
                and str(r.get("valore_key") or "").strip() == str(exclude_signature.get("valore_key") or "").strip()
            ):
                continue

        r_start_s = str(r.get("dal_iso") or "").strip()
        r_end_s = str(r.get("al_iso") or "").strip()
        if not (r_start_s and r_end_s and _is_valid_iso_date(r_start_s) and _is_valid_iso_date(r_end_s)):
            continue
        r_start = datetime.strptime(r_start_s, "%Y-%m-%d").date()
        r_end = datetime.strptime(r_end_s, "%Y-%m-%d").date()
        if r_end < r_start:
            continue

        o_start = max(req_start, r_start)
        o_end = min(req_end, r_end)
        if o_end < o_start:
            continue

        for d in _iter_days_iso(o_start.isoformat(), o_end.isoformat()):
            covered.add(d)

        summaries.append({"id": rid, "dal_iso": r_start_s, "al_iso": r_end_s})

    return sorted(covered), summaries


def _money_to_decimal(v: str) -> Decimal:
    s = (v or "").strip()
    if not s:
        return Decimal("0")
    # 1.234,56 -> 1234.56
    if ("." in s) and ("," in s):
        s = s.replace(".", "").replace(",", ".")
    else:
        # 1234.56 (no migliaia) resta ok
        if ("." in s) and ("," not in s):
            s = s
        else:
            s = s.replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _round2(v: Decimal) -> Decimal:
    return (v or Decimal("0")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _iso_to_date(s: str):
    return datetime.strptime(s, "%Y-%m-%d").date()


def _daterange_iso(start_iso: str, end_iso: str):
    d1 = _iso_to_date(start_iso)
    d2 = _iso_to_date(end_iso)
    if d2 < d1:
        d1, d2 = d2, d1
    cur = d1
    while cur <= d2:
        yield cur.isoformat()
        cur += timedelta(days=1)


def _match_exclude_versamento(row: dict, exclude: dict) -> bool:
    """True se row rappresenta lo stesso versamento da escludere (lock)."""
    if not exclude:
        return False
    ex_id = (exclude.get("id") or "").strip()
    if ex_id and (row.get("id") or "").strip() == ex_id:
        return True

    # fallback chiave composta
    keys = {
        "data_versamento_iso": (exclude.get("orig_data_vers") or "").strip(),
        "dal_iso": (exclude.get("orig_dal") or "").strip(),
        "al_iso": (exclude.get("orig_al") or "").strip(),
        "nome_raw": (exclude.get("orig_nome") or "").strip(),
        "tipo_raw": (exclude.get("orig_tipo") or "").strip(),
        "tessera_raw": (exclude.get("orig_tessera") or "").strip(),
        "riferimento_raw": (exclude.get("orig_riferimento") or "").strip(),
        "valore_key": (exclude.get("orig_valore") or "").strip(),
    }
    # se non c'Ã¨ almeno una parte significativa, non matchiamo
    if not any(keys.values()):
        return False

    def _norm(s: str) -> str:
        return (s or "").strip().lower()

    # tessera: normalizza a sole cifre
    def _digits(s: str) -> str:
        return "".join(ch for ch in (s or "") if ch.isdigit())[:16] or "0"

    return (
        _norm(row.get("data_versamento_iso") or "") == _norm(keys["data_versamento_iso"]) and
        _norm(row.get("dal_iso") or "") == _norm(keys["dal_iso"]) and
        _norm(row.get("al_iso") or "") == _norm(keys["al_iso"]) and
        _norm(row.get("nome_raw") or "") == _norm(keys["nome_raw"]) and
        _norm(row.get("tipo_raw") or "") == _norm(keys["tipo_raw"]) and
        _digits(row.get("tessera_raw") or "") == _digits(keys["tessera_raw"]) and
        _norm(row.get("riferimento_raw") or "") == _norm(keys["riferimento_raw"]) and
        (row.get("valore_key") or "").strip() == (keys["valore_key"] or "").strip()
    )


def _form_flag(name: str) -> bool:
    raw = (request.form.get(name) or "").strip().lower()
    return raw in ("1", "true", "yes", "y", "on")


def _session_role_l() -> str:
    return str(session.get("role") or "").strip().lower()


def _is_admin_role() -> bool:
    return _session_role_l() == "admin"


def _can_validate_rendiconto_period() -> bool:
    return _session_role_l() in {"admin", "supervisor"}


def _session_display_name() -> str:
    return str(session.get("name") or session.get("uid") or "").strip()


def _validated_days_in_range(*, store_code: str, start_iso: str, end_iso: str) -> set[str]:
    try:
        rows = list_convalide_overlapping(store_code=str(store_code), start_iso=start_iso, end_iso=end_iso)
    except Exception:
        return set()

    locked: set[str] = set()
    for r in rows or []:
        dal = str(r.get("dal_iso") or "").strip()
        al = str(r.get("al_iso") or "").strip()
        if not (dal and al and _is_valid_iso_date(dal) and _is_valid_iso_date(al)):
            continue
        for d_iso in _daterange_iso(dal, al):
            locked.add(d_iso)
    return locked


def _blocked_distinte_days_in_range(*, store_code: str, start_iso: str, end_iso: str, exclude: dict | None = None) -> set[str]:
    locked = set(_locked_days_in_range(store_code=store_code, start_iso=start_iso, end_iso=end_iso, exclude=exclude))
    if not _is_admin_role():
        locked |= _validated_days_in_range(store_code=store_code, start_iso=start_iso, end_iso=end_iso)
    return locked


def _versamento_photo_waived() -> tuple[bool, bool, bool]:
    no_receipt = _form_flag("no_receipt_flag")
    lost_receipt = _form_flag("lost_receipt_flag")
    return (no_receipt or lost_receipt), no_receipt, lost_receipt


def _locked_days_in_range(*, store_code: str, start_iso: str, end_iso: str, exclude: dict | None = None) -> set[str]:
    """Ritorna l'insieme dei giorni (ISO) inclusi nel periodo competenza di *altri* versamenti."""
    exclude = exclude or {}

    try:
        periods = list_versamenti_periods_overlapping(store_code=str(store_code), start_iso=start_iso, end_iso=end_iso)
        rows = periods.get("rows") or []
    except Exception:
        rows = []

    d_start = _iso_to_date(start_iso)
    d_end = _iso_to_date(end_iso)
    if d_end < d_start:
        d_start, d_end = d_end, d_start

    locked: set[str] = set()
    for r in rows:
        if _match_exclude_versamento(r, exclude):
            continue
        dal = (r.get("dal_iso") or "").strip()
        al = (r.get("al_iso") or "").strip()
        if not dal or not al:
            continue
        try:
            rd1 = _iso_to_date(dal)
            rd2 = _iso_to_date(al)
        except Exception:
            continue
        if rd2 < rd1:
            rd1, rd2 = rd2, rd1
        a = max(rd1, d_start)
        b = min(rd2, d_end)
        cur = a
        while cur <= b:
            locked.add(cur.isoformat())
            cur += timedelta(days=1)
    return locked


def _build_dashboard_day_snapshot(*, store_code: str, d_iso: str) -> Dict[str, Any]:
    rows = load_primanota_day(
        store_code=str(store_code),
        data_iso=d_iso,
        categories=_distinta_entry_categories(),
    )
    try:
        spese = sum_spese_day(store_code=str(store_code), data_iso=d_iso)
    except Exception:
        spese = {"total": 0.0, "note_credito": 0.0, "net": 0.0}

    chiusura = {k: 0.0 for (k, _lbl, _t) in _CHIUSURA_FIELDS}
    distinte = 0.0
    ticket_si = 0.0
    delivery_si = 0.0
    delivery_no = 0.0
    coupon_si = 0.0
    entry_values: Dict[tuple[str, str, str], float] = {}

    for r in rows or []:
        cat = str(r.get("categoria") or "").strip()
        voce = str(r.get("voce") or "").strip()
        tipo = str(r.get("tipo") or "SI").strip().upper()
        try:
            val = float(r.get("valore") or 0)
        except Exception:
            val = 0.0
        entry_values[(cat, voce, tipo)] = float(entry_values.get((cat, voce, tipo), 0.0)) + val

        if cat == "Dati chiusura":
            key = _VOICE_TO_KEY.get(voce)
            if key:
                chiusura[key] = float(chiusura.get(key, 0.0)) + val
        elif cat == "Distinte":
            distinte += val
        elif cat == "Ticket":
            if tipo == "SI":
                ticket_si += val
        elif cat == "Delivery":
            if tipo == "SI":
                delivery_si += val
            else:
                delivery_no += val
        elif cat == "Coupon":
            if tipo == "SI":
                coupon_si += val

    vendite_lorde = float(chiusura.get("vendite_lorde", 0.0))
    annullati = float(chiusura.get("annullati", 0.0))
    pos = float(chiusura.get("pos", 0.0))

    giro = vendite_lorde - annullati
    spnet = float(spese.get("net") or 0.0)
    diff = distinte + ticket_si + delivery_si + coupon_si + pos + spnet - giro

    try:
        from tenant_config_repository import current_tenant_key

        tenant_key = str(session.get("tenant_key") or current_tenant_key()).strip() or current_tenant_key()
        sales_row = get_daily_sales_day(store_code=str(store_code), data_iso=d_iso, tenant_key=tenant_key)
    except Exception:
        sales_row = None

    if sales_row and str(sales_row.get("source") or "").strip().lower() != "legacy_datidatabase":
        giro = float(sales_row.get("gross_revenue") or giro)
        distinte = float(sales_row.get("cash_deposits_total") or distinte)
        ticket_si = float(sales_row.get("ticket_cash_effect") or ticket_si)
        delivery_si = float(sales_row.get("delivery_cash_effect") or delivery_si)
        delivery_no = float(sales_row.get("delivery_cash_amount") or delivery_no)
        coupon_si = float(sales_row.get("coupon_cash_effect") or coupon_si)
        pos = float(sales_row.get("pos_amount") or pos)
        # Le spese possono essere state modificate dopo il salvataggio della distinta:
        # ricalcoliamo la differenza usando sempre il netto spese corrente.
        diff = distinte + ticket_si + delivery_si + coupon_si + pos + spnet - giro

    chiusura_rows = []
    for (key, lbl, t) in _CHIUSURA_FIELDS:
        v = chiusura.get(key, 0.0)
        if t == "int":
            try:
                v = int(round(float(v)))
            except Exception:
                v = 0
        chiusura_rows.append({"key": key, "label": lbl, "type": t, "value": v})

    custom_sections = []
    try:
        from tenant_config_repository import current_tenant_key

        tenant_key = str(session.get("tenant_key") or current_tenant_key()).strip() or current_tenant_key()
        for section in list_cash_statement_dashboard_customizations(tenant_key=tenant_key):
            detail_rows = []
            for field in section.get("fields") or []:
                cat = str(field.get("legacy_category") or section.get("legacy_category") or section.get("label") or "").strip()
                voce = str(field.get("legacy_voce") or field.get("label") or "").strip()
                tipo = str(field.get("legacy_tipo") or "").strip().upper()
                if not cat or not voce:
                    continue
                if tipo:
                    value = entry_values.get((cat, voce, tipo), 0.0)
                else:
                    value = sum(v for (r_cat, r_voce, _r_tipo), v in entry_values.items() if r_cat == cat and r_voce == voce)
                detail_rows.append(
                    {
                        "key": field.get("field_key"),
                        "label": field.get("label"),
                        "label_en": field.get("label_en") or field.get("label"),
                        "label_fr": field.get("label_fr") or field.get("label"),
                        "label_es": field.get("label_es") or field.get("label"),
                        "type": "int" if str(field.get("value_type") or "").lower() == "int" else "money",
                        "value": value,
                    }
                )
            if detail_rows:
                custom_sections.append(
                    {
                        "section_key": section.get("section_key"),
                        "label": section.get("label"),
                        "label_en": section.get("label_en") or section.get("label"),
                        "label_fr": section.get("label_fr") or section.get("label"),
                        "label_es": section.get("label_es") or section.get("label"),
                        "rows": detail_rows,
                    }
                )
    except Exception:
        current_app.logger.exception("Errore lettura personalizzazioni distinta per dashboard")

    foto_file = None
    photo_url = ""
    try:
        foto_file = get_distinta_cassa_photo_file(store_code=str(store_code), data_iso=d_iso)
    except Exception:
        foto_file = None
    if foto_file:
        photo_url = url_for("rendiconto.distinta_cassa_photo_scoped", store_code=str(store_code), filename=str(foto_file))

    return {
        "date": d_iso,
        "giro": giro,
        "diff": diff,
        "pos": pos,
        "annullati": annullati,
        "scontrini": float(chiusura.get("scontrini", 0.0)),
        "distinte": distinte,
        "ticket_si": ticket_si,
        "delivery_si": delivery_si,
        "delivery_no": delivery_no,
        "coupon_si": coupon_si,
        "spese_net": float(spese.get("net") or 0.0),
        "spese_total": float(spese.get("total") or 0.0),
        "note_credito": float(spese.get("note_credito") or 0.0),
        "chiusura_rows": chiusura_rows,
        "custom_sections": custom_sections,
        "has_photo": bool(foto_file),
        "photo_url": photo_url,
    }


def _build_versamenti_status_for_store(*, store_code: str, year: int | None = None, today: _date | None = None) -> Dict[str, Any]:
    status = {
        "ultimo_al_iso": "",
        "ultimo_al_disp": "",
        "giorni_da_ultimo": None,
        "distinte_non_versate": 0.0,
        "giorni_non_versati": 0,
        "is_fragmented": False,
        "should_alert": False,
        "totale_distinte_anno": 0.0,
        "totale_versato_anno": 0.0,
        "totale_da_versare": 0.0,
    }

    try:
        today = today or _date.today()
        target_year = int(year or today.year)
        year_start = _date(target_year, 1, 1)
        year_end = _date(target_year, 12, 31)
        upper = min(today, year_end)

        if upper < year_start:
            return status

        tot_distinte_anno = sum_categoria_period(
            store_code=str(store_code),
            start_iso=year_start.isoformat(),
            end_iso=upper.isoformat(),
            categoria="Distinte",
        )

        vres = list_versamenti_periods_overlapping(
            store_code=str(store_code),
            start_iso=year_start.isoformat(),
            end_iso=upper.isoformat(),
        )
        vrows = (vres or {}).get("rows") or []

        total_versato = 0.0
        last_al = None

        for r in vrows:
            r = r or {}
            al_iso = str(r.get("al_iso") or "").strip()
            d_al = None
            if _is_valid_iso_date(al_iso):
                try:
                    d_al = _date.fromisoformat(al_iso)
                except Exception:
                    d_al = None

            if d_al is None or d_al < year_start or d_al > upper:
                continue

            if (last_al is None) or (d_al > last_al):
                last_al = d_al

            try:
                v_key = r.get("valore_key")
                v_dec = v_key if isinstance(v_key, Decimal) else Decimal(str(v_key))
            except Exception:
                v_dec = Decimal("0")

            total_versato += float(v_dec)

        totale_da_versare = float(tot_distinte_anno) - float(total_versato)

        status["distinte_non_versate"] = float(totale_da_versare)
        status["totale_distinte_anno"] = float(tot_distinte_anno)
        status["totale_versato_anno"] = float(total_versato)
        status["totale_da_versare"] = float(totale_da_versare)

        if last_al:
            status["ultimo_al_iso"] = last_al.isoformat()
            status["ultimo_al_disp"] = last_al.strftime("%d/%m/%Y")

            giorni_da_ultimo = (today - last_al).days
            if giorni_da_ultimo < 0:
                giorni_da_ultimo = 0

            status["giorni_da_ultimo"] = int(giorni_da_ultimo)
            status["giorni_non_versati"] = int(giorni_da_ultimo)
            status["should_alert"] = bool(
                giorni_da_ultimo > 7 and abs(totale_da_versare) > 1e-9
            )
    except Exception:
        current_app.logger.exception("Errore calcolo stato versamenti dashboard")

    return status


@rendiconto_bp.get("/")
def home():
    return redirect(url_for("rendiconto.distinta_cassa"))


@rendiconto_bp.route("/spese", methods=["GET", "POST"])
def spese():
    _ensure_session_keys()

    # Richiediamo login almeno per scrivere (POST). In GET lasciamo comunque
    # la pagina navigabile, ma senza store selezionato la modale blocca.
    if request.method == "POST":
        r = _require_login()
        if r is not None:
            return r

    store_code = session.get("store_code")
    store_name = session.get("store_name")

    ym = request.args.get("ym") or request.form.get("ym") or ""
    year, month, ym_norm = _parse_ym(ym)

    if request.method == "POST":
        if not store_code:
            flash("Seleziona prima uno store.", "warning")
            return redirect(url_for("rendiconto.spese", ym=ym_norm))

        data_iso = (request.form.get("data") or "").strip()
        tipo = (request.form.get("tipo_operazione") or "").strip()
        forn = (request.form.get("fornitore_spesa") or "").strip()
        doc = (request.form.get("documento") or "").strip()
        imp = (request.form.get("importo") or "").strip()

        if not (data_iso and tipo and forn and doc and imp):
            flash("Compila tutti i campi prima di salvare.", "warning")
            return redirect(url_for("rendiconto.spese", ym=ym_norm))

        foto_file = None
        foto = request.files.get("foto")
        if foto is not None and getattr(foto, "filename", ""):
            try:
                foto_file = upload_spesa_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    file_storage=foto,
                    data_iso=data_iso,
                )
            except Exception as e:
                current_app.logger.exception("Errore upload foto spesa")
                flash(_friendly_photo_upload_error(e, subject="foto spesa") + " La spesa verrÃ  salvata senza foto.", "warning")

        try:
            insert_spesa(
                store_code=str(store_code),
                data_iso=data_iso,
                tipo_operazione=tipo,
                fornitore_spesa=forn,
                documento=doc,
                importo_euro=imp,
                foto_file=foto_file,
            )
            flash("Spesa salvata.", "success")
        except Exception as e:
            # se la spesa non viene salvata, prova a pulire la foto caricata
            if foto_file:
                try:
                    delete_spesa_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(foto_file),
                    )
                except Exception:
                    pass
            current_app.logger.exception("Errore salvataggio spesa")
            flash(f"Errore salvataggio spesa: {e}", "danger")

        return redirect(url_for("rendiconto.spese", ym=ym_norm))

    rows = []
    total = 0.0
    if store_code:
        try:
            res = list_spese_month(store_code=str(store_code), year=year, month=month)
            rows = res.get("rows") or []
            total = float(res.get("total") or 0)
        except Exception as e:
            current_app.logger.exception("Errore lettura spese")
            flash(f"Errore lettura spese: {e}", "danger")

    today_iso = _date.today().isoformat()

    return render_template(
        "rendiconto_spese.html",
        store_code=store_code,
        store_name=store_name,
        ym=ym_norm,
        year=year,
        month=month,
        rows=rows,
        total=total,
        today_iso=today_iso,
    )


@rendiconto_bp.post("/spese/delete")
def spese_delete():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    ym = request.form.get("ym") or ""
    _, _, ym_norm = _parse_ym(ym)

    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.spese", ym=ym_norm))

    orig_data = (request.form.get("orig_data") or "").strip()
    orig_tipo = (request.form.get("orig_tipo") or "").strip()
    orig_forn = (request.form.get("orig_fornitore") or "").strip()
    orig_doc = (request.form.get("orig_documento") or "").strip()
    orig_imp = (request.form.get("orig_importo") or "").strip()

    foto_to_delete = None
    try:
        foto_to_delete = get_spesa_photo_file(
            store_code=str(store_code),
            orig_data_iso=orig_data,
            orig_tipo=orig_tipo,
            orig_fornitore=orig_forn,
            orig_documento=orig_doc,
            orig_importo_key=orig_imp,
        )
    except Exception:
        foto_to_delete = None

    try:
        n = delete_spesa(
            store_code=str(store_code),
            orig_data_iso=orig_data,
            orig_tipo=orig_tipo,
            orig_fornitore=orig_forn,
            orig_documento=orig_doc,
            orig_importo_key=orig_imp,
        )
        if n > 0:
            flash("Spesa eliminata.", "success")
            if foto_to_delete:
                try:
                    delete_spesa_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(foto_to_delete),
                    )
                except Exception as e:
                    current_app.logger.warning("Delete foto spesa fallita: %s", e)
        else:
            flash("Nessuna riga eliminata (record non trovato).", "warning")
    except Exception as e:
        current_app.logger.exception("Errore eliminazione spesa")
        flash(f"Errore eliminazione spesa: {e}", "danger")

    return redirect(url_for("rendiconto.spese", ym=ym_norm))


@rendiconto_bp.post("/spese/update")
def spese_update():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    ym = request.form.get("ym") or ""
    _, _, ym_norm = _parse_ym(ym)

    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.spese", ym=ym_norm))

    new_data = (request.form.get("data") or "").strip()
    new_tipo = (request.form.get("tipo_operazione") or "").strip()
    new_forn = (request.form.get("fornitore_spesa") or "").strip()
    new_doc = (request.form.get("documento") or "").strip()
    new_imp = (request.form.get("importo") or "").strip()

    if not (new_data and new_tipo and new_forn and new_doc and new_imp):
        flash("Compila tutti i campi prima di salvare.", "warning")
        return redirect(url_for("rendiconto.spese", ym=ym_norm))

    orig_data = (request.form.get("orig_data") or "").strip()
    orig_tipo = (request.form.get("orig_tipo") or "").strip()
    orig_forn = (request.form.get("orig_fornitore") or "").strip()
    orig_doc = (request.form.get("orig_documento") or "").strip()
    orig_imp = (request.form.get("orig_importo") or "").strip()

    old_foto = None
    try:
        old_foto = get_spesa_photo_file(
            store_code=str(store_code),
            orig_data_iso=orig_data,
            orig_tipo=orig_tipo,
            orig_fornitore=orig_forn,
            orig_documento=orig_doc,
            orig_importo_key=orig_imp,
        )
    except Exception:
        old_foto = None

    new_foto_file = None
    foto = request.files.get("foto")
    if foto is not None and getattr(foto, "filename", ""):
        try:
            new_foto_file = upload_spesa_photo(
                sb_jwt=str(session.get("sb_token") or ""),
                user_id=str(session.get("uid") or ""),
                store_code=str(store_code),
                file_storage=foto,
                data_iso=new_data,
            )
        except Exception as e:
            current_app.logger.exception("Errore upload foto spesa (update)")
            flash(_friendly_photo_upload_error(e, subject="foto spesa") + " La spesa verrÃ  aggiornata senza cambiare la foto.", "warning")
            new_foto_file = None

    try:
        n = update_spesa(
            store_code=str(store_code),
            orig_data_iso=orig_data,
            orig_tipo=orig_tipo,
            orig_fornitore=orig_forn,
            orig_documento=orig_doc,
            orig_importo_key=orig_imp,
            new_data_iso=new_data,
            new_tipo=new_tipo,
            new_fornitore=new_forn,
            new_documento=new_doc,
            new_importo_euro=new_imp,
            new_foto_file=new_foto_file,
        )
        if n > 0:
            flash("Spesa aggiornata.", "success")
            # se Ã¨ stata sostituita la foto, puliamo la vecchia
            if new_foto_file and old_foto and old_foto != new_foto_file:
                try:
                    delete_spesa_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(old_foto),
                    )
                except Exception as e:
                    current_app.logger.warning("Delete vecchia foto spesa fallita: %s", e)
        else:
            flash("Nessuna riga aggiornata (record non trovato).", "warning")
            # update fallito: evita file orfani
            if new_foto_file:
                try:
                    delete_spesa_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(new_foto_file),
                    )
                except Exception:
                    pass
    except Exception as e:
        current_app.logger.exception("Errore aggiornamento spesa")
        if new_foto_file:
            try:
                delete_spesa_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    filename=str(new_foto_file),
                )
            except Exception:
                pass
        flash(f"Errore aggiornamento spesa: {e}", "danger")

    return redirect(url_for("rendiconto.spese", ym=ym_norm))



@rendiconto_bp.get("/spese/photo/<store_code>/<path:filename>")
def spese_photo_scoped(store_code: str, filename: str):
    """Serve la foto collegata ad una spesa per lo store indicato (stream dal repository SharePoint)."""
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    sc = (store_code or "").strip()
    if not sc or "/" in sc or "\\" in sc:
        abort(404)

    fname = (filename or "").strip()
    # sicurezza: niente path traversal
    if "/" in fname or "\\" in fname:
        abort(404)

    try:
        content = download_spesa_photo(
            sb_jwt=str(session.get("sb_token") or ""),
            user_id=str(session.get("uid") or ""),
            store_code=sc,
            filename=fname,
        )
        return send_file(
            io.BytesIO(content),
            mimetype="image/jpeg",
            download_name=fname,
            as_attachment=False,
            max_age=300,
        )
    except Exception as e:
        current_app.logger.warning("Foto spesa non disponibile (scoped): %s", e)
        abort(404)


@rendiconto_bp.get("/spese/photo/<path:filename>")
def spese_photo(filename: str):
    """Serve la foto collegata ad una spesa (stream dal repository SharePoint)."""
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    if not store_code:
        abort(404)

    fname = (filename or "").strip()
    # sicurezza: niente path traversal
    if "/" in fname or "\\" in fname:
        abort(404)

    try:
        content = download_spesa_photo(
            sb_jwt=str(session.get("sb_token") or ""),
            user_id=str(session.get("uid") or ""),
            store_code=str(store_code),
            filename=fname,
        )
        return send_file(
            io.BytesIO(content),
            mimetype="image/jpeg",
            download_name=fname,
            as_attachment=False,
            max_age=300,
        )
    except Exception as e:
        current_app.logger.warning("Foto spesa non disponibile: %s", e)
        abort(404)


# -------------------------
# Versamenti
# -------------------------


@rendiconto_bp.route("/versamenti", methods=["GET", "POST"])
def versamenti():
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")

    y, m, ym_norm = _parse_ym(request.values.get("ym") or "")
    today_iso = _date.today().isoformat()

    if request.method == "POST":
        r = _require_login()
        if r is not None:
            return r

        if not store_code:
            flash("Seleziona prima uno store.", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        raw_data_vers = (request.form.get("data_versamento") or "").strip()
        raw_dal = (request.form.get("periodo_dal") or "").strip()
        raw_al = (request.form.get("periodo_al") or "").strip()

        if not (_is_valid_iso_date(raw_data_vers) and _is_valid_iso_date(raw_dal) and _is_valid_iso_date(raw_al)):
            flash("Periodo non valido: usa date valide (YYYY-MM-DD).", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        if not _same_month_iso(raw_dal, raw_al):
            flash("Il periodo di competenza deve essere all'interno dello stesso mese.", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        if raw_al < raw_dal:
            flash("Periodo non valido: la data 'Al' non puÃ² essere precedente alla data 'Dal'.", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        data_vers = _parse_date_iso(raw_data_vers)
        dal = _parse_date_iso(raw_dal)
        al = _parse_date_iso(raw_al)

        # Non permettere giorni giÃ  inclusi in altri versamenti
        try:
            covered_days, _overlaps = _overlap_days_with_existing_versamenti(
                store_code=str(store_code),
                dal_iso=dal,
                al_iso=al,
            )
            if covered_days:
                preview = ", ".join(covered_days[:6])
                if len(covered_days) > 6:
                    preview += f" (+{len(covered_days) - 6})"
                flash(f"Nel periodo selezionato ci sono giorni giÃ  versati: {preview}. Modifica il periodo.", "warning")
                return redirect(url_for("rendiconto.versamenti", ym=ym_norm))
        except Exception:
            # se la verifica overlap fallisce, non blocchiamo: la UI gestisce comunque il check via /compute
            pass

        nome = (request.form.get("nome_cognome") or "").strip()
        tipo = (request.form.get("tipo_versamento") or "").strip()
        tessera = (request.form.get("tessera") or "").strip()
        rif = (request.form.get("riferimento") or "").strip()
        valore = (request.form.get("valore") or "").strip()
        photo_waived, no_receipt_flag, lost_receipt_flag = _versamento_photo_waived()

        if not (data_vers and dal and al and nome and tipo and rif and valore):
            flash("Compila i campi obbligatori prima di salvare.", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        tessera_digits = "".join(ch for ch in tessera if ch.isdigit())
        if tipo.strip().lower() == "tessera" and not tessera_digits:
            flash("Compila il campo Tessera.", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))
        if tessera_digits and len(tessera_digits) > 16:
            flash("Tessera: massimo 16 cifre.", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        # Blocco: il versamento puÃ² essere salvato solo se la differenza Ã¨ zero.
        try:
            valore_dec = _round2(_money_to_decimal(valore))
            dist = Decimal(str(sum_categoria_period(store_code=str(store_code), start_iso=dal, end_iso=al, categoria="Distinte")))
            diff = _round2(valore_dec - _round2(dist))
        except Exception as e:
            current_app.logger.exception("Errore verifica differenza versamento")
            flash(f"Errore calcolo differenza: {e}", "danger")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        if diff != Decimal("0.00"):
            flash(
                f"Differenza diversa da zero ({float(diff):.2f} â‚¬). Correggi le distinte del periodo prima di salvare.",
                "warning",
            )
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        foto_file = None
        foto = request.files.get("foto")
        has_uploaded_photo = foto is not None and bool(getattr(foto, "filename", ""))
        if not photo_waived and not has_uploaded_photo:
            flash("Carica la foto oppure seleziona una delle dichiarazioni previste.", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))
        if has_uploaded_photo:
            try:
                foto_file = upload_versamento_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    file_storage=foto,
                    data_iso=data_vers,
                )
            except Exception as e:
                current_app.logger.exception("Errore upload foto versamento (insert)")
                flash(_friendly_photo_upload_error(e, subject="foto versamento") + " Il versamento verrÃ  salvato senza foto.", "warning")
                foto_file = None

        try:
            insert_versamento(
                store_code=str(store_code),
                data_versamento_iso=data_vers,
                periodo_dal_iso=dal,
                periodo_al_iso=al,
                nome_cognome=nome,
                tipo_versamento=tipo,
                tessera=tessera_digits,
                riferimento=rif,
                valore_euro=valore,
                foto_file=foto_file,
                no_receipt_flag=no_receipt_flag,
                lost_receipt_flag=lost_receipt_flag,
            )
            flash("Versamento salvato.", "success")
        except Exception as e:
            current_app.logger.exception("Errore salvataggio versamento")
            if foto_file:
                try:
                    delete_versamento_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(foto_file),
                    )
                except Exception:
                    pass
            flash(f"Errore salvataggio versamento: {e}", "danger")

        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    # GET
    rows = []
    total = 0.0
    has_id = False

    if store_code:
        try:
            info = list_versamenti_month(store_code=str(store_code), year=y, month=m)
            rows = info.get("rows") or []
            total = float(info.get("total") or 0.0)
            has_id = bool(info.get("has_id"))
        except Exception as e:
            current_app.logger.exception("Errore lettura VERSAMENTI_APP")
            flash(f"Errore lettura versamenti: {e}", "danger")

    # Calcolo differenza e distinte periodo (per ogni riga) in modo efficiente
    if store_code and rows:
        try:
            min_dal = min(r.get("dal_iso") for r in rows if r.get("dal_iso"))
            max_al = max(r.get("al_iso") for r in rows if r.get("al_iso"))

            dist_by_day = sum_categoria_by_day_range(
                store_code=str(store_code),
                start_iso=min_dal,
                end_iso=max_al,
                categoria="Distinte",
            )

            d0 = datetime.strptime(min_dal, "%Y-%m-%d").date()
            d1 = datetime.strptime(max_al, "%Y-%m-%d").date()
            n = (d1 - d0).days + 1

            prefix = [0.0] * (n + 1)
            for i in range(n):
                di = d0 + timedelta(days=i)
                prefix[i + 1] = prefix[i] + float(dist_by_day.get(di.isoformat(), 0.0) or 0.0)

            def _sum_period(dal_iso: str, al_iso: str) -> float:
                try:
                    a = datetime.strptime(dal_iso, "%Y-%m-%d").date()
                    b = datetime.strptime(al_iso, "%Y-%m-%d").date()
                except Exception:
                    return 0.0
                if b < a:
                    a, b = b, a
                ia = (a - d0).days
                ib = (b - d0).days
                if ia < 0:
                    ia = 0
                if ib >= n:
                    ib = n - 1
                if ib < ia:
                    return 0.0
                return prefix[ib + 1] - prefix[ia]

            for r in rows:
                dist_tot = _sum_period(r.get("dal_iso") or "", r.get("al_iso") or "")
                r["distinte_periodo"] = float(dist_tot)
                try:
                    v = float(r.get("valore") or 0.0)
                except Exception:
                    v = 0.0
                r["differenza"] = v - float(dist_tot)
        except Exception:
            current_app.logger.exception("Errore calcolo differenza versamenti")

    return render_template(
        "rendiconto_versamenti.html",
        store_code=store_code,
        store_name=store_name,
        ym=ym_norm,
        today_iso=today_iso,
        rows=rows,
        total=total,
        has_id=has_id,
    )


@rendiconto_bp.post("/versamenti/delete")
def versamenti_delete():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    ym_norm = _parse_ym(request.form.get("ym") or "")[2]
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    record_id = (request.form.get("id") or "").strip()
    orig_data = (request.form.get("orig_data_vers") or "").strip()
    orig_dal = (request.form.get("orig_dal") or "").strip()
    orig_al = (request.form.get("orig_al") or "").strip()
    orig_nome = (request.form.get("orig_nome") or "").strip()
    orig_tipo = (request.form.get("orig_tipo") or "").strip()
    orig_tess = (request.form.get("orig_tessera") or "").strip()
    orig_rif = (request.form.get("orig_riferimento") or "").strip()
    orig_val = (request.form.get("orig_valore") or "").strip()

    # Se la tabella non ha ID, escludiamo questo record tramite signature basata sui campi originali
    if not record_id:
        try:
            exclude_sig = {
                "data_versamento_iso": orig_data,
                "dal_iso": orig_dal,
                "al_iso": orig_al,
                "valore_key": orig_val,
                "nome_raw": orig_nome,
            }
            covered_days, _overlaps = _overlap_days_with_existing_versamenti(
                store_code=str(store_code),
                dal_iso=new_dal,
                al_iso=new_al,
                exclude_signature=exclude_sig,
            )
            if covered_days:
                preview = ", ".join(covered_days[:6])
                if len(covered_days) > 6:
                    preview += f" (+{len(covered_days) - 6})"
                flash(f"Nel periodo selezionato ci sono giorni giÃ  versati: {preview}. Modifica il periodo.", "warning")
                return redirect(url_for("rendiconto.versamenti", ym=ym_norm))
        except Exception:
            pass

    old_foto = None
    try:
        old_foto = get_versamento_photo_file(
            store_code=str(store_code),
            record_id=str(record_id or ""),
            orig_data_vers_iso=orig_data,
            orig_dal_iso=orig_dal,
            orig_al_iso=orig_al,
            orig_nome=orig_nome,
            orig_tipo=orig_tipo,
            orig_tessera=orig_tess,
            orig_riferimento=orig_rif,
            orig_valore_key=orig_val,
        )
    except Exception:
        old_foto = None

    try:
        n = delete_versamento(
            store_code=str(store_code),
            record_id=record_id,
            orig_data_vers_iso=orig_data,
            orig_dal_iso=orig_dal,
            orig_al_iso=orig_al,
            orig_nome=orig_nome,
            orig_tipo=orig_tipo,
            orig_tessera=orig_tess,
            orig_riferimento=orig_rif,
            orig_valore_key=orig_val,
        )
        if n > 0:
            flash("Versamento eliminato.", "success")
            if old_foto:
                try:
                    delete_versamento_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(old_foto),
                    )
                except Exception as e:
                    current_app.logger.warning("Delete foto versamento fallita: %s", e)
        else:
            flash("Nessuna riga eliminata (record non trovato).", "warning")
    except Exception as e:
        current_app.logger.exception("Errore eliminazione versamento")
        flash(f"Errore eliminazione versamento: {e}", "danger")

    return redirect(url_for("rendiconto.versamenti", ym=ym_norm))


@rendiconto_bp.post("/versamenti/update")
def versamenti_update():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    ym_norm = _parse_ym(request.form.get("ym") or "")[2]
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    record_id = (request.form.get("id") or "").strip()

    raw_new_data_vers = (request.form.get("data_versamento") or "").strip()
    raw_new_dal = (request.form.get("periodo_dal") or "").strip()
    raw_new_al = (request.form.get("periodo_al") or "").strip()

    if not (_is_valid_iso_date(raw_new_data_vers) and _is_valid_iso_date(raw_new_dal) and _is_valid_iso_date(raw_new_al)):
        flash("Periodo non valido: usa date valide (YYYY-MM-DD).", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    if not _same_month_iso(raw_new_dal, raw_new_al):
        flash("Il periodo di competenza deve essere all'interno dello stesso mese.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    if raw_new_al < raw_new_dal:
        flash("Periodo non valido: la data 'Al' non puÃ² essere precedente alla data 'Dal'.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    new_data_vers = _parse_date_iso(raw_new_data_vers)
    new_dal = _parse_date_iso(raw_new_dal)
    new_al = _parse_date_iso(raw_new_al)

    # Non permettere giorni giÃ  inclusi in altri versamenti (escludendo questo record, se ha ID)
    if record_id:
        try:
            covered_days, _overlaps = _overlap_days_with_existing_versamenti(
                store_code=str(store_code),
                dal_iso=new_dal,
                al_iso=new_al,
                exclude_id=str(record_id),
            )
            if covered_days:
                preview = ", ".join(covered_days[:6])
                if len(covered_days) > 6:
                    preview += f" (+{len(covered_days) - 6})"
                flash(f"Nel periodo selezionato ci sono giorni giÃ  versati: {preview}. Modifica il periodo.", "warning")
                return redirect(url_for("rendiconto.versamenti", ym=ym_norm))
        except Exception:
            # la UI fa comunque il check via /compute
            pass

    new_nome = (request.form.get("nome_cognome") or "").strip()
    new_tipo = (request.form.get("tipo_versamento") or "").strip()
    new_tess = (request.form.get("tessera") or "").strip()
    new_rif = (request.form.get("riferimento") or "").strip()
    new_val = (request.form.get("valore") or "").strip()
    photo_waived, no_receipt_flag, lost_receipt_flag = _versamento_photo_waived()

    if not (new_data_vers and new_dal and new_al and new_nome and new_tipo and new_rif and new_val):
        flash("Compila i campi obbligatori prima di salvare.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    tessera_digits = "".join(ch for ch in new_tess if ch.isdigit())
    if new_tipo.strip().lower() == "tessera" and not tessera_digits:
        flash("Compila il campo Tessera.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))
    if tessera_digits and len(tessera_digits) > 16:
        flash("Tessera: massimo 16 cifre.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    # Blocco: aggiornamento consentito solo se la differenza Ã¨ zero.
    try:
        valore_dec = _round2(_money_to_decimal(new_val))
        dist = Decimal(str(sum_categoria_period(store_code=str(store_code), start_iso=new_dal, end_iso=new_al, categoria="Distinte")))
        diff = _round2(valore_dec - _round2(dist))
    except Exception as e:
        current_app.logger.exception("Errore verifica differenza versamento (update)")
        flash(f"Errore calcolo differenza: {e}", "danger")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    if diff != Decimal("0.00"):
        flash(
            f"Differenza diversa da zero ({float(diff):.2f} â‚¬). Correggi le distinte del periodo prima di salvare.",
            "warning",
        )
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    orig_data = (request.form.get("orig_data_vers") or "").strip()
    orig_dal = (request.form.get("orig_dal") or "").strip()
    orig_al = (request.form.get("orig_al") or "").strip()
    orig_nome = (request.form.get("orig_nome") or "").strip()
    orig_tipo = (request.form.get("orig_tipo") or "").strip()
    orig_tess = (request.form.get("orig_tessera") or "").strip()
    orig_rif = (request.form.get("orig_riferimento") or "").strip()
    orig_val = (request.form.get("orig_valore") or "").strip()

    # Se non abbiamo ID (tabella senza colonna ID), proviamo a escludere il record tramite firma dei campi originali
    if not record_id:
        try:
            exclude_sig = {
                "data_versamento_iso": orig_data,
                "dal_iso": orig_dal,
                "al_iso": orig_al,
                "valore_key": orig_val,
                "nome_raw": orig_nome,
                "tipo_raw": orig_tipo,
            }
            covered_days, _overlaps = _overlap_days_with_existing_versamenti(
                store_code=str(store_code),
                dal_iso=new_dal,
                al_iso=new_al,
                exclude_signature=exclude_sig,
            )
            if covered_days:
                preview = ", ".join(covered_days[:6])
                if len(covered_days) > 6:
                    preview += f" (+{len(covered_days) - 6})"
                flash(f"Nel periodo selezionato ci sono giorni giÃ  versati: {preview}. Modifica il periodo.", "warning")
                return redirect(url_for("rendiconto.versamenti", ym=ym_norm))
        except Exception:
            pass

    old_foto = None
    try:
        old_foto = get_versamento_photo_file(
            store_code=str(store_code),
            record_id=str(record_id or ""),
            orig_data_vers_iso=orig_data,
            orig_dal_iso=orig_dal,
            orig_al_iso=orig_al,
            orig_nome=orig_nome,
            orig_tipo=orig_tipo,
            orig_tessera=orig_tess,
            orig_riferimento=orig_rif,
            orig_valore_key=orig_val,
        )
    except Exception:
        old_foto = None

    new_foto_file = None
    foto = request.files.get("foto")
    has_uploaded_photo = foto is not None and bool(getattr(foto, "filename", ""))
    has_existing_photo = bool(str(old_foto or "").strip())
    if not photo_waived and not has_existing_photo and not has_uploaded_photo:
        flash("Carica la foto oppure seleziona una delle dichiarazioni previste.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))
    if has_uploaded_photo:
        try:
            new_foto_file = upload_versamento_photo(
                sb_jwt=str(session.get("sb_token") or ""),
                user_id=str(session.get("uid") or ""),
                store_code=str(store_code),
                file_storage=foto,
                data_iso=new_data_vers,
            )
        except Exception as e:
            current_app.logger.exception("Errore upload foto versamento (update)")
            flash(_friendly_photo_upload_error(e, subject="foto versamento") + " Il versamento verrÃ  aggiornato senza cambiare la foto.", "warning")
            new_foto_file = None

    try:
        n = update_versamento(
            store_code=str(store_code),
            record_id=record_id,
            orig_data_vers_iso=orig_data,
            orig_dal_iso=orig_dal,
            orig_al_iso=orig_al,
            orig_nome=orig_nome,
            orig_tipo=orig_tipo,
            orig_tessera=orig_tess,
            orig_riferimento=orig_rif,
            orig_valore_key=orig_val,
            new_data_vers_iso=new_data_vers,
            new_dal_iso=new_dal,
            new_al_iso=new_al,
            new_nome=new_nome,
            new_tipo=new_tipo,
            new_tessera=tessera_digits,
            new_riferimento=new_rif,
            new_valore_euro=new_val,
            new_foto_file=new_foto_file,
            new_no_receipt_flag=no_receipt_flag,
            new_lost_receipt_flag=lost_receipt_flag,
        )
        if n > 0:
            flash("Versamento aggiornato.", "success")
            if new_foto_file and old_foto and old_foto != new_foto_file:
                try:
                    delete_versamento_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(old_foto),
                    )
                except Exception as e:
                    current_app.logger.warning("Delete vecchia foto versamento fallita: %s", e)
        else:
            flash("Nessuna riga aggiornata (record non trovato).", "warning")
            if new_foto_file:
                try:
                    delete_versamento_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(new_foto_file),
                    )
                except Exception:
                    pass
    except Exception as e:
        current_app.logger.exception("Errore aggiornamento versamento")
        if new_foto_file:
            try:
                delete_versamento_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    filename=str(new_foto_file),
                )
            except Exception:
                pass
        flash(f"Errore aggiornamento versamento: {e}", "danger")

    return redirect(url_for("rendiconto.versamenti", ym=ym_norm))



@rendiconto_bp.get("/versamenti/photo/<store_code>/<path:filename>")
def versamenti_photo_scoped(store_code: str, filename: str):
    """Serve la foto collegata ad un versamento per lo store indicato (stream dal repository SharePoint)."""
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    sc = (store_code or "").strip()
    if not sc or "/" in sc or "\\" in sc:
        abort(404)

    fname = (filename or "").strip()
    # sicurezza: niente path traversal
    if "/" in fname or "\\" in fname:
        abort(404)

    try:
        content = download_versamento_photo(
            sb_jwt=str(session.get("sb_token") or ""),
            user_id=str(session.get("uid") or ""),
            store_code=sc,
            filename=fname,
        )
        return send_file(
            io.BytesIO(content),
            mimetype="image/jpeg",
            download_name=fname,
            as_attachment=False,
            max_age=300,
        )
    except Exception as e:
        current_app.logger.warning("Foto versamento non disponibile (scoped): %s", e)
        abort(404)


@rendiconto_bp.get("/versamenti/photo/<path:filename>")
def versamenti_photo(filename: str):
    """Serve la foto collegata ad un versamento (stream dal repository SharePoint)."""
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    if not store_code:
        abort(404)

    fname = (filename or "").strip()
    if "/" in fname or "\\" in fname:
        abort(404)

    try:
        content = download_versamento_photo(
            sb_jwt=str(session.get("sb_token") or ""),
            user_id=str(session.get("uid") or ""),
            store_code=str(store_code),
            filename=fname,
        )
        return send_file(
            io.BytesIO(content),
            mimetype="image/jpeg",
            download_name=fname,
            as_attachment=False,
            max_age=300,
        )
    except Exception as e:
        current_app.logger.warning("Foto versamento non disponibile: %s", e)
        abort(404)


@rendiconto_bp.get("/api/versamenti/compute")
def api_versamenti_compute():
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    dal = (request.args.get("dal") or "").strip()
    al = (request.args.get("al") or "").strip()
    valore_raw = (request.args.get("valore") or "").strip()
    exclude_id = (request.args.get("exclude_id") or "").strip()
    # Quando si modifica un versamento, il controllo overlap deve ignorare quel versamento stesso.
    # Se la tabella non ha un ID, usiamo una "signature" basata sui valori originali del record.
    exclude_sig_data = (request.args.get("exclude_sig_data_vers") or "").strip()
    exclude_sig_dal = (request.args.get("exclude_sig_dal") or "").strip()
    exclude_sig_al = (request.args.get("exclude_sig_al") or "").strip()
    exclude_sig_val = (request.args.get("exclude_sig_valore") or "").strip()

    exclude_signature = None
    if exclude_sig_data and exclude_sig_dal and exclude_sig_al and exclude_sig_val:
        exclude_signature = {
            "data_versamento_iso": exclude_sig_data,
            "dal_iso": exclude_sig_dal,
            "al_iso": exclude_sig_al,
            "valore_key": exclude_sig_val,
        }


    if not dal or not al:
        return jsonify(error="Periodo non valido"), 400

    if not (_is_valid_iso_date(dal) and _is_valid_iso_date(al)):
        return jsonify(error="Periodo non valido"), 400

    if not _same_month_iso(dal, al):
        return jsonify(error="Il periodo di competenza deve essere all'interno dello stesso mese."), 400

    if al < dal:
        return jsonify(error="Periodo non valido: la data 'Al' non puÃ² essere precedente alla data 'Dal'."), 400

    dal_iso = _parse_date_iso(dal)
    al_iso = _parse_date_iso(al)
    valore = _parse_float(valore_raw)

    # Non permettere di selezionare giorni giÃ  inclusi in altri versamenti.
    try:
        overlap_days, overlap_summaries = _overlap_days_with_existing_versamenti(
            store_code=str(store_code),
            dal_iso=str(dal_iso),
            al_iso=str(al_iso),
            exclude_id=exclude_id or None,
            exclude_signature=exclude_signature,
        )
        if overlap_days:
            preview_days = ", ".join(overlap_days[:6])
            if len(overlap_days) > 6:
                preview_days += f" (+{len(overlap_days) - 6})"
            msg = f"Nel periodo selezionato ci sono giorni giÃ  versati: {preview_days}. Modifica il periodo."
            return (
                jsonify(
                    {
                        "error": msg,
                        "error_code": "period_overlaps_existing",
                        "days": overlap_days,
                        "overlaps": overlap_summaries,
                    }
                ),
                409,
            )
    except Exception:
        # se fallisce la verifica overlap, non blocchiamo l'utente: continuerÃ  il calcolo
        pass

    try:
        dist = float(
            sum_categoria_period(
                store_code=str(store_code),
                start_iso=dal_iso,
                end_iso=al_iso,
                categoria="Distinte",
            )
        )
    except Exception as e:
        current_app.logger.exception("Errore somma distinte periodo")
        msg = str(e or "")
        low = msg.lower()
        if ("hy000" in msg) or ("driver did not supply" in low):
            return jsonify(error="Database occupato o file in uso. Riprova.", error_code="db_busy"), 503
        return jsonify(error=msg), 500

    diff = float(valore) - dist
    return jsonify({"distinte": dist, "diff": diff})


_DISTINTE_TAGLI = [5, 10, 20, 50, 100, 200, 500]


def _distinte_payload_from_rows(rows: list[dict]) -> dict:
    d1_map = {t: 0 for t in _DISTINTE_TAGLI}
    d2_map = {t: 0 for t in _DISTINTE_TAGLI}
    d1_mon = 0.0
    d2_mon = 0.0
    total = 0.0

    for r in rows or []:
        voce = str(r.get("voce") or "").strip()
        try:
            val_f = float(r.get("valore") or 0.0)
        except Exception:
            val_f = 0.0
        total += val_f

        v_up = voce.upper()
        if v_up.startswith("D1|MONETE"):
            d1_mon = float(val_f)
            continue
        if v_up.startswith("D2|MONETE"):
            d2_mon = float(val_f)
            continue

        # D1|TAGLIO=5|QTA=2
        if "TAGLIO=" in v_up and "QTA=" in v_up:
            try:
                which = "D1" if v_up.startswith("D1|") else "D2" if v_up.startswith("D2|") else ""
                if not which:
                    continue
                parts = voce.split("|")
                taglio = None
                qta = None
                for p in parts:
                    p2 = p.strip()
                    if p2.upper().startswith("TAGLIO="):
                        taglio = float(p2.split("=", 1)[1].replace(",", "."))
                    elif p2.upper().startswith("QTA="):
                        qta = int(float(p2.split("=", 1)[1]))
                if taglio is None or qta is None:
                    continue
                taglio_i = int(round(taglio))
                if taglio_i not in _DISTINTE_TAGLI:
                    continue
                if which == "D1":
                    d1_map[taglio_i] = int(qta)
                else:
                    d2_map[taglio_i] = int(qta)
            except Exception:
                continue

    return {
        "distinte1": [{"taglio": t, "qta": int(d1_map.get(t, 0))} for t in _DISTINTE_TAGLI],
        "distinte2": [{"taglio": t, "qta": int(d2_map.get(t, 0))} for t in _DISTINTE_TAGLI],
        "d1_monete": d1_mon,
        "d2_monete": d2_mon,
        "total": float(total),
    }


@rendiconto_bp.get("/api/versamenti/days")
def api_versamenti_days():
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    dal = (request.args.get("dal") or "").strip()
    al = (request.args.get("al") or "").strip()
    valore_raw = (request.args.get("valore") or "").strip()

    if not dal or not al:
        return jsonify(error="Periodo non valido"), 400

    if not (_is_valid_iso_date(dal) and _is_valid_iso_date(al)):
        return jsonify(error="Periodo non valido"), 400

    if not _same_month_iso(dal, al):
        return jsonify(error="Il periodo di competenza deve essere all'interno dello stesso mese."), 400

    if al < dal:
        return jsonify(error="Periodo non valido: la data 'Al' non puÃ² essere precedente alla data 'Dal'."), 400

    dal_iso = _parse_date_iso(dal)
    al_iso = _parse_date_iso(al)
    valore_dec = _money_to_decimal(valore_raw)

    exclude = {
        "id": (request.args.get("exclude_id") or "").strip(),
        "orig_data_vers": (request.args.get("exclude_orig_data_vers") or "").strip(),
        "orig_dal": (request.args.get("exclude_orig_dal") or "").strip(),
        "orig_al": (request.args.get("exclude_orig_al") or "").strip(),
        "orig_nome": (request.args.get("exclude_orig_nome") or "").strip(),
        "orig_tipo": (request.args.get("exclude_orig_tipo") or "").strip(),
        "orig_tessera": (request.args.get("exclude_orig_tessera") or "").strip(),
        "orig_riferimento": (request.args.get("exclude_orig_riferimento") or "").strip(),
        "orig_valore": (request.args.get("exclude_orig_valore") or "").strip(),
    }

    # Totali distinte per giorno nel range
    try:
        by_day = sum_categoria_by_day_range(
            store_code=str(store_code),
            start_iso=dal_iso,
            end_iso=al_iso,
            categoria="Distinte",
        )
    except Exception as e:
        current_app.logger.exception("Errore somma distinte per giorno")
        return jsonify(error=str(e)), 500

    # Giorni lockati da altri versamenti
    locked = _blocked_distinte_days_in_range(store_code=str(store_code), start_iso=dal_iso, end_iso=al_iso, exclude=exclude)

    days = []
    total_dist = Decimal("0")
    for day_iso in _daterange_iso(dal_iso, al_iso):
        tot = float(by_day.get(day_iso) or 0.0)
        total_dist += Decimal(str(tot))
        # Nota: manteniamo sia "total" (atteso dal popup) sia "distinte" (retro-compatibilitÃ ).
        days.append({
            "date": day_iso,
            "total": float(tot),
            "distinte": float(tot),
            "locked": day_iso in locked,
        })

    diff = _round2(valore_dec - _round2(total_dist))
    return jsonify({"days": days, "distinte_periodo": float(_round2(total_dist)), "diff": float(diff)})


@rendiconto_bp.get("/api/distinte/day")
def api_distinte_day():
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    d = (request.args.get("d") or "").strip()
    if not d:
        return jsonify(error="Data non valida"), 400
    d_iso = _parse_date_iso(d)

    try:
        rows = load_primanota_day(store_code=str(store_code), data_iso=d_iso, categories=["Distinte"])
    except Exception as e:
        current_app.logger.exception("Errore lettura distinte giorno")
        return jsonify(error=str(e)), 500

    payload = _distinte_payload_from_rows(rows)
    payload["d"] = d_iso
    return jsonify(payload)


@rendiconto_bp.post("/api/distinte/day/save")
def api_distinte_day_save():
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    try:
        data = request.get_json(force=True, silent=False) or {}
    except Exception:
        return jsonify(error="Payload non valido"), 400

    d_iso = _parse_date_iso(str(data.get("d") or ""))
    exclude = data.get("exclude") or {}

    # lock: giornata inclusa in altri versamenti (escludendo, se serve, quello in modifica)
    locked = d_iso in _blocked_distinte_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso, exclude=exclude)
    if locked:
        return jsonify(error="Giornata bloccata: non puoi modificare le distinte."), 409

    def _to_items(x):
        return x if isinstance(x, list) else []

    distinte1 = _to_items(data.get("distinte1"))
    distinte2 = _to_items(data.get("distinte2"))

    d1_mon = _parse_float(str(data.get("d1_monete") or ""))
    d2_mon = _parse_float(str(data.get("d2_monete") or ""))

    entries: List[Dict[str, Any]] = []

    def _add_distinte(which: str, items: List[Dict[str, Any]]):
        for it in items or []:
            try:
                taglio = float(it.get("taglio"))
                qta = int(it.get("qta"))
            except Exception:
                continue
            if taglio <= 0 or qta < 0:
                continue
            if qta == 0:
                continue
            voce = f"{which}|TAGLIO={taglio}|QTA={qta}"
            entries.append({"categoria": "Distinte", "voce": voce, "tipo": "SI", "valore": float(taglio * qta)})

    _add_distinte("D1", distinte1)
    _add_distinte("D2", distinte2)

    if d1_mon != 0.0 or str(data.get("d1_monete") or "").strip() != "":
        entries.append({"categoria": "Distinte", "voce": "D1|MONETE", "tipo": "SI", "valore": float(d1_mon)})
    if d2_mon != 0.0 or str(data.get("d2_monete") or "").strip() != "":
        entries.append({"categoria": "Distinte", "voce": "D2|MONETE", "tipo": "SI", "valore": float(d2_mon)})

    try:
        replace_primanota_day(store_code=str(store_code), data_iso=d_iso, entries=entries, categories=["Distinte"])
    except Exception as e:
        current_app.logger.exception("Errore salvataggio distinte giorno")
        return jsonify(error=str(e)), 500

    # ritorna nuovo totale giorno
    try:
        rows2 = load_primanota_day(store_code=str(store_code), data_iso=d_iso, categories=["Distinte"])
        payload = _distinte_payload_from_rows(rows2)
        return jsonify({"ok": True, "d": d_iso, "total": payload.get("total", 0.0)})
    except Exception:
        return jsonify({"ok": True, "d": d_iso, "total": 0.0})


@rendiconto_bp.post("/api/versamenti/commit")
def api_versamenti_commit():
    """Salva un versamento SOLO se la differenza Ã¨ zero (dopo eventuale correzione distinte)."""
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    mode = (request.form.get("mode") or "insert").strip().lower()
    ym = (request.form.get("ym") or request.args.get("ym") or "").strip()
    _y, _m, ym_norm = _parse_ym(ym)

    data_versamento = (request.form.get("data_versamento") or "").strip()
    periodo_dal = (request.form.get("periodo_dal") or "").strip()
    periodo_al = (request.form.get("periodo_al") or "").strip()
    nome_cognome = (request.form.get("nome_cognome") or "").strip()
    tipo_versamento = (request.form.get("tipo_versamento") or "").strip()
    tessera = (request.form.get("tessera") or "").strip()
    riferimento = (request.form.get("riferimento") or "").strip()
    valore_raw = (request.form.get("valore") or "").strip()
    photo_waived, no_receipt_flag, lost_receipt_flag = _versamento_photo_waived()

    if not (data_versamento and periodo_dal and periodo_al and nome_cognome and tipo_versamento and riferimento and valore_raw):
        return jsonify(error="Compila tutti i campi obbligatori."), 400
    tessera_digits = "".join(ch for ch in tessera if ch.isdigit())
    if tipo_versamento.strip().lower() == "tessera" and not tessera_digits:
        return jsonify(error="Compila il campo Tessera."), 400
    if tessera_digits and len(tessera_digits) > 16:
        return jsonify(error="Tessera: massimo 16 cifre."), 400

    if not (_is_valid_iso_date(data_versamento) and _is_valid_iso_date(periodo_dal) and _is_valid_iso_date(periodo_al)):
        return jsonify(error="Periodo non valido"), 400

    if not _same_month_iso(periodo_dal, periodo_al):
        return jsonify(error="Il periodo di competenza deve essere all'interno dello stesso mese."), 400

    if periodo_al < periodo_dal:
        return jsonify(error="Periodo non valido: la data 'Al' non puÃ² essere precedente alla data 'Dal'."), 400

    dv_iso = _parse_date_iso(data_versamento)
    dal_iso = _parse_date_iso(periodo_dal)
    al_iso = _parse_date_iso(periodo_al)
    valore_dec = _round2(_money_to_decimal(valore_raw))

    # Blocco: non permettere salvataggi su periodi giÃ  versati (anche tramite endpoint /api/versamenti/commit).
    # Nota: in update escludiamo il record corrente (se identificabile).
    try:
        ex_id = None
        ex_sig = None
        if mode == "update":
            _rid = (request.form.get("id") or "").strip()
            if _rid:
                ex_id = _rid
            else:
                _odv = (request.form.get("orig_data_vers") or "").strip()
                _odal = (request.form.get("orig_dal") or "").strip()
                _oal = (request.form.get("orig_al") or "").strip()
                _oval = (request.form.get("orig_valore") or "").strip()
                if _odv and _odal and _oal and _oval:
                    ex_sig = {
                        "data_versamento_iso": _odv,
                        "dal_iso": _odal,
                        "al_iso": _oal,
                        "valore_key": _oval,
                    }

        covered_days, _overlaps = _overlap_days_with_existing_versamenti(
            store_code=str(store_code),
            dal_iso=dal_iso,
            al_iso=al_iso,
            exclude_id=ex_id,
            exclude_signature=ex_sig,
        )
        if covered_days:
            preview = ", ".join(covered_days[:6])
            if len(covered_days) > 6:
                preview += f" (+{len(covered_days) - 6})"
            return jsonify(error=f"Nel periodo selezionato ci sono giorni giÃ  versati: {preview}. Modifica il periodo."), 409
    except Exception:
        current_app.logger.exception("Errore verifica giorni giÃ  versati (commit)")

    # Verifica differenza
    try:
        dist = Decimal(str(sum_categoria_period(store_code=str(store_code), start_iso=dal_iso, end_iso=al_iso, categoria="Distinte")))
    except Exception as e:
        current_app.logger.exception("Errore somma distinte periodo")
        return jsonify(error=str(e)), 500

    diff = _round2(valore_dec - _round2(dist))
    if diff != Decimal("0.00"):
        return jsonify({"needs_adjustment": True, "diff": float(diff), "distinte_periodo": float(_round2(dist))}), 409

    foto_file = None
    foto = request.files.get("foto")

    if mode == "update":
        record_id = (request.form.get("id") or "").strip()

        orig_data_vers = (request.form.get("orig_data_vers") or "").strip()
        orig_dal = (request.form.get("orig_dal") or "").strip()
        orig_al = (request.form.get("orig_al") or "").strip()
        orig_nome = (request.form.get("orig_nome") or "").strip()
        orig_tipo = (request.form.get("orig_tipo") or "").strip()
        orig_tessera = (request.form.get("orig_tessera") or "").strip()
        orig_riferimento = (request.form.get("orig_riferimento") or "").strip()
        orig_valore = (request.form.get("orig_valore") or "").strip()

        old_foto = None
        try:
            old_foto = get_versamento_photo_file(
                store_code=str(store_code),
                record_id=record_id,
                orig_data_vers_iso=_parse_date_iso(orig_data_vers),
                orig_dal_iso=_parse_date_iso(orig_dal),
                orig_al_iso=_parse_date_iso(orig_al),
                orig_nome=orig_nome,
                orig_tipo=orig_tipo,
                orig_tessera=orig_tessera,
                orig_riferimento=orig_riferimento,
                orig_valore_key=orig_valore,
            )
        except Exception:
            old_foto = None

        new_foto_file = None
        has_uploaded_photo = foto is not None and bool(getattr(foto, "filename", ""))
        has_existing_photo = bool(str(old_foto or "").strip())
        if not photo_waived and not has_existing_photo and not has_uploaded_photo:
            return jsonify(error="Carica la foto oppure seleziona una delle dichiarazioni previste."), 400
        if has_uploaded_photo:
            try:
                new_foto_file = upload_versamento_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    file_storage=foto,
                    data_iso=dv_iso,
                )
            except Exception as e:
                return jsonify(error=_friendly_photo_upload_error(e, subject="foto versamento")), 500

        try:
            update_versamento(
                store_code=str(store_code),
                record_id=record_id,
                orig_data_vers_iso=_parse_date_iso(orig_data_vers),
                orig_dal_iso=_parse_date_iso(orig_dal),
                orig_al_iso=_parse_date_iso(orig_al),
                orig_nome=orig_nome,
                orig_tipo=orig_tipo,
                orig_tessera=orig_tessera,
                orig_riferimento=orig_riferimento,
                orig_valore_key=orig_valore,
                new_data_vers_iso=dv_iso,
                new_dal_iso=dal_iso,
                new_al_iso=al_iso,
                new_nome=nome_cognome,
                new_tipo=tipo_versamento,
                new_tessera=tessera_digits,
                new_riferimento=riferimento,
                new_valore_euro=str(valore_dec),
                new_foto_file=new_foto_file,
                new_no_receipt_flag=no_receipt_flag,
                new_lost_receipt_flag=lost_receipt_flag,
            )
        except Exception as e:
            if new_foto_file:
                try:
                    delete_versamento_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=new_foto_file,
                    )
                except Exception:
                    pass
            return jsonify(error=f"Errore aggiornamento versamento: {e}"), 500

        if new_foto_file and old_foto and old_foto != new_foto_file:
            try:
                delete_versamento_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    filename=old_foto,
                )
            except Exception:
                pass

        return jsonify({"ok": True, "ym": ym_norm})

    # INSERT
    has_uploaded_photo = foto is not None and bool(getattr(foto, "filename", ""))
    if not photo_waived and not has_uploaded_photo:
        return jsonify(error="Carica la foto oppure seleziona una delle dichiarazioni previste."), 400
    if has_uploaded_photo:
        try:
            foto_file = upload_versamento_photo(
                sb_jwt=str(session.get("sb_token") or ""),
                user_id=str(session.get("uid") or ""),
                store_code=str(store_code),
                file_storage=foto,
                data_iso=dv_iso,
            )
        except Exception as e:
            return jsonify(error=_friendly_photo_upload_error(e, subject="foto versamento")), 500

    try:
        insert_versamento(
            store_code=str(store_code),
            data_versamento_iso=dv_iso,
            periodo_dal_iso=dal_iso,
            periodo_al_iso=al_iso,
            nome_cognome=nome_cognome,
            tipo_versamento=tipo_versamento,
            tessera=tessera_digits,
            riferimento=riferimento,
            valore_euro=str(valore_dec),
            foto_file=foto_file,
            no_receipt_flag=no_receipt_flag,
            lost_receipt_flag=lost_receipt_flag,
        )
    except Exception as e:
        if foto_file:
            try:
                delete_versamento_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    filename=foto_file,
                )
            except Exception:
                pass
        return jsonify(error=f"Errore salvataggio versamento: {e}"), 500

    return jsonify({"ok": True, "ym": ym_norm})


# -------------------------
# Distinta cassa (Prima Nota)
# -------------------------


_CHIUSURA_FIELDS = [
    ("vendite_lorde", "VENDITE LORDE", "money"),
    ("annullati", "ANNULLATI", "money"),
    ("scontrini", "SCONTRINI", "int"),
    ("pos", "POS", "money"),
    ("contanti", "CONTANTI", "money"),
    ("ticket", "TICKET", "money"),
    ("fatture", "FATTURE", "money"),
    ("numero_fatture", "NUMERO FATTURE", "int"),
    ("omaggi", "OMAGGI", "money"),
    ("vendite_iva_4", "VENDITE IVA 4%", "money"),
    ("vendite_iva_22", "VENDITE IVA 22%", "money"),
]


_VOICE_TO_KEY = {lbl: key for (key, lbl, _t) in _CHIUSURA_FIELDS}
_KEY_TO_TYPE = {key: t for (key, _lbl, t) in _CHIUSURA_FIELDS}


def _get_cash_statement_config_cached() -> Dict[str, List[Dict[str, Any]]]:
    tenant_key = _session_tenant_key()
    language_code = _session_ui_language()
    cache_key = f"cash_statement_config:{tenant_key}:{language_code}"
    return _rendiconto_ttl_cached(cache_key, 300, list_cash_statement_config) or {"sections": [], "fields": []}


def _distinta_custom_config_from_cfg(cfg: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    sections = {
        str(s.get("section_key") or ""): s
        for s in (cfg.get("sections") or [])
        if s.get("is_active") and s.get("is_visible") and str(s.get("section_kind") or "fields") == "fields"
    }
    out: List[Dict[str, Any]] = []
    fixed_keys = set(_KEY_TO_TYPE.keys()) | {"cash_deposit_line", "ticket_line", "delivery_line", "coupon_line"}
    by_section: Dict[str, List[Dict[str, Any]]] = {}
    for field in (cfg.get("fields") or []):
        section_key = str(field.get("section_key") or "")
        field_key = str(field.get("field_key") or "")
        if section_key not in sections:
            continue
        if not field.get("is_active") or not field.get("is_visible"):
            continue
        if field_key in fixed_keys:
            continue
        by_section.setdefault(section_key, []).append(field)
    for section_key, fields in by_section.items():
        section = sections[section_key]
        out.append(
            {
                "section_key": section_key,
                "label": str(section.get("label") or section_key),
                "legacy_category": str(section.get("legacy_category") or section.get("label") or section_key),
                "fields": sorted(fields, key=lambda f: (int(f.get("sort_order") or 0), str(f.get("label") or ""))),
            }
        )
    return sorted(out, key=lambda s: int(sections[s["section_key"]].get("sort_order") or 0))


def _distinta_custom_config() -> List[Dict[str, Any]]:
    try:
        return _distinta_custom_config_from_cfg(_get_cash_statement_config_cached())
    except Exception:
        current_app.logger.exception("Errore lettura configurazione dinamica Distinta cassa")
        return []


def _distinta_custom_option_list_config_from_cfg(cfg: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    sections = [
        s for s in (cfg.get("sections") or [])
        if s.get("is_active") and s.get("is_visible") and str(s.get("section_kind") or "") == "option_list"
    ]
    standard = {"ticket", "delivery", "coupon"}
    out = []
    for section in sections:
        section_key = str(section.get("section_key") or "")
        if section_key in standard:
            continue
        options = [
            {
                "value": str(f.get("label") or f.get("legacy_voce") or "").strip(),
                "tipo": str(f.get("legacy_tipo") or "SI").strip().upper() or "SI",
                "cash_difference_sign": float(f.get("cash_difference_sign") or 0.0) if f.get("affects_cash_difference") else 0.0,
            }
            for f in (cfg.get("fields") or [])
            if str(f.get("section_key") or "") == section_key and f.get("is_active") and f.get("is_visible")
        ]
        options = [o for o in options if o["value"]]
        out.append(
            {
                "section_key": section_key,
                "label": str(section.get("label") or section_key),
                "legacy_category": str(section.get("legacy_category") or section.get("label") or section_key),
                "sort_order": int(section.get("sort_order") or 0),
                "options": sorted(options, key=lambda o: o["value"].lower()),
            }
        )
    return sorted(out, key=lambda s: (s["sort_order"], s["label"]))


def _distinta_custom_option_list_config() -> List[Dict[str, Any]]:
    try:
        return _distinta_custom_option_list_config_from_cfg(_get_cash_statement_config_cached())
    except Exception:
        current_app.logger.exception("Errore lettura configurazione elenchi dinamici Distinta cassa")
        return []


def _custom_field_form_name(section_key: str, field_key: str) -> str:
    return f"custom_field__{section_key}__{field_key}"


def _custom_list_form_name(section_key: str) -> str:
    return f"custom_list__{section_key}_json"


def _distinta_entry_categories(
    custom_config: List[Dict[str, Any]] | None = None,
    custom_option_lists: List[Dict[str, Any]] | None = None,
) -> List[str]:
    cats = ["Dati chiusura", "Distinte", "Ticket", "Delivery", "Coupon"]

    def add(value: Any) -> None:
        cat = str(value or "").strip()
        if cat and cat not in cats:
            cats.append(cat)

    if custom_config is None:
        custom_config = _distinta_custom_config()
    if custom_option_lists is None:
        custom_option_lists = _distinta_custom_option_list_config()

    for section in custom_config or []:
        add(section.get("legacy_category"))
        add(section.get("label"))
        for field in section.get("fields") or []:
            add(field.get("legacy_category"))

    for section in custom_option_lists or []:
        add(section.get("legacy_category"))
        add(section.get("label"))

    return cats


def _cash_statement_labels_from_cfg(cfg: Dict[str, List[Dict[str, Any]]]) -> Dict[str, Dict[str, str]]:
    return {
        "sections": {
            str(section.get("section_key") or ""): str(section.get("label") or section.get("section_key") or "")
            for section in (cfg.get("sections") or [])
        },
        "fields": {
            str(field.get("field_key") or ""): str(field.get("label") or field.get("field_key") or "")
            for field in (cfg.get("fields") or [])
            if str(field.get("section_key") or "") == "dati_chiusura"
        },
    }


def _get_ipratico_enabled_cached() -> bool:
    tenant_key = _session_tenant_key()
    cache_key = f"ipratico_enabled:{tenant_key}"
    return bool(
        _rendiconto_ttl_cached(
            cache_key,
            300,
            lambda: bool(get_ipratico_config().get("enabled")),
        )
    )


def _get_elenchi_options_cached(*, store_code: str) -> Dict[str, Any]:
    tenant_key = _session_tenant_key()
    cache_key = f"distinta_elenchi:{tenant_key}:{str(store_code).strip()}"
    return _rendiconto_ttl_cached(cache_key, 120, lambda: get_elenchi_options(store_code=str(store_code))) or {
        "tickets": [],
        "deliveries": [],
        "coupons": [],
    }


def _label_norm(v: str) -> str:
    return "".join(ch for ch in str(v or "").strip().lower() if ch.isalnum())


def _delivery_voce_from_import(options: List[Dict[str, Any]], provider: str, is_cash: bool, fallback_label: str) -> tuple[str, bool]:
    values = [str((o or {}).get("value") or "").strip() for o in (options or [])]
    values = [v for v in values if v]
    if not values:
        return fallback_label, False

    provider_norm = _label_norm(provider)
    best_val = ""
    best_score = -10**9
    for val in values:
        norm = _label_norm(val)
        score = 0
        if provider_norm and provider_norm in norm:
            score += 100
        elif provider_norm and norm in provider_norm:
            score += 40
        if any(k in norm for k in ("deliveroo", "glovo", "ubereats", "ubereat", "justeat", "just eat", "wol t", "wolt")):
            score += 5
        option_cash = any(k in norm for k in ("contanti", "cash", "cod"))
        if is_cash == option_cash:
            score += 20
        else:
            score -= 20
        if score > best_score:
            best_score = score
            best_val = val

    if best_val and best_score >= 60:
        return best_val, True
    return fallback_label, False


def _coupon_voce_satispay(options: List[Dict[str, Any]]) -> tuple[str, bool]:
    for opt in (options or []):
        val = str((opt or {}).get("value") or "").strip()
        if "satispay" in _label_norm(val):
            return val, True
    return "Satispay", False


def _ticket_voce_from_import(options: List[Dict[str, Any]]) -> tuple[str, bool]:
    values = [str((o or {}).get("value") or "").strip() for o in (options or [])]
    values = [v for v in values if v]
    if not values:
        return "Ticket", False
    if len(values) == 1:
        return values[0], True
    for val in values:
        if "ticket" in _label_norm(val):
            return val, True
    return values[0], False


def _map_ipratico_import_for_distinta(
    *,
    imported: Dict[str, Any],
    options: Dict[str, List[Dict[str, Any]]],
    store_code: str,
    d_iso: str,
) -> Dict[str, Any]:
    warnings = list(imported.get("warnings") or [])

    deliveries_map: Dict[tuple[str, str], float] = {}
    for row in (imported.get("deliveries") or []):
        provider = str((row or {}).get("provider") or "").strip()
        is_cash = bool((row or {}).get("is_cash"))
        fallback_label = str((row or {}).get("label") or "").strip() or "Delivery"
        voce, matched = _delivery_voce_from_import(options.get("deliveries") or [], provider, is_cash, fallback_label)
        if not matched:
            warnings.append(f"Voce delivery non trovata in ELENCHI per provider '{provider}': uso '{voce}'.")
        tipo = "NO" if is_cash else "SI"
        key = (voce, tipo)
        deliveries_map[key] = float(deliveries_map.get(key, 0.0)) + float((row or {}).get("valore") or 0.0)

    deliveries = [
        {"voce": voce, "tipo": tipo, "valore": round(val, 2)}
        for (voce, tipo), val in sorted(deliveries_map.items(), key=lambda x: (_label_norm(x[0][0]), x[0][1]))
        if abs(val) > 1e-9
    ]

    coupons = []
    satispay_total = float((imported.get("satispay") or 0.0))
    if abs(satispay_total) > 1e-9:
        voce, matched = _coupon_voce_satispay(options.get("coupons") or [])
        if not matched:
            warnings.append(f"Voce coupon Satispay non trovata in ELENCHI: uso '{voce}'.")
        coupons.append({"voce": voce, "tipo": "SI", "valore": round(satispay_total, 2)})

    return {
        "ok": True,
        "chiusura": imported.get("chiusura") or {},
        "tickets": [],
        "deliveries": deliveries,
        "coupons": coupons,
        "meta": {
            "records_count": int(imported.get("records_count") or 0),
            "raw_records_count": int(imported.get("raw_records_count") or 0),
            "skipped_unmatched": int(imported.get("skipped_unmatched") or 0),
            "store_code": str(store_code),
            "date": d_iso,
        },
        "warnings": warnings,
        "satispay": round(satispay_total, 2),
    }


def _ipratico_rule_key(value: object) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _ipratico_mapping_candidates(fact: Dict[str, Any]) -> List[str]:
    method_key = str((fact or {}).get("method_key") or "").strip()
    classified = str((fact or {}).get("classified_as") or "").strip().lower()
    raw_label = str((fact or {}).get("raw_payment_label") or "")
    provider = _ipratico_rule_key((fact or {}).get("delivery_provider") or (fact or {}).get("delivery_source_app") or "")
    delivery_bucket = str((fact or {}).get("delivery_bucket") or "").strip().lower()

    out: List[str] = []
    if method_key:
        out.append(method_key)
    if provider and delivery_bucket:
        out.append(f"rule:delivery:{provider}:{delivery_bucket}")
    if classified.startswith("delivery_"):
        out.append(f"rule:payment:{classified.replace('delivery_', '')}")
    elif classified:
        out.append(f"rule:payment:{classified}")

    label_norm = _ipratico_rule_key(raw_label)
    for token in ("cashondelivery", "contanti", "cash", "pos", "card", "bancomat", "visa", "mastercard", "nexi", "ticket", "satispay"):
        if token in label_norm:
            out.append(f"rule:payment:{token}")

    seen = set()
    return [x for x in out if x and not (x in seen or seen.add(x))]


def _map_ipratico_import_for_distinta_with_mappings(
    *,
    imported: Dict[str, Any],
    mappings: List[Dict[str, Any]],
    store_code: str,
    d_iso: str,
) -> Dict[str, Any]:
    active = [
        m for m in (mappings or [])
        if m and bool(m.get("is_active")) and str(m.get("target_section_key") or "").strip() and str(m.get("target_field_key") or "").strip()
    ]
    mapping_by_key = {str(m.get("method_key") or "").strip(): m for m in active if str(m.get("method_key") or "").strip()}
    warnings = list(imported.get("warnings") or [])
    chiusura = dict(imported.get("chiusura") or {})
    for key in ("contanti", "pos", "ticket"):
        chiusura[key] = 0.0

    tickets_map: Dict[tuple[str, str], float] = {}
    deliveries_map: Dict[tuple[str, str], float] = {}
    coupons_map: Dict[tuple[str, str], float] = {}
    custom_lists_map: Dict[str, Dict[tuple[str, str], float]] = {}
    mapped_count = 0
    unmapped_count = 0

    def add_line(bucket: Dict[tuple[str, str], float], label: str, tipo: str, amount: float) -> None:
        key = (str(label or "").strip() or "Voce iPratico", str(tipo or "").strip().upper() or "SI")
        bucket[key] = float(bucket.get(key, 0.0)) + float(amount or 0.0)

    for fact in imported.get("payment_facts") or []:
        amount = float((fact or {}).get("amount") or 0.0)
        if abs(amount) <= 1e-9:
            continue
        mapping = None
        for key in _ipratico_mapping_candidates(fact):
            mapping = mapping_by_key.get(key)
            if mapping:
                break
        if not mapping:
            unmapped_count += 1
            label = str((fact or {}).get("method_label") or (fact or {}).get("raw_payment_label") or "Senza etichetta").strip()
            warnings.append(f"Metodo iPratico non mappato: {label} ({round(amount, 2)}).")
            continue

        mapped_count += 1
        section = str(mapping.get("target_section_key") or "").strip()
        field = str(mapping.get("target_field_key") or "").strip()
        label = str(mapping.get("target_label") or "").strip() or str((fact or {}).get("method_label") or "").strip()
        tipo = str(mapping.get("target_tipo") or "").strip().upper()

        if section == "dati_chiusura" and field:
            chiusura[field] = round(float(chiusura.get(field) or 0.0) + amount, 2)
        elif section == "ticket" and field == "ticket_line":
            add_line(tickets_map, label, tipo or "SI", amount)
        elif section == "delivery" and field == "delivery_line":
            add_line(deliveries_map, label, tipo or ("NO" if str((fact or {}).get("delivery_bucket") or "") == "cash" else "SI"), amount)
        elif section == "coupon" and field == "coupon_line":
            add_line(coupons_map, label, tipo or "SI", amount)
        elif section:
            custom_bucket = custom_lists_map.setdefault(section, {})
            add_line(custom_bucket, label, tipo or "SI", amount)

    def rows_from(bucket: Dict[tuple[str, str], float]) -> List[Dict[str, Any]]:
        return [
            {"voce": voce, "tipo": tipo, "valore": round(val, 2)}
            for (voce, tipo), val in sorted(bucket.items(), key=lambda x: (_label_norm(x[0][0]), x[0][1]))
            if abs(val) > 1e-9
        ]

    custom_lists = {section: rows_from(bucket) for section, bucket in custom_lists_map.items()}
    return {
        "ok": True,
        "chiusura": chiusura,
        "tickets": rows_from(tickets_map),
        "deliveries": rows_from(deliveries_map),
        "coupons": rows_from(coupons_map),
        "custom_lists": custom_lists,
        "meta": {
            "records_count": int(imported.get("records_count") or 0),
            "raw_records_count": int(imported.get("raw_records_count") or 0),
            "skipped_unmatched": int(imported.get("skipped_unmatched") or 0),
            "mapped_payments": int(mapped_count),
            "unmapped_payments": int(unmapped_count),
            "store_code": str(store_code),
            "date": d_iso,
            "mapping_mode": "tenant",
        },
        "warnings": warnings,
        "satispay": 0.0,
    }


@rendiconto_bp.get("/api/distinta-cassa/import-ipratico")
def distinta_cassa_import_ipratico():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return jsonify({"ok": False, "error": "Login richiesto."}), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify({"ok": False, "error": "Seleziona prima uno store."}), 400

    try:
        if not get_ipratico_config().get("enabled"):
            return jsonify({"ok": False, "error": "Integrazione iPratico disattivata."}), 409
    except Exception as e:
        current_app.logger.exception("Errore lettura configurazione iPratico")
        return jsonify({"ok": False, "error": f"Configurazione iPratico non disponibile: {e}"}), 500

    d_iso = _parse_date_iso(request.args.get("d") or "")
    if (not _is_admin_role()) and (d_iso in _validated_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso)):
        return jsonify({"ok": False, "error": "Giornata convalidata: solo l'amministratore puÃ² modificare la distinta."}), 409

    try:
        options = get_elenchi_options(store_code=str(store_code))
    except Exception:
        options = {"tickets": [], "deliveries": [], "coupons": []}

    try:
        mappings = list_ipratico_payment_mappings()
    except Exception:
        current_app.logger.exception("Errore lettura mappature iPratico")
        mappings = []

    active_mappings = [m for m in (mappings or []) if bool((m or {}).get("is_active"))]
    try:
        if active_mappings:
            imported = import_ipratico_distinta_day_detailed(str(store_code), d_iso)
            mapped = _map_ipratico_import_for_distinta_with_mappings(
                imported=imported,
                mappings=active_mappings,
                store_code=str(store_code),
                d_iso=d_iso,
            )
        else:
            imported = import_ipratico_distinta_day(str(store_code), d_iso)
            mapped = _map_ipratico_import_for_distinta(
                imported=imported,
                options=options,
                store_code=str(store_code),
                d_iso=d_iso,
            )
            mapped.setdefault("meta", {})["mapping_mode"] = "legacy"
    except Exception as e:
        current_app.logger.exception("Errore import iPratico Distinta cassa")
        return jsonify({"ok": False, "error": str(e)}), 500
    try:
        upsert_distinta_cassa_ipratico_snapshot(
            store_code=str(store_code),
            data_iso=d_iso,
            imported_payload=imported,
            mapped_payload=mapped,
            save_origin="manual_import",
            imported_manually=True,
        )
    except Exception as e:
        current_app.logger.exception("Errore salvataggio snapshot iPratico Distinta cassa")
        mapped["warnings"] = list(mapped.get("warnings") or [])
        mapped["warnings"].append(f"Snapshot iPratico non salvato su database: {e}")
    return jsonify(mapped)


@rendiconto_bp.get("/distinta-cassa")
def distinta_cassa():
    route_started_at = time.perf_counter()
    timings: Dict[str, float] = {}
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")

    d_iso = _parse_date_iso(request.args.get("d") or "")

    # Se la giornata Ã¨ giÃ  inclusa nel periodo competenza di un versamento,
    # blocchiamo la modifica delle distinte contanti (tagli/monete).
    locked_distinte = False
    locked_validated = False
    if store_code:
        locked_set = _locked_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso)
        locked_distinte = d_iso in locked_set
        locked_validated = (not _is_admin_role()) and (d_iso in _validated_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso))

    options = {"tickets": [], "deliveries": [], "coupons": []}
    existing_rows: List[Dict[str, Any]] = []
    cash_statement_labels = {"sections": {}, "fields": {}}
    cash_statement_cfg: Dict[str, List[Dict[str, Any]]] = {"sections": [], "fields": []}
    try:
        started_at = time.perf_counter()
        cash_statement_cfg = _get_cash_statement_config_cached()
        cash_statement_labels = _cash_statement_labels_from_cfg(cash_statement_cfg)
        timings["cash_cfg"] = time.perf_counter() - started_at
    except Exception:
        current_app.logger.exception("Errore lettura etichette tradotte Distinta cassa")
    custom_config = _distinta_custom_config_from_cfg(cash_statement_cfg)
    custom_option_lists = _distinta_custom_option_list_config_from_cfg(cash_statement_cfg)
    distinta_categories = _distinta_entry_categories(custom_config, custom_option_lists)
    custom_field_index = {}
    for section in custom_config:
        for field in section.get("fields") or []:
            voce_key = str(field.get("legacy_voce") or field.get("label") or "")
            for cat_key in (
                str(field.get("legacy_category") or ""),
                str(section.get("legacy_category") or ""),
                str(section.get("label") or ""),
            ):
                if cat_key and voce_key:
                    custom_field_index[(cat_key, voce_key)] = field
    spese_info = {"total": 0.0, "note_credito": 0.0, "net": 0.0}
    ipratico_enabled = True
    try:
        started_at = time.perf_counter()
        ipratico_enabled = _get_ipratico_enabled_cached()
        timings["ipratico_cfg"] = time.perf_counter() - started_at
    except Exception:
        current_app.logger.exception("Errore lettura configurazione iPratico")
        ipratico_enabled = False

    init = {
        "chiusura": {k: "" for (k, _lbl, _t) in _CHIUSURA_FIELDS},
        "distinte1": [],
        "distinte2": [],
        "d1_monete": "",
        "d2_monete": "",
        "tickets": [],
        "deliveries": [],
        "coupons": [],
        "custom_fields": {},
        "custom_lists": {},
    }
    distinta_photo_file = None
    ipratico_snapshot = None

    if store_code:
        try:
            started_at = time.perf_counter()
            options = _get_elenchi_options_cached(store_code=str(store_code))
            timings["elenchi"] = time.perf_counter() - started_at
        except Exception as e:
            current_app.logger.exception("Errore lettura ELENCHI")
            flash(f"Errore lettura ELENCHI: {e}", "danger")

        try:
            started_at = time.perf_counter()
            existing_rows = load_primanota_day(
                store_code=str(store_code),
                data_iso=d_iso,
                categories=distinta_categories,
            )
            timings["primanota_day"] = time.perf_counter() - started_at
        except Exception as e:
            current_app.logger.exception("Errore lettura DATIPRIMANOTA")
            flash(f"Errore lettura DATIPRIMANOTA: {e}", "danger")

        try:
            started_at = time.perf_counter()
            spese_info = sum_spese_day(store_code=str(store_code), data_iso=d_iso)
            timings["spese_day"] = time.perf_counter() - started_at
        except Exception as e:
            current_app.logger.exception("Errore calcolo spese giorno")
            flash(f"Errore calcolo spese giorno: {e}", "danger")

        try:
            started_at = time.perf_counter()
            distinta_photo_file = get_distinta_cassa_photo_file(store_code=str(store_code), data_iso=d_iso)
            timings["distinta_photo"] = time.perf_counter() - started_at
        except Exception as e:
            current_app.logger.exception("Errore lettura foto Distinta cassa")
            flash(f"Errore lettura foto Distinta cassa: {e}", "danger")

        try:
            started_at = time.perf_counter()
            ipratico_snapshot = get_distinta_cassa_ipratico_snapshot(store_code=str(store_code), data_iso=d_iso)
            timings["ipratico_snapshot"] = time.perf_counter() - started_at
        except Exception as e:
            current_app.logger.exception("Errore lettura snapshot iPratico Distinta cassa")
            flash(f"Errore lettura snapshot iPratico: {e}", "warning")

        # parse existing
        for r in existing_rows:
            cat = str(r.get("categoria") or "")
            voce = str(r.get("voce") or "")
            tipo = str(r.get("tipo") or "SI")
            val = r.get("valore")
            try:
                val_f = float(val)
            except Exception:
                val_f = 0.0

            if cat == "Dati chiusura":
                key = _VOICE_TO_KEY.get(voce)
                if key:
                    t = _KEY_TO_TYPE.get(key, "money")
                    if t == "int":
                        try:
                            init["chiusura"][key] = str(int(round(val_f)))
                        except Exception:
                            init["chiusura"][key] = "0"
                    else:
                        init["chiusura"][key] = str(val_f)

            elif cat == "Distinte":
                # Voce formato: D1|TAGLIO=20|QTA=3 oppure D1|MONETE
                v = voce.strip()
                if v.startswith("D1|MONETE"):
                    init["d1_monete"] = str(val_f)
                elif v.startswith("D2|MONETE"):
                    init["d2_monete"] = str(val_f)
                elif v.startswith("D1|TAGLIO="):
                    try:
                        parts = v.split("|")
                        taglio = float(parts[1].split("=", 1)[1])
                        qta = int(parts[2].split("=", 1)[1])
                        init["distinte1"].append({"taglio": taglio, "qta": qta})
                    except Exception:
                        pass
                elif v.startswith("D2|TAGLIO="):
                    try:
                        parts = v.split("|")
                        taglio = float(parts[1].split("=", 1)[1])
                        qta = int(parts[2].split("=", 1)[1])
                        init["distinte2"].append({"taglio": taglio, "qta": qta})
                    except Exception:
                        pass

            elif cat == "Ticket":
                init["tickets"].append({"voce": voce, "tipo": tipo, "valore": val_f})
            elif cat == "Delivery":
                init["deliveries"].append({"voce": voce, "tipo": tipo, "valore": val_f})
            elif cat == "Coupon":
                init["coupons"].append({"voce": voce, "tipo": tipo, "valore": val_f})
            else:
                field = custom_field_index.get((cat, voce))
                if field:
                    init["custom_fields"][_custom_field_form_name(str(field.get("section_key") or ""), str(field.get("field_key") or ""))] = str(val_f)
                else:
                    for section in custom_option_lists:
                        if cat == section.get("legacy_category") or cat == section.get("label"):
                            init["custom_lists"].setdefault(section.get("section_key"), []).append({"voce": voce, "tipo": tipo, "valore": val_f})

    has_saved_data = bool(existing_rows)
    total_elapsed = time.perf_counter() - route_started_at
    if total_elapsed > 0.6:
        current_app.logger.info(
            "Distinta cassa load store=%s date=%s total=%.3fs timings=%s",
            store_code,
            d_iso,
            total_elapsed,
            {k: round(v, 3) for k, v in timings.items()},
        )

    return render_template(
        "rendiconto_distinta_cassa.html",
        store_code=store_code,
        store_name=store_name,
        d=d_iso,
        locked_distinte=locked_distinte,
        options=options,
        spese_info=spese_info,
        init=init,
        has_saved_data=has_saved_data,
        chiusura_fields=_CHIUSURA_FIELDS,
        cash_statement_labels=cash_statement_labels,
        distinta_photo_file=distinta_photo_file,
        ipratico_snapshot=ipratico_snapshot,
        ipratico_enabled=ipratico_enabled,
        custom_config=custom_config,
        custom_option_lists=custom_option_lists,
        locked_validated=locked_validated,
    )


@rendiconto_bp.post("/distinta-cassa/save")
def distinta_cassa_save():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.distinta_cassa"))

    d_iso = _parse_date_iso(request.form.get("d") or "")
    if (not _is_admin_role()) and (d_iso in _validated_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso)):
        flash("Giornata convalidata: solo l'amministratore puÃ² modificare la distinta di cassa.", "warning")
        return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))

    new_photo_file = None
    foto = request.files.get("foto")
    if foto is not None and getattr(foto, "filename", ""):
        try:
            new_photo_file = upload_azzeramento_photo(
                sb_jwt=str(session.get("sb_token") or ""),
                user_id=str(session.get("uid") or ""),
                store_code=str(store_code),
                file_storage=foto,
                data_iso=d_iso,
            )
        except Exception as e:
            current_app.logger.exception("Errore upload foto Distinta cassa")
            flash(_friendly_photo_upload_error(e, subject="foto distinta") + " La distinta verra' salvata senza cambiare la foto.", "warning")
            new_photo_file = None

    # campi chiusura
    required_keys = {"vendite_lorde", "scontrini", "contanti", "pos"}
    chiusura_vals: Dict[str, float] = {}
    for (key, _lbl, _t) in _CHIUSURA_FIELDS:
        raw = (request.form.get(key) or "").strip()
        if key in required_keys and raw == "":
            flash("Compila i campi obbligatori: Vendite lorde, Scontrini, Contanti, POS.", "warning")
            return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))
        if raw == "":
            chiusura_vals[key] = 0.0
            continue

        # parsing coerente: money -> float, int -> intero
        if _KEY_TO_TYPE.get(key) == "int":
            try:
                chiusura_vals[key] = float(int(round(_parse_float(raw))))
            except Exception:
                chiusura_vals[key] = 0.0
        else:
            chiusura_vals[key] = _parse_float(raw)

    # json lists
    def _load_json(name: str):
        s = (request.form.get(name) or "").strip()
        if not s:
            return []
        try:
            v = json.loads(s)
            return v if isinstance(v, list) else []
        except Exception:
            return []

    distinte1 = _load_json("distinte1_json")
    distinte2 = _load_json("distinte2_json")
    tickets = _load_json("tickets_json")
    deliveries = _load_json("deliveries_json")
    coupons = _load_json("coupons_json")

    d1_monete = _parse_float(request.form.get("d1_monete") or "")
    d2_monete = _parse_float(request.form.get("d2_monete") or "")

    # opzioni per tipo SI/NO (per sicurezza lato server)
    try:
        opt = get_elenchi_options(store_code=str(store_code))
    except Exception:
        opt = {"tickets": [], "deliveries": [], "coupons": []}

    ticket_tipo_map = {o.get("value"): (o.get("tipo") or "SI") for o in (opt.get("tickets") or [])}
    delivery_tipo_map = {o.get("value"): (o.get("tipo") or "SI") for o in (opt.get("deliveries") or [])}
    coupon_tipo_map = {o.get("value"): (o.get("tipo") or "SI") for o in (opt.get("coupons") or [])}
    custom_config = _distinta_custom_config()
    custom_option_lists = _distinta_custom_option_list_config()
    distinta_categories = _distinta_entry_categories(custom_config, custom_option_lists)
    custom_field_vals: Dict[str, float] = {}

    try:
        ipratico_snapshot = get_distinta_cassa_ipratico_snapshot(store_code=str(store_code), data_iso=d_iso)
    except Exception as ex_snap:
        current_app.logger.exception("Errore lettura snapshot iPratico Distinta cassa")
        ipratico_snapshot = None
        flash(f"Errore lettura snapshot iPratico: {ex_snap}", "warning")

    ipratico_enabled = True
    try:
        ipratico_enabled = bool(get_ipratico_config().get("enabled"))
    except Exception:
        current_app.logger.exception("Errore lettura configurazione iPratico")
        ipratico_enabled = False

    if ipratico_enabled and not ipratico_snapshot:
        try:
            mappings = list_ipratico_payment_mappings()
            active_mappings = [m for m in (mappings or []) if bool((m or {}).get("is_active"))]
            if active_mappings:
                imported = import_ipratico_distinta_day_detailed(str(store_code), d_iso)
                mapped = _map_ipratico_import_for_distinta_with_mappings(
                    imported=imported,
                    mappings=active_mappings,
                    store_code=str(store_code),
                    d_iso=d_iso,
                )
            else:
                imported = import_ipratico_distinta_day(str(store_code), d_iso)
                mapped = _map_ipratico_import_for_distinta(
                    imported=imported,
                    options=opt,
                    store_code=str(store_code),
                    d_iso=d_iso,
                )
                mapped.setdefault("meta", {})["mapping_mode"] = "legacy"
            upsert_distinta_cassa_ipratico_snapshot(
                store_code=str(store_code),
                data_iso=d_iso,
                imported_payload=imported,
                mapped_payload=mapped,
                save_origin="auto_on_save",
                imported_manually=False,
            )
        except Exception as ex_ipr:
            current_app.logger.exception("Errore auto-import iPratico al salvataggio Distinta cassa")
            flash(f"Distinta salvata senza snapshot iPratico: {ex_ipr}", "warning")

    entries: List[Dict[str, Any]] = []

    # Dati chiusura (solo manuali)
    for (key, lbl, _t) in _CHIUSURA_FIELDS:
        entries.append(
            {
                "categoria": "Dati chiusura",
                "voce": lbl,
                "tipo": "SI",
                "valore": float(chiusura_vals.get(key, 0.0)),
            }
        )

    # Distinte (contanti). Se la giornata Ã¨ inclusa in un periodo competenza di un versamento,
    # NON consentiamo modifiche su tagli/monete e preserviamo quanto giÃ  salvato.
    locked_distinte = d_iso in _blocked_distinte_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso)

    if locked_distinte:
        try:
            existing_dist = load_primanota_day(store_code=str(store_code), data_iso=d_iso, categories=["Distinte"])
        except Exception:
            existing_dist = []
        for e in existing_dist:
            entries.append(
                {
                    "categoria": "Distinte",
                    "voce": str(e.get("voce") or ""),
                    "tipo": str(e.get("tipo") or "SI"),
                    "valore": float(e.get("valore") or 0.0),
                }
            )
    else:
        def _add_distinte(which: str, items: List[Dict[str, Any]]):
            for it in items or []:
                try:
                    taglio = float(it.get("taglio"))
                    qta = int(it.get("qta"))
                except Exception:
                    continue
                if taglio <= 0 or qta <= 0:
                    continue
                voce = f"{which}|TAGLIO={taglio}|QTA={qta}"
                entries.append(
                    {
                        "categoria": "Distinte",
                        "voce": voce,
                        "tipo": "SI",
                        "valore": float(taglio * qta),
                    }
                )

        _add_distinte("D1", distinte1)
        _add_distinte("D2", distinte2)

        if d1_monete != 0.0 or (request.form.get("d1_monete") or "").strip() != "":
            entries.append({"categoria": "Distinte", "voce": "D1|MONETE", "tipo": "SI", "valore": float(d1_monete)})
        if d2_monete != 0.0 or (request.form.get("d2_monete") or "").strip() != "":
            entries.append({"categoria": "Distinte", "voce": "D2|MONETE", "tipo": "SI", "valore": float(d2_monete)})

    # Ticket / Delivery / Coupon
    def _add_list(cat: str, items: List[Dict[str, Any]], tipo_map: Dict[str, str]):
        for it in items or []:
            voce = str(it.get("voce") or "").strip()
            if not voce:
                continue
            val = it.get("valore")
            try:
                val_f = float(val)
            except Exception:
                val_f = 0.0
            tipo = tipo_map.get(voce, str(it.get("tipo") or "SI"))
            entries.append({"categoria": cat, "voce": voce, "tipo": tipo, "valore": float(val_f)})

    _add_list("Ticket", tickets, ticket_tipo_map)
    _add_list("Delivery", deliveries, delivery_tipo_map)
    _add_list("Coupon", coupons, coupon_tipo_map)

    for section in custom_config:
        for field in section.get("fields") or []:
            if str(field.get("value_type") or "").strip().lower() == "calculated":
                continue
            field_key = str(field.get("field_key") or "").strip()
            if not field_key:
                continue
            form_name = _custom_field_form_name(str(section.get("section_key") or ""), field_key)
            raw = request.form.get(form_name)
            value_type = str(field.get("value_type") or "").strip().lower()
            if raw is None or str(raw).strip() == "":
                val_f = 0.0
            elif value_type == "int":
                try:
                    val_f = float(int(round(_parse_float(str(raw)))))
                except Exception:
                    val_f = 0.0
            else:
                try:
                    val_f = float(_parse_money_decimal_value(str(raw)))
                except Exception:
                    val_f = 0.0
            custom_field_vals[field_key] = val_f
            entries.append(
                {
                    "categoria": str(field.get("legacy_category") or section.get("label") or ""),
                    "voce": str(field.get("legacy_voce") or field.get("label") or ""),
                    "tipo": str(field.get("legacy_tipo") or "SI"),
                    "valore": val_f,
                }
            )

    for section in custom_option_lists:
        raw_json = request.form.get(_custom_list_form_name(str(section.get("section_key") or ""))) or "[]"
        try:
            rows = json.loads(raw_json)
        except Exception:
            rows = []
        if not isinstance(rows, list):
            rows = []
        for row in rows:
            voce = str((row or {}).get("voce") or "").strip()
            if not voce:
                continue
            try:
                val_f = float(_parse_money_decimal_value(str((row or {}).get("valore") or "0")))
            except Exception:
                val_f = 0.0
            entries.append(
                {
                    "categoria": str(section.get("legacy_category") or section.get("label") or ""),
                    "voce": voce,
                    "tipo": str((row or {}).get("tipo") or "SI"),
                    "valore": val_f,
                }
            )

    # Calcoli per scritture giornaliere report-ready
    giro_affari = float(chiusura_vals.get("vendite_lorde", 0.0)) - float(chiusura_vals.get("annullati", 0.0))
    totale_delivery_si = 0.0
    for e in entries:
        if str(e.get("categoria") or "") == "Delivery" and str(e.get("tipo") or "").strip().upper() == "SI":
            try:
                totale_delivery_si += float(e.get("valore") or 0.0)
            except Exception:
                pass
    try:
        scontrini_int = int(round(float(chiusura_vals.get("scontrini", 0.0) or 0.0)))
    except Exception:
        scontrini_int = 0

    try:
        spese_day_info = sum_spese_day(store_code=str(store_code), data_iso=d_iso)
    except Exception as ex_spese:
        current_app.logger.exception("Errore calcolo spese giorno per StoreHubDailySales")
        spese_day_info = {"net": 0.0}
        flash(f"Distinta salvata, ma errore lettura spese per riepilogo StoreHub: {ex_spese}", "warning")

    try:
        replace_primanota_day(
            store_code=str(store_code),
            data_iso=d_iso,
            entries=entries,
            categories=distinta_categories,
        )
        # Scrittura StoreHub-native, pensata per report futuri e tenant separati.
        try:
            upsert_daily_sales_from_distinta(
                store_code=str(store_code),
                data_iso=d_iso,
                chiusura_vals={**chiusura_vals, **custom_field_vals},
                entries=entries,
                expenses_net=float(spese_day_info.get("net") or 0.0),
            )
        except Exception as ex_daily:
            current_app.logger.exception("Errore scrittura StoreHubDailySales da Distinta")
            flash(f"Distinta salvata, ma errore scrittura riepilogo StoreHub: {ex_daily}", "warning")

        # Scrittura aggiuntiva su DatiDatabase (non deve bloccare la Distinta)
        try:
            upsert_datidatabase_from_distinta(
                store_code=str(store_code),
                data_iso=d_iso,
                giro_affari=giro_affari,
                totale_delivery=totale_delivery_si,
                scontrini=scontrini_int,
            )
        except Exception as ex2:
            current_app.logger.exception("Errore scrittura DatiDatabase da Distinta")
            flash(f"Distinta salvata, ma errore scrittura DatiDatabase: {ex2}", "warning")

        if new_photo_file:
            try:
                upsert_distinta_cassa_photo_file(store_code=str(store_code), data_iso=d_iso, foto_file=new_photo_file)
            except Exception as ex3:
                current_app.logger.exception("Errore salvataggio foto Distinta cassa su SQL")
                try:
                    delete_azzeramento_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(new_photo_file),
                    )
                except Exception:
                    pass
                flash(f"Distinta salvata, ma errore salvataggio foto: {ex3}", "warning")

        flash("Distinta cassa salvata.", "success")
    except Exception as e:
        if new_photo_file:
            try:
                delete_azzeramento_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    filename=str(new_photo_file),
                )
            except Exception:
                pass
        current_app.logger.exception("Errore salvataggio Distinta cassa")
        flash(f"Errore salvataggio Distinta cassa: {e}", "danger")

    return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))




@rendiconto_bp.post("/distinta-cassa/delete")
def distinta_cassa_delete():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.distinta_cassa"))

    d_iso = _parse_date_iso(request.form.get("d") or "")
    delete_photo = str(request.form.get("delete_photo") or "").strip().lower() in {"1", "true", "on", "yes"}

    if (not _is_admin_role()) and (d_iso in _validated_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso)):
        flash("Giornata convalidata: solo l'amministratore puÃ² eliminare o modificare la distinta.", "warning")
        return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))

    # Cancellazione non consentita se la giornata Ã¨ inclusa nel periodo competenza di un versamento
    locked = d_iso in _locked_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso)
    if locked:
        flash("Giornata bloccata da un versamento: non puoi eliminare la distinta di cassa.", "warning")
        return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))

    try:
        custom_config = _distinta_custom_config()
        custom_option_lists = _distinta_custom_option_list_config()
        delete_primanota_day(
            store_code=str(store_code),
            data_iso=d_iso,
            categories=_distinta_entry_categories(custom_config, custom_option_lists),
        )

        # Cancellazione aggiuntiva su DatiDatabase (se presente)
        try:
            res = delete_datidatabase_day(store_code=str(store_code), data_iso=d_iso)
            if not res.get("ok"):
                raise Exception(res.get("error") or "Errore cancellazione DatiDatabase")
        except Exception as ex2:
            current_app.logger.exception("Errore cancellazione DatiDatabase da Distinta")
            flash(f"Distinta eliminata, ma errore cancellazione DatiDatabase: {ex2}", "warning")

        try:
            delete_daily_sales_day(store_code=str(store_code), data_iso=d_iso)
        except Exception as ex_daily:
            current_app.logger.exception("Errore cancellazione StoreHubDailySales da Distinta")
            flash(f"Distinta eliminata, ma errore cancellazione riepilogo StoreHub: {ex_daily}", "warning")

        if delete_photo:
            try:
                foto_file = get_distinta_cassa_photo_file(store_code=str(store_code), data_iso=d_iso)
            except Exception:
                foto_file = None
            if foto_file:
                try:
                    delete_azzeramento_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(foto_file),
                    )
                    delete_distinta_cassa_photo_assoc(store_code=str(store_code), data_iso=d_iso)
                except Exception as ex3:
                    current_app.logger.exception("Errore cancellazione foto Distinta cassa")
                    flash(f"Distinta eliminata, ma errore cancellazione foto: {ex3}", "warning")

        flash("Distinta cassa eliminata.", "success")
    except Exception as e:
        current_app.logger.exception("Errore eliminazione Distinta cassa")
        flash(f"Errore eliminazione Distinta cassa: {e}", "danger")

    return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))


@rendiconto_bp.post("/distinta-cassa/photo/delete")
def distinta_cassa_photo_delete():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.distinta_cassa"))

    d_iso = _parse_date_iso(request.form.get("d") or "")
    if (not _is_admin_role()) and (d_iso in _validated_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso)):
        flash("Giornata convalidata: solo l'amministratore puÃ² modificare la foto associata.", "warning")
        return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))

    try:
        foto_file = get_distinta_cassa_photo_file(store_code=str(store_code), data_iso=d_iso)
        if not foto_file:
            flash("Nessuna foto associata a questa giornata.", "warning")
            return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))

        try:
            delete_azzeramento_photo(
                sb_jwt=str(session.get("sb_token") or ""),
                user_id=str(session.get("uid") or ""),
                store_code=str(store_code),
                filename=str(foto_file),
            )
        finally:
            delete_distinta_cassa_photo_assoc(store_code=str(store_code), data_iso=d_iso)

        flash("Foto distinta eliminata.", "success")
    except Exception as e:
        current_app.logger.exception("Errore eliminazione foto Distinta cassa")
        flash(f"Errore eliminazione foto Distinta cassa: {e}", "danger")

    return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))


@rendiconto_bp.get("/distinta-cassa/photo/<store_code>/<path:filename>")
def distinta_cassa_photo_scoped(store_code: str, filename: str):
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    sc = (store_code or "").strip()
    if not sc or "/" in sc or "\\" in sc:
        abort(404)

    fname = (filename or "").strip()
    if "/" in fname or "\\" in fname:
        abort(404)

    try:
        content = download_azzeramento_photo(
            sb_jwt=str(session.get("sb_token") or ""),
            user_id=str(session.get("uid") or ""),
            store_code=sc,
            filename=fname,
        )
        return send_file(io.BytesIO(content), mimetype="image/jpeg", download_name=fname, as_attachment=False, max_age=300)
    except Exception as e:
        current_app.logger.warning("Foto distinta cassa non disponibile (scoped): %s", e)
        abort(404)


@rendiconto_bp.get("/distinta-cassa/photo/<path:filename>")
def distinta_cassa_photo(filename: str):
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    if not store_code:
        abort(404)

    fname = (filename or "").strip()
    if "/" in fname or "\\" in fname:
        abort(404)

    try:
        content = download_azzeramento_photo(
            sb_jwt=str(session.get("sb_token") or ""),
            user_id=str(session.get("uid") or ""),
            store_code=str(store_code),
            filename=fname,
        )
        return send_file(io.BytesIO(content), mimetype="image/jpeg", download_name=fname, as_attachment=False, max_age=300)
    except Exception as e:
        current_app.logger.warning("Foto distinta cassa non disponibile: %s", e)
        abort(404)


# -------------------------
# Dashboard Rendiconto (API)
# -------------------------


@rendiconto_bp.get("/api/dashboard/month")
def api_dashboard_month():
    """Ritorna mappa giorni del mese con {giro, diff} e summary per la dashboard, piÃ¹ lo stato versamenti."""
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    try:
        year = int(request.args.get("year") or 0)
        month = int(request.args.get("month") or 0)
        if year < 2000 or not (1 <= month <= 12):
            raise ValueError
    except Exception:
        today = _date.today()
        year, month = today.year, today.month

    try:
        prim_rows = load_primanota_month_agg(str(store_code), year=year, month=month)
    except Exception as e:
        current_app.logger.exception("Errore lettura DATIPRIMANOTA (month agg)")
        return jsonify(error=f"Errore lettura prima nota: {e}"), 500

    try:
        spese_by_day = sum_spese_month_by_day(store_code=str(store_code), year=year, month=month)
    except Exception:
        current_app.logger.exception("Errore lettura SPESE (month agg)")
        spese_by_day = {}

    try:
        photo_days = list_distinta_cassa_photo_days(store_code=str(store_code), year=year, month=month)
    except Exception:
        current_app.logger.exception("Errore lettura foto Distinta cassa (month agg)")
        photo_days = {}

    try:
        validated_days = list_convalida_days_month(store_code=str(store_code), year=year, month=month)
    except Exception:
        current_app.logger.exception("Errore lettura convalide Distinta cassa (month agg)")
        validated_days = {}

    try:
        from tenant_config_repository import current_tenant_key

        tenant_key = str(session.get("tenant_key") or current_tenant_key()).strip() or current_tenant_key()
    except Exception:
        tenant_key = str(session.get("tenant_key") or "").strip() or "default"

    start_day = _date(year, month, 1)
    end_day = _date(year + 1, 1, 1) - timedelta(days=1) if month == 12 else _date(year, month + 1, 1) - timedelta(days=1)
    try:
        daily_sales_map = list_daily_sales_range(
            store_code=str(store_code),
            start_day=start_day,
            end_day=end_day,
            tenant_key=tenant_key,
            include_legacy_fallback=True,
        )
    except Exception:
        current_app.logger.exception("Errore lettura StoreHubDailySales (dashboard month)")
        daily_sales_map = {}

    # Aggregazione per giorno
    by_day = {}
    for r in prim_rows or []:
        d_iso = str(r.get("date") or "").strip()
        if not d_iso:
            continue
        cat = str(r.get("categoria") or "").strip()
        voce = str(r.get("voce") or "").strip()
        tipo = str(r.get("tipo") or "SI").strip().upper()
        try:
            s = float(r.get("sum") or 0)
        except Exception:
            s = 0.0

        agg = by_day.setdefault(
            d_iso,
            {
                "vendite_lorde": 0.0,
                "annullati": 0.0,
                "pos": 0.0,
                "scontrini": 0.0,
                "distinte": 0.0,
                "ticket_si": 0.0,
                "delivery_si": 0.0,
                "coupon_si": 0.0,
            },
        )

        if cat == "Dati chiusura":
            key = _VOICE_TO_KEY.get(voce)
            if key in {"vendite_lorde", "annullati", "pos", "scontrini"}:
                agg[key] += s
        elif cat == "Distinte":
            agg["distinte"] += s
        elif cat == "Ticket":
            if tipo == "SI":
                agg["ticket_si"] += s
        elif cat == "Delivery":
            if tipo == "SI":
                agg["delivery_si"] += s
        elif cat == "Coupon":
            if tipo == "SI":
                agg["coupon_si"] += s

    days = {}
    all_days = (
        set(by_day.keys())
        | set((photo_days or {}).keys())
        | set((validated_days or {}).keys())
        | set((spese_by_day or {}).keys())
        | set((daily_sales_map or {}).keys())
    )
    for d_iso in sorted(all_days):
        a = by_day.get(d_iso) or {
            "vendite_lorde": 0.0,
            "annullati": 0.0,
            "pos": 0.0,
            "scontrini": 0.0,
            "distinte": 0.0,
            "ticket_si": 0.0,
            "delivery_si": 0.0,
            "coupon_si": 0.0,
        }
        giro = float(a.get("vendite_lorde", 0.0)) - float(a.get("annullati", 0.0))
        spnet = float((spese_by_day.get(d_iso) or {}).get("net") or 0.0)
        diff = (
            float(a.get("distinte", 0.0))
            + float(a.get("ticket_si", 0.0))
            + float(a.get("delivery_si", 0.0))
            + float(a.get("coupon_si", 0.0))
            + float(a.get("pos", 0.0))
            + spnet
            - giro
        )
        sales_row = daily_sales_map.get(d_iso) or {}
        if sales_row and str(sales_row.get("source_family") or "").strip().lower() == "storehub":
            giro = float(sales_row.get("gross_revenue") or giro)
            pos = float(sales_row.get("pos_amount") or a.get("pos", 0.0))
            distinte = float(sales_row.get("cash_deposits_total") or a.get("distinte", 0.0))
            ticket_si = float(sales_row.get("ticket_cash_effect") or a.get("ticket_si", 0.0))
            delivery_si = float(sales_row.get("delivery_cash_effect") or a.get("delivery_si", 0.0))
            coupon_si = float(sales_row.get("coupon_cash_effect") or a.get("coupon_si", 0.0))
            annullati = float(sales_row.get("cancelled_amount") or a.get("annullati", 0.0))
            scontrini = float(sales_row.get("receipts_count") or a.get("scontrini", 0.0))
            diff = distinte + ticket_si + delivery_si + coupon_si + pos + spnet - giro
        else:
            pos = float(a.get("pos", 0.0))
            distinte = float(a.get("distinte", 0.0))
            ticket_si = float(a.get("ticket_si", 0.0))
            delivery_si = float(a.get("delivery_si", 0.0))
            coupon_si = float(a.get("coupon_si", 0.0))
            annullati = float(a.get("annullati", 0.0))
            scontrini = float(a.get("scontrini", 0.0))
        foto_file = str((photo_days or {}).get(d_iso) or "").strip()
        days[d_iso] = {
            "giro": giro,
            "diff": diff,
            "pos": pos,
            "annullati": annullati,
            "scontrini": scontrini,
            "distinte": distinte,
            "ticket_si": ticket_si,
            "delivery_si": delivery_si,
            "coupon_si": coupon_si,
            "spese_net": spnet,
            "has_photo": bool(foto_file),
            "photo_url": url_for("rendiconto.distinta_cassa_photo_scoped", store_code=str(store_code), filename=foto_file) if foto_file else "",
            "validated": bool((validated_days or {}).get(d_iso)),
        }

    # Totali mese (riepilogo) ricostruiti dai valori giornalieri gia normalizzati.
    giro_mese = sum(float((row or {}).get("giro") or 0.0) for row in days.values())
    diff_mese = sum(float((row or {}).get("diff") or 0.0) for row in days.values())
    tot_pos = sum(float((row or {}).get("pos") or 0.0) for row in days.values())
    tot_distinte = sum(float((row or {}).get("distinte") or 0.0) for row in days.values())
    tot_annullati = sum(float((row or {}).get("annullati") or 0.0) for row in days.values())
    tot_scontrini = sum(float((row or {}).get("scontrini") or 0.0) for row in days.values())

    try:
        scontrini_mese = int(round(float(tot_scontrini)))
    except Exception:
        scontrini_mese = 0

    summary = {
        "giro": giro_mese,
        "scontrini": scontrini_mese,
        "pos": float(tot_pos),
        "distinte": float(tot_distinte),
        "annullati": float(tot_annullati),
        "diff": float(diff_mese),
    }

        # ---- Versamenti: stato annuale (Distinte anno vs Versato) ----
    versamenti_status = _build_versamenti_status_for_store(
        store_code=str(store_code),
        year=year,
    )

    return jsonify(
        {
            "year": year,
            "month": month,
            "days": days,
            "summary": summary,
            "versamenti_status": versamenti_status,
        }
    )


@rendiconto_bp.get("/api/dashboard/day")
def api_dashboard_day():
    """Ritorna dettaglio del giorno per il popup rendiconto in dashboard."""
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    d_iso = _parse_date_iso(request.args.get("date") or "")

    try:
        payload = _build_dashboard_day_snapshot(store_code=str(store_code), d_iso=d_iso)
    except Exception as e:
        current_app.logger.exception("Errore lettura dashboard day")
        return jsonify(error=f"Errore lettura prima nota: {e}"), 500

    payload["validated"] = bool(_validated_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso))
    return jsonify(payload)


@rendiconto_bp.get("/api/dashboard/validate-period/preview")
def api_dashboard_validate_period_preview():
    _ensure_session_keys()
    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401
    if not _can_validate_rendiconto_period():
        return jsonify(error="Operazione consentita solo a supervisor e admin."), 403

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    dal = (request.args.get("dal") or "").strip()
    al = (request.args.get("al") or "").strip()
    if not (_is_valid_iso_date(dal) and _is_valid_iso_date(al)):
        return jsonify(error="Periodo non valido."), 400
    if al < dal:
        dal, al = al, dal

    overlaps = list_convalide_overlapping(store_code=str(store_code), start_iso=dal, end_iso=al)
    if overlaps:
        return jsonify(error="Esiste giÃ  una convalida che si sovrappone al periodo selezionato.", overlaps=overlaps), 409

    days = []
    total_diff = 0.0
    for day_iso in _daterange_iso(dal, al):
        snap = _build_dashboard_day_snapshot(store_code=str(store_code), d_iso=day_iso)
        diff = float(snap.get("diff") or 0.0)
        total_diff += diff
        days.append({"date": day_iso, "diff": diff})

    return jsonify({"ok": True, "dal": dal, "al": al, "days": days, "total_diff": float(total_diff)})


@rendiconto_bp.post("/api/dashboard/validate-period/confirm")
def api_dashboard_validate_period_confirm():
    _ensure_session_keys()
    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401
    if not _can_validate_rendiconto_period():
        return jsonify(error="Operazione consentita solo a supervisor e admin."), 403

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    data = request.get_json(silent=True) or {}
    dal = str(data.get("dal") or "").strip()
    al = str(data.get("al") or "").strip()
    if not (_is_valid_iso_date(dal) and _is_valid_iso_date(al)):
        return jsonify(error="Periodo non valido."), 400
    if al < dal:
        dal, al = al, dal

    overlaps = list_convalide_overlapping(store_code=str(store_code), start_iso=dal, end_iso=al)
    if overlaps:
        return jsonify(error="Esiste giÃ  una convalida che si sovrappone al periodo selezionato.", overlaps=overlaps), 409

    rec = insert_convalida_periodo(
        store_code=str(store_code),
        dal_iso=dal,
        al_iso=al,
        total_diff=float(data.get("total_diff") or 0.0),
        created_by=str(session.get("uid") or ""),
        created_name=_session_display_name(),
        created_role=_session_role_l(),
    )
    return jsonify({"ok": True, "record": rec})


@rendiconto_bp.get("/api/dashboard/validate-period/export.xlsx")
def api_dashboard_validate_period_export_excel():
    _ensure_session_keys()
    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401
    if not _is_admin_role():
        return jsonify(error="Operazione consentita solo all'amministratore."), 403

    store_code = str(session.get("store_code") or "").strip()
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    rows = list_convalide_store(store_code=store_code)
    wb = Workbook()
    ws = wb.active
    ws.title = "Convalide"
    headers = ["Store", "Dal", "Al", "Diff. cassa totale", "Richiesto da", "Ruolo", "Creato il"]
    ws.append(headers)
    for cell in ws[1]:
        if Font:
            cell.font = Font(bold=True)
        if Alignment:
            cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([
            row.get("site") or "",
            row.get("dal_iso") or "",
            row.get("al_iso") or "",
            float(row.get("total_diff") or 0.0),
            row.get("created_name") or row.get("created_by") or "",
            row.get("created_role") or "",
            row.get("created_at") or "",
        ])
    for col in ws.columns:
        letter = col[0].column_letter
        width = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[letter].width = min(max(width, 12), 28)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name=f"convalide_distinte_{store_code}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@rendiconto_bp.post("/api/dashboard/validate-period/delete")
def api_dashboard_validate_period_delete():
    _ensure_session_keys()
    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401
    if not _is_admin_role():
        return jsonify(error="Operazione consentita solo all'amministratore."), 403

    store_code = str(session.get("store_code") or "").strip()
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    raw_id = (request.form.get("convalida_id") or "").strip()
    try:
        convalida_id = int(raw_id)
    except Exception:
        flash("Convalida non valida.", "warning")
        return redirect(url_for("dashboard"))

    try:
        ok = delete_convalida_by_id(convalida_id=convalida_id, store_code=store_code)
    except Exception as e:
        current_app.logger.exception("Errore eliminazione convalida dashboard")
        flash(f"Errore eliminazione convalida: {e}", "danger")
        return redirect(url_for("dashboard"))

    if ok:
        flash("Convalida eliminata.", "success")
    else:
        flash("Convalida non trovata per lo store corrente.", "warning")
    return redirect(url_for("dashboard"))
# ------------------------------------------------------------
# RIEPILOGO (multi-store, mensile) - sezione Rendiconto
# ------------------------------------------------------------

def _parse_month_yyyy_mm(value: str):
    """Parse YYYY-MM -> (year, month, start_date, end_date)."""
    raw = (value or "").strip()
    y = 0
    m = 0
    if raw and "-" in raw:
        try:
            p = raw.split("-", 1)
            y = int(p[0])
            m = int(p[1])
        except Exception:
            y = 0
            m = 0

    if y < 2000 or not (1 <= m <= 12):
        today = _date.today()
        y, m = today.year, today.month

    start = _date(y, m, 1)
    # last day of month
    if m == 12:
        end = _date(y + 1, 1, 1) - timedelta(days=1)
    else:
        end = _date(y, m + 1, 1) - timedelta(days=1)
    return y, m, start, end


def _parse_month_yyyy_mm_strict_iso(value: str):
    """Parse strict YYYY-MM -> (month_norm, year, month, start_iso, end_iso).

    Se value Ã¨ mancante o non valido, ritorna ("", 0, 0, "", "").
    """
    raw = (value or "").strip()
    if not raw:
        return "", 0, 0, "", ""

    try:
        y_s, m_s = raw.split("-", 1)
        y = int(y_s)
        m = int(m_s)
    except Exception:
        return "", 0, 0, "", ""

    if y < 2000 or y > 2100 or not (1 <= m <= 12):
        return "", 0, 0, "", ""

    start = _date(y, m, 1)
    if m == 12:
        end = _date(y + 1, 1, 1) - timedelta(days=1)
    else:
        end = _date(y, m + 1, 1) - timedelta(days=1)

    month_norm = f"{y:04d}-{m:02d}"
    return month_norm, y, m, start.isoformat(), end.isoformat()


def _available_stores_for_user(user_id: str | None):
    """Lista store visibili all'utente (stessa logica Magazzino)."""
    try:
        role = str(session.get("role") or "").lower()

        # refresh role ogni ~10 minuti (o se admin)
        try:
            if user_id:
                now = int(time.time())
                cached_for = session.get("role_verified_for")
                cached_at = int(session.get("role_verified_at") or 0)
                need_check = (role == "admin") or (cached_for != user_id) or ((now - cached_at) > 600)
                if need_check:
                    from db_integration import get_profile_role_by_id

                    srv_role = get_profile_role_by_id(user_id)
                    if srv_role:
                        role = str(srv_role).lower()
                        session["role"] = srv_role
                        session["role_verified_for"] = user_id
                        session["role_verified_at"] = now
        except Exception:
            pass

        if role == "admin":
            stores = get_warehouse_stores() or []
        else:
            assigned = []
            if user_id:
                try:
                    assigned = get_user_warehouse_stores(str(user_id))
                except Exception:
                    assigned = []

            all_stores = get_warehouse_stores() or []
            allowed_codes = {str(row.get("store_code")) for row in assigned if row.get("store_code")}

            if allowed_codes:
                stores = [s for s in all_stores if str((s or {}).get("code")) in allowed_codes]
            else:
                cur_code = session.get("store_code")
                if cur_code:
                    stores = [s for s in all_stores if str((s or {}).get("code")) == str(cur_code)]
                else:
                    stores = []
    except Exception:
        stores = []

    try:
        stores = sorted(
            stores or [],
            key=lambda s: (
                str((s or {}).get("name") or "").strip().lower(),
                str((s or {}).get("code") or "").strip().lower(),
            ),
        )
    except Exception:
        pass

    return stores


_CHIUSURA_VOICE_TO_KEY = {
    "VENDITE LORDE": "vendite_lorde",
    "ANNULLATI": "annullati",
    "POS": "pos",
    "SCONTRINI": "scontrini",
}


@rendiconto_bp.get("/riepilogo")
def riepilogo():
    _ensure_session_keys()
    return render_template("rendiconto_riepilogo.html")


def _empty_rendiconto_detail_map() -> Dict[str, Dict[str, Dict[str, Any]]]:
    return {"Ticket": {}, "Delivery": {}, "Coupon": {}}


def _add_rendiconto_detail_value(detail_map: Dict[str, Dict[str, Dict[str, Any]]], *, categoria: str, voce: str, tipo: str, valore: float) -> None:
    cat = str(categoria or "").strip()
    if cat not in detail_map:
        return
    voice = str(voce or "").strip() or cat
    typ = str(tipo or "SI").strip().upper() or "SI"
    if typ not in {"SI", "NO"}:
        typ = "SI"
    key = f"{voice}||{typ}"
    bucket = detail_map[cat].setdefault(key, {"voce": voice, "tipo": typ, "valore": 0.0})
    bucket["valore"] = float(bucket.get("valore") or 0.0) + float(valore or 0.0)


def _rendiconto_detail_label(categoria: str, voce: str, tipo: str) -> str:
    cat = str(categoria or "").strip()
    voice = str(voce or "").strip() or cat
    typ = str(tipo or "SI").strip().upper()
    if cat == "Delivery":
        suffix = "CONTANTI" if typ == "NO" else "ONLINE"
        return f"{voice} {suffix}".strip()
    if typ == "NO":
        return f"{voice} NO".strip()
    return voice


def _rendiconto_detail_list(detail_map: Dict[str, Dict[str, Dict[str, Any]]], categoria: str) -> List[Dict[str, Any]]:
    cat = str(categoria or "").strip()
    rows: List[Dict[str, Any]] = []
    for item in (detail_map.get(cat) or {}).values():
        val = float(item.get("valore") or 0.0)
        if abs(val) < 0.000001:
            continue
        voce = str(item.get("voce") or "").strip()
        tipo = str(item.get("tipo") or "SI").strip().upper()
        rows.append(
            {
                "voce": voce,
                "tipo": tipo,
                "label": _rendiconto_detail_label(cat, voce, tipo),
                "valore": val,
            }
        )
    return sorted(rows, key=lambda r: (str(r.get("label") or "").lower(), str(r.get("tipo") or "")))


def _append_rendiconto_detail_sheet(wb, rows: List[Dict[str, Any]], *, sheet_name: str = "Dettaglio") -> None:
    ws_detail = wb.create_sheet(sheet_name)
    headers = ["Store", "Categoria", "Dettaglio", "Tipo", "Importo"]
    ws_detail.append(headers)
    if Font and Alignment:
        for cell in ws_detail[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    for rr in rows or []:
        store_label = str(rr.get("store_name") or "").strip() or str(rr.get("store_code") or "").strip()
        for cat, key in (("Ticket", "ticket_detail"), ("Delivery", "delivery_detail"), ("Coupon", "coupon_detail")):
            for item in rr.get(key) or []:
                ws_detail.append(
                    [
                        store_label,
                        cat,
                        str(item.get("label") or item.get("voce") or ""),
                        str(item.get("tipo") or ""),
                        float(item.get("valore") or 0.0),
                    ]
                )

    eur_fmt = "#,##0.00"
    for row_idx in range(2, ws_detail.max_row + 1):
        ws_detail.cell(row=row_idx, column=5).number_format = eur_fmt
        ws_detail.cell(row=row_idx, column=5).alignment = Alignment(horizontal="right")

    widths = [32, 14, 34, 10, 14]
    for i, w in enumerate(widths, start=1):
        ws_detail.column_dimensions[chr(64 + i)].width = w
    ws_detail.freeze_panes = "A2"


@rendiconto_bp.get("/ricerca")
def ricerca():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    user_id = session.get("uid")
    stores = _available_stores_for_user(str(user_id) if user_id else None)
    return render_template(
        "rendiconto_ricerca.html",
        stores_count=len(stores or []),
    )


@rendiconto_bp.route("/gestione-delivery", methods=["GET", "POST"])
def gestione_delivery():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    store_name = session.get("store_name")

    def parse_date(value: str) -> _date:
        value = (value or "").strip()
        if not value:
            return _date.today()
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except Exception:
            try:
                year_s, week_s = value.split("-W", 1)
                return _date.fromisocalendar(int(year_s), int(week_s), 1)
            except Exception:
                return _date.today()

    # Week selector (always normalized to Monday)
    if request.method == "POST":
        week_start = parse_date(str(request.form.get("week_start") or ""))
    else:
        week_start = parse_date(str(request.args.get("week_start") or ""))
    week_start = week_monday(week_start)

    def parse_money_decimal(v: str) -> Decimal:
        x = Decimal(str(_parse_float(v or "0"))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return x

    def parse_int(v: str) -> int:
        s = (v or "").strip()
        try:
            return int(float(s.replace(",", ".")))
        except Exception:
            return 0

    def parse_rating(v: str) -> Optional[Decimal]:
        s = (v or "").strip()
        if not s:
            return None
        try:
            return Decimal(str(_parse_float(s))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        except Exception:
            return None

    if request.method == "POST" and store_code:
        save_scope = str(request.form.get("save_scope") or "all").strip().lower()
        providers_config = {
            str(p.get("provider_key") or "").strip().lower(): p
            for p in list_delivery_providers(active_only=True)
            if str(p.get("provider_key") or "").strip()
        }
        allowed_scopes = {"all"} | set(providers_config.keys())
        if save_scope not in allowed_scopes:
            save_scope = "all"

        def save_provider(provider_key: str) -> None:
            cfg = providers_config[provider_key]
            prefix = provider_key
            upsert_delivery_weekly(
                str(store_code),
                str(cfg["platform"]),
                week_start,
                payment_online=parse_money_decimal(str(request.form.get(f"{prefix}_payment_online") or "0")),
                payment_cash=parse_money_decimal(str(request.form.get(f"{prefix}_payment_cash") or "0")),
                orders=parse_int(str(request.form.get(f"{prefix}_orders") or "0")),
                cancelled_orders=parse_int(str(request.form.get(f"{prefix}_cancelled_orders") or "0")),
                complaints_received=parse_int(str(request.form.get(f"{prefix}_complaints_received") or "0")),
                refund_value=parse_money_decimal(str(request.form.get(f"{prefix}_refund_value") or "0")),
                complaints_contested=parse_int(str(request.form.get(f"{prefix}_complaints_contested") or "0")),
                appeals_accepted=parse_int(str(request.form.get(f"{prefix}_appeals_accepted") or "0")),
                refunds_cancelled_value=parse_money_decimal(str(request.form.get(f"{prefix}_refunds_cancelled_value") or "0")),
                opening_pct=parse_rating(str(request.form.get(f"{prefix}_opening_pct") or "")),
                rating_value=parse_rating(str(request.form.get(f"{prefix}_rating") or "")),
                rating_unit=str(cfg["rating_unit"]),
            )

        try:
            if save_scope == "all":
                providers_to_save = list(providers_config.keys())
            else:
                providers_to_save = [save_scope]

            for provider_key in providers_to_save:
                save_provider(provider_key)

            if save_scope == "all":
                flash("Dati delivery salvati.", "success")
            else:
                flash(f"Dati {providers_config[save_scope]['label']} salvati.", "success")
        except Exception as e:
            if save_scope == "all":
                flash(f"Errore salvataggio delivery: {e}", "danger")
            else:
                flash(f"Errore salvataggio {providers_config[save_scope]['label']}: {e}", "danger")

        return redirect(url_for("rendiconto.gestione_delivery", week_start=week_start.isoformat()))

    # Load current week rows + prev week rating (for delta)
    providers = list_delivery_providers(active_only=True)
    provider_rows: list[dict[str, Any]] = []
    if store_code:
        try:
            for provider in providers:
                platform = str(provider.get("platform") or "").strip().upper()
                item = dict(provider)
                item["data_row"] = get_delivery_weekly(str(store_code), platform, week_start)
                item["prev_rating_raw"] = get_delivery_prev_rating(str(store_code), platform, week_start)
                provider_rows.append(item)
        except Exception as e:
            flash(
                "Tabella delivery non disponibile o errore lettura. "
                "Crea la tabella con lo script SQL incluso nella patch e riprova. "
                f"Dettaglio: {e}",
                "warning",
            )
    else:
        provider_rows = [dict(p, data_row=None, prev_rating_raw=None) for p in providers]

    def dec_to_str(v: Any) -> str:
        if v is None:
            return ""
        try:
            return f"{float(v):.2f}"
        except Exception:
            return str(v)

    def row_to_dict(row: Any) -> Dict[str, Any]:
        if not row:
            return {
                "payment_online": "0.00",
                "payment_cash": "0.00",
                "orders": 0,
                "cancelled_orders": 0,
                "complaints_received": 0,
                "refund_value": "0.00",
                "complaints_contested": 0,
                "appeals_accepted": 0,
                "refunds_cancelled_value": "0.00",
                "opening_pct": "",
                "rating": "",
            }
        return {
            "payment_online": dec_to_str(row.payment_online),
            "payment_cash": dec_to_str(row.payment_cash),
            "orders": int(row.orders or 0),
            "cancelled_orders": int(getattr(row, "cancelled_orders", 0) or 0),
            "complaints_received": int(row.complaints_received or 0),
            "refund_value": dec_to_str(row.refund_value),
            "complaints_contested": int(row.complaints_contested or 0),
            "appeals_accepted": int(row.appeals_accepted or 0),
            "refunds_cancelled_value": dec_to_str(row.refunds_cancelled_value),
            "opening_pct": dec_to_str(getattr(row, "opening_pct", None)) if getattr(row, "opening_pct", None) is not None else "",
            "rating": dec_to_str(row.rating_value) if getattr(row, "rating_value", None) is not None else "",
        }

    return render_template(
        "rendiconto_gestione_delivery.html",
        store_code=store_code,
        store_name=store_name,
        week_start=week_start.isoformat(),
        week_value=f"{week_start.isocalendar().year:04d}-W{week_start.isocalendar().week:02d}",
        providers=[
            dict(
                p,
                data=row_to_dict(p.get("data_row")),
                prev_rating=dec_to_str(p.get("prev_rating_raw")) if p.get("prev_rating_raw") is not None else "",
            )
            for p in provider_rows
        ],
    )


@rendiconto_bp.get("/api/delivery-distinta-totals")
def api_delivery_distinta_totals():
    _ensure_session_keys()
    if not session.get("uid"):
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify({"ok": False, "error": "Seleziona uno store."}), 400

    week_start_raw = str(request.args.get("week_start") or "").strip()
    if not _is_valid_iso_date(week_start_raw):
        return jsonify({"ok": False, "error": "Param 'week_start' non valido (YYYY-MM-DD)."}), 400

    try:
        d = datetime.strptime(week_start_raw, "%Y-%m-%d").date()
    except Exception:
        return jsonify({"ok": False, "error": "Param 'week_start' non valido (YYYY-MM-DD)."}), 400

    week_start = week_monday(d)
    week_end = week_start + timedelta(days=6)

    try:
        voce_sums = sum_delivery_voce_range(
            store_code=str(store_code),
            start_iso=week_start.isoformat(),
            end_iso=week_end.isoformat(),
        )
    except Exception as e:
        return jsonify({"ok": False, "error": f"Errore lettura distinte: {e}"}), 500

    def norm_provider(s: str) -> str:
        return " ".join(str(s or "").strip().upper().split())

    provider_aliases: Dict[str, str] = {}
    try:
        for provider in list_delivery_providers(active_only=True):
            platform = norm_provider(provider.get("platform") or "")
            if not platform:
                continue
            provider_aliases[platform] = platform
            label = norm_provider(provider.get("label") or "")
            if label:
                provider_aliases[label] = platform
            key = norm_provider(str(provider.get("provider_key") or "").replace("_", " "))
            if key:
                provider_aliases[key] = platform
    except Exception:
        provider_aliases = {}

    providers: Dict[str, Dict[str, float]] = {}
    for voce, val in (voce_sums or {}).items():
        v = norm_provider(voce)
        if not v:
            continue

        is_cash = v.endswith(" CONTANTI")
        base = v[:-len(" CONTANTI")].strip() if is_cash else v
        if not base:
            continue
        base = provider_aliases.get(base, base)

        dct = providers.setdefault(base, {"online": 0.0, "cash": 0.0})
        try:
            f = float(val or 0.0)
        except Exception:
            f = 0.0
        if is_cash:
            dct["cash"] += f
        else:
            dct["online"] += f

    for _k, dct in providers.items():
        dct["online"] = float(Decimal(str(dct.get("online") or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
        dct["cash"] = float(Decimal(str(dct.get("cash") or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

    return jsonify(
        {
            "ok": True,
            "week_start": week_start.isoformat(),
            "week_end": week_end.isoformat(),
            "providers": providers,
        }
    )


@rendiconto_bp.get("/api/ricerca")
def api_ricerca():
    _ensure_session_keys()
    if not session.get("uid"):
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    kind = str(request.args.get("kind") or "VERSAMENTI").strip().upper()
    start_iso = str(request.args.get("start") or "").strip()
    end_iso = str(request.args.get("end") or "").strip()

    try:
        start_d = datetime.strptime(start_iso, "%Y-%m-%d").date()
        end_d = datetime.strptime(end_iso, "%Y-%m-%d").date()
    except Exception:
        return jsonify({"ok": False, "error": "Intervallo date non valido. Usa YYYY-MM-DD."}), 400

    if end_d < start_d:
        return jsonify({"ok": False, "error": "Intervallo date non valido: 'end' < 'start'."}), 400

    user_id = session.get("uid")
    stores = _available_stores_for_user(str(user_id) if user_id else None) or []
    store_map = {str((s or {}).get("code") or "").strip(): str((s or {}).get("name") or "").strip() for s in stores or []}
    store_codes = [c for c in store_map.keys() if c]

    def fmt_money(v: float) -> str:
        try:
            x = float(v or 0.0)
        except Exception:
            x = 0.0
        s = f"{x:,.2f}"
        # python usa ',' migliaia e '.' decimale; invertiamo per IT
        s = s.replace(",", "X").replace(".", ",").replace("X", ".")
        return s

    warnings: List[str] = []
    total = 0.0
    rows_out: List[Dict[str, Any]] = []
    columns: List[Dict[str, Any]] = []

    try:
        if kind == "SPESE":
            res = search_spese_range_multi(store_codes, start=start_d, end=end_d) if store_codes else {"rows": [], "total": 0.0, "warnings": []}
            warnings = list(res.get("warnings") or [])
            total = float(res.get("total") or 0.0)

            columns = [
                {"key": "store", "label": "Store"},
                {"key": "data", "label": "Data"},
                {"key": "tipo", "label": "Tipo"},
                {"key": "fornitore", "label": "Fornitore"},
                {"key": "documento", "label": "Documento"},
                {"key": "importo", "label": "Importo"},
                {"key": "foto", "label": "Foto", "type": "photo"},
            ]

            for r in (res.get("rows") or []):
                site = str((r or {}).get("site") or "").strip()
                name = store_map.get(site) or ""
                store_disp = site + ((" - " + name) if name else "")
                foto_file = (r or {}).get("foto_file")
                foto_url = url_for("rendiconto.spese_photo_scoped", store_code=site, filename=foto_file) if foto_file else ""
                rows_out.append(
                    {
                        "store": store_disp,
                        "data": (r or {}).get("data") or "",
                        "tipo": (r or {}).get("tipo") or "",
                        "fornitore": (r or {}).get("fornitore") or "",
                        "documento": (r or {}).get("documento") or "",
                        "importo": fmt_money((r or {}).get("importo") or 0.0),
                        "foto_url": foto_url,
                        "foto_file": str(foto_file or "") if foto_file else "",
                    }
                )

        else:
            # default: VERSAMENTI
            res = search_versamenti_range_multi(store_codes, start=start_d, end=end_d) if store_codes else {"rows": [], "total": 0.0, "warnings": []}
            warnings = list(res.get("warnings") or [])
            total = float(res.get("total") or 0.0)

            columns = [
                {"key": "store", "label": "Store"},
                {"key": "data_versamento", "label": "Data versamento"},
                {"key": "dal", "label": "Dal"},
                {"key": "al", "label": "Al"},
                {"key": "nome", "label": "Nome"},
                {"key": "tipo", "label": "Tipo"},
                {"key": "tessera", "label": "Tessera"},
                {"key": "riferimento", "label": "Riferimento"},
                {"key": "valore", "label": "Valore"},
                {"key": "foto", "label": "Foto", "type": "photo"},
            ]

            for r in (res.get("rows") or []):
                site = str((r or {}).get("site") or "").strip()
                name = store_map.get(site) or ""
                store_disp = site + ((" - " + name) if name else "")
                foto_file = (r or {}).get("foto_file")
                foto_url = url_for("rendiconto.versamenti_photo_scoped", store_code=site, filename=foto_file) if foto_file else ""
                rows_out.append(
                    {
                        "store": store_disp,
                        "data_versamento": (r or {}).get("data_versamento") or "",
                        "dal": (r or {}).get("dal") or "",
                        "al": (r or {}).get("al") or "",
                        "nome": (r or {}).get("nome") or "",
                        "tipo": (r or {}).get("tipo") or "",
                        "tessera": (r or {}).get("tessera") or "",
                        "riferimento": (r or {}).get("riferimento") or "",
                        "valore": fmt_money((r or {}).get("valore") or 0.0),
                        "foto_url": foto_url,
                        "foto_file": str(foto_file or "") if foto_file else "",
                    }
                )
    except Exception as ex:
        current_app.logger.exception("Errore api_ricerca")
        return jsonify({"ok": False, "error": str(ex)}), 500

    return jsonify(
        {
            "ok": True,
            "kind": kind,
            "start": start_iso,
            "end": end_iso,
            "columns": columns,
            "rows": rows_out,
            "total": total,
            "total_display": fmt_money(total),
            "warnings": warnings,
        }
    )


@rendiconto_bp.post("/api/ricerca/export.xlsx")
def api_ricerca_export_xlsx():
    _ensure_session_keys()
    if not session.get("uid"):
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    if Workbook is None:
        return jsonify({"ok": False, "error": "openpyxl non disponibile"}), 500

    data = request.get_json(silent=True) or {}
    kind = str(data.get("kind") or "RICERCA").strip().lower()
    start_iso = str(data.get("start") or "").strip()
    end_iso = str(data.get("end") or "").strip()
    columns = data.get("columns") or []
    rows = data.get("rows") or []

    # Normalizza colonne
    cols = []
    for c in columns:
        k = str((c or {}).get("key") or "").strip()
        lbl = str((c or {}).get("label") or k).strip()
        typ = str((c or {}).get("type") or "").strip().lower()
        if k:
            cols.append({"key": k, "label": lbl, "type": typ})

    wb = Workbook()
    ws = wb.active
    ws.title = "Ricerca"

    # Header
    header = []
    for c in cols:
        if c.get("type") == "photo":
            header.append(c.get("label") + " URL")
            header.append(c.get("label") + " FILE")
        else:
            header.append(c.get("label"))
    ws.append(header)

    if Font and Alignment:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # Rows
    for r in rows:
        out = []
        for c in cols:
            k = c.get("key")
            if c.get("type") == "photo":
                out.append(str((r or {}).get(k + "_url") or ""))
                out.append(str((r or {}).get(k + "_file") or ""))
            else:
                out.append(str((r or {}).get(k) or ""))
        ws.append(out)

    # Autofit (best-effort)
    try:
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[chr(64 + col_idx)].width = min(max(10, max_len + 2), 60)
    except Exception:
        pass

    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = f"rendiconto_ricerca_{kind}_{start_iso}_{end_iso}.xlsx".replace("__", "_")

    return send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@rendiconto_bp.get("/api/riepilogo/mensile")
def api_riepilogo_mensile():
    """Riepilogo mensile multi-store (Rendiconto)."""
    _ensure_session_keys()

    user_id = session.get("uid")
    if not user_id:
        return jsonify({"ok": False, "error": "not_logged_in"}), 401

    month = (request.args.get("month") or "").strip()  # YYYY-MM
    y, m, start_d, end_d = _parse_month_yyyy_mm(month)
    month_norm = f"{y:04d}-{m:02d}"

    try:
        stores = _available_stores_for_user(str(user_id)) or []
    except Exception:
        stores = []


    rows = []
    warnings_all = []

    prim_map = {}
    if get_backend() == "sqlserver":
        try:
            store_codes = [str((s or {}).get("code") or "").strip() for s in (stores or []) if (s or {}).get("code")]
            store_codes = [c for c in store_codes if c]
            prim_map = load_primanota_month_agg_totals_multi(store_codes, year=y, month=m) if store_codes else {}
        except Exception as ex:
            prim_map = {}
            warnings_all.append(f"Errore lettura PrimaNota multi-store: {ex}")

    spese_map = {}
    vers_map = {}
    ipratico_cash_map = {}
    if get_backend() == "sqlserver":
        try:
            store_codes = [str((s or {}).get("code") or "").strip() for s in (stores or []) if (s or {}).get("code")]
            store_codes = [c for c in store_codes if c]
            spese_map = sum_spese_month_total_net_multi(store_codes, year=y, month=m) if store_codes else {}
        except Exception as ex:
            spese_map = {}
            warnings_all.append(f"Errore lettura Spese multi-store: {ex}")
        try:
            store_codes = [str((s or {}).get("code") or "").strip() for s in (stores or []) if (s or {}).get("code")]
            store_codes = [c for c in store_codes if c]
            vers_map = sum_versamenti_month_total_multi(store_codes, year=y, month=m) if store_codes else {}
        except Exception as ex:
            vers_map = {}
            warnings_all.append(f"Errore lettura Versamenti multi-store: {ex}")
        try:
            store_codes = [str((s or {}).get("code") or "").strip() for s in (stores or []) if (s or {}).get("code")]
            store_codes = [c for c in store_codes if c]
            ipratico_cash_map = sum_distinta_cassa_ipratico_contanti_month_multi(store_codes, year=y, month=m) if store_codes else {}
        except Exception as ex:
            ipratico_cash_map = {}
            warnings_all.append(f"Errore lettura Contanti iPratico multi-store: {ex}")



    for s in stores or []:

        code = (s or {}).get("code")
        name = (s or {}).get("name")
        if not code:
            continue

        row = {
            "store_code": str(code),
            "store_name": str(name or ""),
            "giro_affari": 0.0,
            "diff_cassa": 0.0,
            "distinte": 0.0,
            "contanti_ipratico": 0.0,
            "giorni_non_versati": 0,
            "pos": 0.0,
            "scontrini": 0.0,
            "spese": 0.0,
            "versamenti": 0.0,
            "ticket_si": 0.0,
            "ticket_no": 0.0,
            "delivery_si": 0.0,
            "delivery_no": 0.0,
            "coupon_si": 0.0,
            "coupon_no": 0.0,
            "ticket_detail": [],
            "delivery_detail": [],
            "coupon_detail": [],
            "warnings": [],
        }

        # --- Prima nota (aggregati mese) ---
        try:
            if prim_map:
                prim_rows = prim_map.get(str(code), [])
            else:
                prim_rows = load_primanota_month_agg_totals(str(code), year=y, month=m)

            tot_vendite_lorde = 0.0
            tot_annullati = 0.0
            tot_pos = 0.0
            tot_scontrini = 0.0
            tot_distinte = 0.0
            t_si = 0.0
            t_no = 0.0
            d_si = 0.0
            d_no = 0.0
            c_si = 0.0
            c_no = 0.0
            detail_map = _empty_rendiconto_detail_map()

            for r in prim_rows or []:
                cat = str(r.get("categoria") or "").strip()
                voce = str(r.get("voce") or "").strip()
                tipo = str(r.get("tipo") or "SI").strip().upper()
                try:
                    s_val = float(r.get("sum") or 0.0)
                except Exception:
                    s_val = 0.0

                if cat == "Dati chiusura":
                    k = _CHIUSURA_VOICE_TO_KEY.get(voce.upper())
                    if k == "vendite_lorde":
                        tot_vendite_lorde += s_val
                    elif k == "annullati":
                        tot_annullati += s_val
                    elif k == "pos":
                        tot_pos += s_val
                    elif k == "scontrini":
                        tot_scontrini += s_val
                elif cat == "Distinte":
                    tot_distinte += s_val
                elif cat == "Ticket":
                    _add_rendiconto_detail_value(detail_map, categoria=cat, voce=voce, tipo=tipo, valore=s_val)
                    if tipo == "SI":
                        t_si += s_val
                    else:
                        t_no += s_val
                elif cat == "Delivery":
                    _add_rendiconto_detail_value(detail_map, categoria=cat, voce=voce, tipo=tipo, valore=s_val)
                    if tipo == "SI":
                        d_si += s_val
                    else:
                        d_no += s_val
                elif cat == "Coupon":
                    _add_rendiconto_detail_value(detail_map, categoria=cat, voce=voce, tipo=tipo, valore=s_val)
                    if tipo == "SI":
                        c_si += s_val
                    else:
                        c_no += s_val

            row["giro_affari"] = float(tot_vendite_lorde) - float(tot_annullati)
            row["pos"] = float(tot_pos)
            row["scontrini"] = float(tot_scontrini)
            row["distinte"] = float(tot_distinte)
            row["ticket_si"] = float(t_si)
            row["ticket_no"] = float(t_no)
            row["delivery_si"] = float(d_si)
            row["delivery_no"] = float(d_no)
            row["coupon_si"] = float(c_si)
            row["coupon_no"] = float(c_no)
            row["ticket_detail"] = _rendiconto_detail_list(detail_map, "Ticket")
            row["delivery_detail"] = _rendiconto_detail_list(detail_map, "Delivery")
            row["coupon_detail"] = _rendiconto_detail_list(detail_map, "Coupon")
        except Exception as e:
            current_app.logger.exception("Errore riepilogo rendiconto (primanota) store %s", code)
            warnings_all.append(f"[{code}] Prima nota: {e}")
            row["warnings"].append("Prima nota non disponibile")

        # --- Spese (totale mese) ---
        try:
            row["spese"] = float(spese_map.get(str(code), 0.0)) if spese_map else float(sum_spese_month_total_net(store_code=str(code), year=y, month=m))
        except Exception as e:
            current_app.logger.exception("Errore riepilogo rendiconto (spese) store %s", code)
            warnings_all.append(f"[{code}] Spese: {e}")
            row["warnings"].append("Spese non disponibili")

        # --- Versamenti (totale mese) ---
        try:
            row["versamenti"] = float(vers_map.get(str(code), 0.0)) if vers_map else float(sum_versamenti_month_total(store_code=str(code), year=y, month=m))
        except Exception as e:
            current_app.logger.exception("Errore riepilogo rendiconto (versamenti) store %s", code)
            warnings_all.append(f"[{code}] Versamenti: {e}")
            row["warnings"].append("Versamenti non disponibili")

        # --- Contanti iPratico (totale mese) ---
        try:
            row["contanti_ipratico"] = float(ipratico_cash_map.get(str(code), 0.0))
        except Exception as e:
            current_app.logger.exception("Errore riepilogo rendiconto (contanti iPratico) store %s", code)
            warnings_all.append(f"[{code}] Contanti iPratico: {e}")
            row["warnings"].append("Contanti iPratico non disponibili")

        # --- Giorni non versati (stessa logica della dashboard) ---
        try:
            status = _build_versamenti_status_for_store(
                store_code=str(code),
                year=y,
            )
            row["giorni_non_versati"] = int(status.get("giorni_non_versati") or 0)
        except Exception as e:
            current_app.logger.exception("Errore riepilogo rendiconto (giorni non versati) store %s", code)
            warnings_all.append(f"[{code}] Giorni non versati: {e}")
            row["warnings"].append("Giorni non versati non disponibili")

        # --- Differenza di cassa (mese) ---
        # Coerente con il calcolo usato nella dashboard: distinte + ticket_si + delivery_si + coupon_si + pos + spese_net - giro
        try:
            row["diff_cassa"] = (
                float(row.get("distinte") or 0.0)
                + float(row.get("ticket_si") or 0.0)
                + float(row.get("delivery_si") or 0.0)
                + float(row.get("coupon_si") or 0.0)
                + float(row.get("pos") or 0.0)
                + float(row.get("spese") or 0.0)
                - float(row.get("giro_affari") or 0.0)
            )
        except Exception:
            row["diff_cassa"] = 0.0

        rows.append(row)

    # Dedup warnings
    uniq_w = []
    seen = set()
    for w in warnings_all:
        if w not in seen:
            uniq_w.append(w)
            seen.add(w)

    # Sort by store name then code
    try:
        rows = sorted(rows, key=lambda r: (str(r.get("store_name") or "").lower(), str(r.get("store_code") or "")))
    except Exception:
        pass

    return jsonify(
        {
            "ok": True,
            "month": month_norm,
            "period": {"start": start_d.isoformat(), "end": end_d.isoformat()},
            "rows": rows,
            "warnings": uniq_w[:200],
        }
    )


@rendiconto_bp.get("/api/riepilogo/mensile.xlsx")
def api_riepilogo_mensile_xlsx():
    """Esporta in Excel la tabella del riepilogo mensile multi-store."""
    _ensure_session_keys()

    if Workbook is None:
        abort(500, description="Dipendenza mancante: openpyxl")

    user_id = session.get("uid")
    if not user_id:
        abort(401)

    month = (request.args.get("month") or "").strip()  # YYYY-MM
    y, m, start_d, end_d = _parse_month_yyyy_mm(month)
    month_norm = f"{y:04d}-{m:02d}"

    try:
        stores = _available_stores_for_user(str(user_id)) or []
    except Exception:
        stores = []


    prim_map = {}
    spese_map = {}
    vers_map = {}
    ipratico_cash_map = {}
    if get_backend() == "sqlserver":
        try:
            store_codes = [str((s or {}).get("code") or "").strip() for s in (stores or []) if (s or {}).get("code")]
            store_codes = [c for c in store_codes if c]
            prim_map = load_primanota_month_agg_totals_multi(store_codes, year=y, month=m) if store_codes else {}
        except Exception:
            prim_map = {}
        try:
            store_codes = [str((s or {}).get("code") or "").strip() for s in (stores or []) if (s or {}).get("code")]
            store_codes = [c for c in store_codes if c]
            spese_map = sum_spese_month_total_net_multi(store_codes, year=y, month=m) if store_codes else {}
        except Exception:
            spese_map = {}
        try:
            store_codes = [str((s or {}).get("code") or "").strip() for s in (stores or []) if (s or {}).get("code")]
            store_codes = [c for c in store_codes if c]
            vers_map = sum_versamenti_month_total_multi(store_codes, year=y, month=m) if store_codes else {}
        except Exception:
            vers_map = {}
        try:
            store_codes = [str((s or {}).get("code") or "").strip() for s in (stores or []) if (s or {}).get("code")]
            store_codes = [c for c in store_codes if c]
            ipratico_cash_map = sum_distinta_cassa_ipratico_contanti_month_multi(store_codes, year=y, month=m) if store_codes else {}
        except Exception:
            ipratico_cash_map = {}

    # Costruiamo le stesse righe dell'API JSON (senza dipendenze dal frontend)
    rows = []
    for s in stores or []:
        code = (s or {}).get("code")
        name = (s or {}).get("name")
        if not code:
            continue

        r = {
            "store_code": str(code),
            "store_name": str(name or ""),
            "giro_affari": 0.0,
            "scontrini": 0.0,
            "distinte": 0.0,
            "contanti_ipratico": 0.0,
            "giorni_non_versati": 0,
            "pos": 0.0,
            "spese": 0.0,
            "versamenti": 0.0,
            "ticket_si": 0.0,
            "ticket_no": 0.0,
            "delivery_si": 0.0,
            "delivery_no": 0.0,
            "coupon_si": 0.0,
            "coupon_no": 0.0,
            "ticket_detail": [],
            "delivery_detail": [],
            "coupon_detail": [],
            "diff_cassa": 0.0,
        }

        # Prima nota (aggregati mese)
        try:
            if prim_map:
                prim_rows = prim_map.get(str(code), [])
            else:
                prim_rows = load_primanota_month_agg_totals(str(code), year=y, month=m)

            tot_vendite_lorde = 0.0
            tot_annullati = 0.0
            tot_pos = 0.0
            tot_scontrini = 0.0
            tot_distinte = 0.0
            t_si = 0.0
            t_no = 0.0
            d_si = 0.0
            d_no = 0.0
            c_si = 0.0
            c_no = 0.0
            detail_map = _empty_rendiconto_detail_map()

            for rr in prim_rows or []:
                cat = str(rr.get("categoria") or "").strip()
                voce = str(rr.get("voce") or "").strip()
                tipo = str(rr.get("tipo") or "SI").strip().upper()
                try:
                    s_val = float(rr.get("sum") or 0.0)
                except Exception:
                    s_val = 0.0

                if cat == "Dati chiusura":
                    k = _CHIUSURA_VOICE_TO_KEY.get(voce.upper())
                    if k == "vendite_lorde":
                        tot_vendite_lorde += s_val
                    elif k == "annullati":
                        tot_annullati += s_val
                    elif k == "pos":
                        tot_pos += s_val
                    elif k == "scontrini":
                        tot_scontrini += s_val
                elif cat == "Distinte":
                    tot_distinte += s_val
                elif cat == "Ticket":
                    _add_rendiconto_detail_value(detail_map, categoria=cat, voce=voce, tipo=tipo, valore=s_val)
                    if tipo == "SI":
                        t_si += s_val
                    else:
                        t_no += s_val
                elif cat == "Delivery":
                    _add_rendiconto_detail_value(detail_map, categoria=cat, voce=voce, tipo=tipo, valore=s_val)
                    if tipo == "SI":
                        d_si += s_val
                    else:
                        d_no += s_val
                elif cat == "Coupon":
                    _add_rendiconto_detail_value(detail_map, categoria=cat, voce=voce, tipo=tipo, valore=s_val)
                    if tipo == "SI":
                        c_si += s_val
                    else:
                        c_no += s_val

            r["giro_affari"] = float(tot_vendite_lorde) - float(tot_annullati)
            r["pos"] = float(tot_pos)
            r["scontrini"] = float(tot_scontrini)
            r["distinte"] = float(tot_distinte)
            r["ticket_si"] = float(t_si)
            r["ticket_no"] = float(t_no)
            r["delivery_si"] = float(d_si)
            r["delivery_no"] = float(d_no)
            r["coupon_si"] = float(c_si)
            r["coupon_no"] = float(c_no)
            r["ticket_detail"] = _rendiconto_detail_list(detail_map, "Ticket")
            r["delivery_detail"] = _rendiconto_detail_list(detail_map, "Delivery")
            r["coupon_detail"] = _rendiconto_detail_list(detail_map, "Coupon")
        except Exception:
            # Se fallisce, lasciamo 0 e continuiamo
            pass

        # Spese
        try:
            r["spese"] = float(spese_map.get(str(code), 0.0)) if spese_map else float(sum_spese_month_total_net(store_code=str(code), year=y, month=m))
        except Exception:
            pass

        # Versamenti
        try:
            r["versamenti"] = float(vers_map.get(str(code), 0.0)) if vers_map else float(sum_versamenti_month_total(store_code=str(code), year=y, month=m))
        except Exception:
            pass

        # Contanti iPratico
        try:
            r["contanti_ipratico"] = float(ipratico_cash_map.get(str(code), 0.0))
        except Exception:
            pass

        # Giorni non versati
        try:
            status = _build_versamenti_status_for_store(
                store_code=str(code),
                year=y,
            )
            r["giorni_non_versati"] = int(status.get("giorni_non_versati") or 0)
        except Exception:
            pass

        # Diff cassa
        try:
            r["diff_cassa"] = float(
                float(r.get("distinte") or 0.0)
                + float(r.get("ticket_si") or 0.0)
                + float(r.get("delivery_si") or 0.0)
                + float(r.get("coupon_si") or 0.0)
                + float(r.get("pos") or 0.0)
                + float(r.get("spese") or 0.0)
                - float(r.get("giro_affari") or 0.0)
            )
        except Exception:
            r["diff_cassa"] = 0.0

        rows.append(r)

    try:
        rows = sorted(rows, key=lambda rr: (str(rr.get("store_name") or "").lower(), str(rr.get("store_code") or "")))
    except Exception:
        pass

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Riepilogo"

    headers = [
        "Store",
        "Giro affari",
        "Scontrini",
        "Differenza di cassa",
        "Distinte",
        "Contanti iPratico",
        "POS",
        "Spese",
        "Versamenti",
        "Giorni non versati",
        "Ticket",
        "Delivery",
        "Coupon",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze: prima riga + prima colonna
    ws.freeze_panes = "B2"

    def _money(v):
        try:
            return float(v)
        except Exception:
            return 0.0

    for rr in rows:
        store_label = (str(rr.get("store_name") or "").strip() or str(rr.get("store_code") or "").strip())
        ticket = _money(rr.get("ticket_si")) + _money(rr.get("ticket_no"))
        delivery = _money(rr.get("delivery_si")) + _money(rr.get("delivery_no"))
        coupon = _money(rr.get("coupon_si")) + _money(rr.get("coupon_no"))
        ws.append(
            [
                store_label,
                _money(rr.get("giro_affari")),
                int(round(_money(rr.get("scontrini")))),
                _money(rr.get("diff_cassa")),
                _money(rr.get("distinte")),
                _money(rr.get("contanti_ipratico")),
                _money(rr.get("pos")),
                _money(rr.get("spese")),
                _money(rr.get("versamenti")),
                int(rr.get("giorni_non_versati") or 0),
                _money(ticket),
                _money(delivery),
                _money(coupon),
            ]
        )

    # Formattazione numeri (Euro)
    eur_fmt = "#,##0.00"
    int_fmt = u"#,##0"
    for row_idx in range(2, ws.max_row + 1):
        # col 2 = Giro affari (EUR)
        ws.cell(row=row_idx, column=2).number_format = eur_fmt
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right")

        # col 3 = Scontrini (INT)
        ws.cell(row=row_idx, column=3).number_format = int_fmt
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right")

        # col 4..13 = EUR, esclusa col 3 che e' intera
        for col_idx in range(4, 14):
            ws.cell(row=row_idx, column=col_idx).number_format = eur_fmt
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="right")

    # Larghezze colonne
    widths = [32, 16, 12, 18, 16, 18, 14, 14, 16, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    _append_rendiconto_detail_sheet(wb, rows)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"riepilogo_{month_norm}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@rendiconto_bp.get("/api/riepilogo/giornaliero.xlsx")
def api_riepilogo_giornaliero_xlsx():
    """Export Excel giornaliero del periodo selezionato (mese).

    Una riga per (data, store) con le stesse metriche del riepilogo mensile.
    """
    if Workbook is None:
        return jsonify({"error": "openpyxl non disponibile"}), 500

    lg = _require_login()
    if lg:
        return lg

    month_norm, y, m, start_iso, end_iso = _parse_month_yyyy_mm_strict_iso(request.args.get("month"))
    if not month_norm:
        return jsonify({"error": "Parametro month mancante o non valido. Formato: YYYY-MM"}), 400

    user_id = session.get("uid")
    try:
        stores = _available_stores_for_user(str(user_id)) or []
    except Exception:
        stores = []

    # Pre-costruzione giorni del mese
    days = list(_iter_days_iso(start_iso, end_iso))

    out_rows = []

    def _num(v):
        try:
            return float(v)
        except Exception:
            return 0.0

    for s in stores or []:
        code = (s or {}).get("code")
        name = (s or {}).get("name")
        if not code:
            continue

        store_label = (str(name or "").strip() or str(code).strip())

        # Base per giorno
        by_day = {
            d: {
                "vendite_lorde": 0.0,
                "annullati": 0.0,
                "pos": 0.0,
                "scontrini": 0.0,
                "distinte": 0.0,
                "ticket_si": 0.0,
                "ticket_no": 0.0,
                "delivery_si": 0.0,
                "delivery_no": 0.0,
                "coupon_si": 0.0,
                "coupon_no": 0.0,
                "detail_map": _empty_rendiconto_detail_map(),
                "spese": 0.0,
                "versato": False,
            }
            for d in days
        }

        # Prima nota per giorno
        try:
            prim = load_primanota_month_agg(str(code), year=y, month=m)
            for r in prim or []:
                d_iso = str(r.get("date") or "").strip()
                if not d_iso or d_iso not in by_day:
                    continue
                cat = str(r.get("categoria") or "").strip()
                voce = str(r.get("voce") or "").strip()
                tipo = str(r.get("tipo") or "SI").strip().upper()
                s_val = _num(r.get("sum"))

                if cat == "Dati chiusura":
                    k = _CHIUSURA_VOICE_TO_KEY.get(voce.upper())
                    if k == "vendite_lorde":
                        by_day[d_iso]["vendite_lorde"] += s_val
                    elif k == "annullati":
                        by_day[d_iso]["annullati"] += s_val
                    elif k == "pos":
                        by_day[d_iso]["pos"] += s_val
                    elif k == "scontrini":
                        by_day[d_iso]["scontrini"] += s_val
                elif cat == "Distinte":
                    by_day[d_iso]["distinte"] += s_val
                elif cat == "Ticket":
                    _add_rendiconto_detail_value(by_day[d_iso]["detail_map"], categoria=cat, voce=voce, tipo=tipo, valore=s_val)
                    if tipo == "SI":
                        by_day[d_iso]["ticket_si"] += s_val
                    else:
                        by_day[d_iso]["ticket_no"] += s_val
                elif cat == "Delivery":
                    _add_rendiconto_detail_value(by_day[d_iso]["detail_map"], categoria=cat, voce=voce, tipo=tipo, valore=s_val)
                    if tipo == "SI":
                        by_day[d_iso]["delivery_si"] += s_val
                    else:
                        by_day[d_iso]["delivery_no"] += s_val
                elif cat == "Coupon":
                    _add_rendiconto_detail_value(by_day[d_iso]["detail_map"], categoria=cat, voce=voce, tipo=tipo, valore=s_val)
                    if tipo == "SI":
                        by_day[d_iso]["coupon_si"] += s_val
                    else:
                        by_day[d_iso]["coupon_no"] += s_val
        except Exception:
            pass

        # Spese per giorno (net)
        try:
            sp_by_day = sum_spese_month_by_day(store_code=str(code), year=y, month=m) or {}
            for d_iso, vals in (sp_by_day or {}).items():
                if d_iso in by_day:
                    by_day[d_iso]["spese"] = _num((vals or {}).get("net"))
        except Exception:
            pass

        # Giorni coperti da versamenti per competenza
        try:
            vres = list_versamenti_periods_overlapping(
                store_code=str(code),
                start_iso=days[0],
                end_iso=days[-1],
            ) or []
            for vr in vres:
                dal_iso = str(vr.get("periodo_dal_iso") or "").strip()
                al_iso = str(vr.get("periodo_al_iso") or "").strip()
                if not dal_iso or not al_iso:
                    continue
                for d_iso in _daterange_iso(dal_iso, al_iso):
                    if d_iso in by_day:
                        by_day[d_iso]["versato"] = True
        except Exception:
            pass

        # Costruzione righe output
        for d_iso in days:
            dd = by_day.get(d_iso) or {}
            giro = _num(dd.get("vendite_lorde")) - _num(dd.get("annullati"))
            ticket = _num(dd.get("ticket_si")) + _num(dd.get("ticket_no"))
            delivery = _num(dd.get("delivery_si")) + _num(dd.get("delivery_no"))
            coupon = _num(dd.get("coupon_si")) + _num(dd.get("coupon_no"))
            diff = (
                _num(dd.get("distinte"))
                + _num(dd.get("ticket_si"))
                + _num(dd.get("delivery_si"))
                + _num(dd.get("coupon_si"))
                + _num(dd.get("pos"))
                + _num(dd.get("spese"))
                - giro
            )

            out_rows.append(
                {
                    "date_iso": d_iso,
                    "store": store_label,
                    "scontrini": int(round(_num(dd.get("scontrini")))),
                    "giro_affari": giro,
                    "diff_cassa": diff,
                    "distinte": _num(dd.get("distinte")),
                    "pos": _num(dd.get("pos")),
                    "spese": _num(dd.get("spese")),
                    "versato": "SI" if bool(dd.get("versato")) else "",
                    "ticket": ticket,
                    "delivery": delivery,
                    "coupon": coupon,
                    "ticket_detail": _rendiconto_detail_list(dd.get("detail_map") or {}, "Ticket"),
                    "delivery_detail": _rendiconto_detail_list(dd.get("detail_map") or {}, "Delivery"),
                    "coupon_detail": _rendiconto_detail_list(dd.get("detail_map") or {}, "Coupon"),
                }
            )

    try:
        out_rows = sorted(out_rows, key=lambda rr: (rr.get("date_iso") or "", str(rr.get("store") or "").lower()))
    except Exception:
        pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Giornaliero"

    headers = [
        "Data",
        "Store",
        "Scontrini",
        "Giro affari",
        "Differenza di cassa",
        "Distinte",
        "POS",
        "Spese",
        "Versato",
        "Ticket",
        "Delivery",
        "Coupon",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for col_idx, _h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "C2"

    for rr in out_rows:
        try:
            d_obj = datetime.strptime(str(rr.get("date_iso") or ""), "%Y-%m-%d").date()
        except Exception:
            d_obj = str(rr.get("date_iso") or "")

        ws.append(
            [
                d_obj,
                rr.get("store"),
                rr.get("scontrini"),
                rr.get("giro_affari"),
                rr.get("diff_cassa"),
                rr.get("distinte"),
                rr.get("pos"),
                rr.get("spese"),
                rr.get("versato"),
                rr.get("ticket"),
                rr.get("delivery"),
                rr.get("coupon"),
            ]
        )

    eur_fmt = "#,##0.00"
    int_fmt = u"#,##0"
    date_fmt = u"DD/MM/YYYY"

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).number_format = date_fmt
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="right")

        ws.cell(row=row_idx, column=3).number_format = int_fmt
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right")

        for col_idx in (4, 5, 6, 7, 8, 10, 11, 12):
            ws.cell(row=row_idx, column=col_idx).number_format = eur_fmt
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="right")

        ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal="center")

    widths = [12, 32, 12, 16, 18, 16, 14, 14, 16, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws_detail = wb.create_sheet("Dettaglio")
    detail_headers = ["Data", "Store", "Categoria", "Dettaglio", "Tipo", "Importo"]
    ws_detail.append(detail_headers)
    if Font and Alignment:
        for cell in ws_detail[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    for rr in out_rows:
        try:
            d_obj = datetime.strptime(str(rr.get("date_iso") or ""), "%Y-%m-%d").date()
        except Exception:
            d_obj = str(rr.get("date_iso") or "")
        for cat, key in (("Ticket", "ticket_detail"), ("Delivery", "delivery_detail"), ("Coupon", "coupon_detail")):
            for item in rr.get(key) or []:
                ws_detail.append(
                    [
                        d_obj,
                        rr.get("store"),
                        cat,
                        str(item.get("label") or item.get("voce") or ""),
                        str(item.get("tipo") or ""),
                        float(item.get("valore") or 0.0),
                    ]
                )

    for row_idx in range(2, ws_detail.max_row + 1):
        ws_detail.cell(row=row_idx, column=1).number_format = date_fmt
        ws_detail.cell(row=row_idx, column=6).number_format = eur_fmt
        ws_detail.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right")
    detail_widths = [12, 32, 14, 34, 10, 14]
    for i, w in enumerate(detail_widths, start=1):
        ws_detail.column_dimensions[chr(64 + i)].width = w
    ws_detail.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"riepilogo_giornaliero_{month_norm}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

