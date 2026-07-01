from __future__ import annotations

from datetime import date as _date, datetime, timedelta
import io
from decimal import Decimal, ROUND_HALF_UP

import time
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app, jsonify, send_file, abort

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore
    Font = None  # type: ignore
    Alignment = None  # type: ignore

from db_integration import get_warehouse_stores, get_user_warehouse_stores

import json

from spese_repository import (
    insert_spesa,
    list_spese_month,
    delete_spesa,
    update_spesa,
    get_spesa_photo_file,
    sum_spese_day,
    sum_spese_month_by_day,
    sum_spese_month_total_net,
)

from sharepoint_photos_repository import (
    SharePointTestError,
    upload_spesa_photo,
    download_spesa_photo,
    delete_spesa_photo,
    upload_versamento_photo,
    download_versamento_photo,
    delete_versamento_photo,
)
from primanota_repository import (
    get_elenchi_options,
    load_primanota_day,
    replace_primanota_day,
    delete_primanota_day,
    load_primanota_month_agg,
    load_primanota_month_agg_totals,
    sum_categoria_period,
    sum_categoria_by_day_range,
)

from dati_database_repository import upsert_datidatabase_from_distinta, delete_datidatabase_day

from versamenti_repository import (
    insert_versamento,
    list_versamenti_month,
    list_versamenti_periods_overlapping,
    sum_versamenti_month_total,
    delete_versamento,
    update_versamento,
    get_versamento_photo_file,
)

rendiconto_bp = Blueprint("rendiconto", __name__, url_prefix="/rendiconto")


def _ensure_session_keys() -> None:
    session.setdefault("store_code", None)
    session.setdefault("store_name", None)


def _require_login():
    if not session.get("uid"):
        nxt = request.full_path if request.query_string else request.path
        return redirect(url_for("login", next=nxt))
    return None


def _parse_ym(value: str):
    value = (value or "").strip()
    if not value:
        today = _date.today()
        return today.year, today.month, f"{today.year:04d}-{today.month:02d}"
    try:
        y, m = value.split("-", 1)
        y = int(y); m = int(m)
        if not (1 <= m <= 12) or y < 2000:
            raise ValueError
        return y, m, f"{y:04d}-{m:02d}"
    except Exception:
        today = _date.today()
        return today.year, today.month, f"{today.year:04d}-{today.month:02d}"


def _parse_date_iso(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return _date.today().isoformat()
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return value
    except Exception:
        return _date.today().isoformat()


def _is_valid_iso_date(value: str) -> bool:
    value = (value or "").strip()
    if not value:
        return False
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except Exception:
        return False


def _same_month_iso(dal_iso: str, al_iso: str) -> bool:
    dal_iso = (dal_iso or "").strip()
    al_iso = (al_iso or "").strip()
    if not (_is_valid_iso_date(dal_iso) and _is_valid_iso_date(al_iso)):
        return False
    return dal_iso[:7] == al_iso[:7]


def _parse_float(v: str) -> float:
    s = (v or "").strip().replace(".", "").replace(",", ".")
    # gestione input come 1.234,56 -> 1234.56
    # se l'utente inserisce 1234.56 resta ok dopo replace sopra ("." tolto),
    # quindi forziamo: se la stringa originale contiene punto come decimale,
    # l'utente può scrivere 1234.56 senza separatori migliaia.
    if v and ("." in v) and ("," not in v):
        s = (v or "").strip().replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _iter_days_iso(start_iso: str, end_iso: str):
    """Yield giorni ISO inclusivi."""
    try:
        d0 = datetime.strptime(start_iso, "%Y-%m-%d").date()
        d1 = datetime.strptime(end_iso, "%Y-%m-%d").date()
    except Exception:
        return
    if d1 < d0:
        return
    cur = d0
    while cur <= d1:
        yield cur.isoformat()
        cur += timedelta(days=1)


def _overlap_days_with_existing_versamenti(
    *,
    store_code: str,
    dal_iso: str,
    al_iso: str,
    exclude_id: str | None = None,
    exclude_signature: dict | None = None,
):
    """Ritorna (giorni_overlap, summaries_versamenti).

    Se exclude_id è valorizzato, il versamento con quell'id viene ignorato.
    Se exclude_signature è valorizzato e le righe non hanno id, prova a escludere la riga matching.
    """
    exclude_id = (exclude_id or "").strip()
    exclude_signature = exclude_signature or None

    vres = list_versamenti_periods_overlapping(store_code=store_code, start_iso=dal_iso, end_iso=al_iso)
    rows = (vres or {}).get("rows") or []
    if not rows:
        return [], []

    req_start = datetime.strptime(dal_iso, "%Y-%m-%d").date()
    req_end = datetime.strptime(al_iso, "%Y-%m-%d").date()

    covered = set()
    summaries = []

    for r in rows:
        rid = str(r.get("id") or "").strip()
        if exclude_id and rid and rid == exclude_id:
            continue

        if not rid and exclude_signature:
            if (
                str(r.get("dal_iso") or "").strip() == str(exclude_signature.get("dal_iso") or "").strip()
                and str(r.get("al_iso") or "").strip() == str(exclude_signature.get("al_iso") or "").strip()
                and str(r.get("data_versamento_iso") or "").strip()
                == str(exclude_signature.get("data_versamento_iso") or "").strip()
                and str(r.get("valore_key") or "").strip() == str(exclude_signature.get("valore_key") or "").strip()
            ):
                continue

        r_start_s = str(r.get("dal_iso") or "").strip()
        r_end_s = str(r.get("al_iso") or "").strip()
        if not (r_start_s and r_end_s and _is_valid_iso_date(r_start_s) and _is_valid_iso_date(r_end_s)):
            continue
        r_start = datetime.strptime(r_start_s, "%Y-%m-%d").date()
        r_end = datetime.strptime(r_end_s, "%Y-%m-%d").date()
        if r_end < r_start:
            continue

        o_start = max(req_start, r_start)
        o_end = min(req_end, r_end)
        if o_end < o_start:
            continue

        for d in _iter_days_iso(o_start.isoformat(), o_end.isoformat()):
            covered.add(d)

        summaries.append({"id": rid, "dal_iso": r_start_s, "al_iso": r_end_s})

    return sorted(covered), summaries


def _money_to_decimal(v: str) -> Decimal:
    s = (v or "").strip()
    if not s:
        return Decimal("0")
    # 1.234,56 -> 1234.56
    if ("." in s) and ("," in s):
        s = s.replace(".", "").replace(",", ".")
    else:
        # 1234.56 (no migliaia) resta ok
        if ("." in s) and ("," not in s):
            s = s
        else:
            s = s.replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _round2(v: Decimal) -> Decimal:
    return (v or Decimal("0")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _iso_to_date(s: str):
    return datetime.strptime(s, "%Y-%m-%d").date()


def _daterange_iso(start_iso: str, end_iso: str):
    d1 = _iso_to_date(start_iso)
    d2 = _iso_to_date(end_iso)
    if d2 < d1:
        d1, d2 = d2, d1
    cur = d1
    while cur <= d2:
        yield cur.isoformat()
        cur += timedelta(days=1)


def _match_exclude_versamento(row: dict, exclude: dict) -> bool:
    """True se row rappresenta lo stesso versamento da escludere (lock)."""
    if not exclude:
        return False
    ex_id = (exclude.get("id") or "").strip()
    if ex_id and (row.get("id") or "").strip() == ex_id:
        return True

    # fallback chiave composta
    keys = {
        "data_versamento_iso": (exclude.get("orig_data_vers") or "").strip(),
        "dal_iso": (exclude.get("orig_dal") or "").strip(),
        "al_iso": (exclude.get("orig_al") or "").strip(),
        "nome_raw": (exclude.get("orig_nome") or "").strip(),
        "tipo_raw": (exclude.get("orig_tipo") or "").strip(),
        "tessera_raw": (exclude.get("orig_tessera") or "").strip(),
        "riferimento_raw": (exclude.get("orig_riferimento") or "").strip(),
        "valore_key": (exclude.get("orig_valore") or "").strip(),
    }
    # se non c'è almeno una parte significativa, non matchiamo
    if not any(keys.values()):
        return False

    def _norm(s: str) -> str:
        return (s or "").strip().lower()

    # tessera: normalizza a sole cifre
    def _digits(s: str) -> str:
        return "".join(ch for ch in (s or "") if ch.isdigit())[:16] or "0"

    return (
        _norm(row.get("data_versamento_iso") or "") == _norm(keys["data_versamento_iso"]) and
        _norm(row.get("dal_iso") or "") == _norm(keys["dal_iso"]) and
        _norm(row.get("al_iso") or "") == _norm(keys["al_iso"]) and
        _norm(row.get("nome_raw") or "") == _norm(keys["nome_raw"]) and
        _norm(row.get("tipo_raw") or "") == _norm(keys["tipo_raw"]) and
        _digits(row.get("tessera_raw") or "") == _digits(keys["tessera_raw"]) and
        _norm(row.get("riferimento_raw") or "") == _norm(keys["riferimento_raw"]) and
        (row.get("valore_key") or "").strip() == (keys["valore_key"] or "").strip()
    )


def _locked_days_in_range(*, store_code: str, start_iso: str, end_iso: str, exclude: dict | None = None) -> set[str]:
    """Ritorna l'insieme dei giorni (ISO) inclusi nel periodo competenza di *altri* versamenti."""
    exclude = exclude or {}

    try:
        periods = list_versamenti_periods_overlapping(store_code=str(store_code), start_iso=start_iso, end_iso=end_iso)
        rows = periods.get("rows") or []
    except Exception:
        rows = []

    d_start = _iso_to_date(start_iso)
    d_end = _iso_to_date(end_iso)
    if d_end < d_start:
        d_start, d_end = d_end, d_start

    locked: set[str] = set()
    for r in rows:
        if _match_exclude_versamento(r, exclude):
            continue
        dal = (r.get("dal_iso") or "").strip()
        al = (r.get("al_iso") or "").strip()
        if not dal or not al:
            continue
        try:
            rd1 = _iso_to_date(dal)
            rd2 = _iso_to_date(al)
        except Exception:
            continue
        if rd2 < rd1:
            rd1, rd2 = rd2, rd1
        a = max(rd1, d_start)
        b = min(rd2, d_end)
        cur = a
        while cur <= b:
            locked.add(cur.isoformat())
            cur += timedelta(days=1)
    return locked


@rendiconto_bp.get("/")
def home():
    return redirect(url_for("rendiconto.distinta_cassa"))


@rendiconto_bp.route("/spese", methods=["GET", "POST"])
def spese():
    _ensure_session_keys()

    # Richiediamo login almeno per scrivere (POST). In GET lasciamo comunque
    # la pagina navigabile, ma senza store selezionato la modale blocca.
    if request.method == "POST":
        r = _require_login()
        if r is not None:
            return r

    store_code = session.get("store_code")
    store_name = session.get("store_name")

    ym = request.args.get("ym") or request.form.get("ym") or ""
    year, month, ym_norm = _parse_ym(ym)

    if request.method == "POST":
        if not store_code:
            flash("Seleziona prima uno store.", "warning")
            return redirect(url_for("rendiconto.spese", ym=ym_norm))

        data_iso = (request.form.get("data") or "").strip()
        tipo = (request.form.get("tipo_operazione") or "").strip()
        forn = (request.form.get("fornitore_spesa") or "").strip()
        doc = (request.form.get("documento") or "").strip()
        imp = (request.form.get("importo") or "").strip()

        if not (data_iso and tipo and forn and doc and imp):
            flash("Compila tutti i campi prima di salvare.", "warning")
            return redirect(url_for("rendiconto.spese", ym=ym_norm))

        foto_file = None
        foto = request.files.get("foto")
        if foto is not None and getattr(foto, "filename", ""):
            try:
                foto_file = upload_spesa_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    file_storage=foto,
                    data_iso=data_iso,
                )
            except Exception as e:
                current_app.logger.exception("Errore upload foto spesa")
                flash(f"Foto: upload fallito ({e}). La spesa verrà salvata senza foto.", "warning")

        try:
            insert_spesa(
                store_code=str(store_code),
                data_iso=data_iso,
                tipo_operazione=tipo,
                fornitore_spesa=forn,
                documento=doc,
                importo_euro=imp,
                foto_file=foto_file,
            )
            flash("Spesa salvata.", "success")
        except Exception as e:
            # se la spesa non viene salvata, prova a pulire la foto caricata
            if foto_file:
                try:
                    delete_spesa_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(foto_file),
                    )
                except Exception:
                    pass
            current_app.logger.exception("Errore salvataggio spesa")
            flash(f"Errore salvataggio spesa: {e}", "danger")

        return redirect(url_for("rendiconto.spese", ym=ym_norm))

    rows = []
    total = 0.0
    if store_code:
        try:
            res = list_spese_month(store_code=str(store_code), year=year, month=month)
            rows = res.get("rows") or []
            total = float(res.get("total") or 0)
        except Exception as e:
            current_app.logger.exception("Errore lettura spese")
            flash(f"Errore lettura spese: {e}", "danger")

    today_iso = _date.today().isoformat()

    return render_template(
        "rendiconto_spese.html",
        store_code=store_code,
        store_name=store_name,
        ym=ym_norm,
        year=year,
        month=month,
        rows=rows,
        total=total,
        today_iso=today_iso,
    )


@rendiconto_bp.post("/spese/delete")
def spese_delete():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    ym = request.form.get("ym") or ""
    _, _, ym_norm = _parse_ym(ym)

    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.spese", ym=ym_norm))

    orig_data = (request.form.get("orig_data") or "").strip()
    orig_tipo = (request.form.get("orig_tipo") or "").strip()
    orig_forn = (request.form.get("orig_fornitore") or "").strip()
    orig_doc = (request.form.get("orig_documento") or "").strip()
    orig_imp = (request.form.get("orig_importo") or "").strip()

    foto_to_delete = None
    try:
        foto_to_delete = get_spesa_photo_file(
            store_code=str(store_code),
            orig_data_iso=orig_data,
            orig_tipo=orig_tipo,
            orig_fornitore=orig_forn,
            orig_documento=orig_doc,
            orig_importo_key=orig_imp,
        )
    except Exception:
        foto_to_delete = None

    try:
        n = delete_spesa(
            store_code=str(store_code),
            orig_data_iso=orig_data,
            orig_tipo=orig_tipo,
            orig_fornitore=orig_forn,
            orig_documento=orig_doc,
            orig_importo_key=orig_imp,
        )
        if n > 0:
            flash("Spesa eliminata.", "success")
            if foto_to_delete:
                try:
                    delete_spesa_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(foto_to_delete),
                    )
                except Exception as e:
                    current_app.logger.warning("Delete foto spesa fallita: %s", e)
        else:
            flash("Nessuna riga eliminata (record non trovato).", "warning")
    except Exception as e:
        current_app.logger.exception("Errore eliminazione spesa")
        flash(f"Errore eliminazione spesa: {e}", "danger")

    return redirect(url_for("rendiconto.spese", ym=ym_norm))


@rendiconto_bp.post("/spese/update")
def spese_update():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    ym = request.form.get("ym") or ""
    _, _, ym_norm = _parse_ym(ym)

    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.spese", ym=ym_norm))

    new_data = (request.form.get("data") or "").strip()
    new_tipo = (request.form.get("tipo_operazione") or "").strip()
    new_forn = (request.form.get("fornitore_spesa") or "").strip()
    new_doc = (request.form.get("documento") or "").strip()
    new_imp = (request.form.get("importo") or "").strip()

    if not (new_data and new_tipo and new_forn and new_doc and new_imp):
        flash("Compila tutti i campi prima di salvare.", "warning")
        return redirect(url_for("rendiconto.spese", ym=ym_norm))

    orig_data = (request.form.get("orig_data") or "").strip()
    orig_tipo = (request.form.get("orig_tipo") or "").strip()
    orig_forn = (request.form.get("orig_fornitore") or "").strip()
    orig_doc = (request.form.get("orig_documento") or "").strip()
    orig_imp = (request.form.get("orig_importo") or "").strip()

    old_foto = None
    try:
        old_foto = get_spesa_photo_file(
            store_code=str(store_code),
            orig_data_iso=orig_data,
            orig_tipo=orig_tipo,
            orig_fornitore=orig_forn,
            orig_documento=orig_doc,
            orig_importo_key=orig_imp,
        )
    except Exception:
        old_foto = None

    new_foto_file = None
    foto = request.files.get("foto")
    if foto is not None and getattr(foto, "filename", ""):
        try:
            new_foto_file = upload_spesa_photo(
                sb_jwt=str(session.get("sb_token") or ""),
                user_id=str(session.get("uid") or ""),
                store_code=str(store_code),
                file_storage=foto,
                data_iso=new_data,
            )
        except Exception as e:
            current_app.logger.exception("Errore upload foto spesa (update)")
            flash(f"Foto: upload fallito ({e}). La spesa verrà aggiornata senza cambiare la foto.", "warning")
            new_foto_file = None

    try:
        n = update_spesa(
            store_code=str(store_code),
            orig_data_iso=orig_data,
            orig_tipo=orig_tipo,
            orig_fornitore=orig_forn,
            orig_documento=orig_doc,
            orig_importo_key=orig_imp,
            new_data_iso=new_data,
            new_tipo=new_tipo,
            new_fornitore=new_forn,
            new_documento=new_doc,
            new_importo_euro=new_imp,
            new_foto_file=new_foto_file,
        )
        if n > 0:
            flash("Spesa aggiornata.", "success")
            # se è stata sostituita la foto, puliamo la vecchia
            if new_foto_file and old_foto and old_foto != new_foto_file:
                try:
                    delete_spesa_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(old_foto),
                    )
                except Exception as e:
                    current_app.logger.warning("Delete vecchia foto spesa fallita: %s", e)
        else:
            flash("Nessuna riga aggiornata (record non trovato).", "warning")
            # update fallito: evita file orfani
            if new_foto_file:
                try:
                    delete_spesa_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(new_foto_file),
                    )
                except Exception:
                    pass
    except Exception as e:
        current_app.logger.exception("Errore aggiornamento spesa")
        if new_foto_file:
            try:
                delete_spesa_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    filename=str(new_foto_file),
                )
            except Exception:
                pass
        flash(f"Errore aggiornamento spesa: {e}", "danger")

    return redirect(url_for("rendiconto.spese", ym=ym_norm))


@rendiconto_bp.get("/spese/photo/<path:filename>")
def spese_photo(filename: str):
    """Serve la foto collegata ad una spesa (stream dal repository SharePoint)."""
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    if not store_code:
        abort(404)

    fname = (filename or "").strip()
    # sicurezza: niente path traversal
    if "/" in fname or "\\" in fname:
        abort(404)

    try:
        content = download_spesa_photo(
            sb_jwt=str(session.get("sb_token") or ""),
            user_id=str(session.get("uid") or ""),
            store_code=str(store_code),
            filename=fname,
        )
        return send_file(
            io.BytesIO(content),
            mimetype="image/jpeg",
            download_name=fname,
            as_attachment=False,
            max_age=300,
        )
    except Exception as e:
        current_app.logger.warning("Foto spesa non disponibile: %s", e)
        abort(404)


# -------------------------
# Versamenti
# -------------------------


@rendiconto_bp.route("/versamenti", methods=["GET", "POST"])
def versamenti():
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")

    y, m, ym_norm = _parse_ym(request.values.get("ym") or "")
    today_iso = _date.today().isoformat()

    if request.method == "POST":
        r = _require_login()
        if r is not None:
            return r

        if not store_code:
            flash("Seleziona prima uno store.", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        raw_data_vers = (request.form.get("data_versamento") or "").strip()
        raw_dal = (request.form.get("periodo_dal") or "").strip()
        raw_al = (request.form.get("periodo_al") or "").strip()

        if not (_is_valid_iso_date(raw_data_vers) and _is_valid_iso_date(raw_dal) and _is_valid_iso_date(raw_al)):
            flash("Periodo non valido: usa date valide (YYYY-MM-DD).", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        if not _same_month_iso(raw_dal, raw_al):
            flash("Il periodo di competenza deve essere all'interno dello stesso mese.", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        if raw_al < raw_dal:
            flash("Periodo non valido: la data 'Al' non può essere precedente alla data 'Dal'.", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        data_vers = _parse_date_iso(raw_data_vers)
        dal = _parse_date_iso(raw_dal)
        al = _parse_date_iso(raw_al)

        # Non permettere giorni già inclusi in altri versamenti
        try:
            covered_days, _overlaps = _overlap_days_with_existing_versamenti(
                store_code=str(store_code),
                dal_iso=dal,
                al_iso=al,
            )
            if covered_days:
                preview = ", ".join(covered_days[:6])
                if len(covered_days) > 6:
                    preview += f" (+{len(covered_days) - 6})"
                flash(f"Nel periodo selezionato ci sono giorni già versati: {preview}. Modifica il periodo.", "warning")
                return redirect(url_for("rendiconto.versamenti", ym=ym_norm))
        except Exception:
            # se la verifica overlap fallisce, non blocchiamo: la UI gestisce comunque il check via /compute
            pass

        nome = (request.form.get("nome_cognome") or "").strip()
        tipo = (request.form.get("tipo_versamento") or "").strip()
        tessera = (request.form.get("tessera") or "").strip()
        rif = (request.form.get("riferimento") or "").strip()
        valore = (request.form.get("valore") or "").strip()

        if not (data_vers and dal and al and nome and tipo and valore):
            flash("Compila i campi obbligatori prima di salvare.", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        tessera_digits = "".join(ch for ch in tessera if ch.isdigit())
        if tessera_digits and len(tessera_digits) > 16:
            flash("Tessera: massimo 16 cifre.", "warning")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        # Blocco: il versamento può essere salvato solo se la differenza è zero.
        try:
            valore_dec = _round2(_money_to_decimal(valore))
            dist = Decimal(str(sum_categoria_period(store_code=str(store_code), start_iso=dal, end_iso=al, categoria="Distinte")))
            diff = _round2(valore_dec - _round2(dist))
        except Exception as e:
            current_app.logger.exception("Errore verifica differenza versamento")
            flash(f"Errore calcolo differenza: {e}", "danger")
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        if diff != Decimal("0.00"):
            flash(
                f"Differenza diversa da zero ({float(diff):.2f} €). Correggi le distinte del periodo prima di salvare.",
                "warning",
            )
            return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

        foto_file = None
        foto = request.files.get("foto")
        if foto is not None and getattr(foto, "filename", ""):
            try:
                foto_file = upload_versamento_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    file_storage=foto,
                    data_iso=data_vers,
                )
            except Exception as e:
                current_app.logger.exception("Errore upload foto versamento (insert)")
                flash(f"Foto: upload fallito ({e}). Il versamento verrà salvato senza foto.", "warning")
                foto_file = None

        try:
            insert_versamento(
                store_code=str(store_code),
                data_versamento_iso=data_vers,
                periodo_dal_iso=dal,
                periodo_al_iso=al,
                nome_cognome=nome,
                tipo_versamento=tipo,
                tessera=tessera_digits,
                riferimento=rif,
                valore_euro=valore,
                foto_file=foto_file,
            )
            flash("Versamento salvato.", "success")
        except Exception as e:
            current_app.logger.exception("Errore salvataggio versamento")
            if foto_file:
                try:
                    delete_versamento_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(foto_file),
                    )
                except Exception:
                    pass
            flash(f"Errore salvataggio versamento: {e}", "danger")

        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    # GET
    rows = []
    total = 0.0
    has_id = False

    if store_code:
        try:
            info = list_versamenti_month(store_code=str(store_code), year=y, month=m)
            rows = info.get("rows") or []
            total = float(info.get("total") or 0.0)
            has_id = bool(info.get("has_id"))
        except Exception as e:
            current_app.logger.exception("Errore lettura VERSAMENTI_APP")
            flash(f"Errore lettura versamenti: {e}", "danger")

    # Calcolo differenza e distinte periodo (per ogni riga) in modo efficiente
    if store_code and rows:
        try:
            min_dal = min(r.get("dal_iso") for r in rows if r.get("dal_iso"))
            max_al = max(r.get("al_iso") for r in rows if r.get("al_iso"))

            dist_by_day = sum_categoria_by_day_range(
                store_code=str(store_code),
                start_iso=min_dal,
                end_iso=max_al,
                categoria="Distinte",
            )

            d0 = datetime.strptime(min_dal, "%Y-%m-%d").date()
            d1 = datetime.strptime(max_al, "%Y-%m-%d").date()
            n = (d1 - d0).days + 1

            prefix = [0.0] * (n + 1)
            for i in range(n):
                di = d0 + timedelta(days=i)
                prefix[i + 1] = prefix[i] + float(dist_by_day.get(di.isoformat(), 0.0) or 0.0)

            def _sum_period(dal_iso: str, al_iso: str) -> float:
                try:
                    a = datetime.strptime(dal_iso, "%Y-%m-%d").date()
                    b = datetime.strptime(al_iso, "%Y-%m-%d").date()
                except Exception:
                    return 0.0
                if b < a:
                    a, b = b, a
                ia = (a - d0).days
                ib = (b - d0).days
                if ia < 0:
                    ia = 0
                if ib >= n:
                    ib = n - 1
                if ib < ia:
                    return 0.0
                return prefix[ib + 1] - prefix[ia]

            for r in rows:
                dist_tot = _sum_period(r.get("dal_iso") or "", r.get("al_iso") or "")
                r["distinte_periodo"] = float(dist_tot)
                try:
                    v = float(r.get("valore") or 0.0)
                except Exception:
                    v = 0.0
                r["differenza"] = v - float(dist_tot)
        except Exception:
            current_app.logger.exception("Errore calcolo differenza versamenti")

    return render_template(
        "rendiconto_versamenti.html",
        store_code=store_code,
        store_name=store_name,
        ym=ym_norm,
        today_iso=today_iso,
        rows=rows,
        total=total,
        has_id=has_id,
    )


@rendiconto_bp.post("/versamenti/delete")
def versamenti_delete():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    ym_norm = _parse_ym(request.form.get("ym") or "")[2]
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    record_id = (request.form.get("id") or "").strip()
    orig_data = (request.form.get("orig_data_vers") or "").strip()
    orig_dal = (request.form.get("orig_dal") or "").strip()
    orig_al = (request.form.get("orig_al") or "").strip()
    orig_nome = (request.form.get("orig_nome") or "").strip()
    orig_tipo = (request.form.get("orig_tipo") or "").strip()
    orig_tess = (request.form.get("orig_tessera") or "").strip()
    orig_rif = (request.form.get("orig_riferimento") or "").strip()
    orig_val = (request.form.get("orig_valore") or "").strip()

    # Se la tabella non ha ID, escludiamo questo record tramite signature basata sui campi originali
    if not record_id:
        try:
            exclude_sig = {
                "data_versamento_iso": orig_data,
                "dal_iso": orig_dal,
                "al_iso": orig_al,
                "valore_key": orig_val,
                "nome_raw": orig_nome,
            }
            covered_days, _overlaps = _overlap_days_with_existing_versamenti(
                store_code=str(store_code),
                dal_iso=new_dal,
                al_iso=new_al,
                exclude_signature=exclude_sig,
            )
            if covered_days:
                preview = ", ".join(covered_days[:6])
                if len(covered_days) > 6:
                    preview += f" (+{len(covered_days) - 6})"
                flash(f"Nel periodo selezionato ci sono giorni già versati: {preview}. Modifica il periodo.", "warning")
                return redirect(url_for("rendiconto.versamenti", ym=ym_norm))
        except Exception:
            pass

    old_foto = None
    try:
        old_foto = get_versamento_photo_file(
            store_code=str(store_code),
            record_id=str(record_id or ""),
            orig_data_vers_iso=orig_data,
            orig_dal_iso=orig_dal,
            orig_al_iso=orig_al,
            orig_nome=orig_nome,
            orig_tipo=orig_tipo,
            orig_tessera=orig_tess,
            orig_riferimento=orig_rif,
            orig_valore_key=orig_val,
        )
    except Exception:
        old_foto = None

    try:
        n = delete_versamento(
            store_code=str(store_code),
            record_id=record_id,
            orig_data_vers_iso=orig_data,
            orig_dal_iso=orig_dal,
            orig_al_iso=orig_al,
            orig_nome=orig_nome,
            orig_tipo=orig_tipo,
            orig_tessera=orig_tess,
            orig_riferimento=orig_rif,
            orig_valore_key=orig_val,
        )
        if n > 0:
            flash("Versamento eliminato.", "success")
            if old_foto:
                try:
                    delete_versamento_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(old_foto),
                    )
                except Exception as e:
                    current_app.logger.warning("Delete foto versamento fallita: %s", e)
        else:
            flash("Nessuna riga eliminata (record non trovato).", "warning")
    except Exception as e:
        current_app.logger.exception("Errore eliminazione versamento")
        flash(f"Errore eliminazione versamento: {e}", "danger")

    return redirect(url_for("rendiconto.versamenti", ym=ym_norm))


@rendiconto_bp.post("/versamenti/update")
def versamenti_update():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    ym_norm = _parse_ym(request.form.get("ym") or "")[2]
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    record_id = (request.form.get("id") or "").strip()

    raw_new_data_vers = (request.form.get("data_versamento") or "").strip()
    raw_new_dal = (request.form.get("periodo_dal") or "").strip()
    raw_new_al = (request.form.get("periodo_al") or "").strip()

    if not (_is_valid_iso_date(raw_new_data_vers) and _is_valid_iso_date(raw_new_dal) and _is_valid_iso_date(raw_new_al)):
        flash("Periodo non valido: usa date valide (YYYY-MM-DD).", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    if not _same_month_iso(raw_new_dal, raw_new_al):
        flash("Il periodo di competenza deve essere all'interno dello stesso mese.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    if raw_new_al < raw_new_dal:
        flash("Periodo non valido: la data 'Al' non può essere precedente alla data 'Dal'.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    new_data_vers = _parse_date_iso(raw_new_data_vers)
    new_dal = _parse_date_iso(raw_new_dal)
    new_al = _parse_date_iso(raw_new_al)

    # Non permettere giorni già inclusi in altri versamenti (escludendo questo record, se ha ID)
    if record_id:
        try:
            covered_days, _overlaps = _overlap_days_with_existing_versamenti(
                store_code=str(store_code),
                dal_iso=new_dal,
                al_iso=new_al,
                exclude_id=str(record_id),
            )
            if covered_days:
                preview = ", ".join(covered_days[:6])
                if len(covered_days) > 6:
                    preview += f" (+{len(covered_days) - 6})"
                flash(f"Nel periodo selezionato ci sono giorni già versati: {preview}. Modifica il periodo.", "warning")
                return redirect(url_for("rendiconto.versamenti", ym=ym_norm))
        except Exception:
            # la UI fa comunque il check via /compute
            pass

    new_nome = (request.form.get("nome_cognome") or "").strip()
    new_tipo = (request.form.get("tipo_versamento") or "").strip()
    new_tess = (request.form.get("tessera") or "").strip()
    new_rif = (request.form.get("riferimento") or "").strip()
    new_val = (request.form.get("valore") or "").strip()

    if not (new_data_vers and new_dal and new_al and new_nome and new_tipo and new_val):
        flash("Compila i campi obbligatori prima di salvare.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    tessera_digits = "".join(ch for ch in new_tess if ch.isdigit())
    if tessera_digits and len(tessera_digits) > 16:
        flash("Tessera: massimo 16 cifre.", "warning")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    # Blocco: aggiornamento consentito solo se la differenza è zero.
    try:
        valore_dec = _round2(_money_to_decimal(new_val))
        dist = Decimal(str(sum_categoria_period(store_code=str(store_code), start_iso=new_dal, end_iso=new_al, categoria="Distinte")))
        diff = _round2(valore_dec - _round2(dist))
    except Exception as e:
        current_app.logger.exception("Errore verifica differenza versamento (update)")
        flash(f"Errore calcolo differenza: {e}", "danger")
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    if diff != Decimal("0.00"):
        flash(
            f"Differenza diversa da zero ({float(diff):.2f} €). Correggi le distinte del periodo prima di salvare.",
            "warning",
        )
        return redirect(url_for("rendiconto.versamenti", ym=ym_norm))

    orig_data = (request.form.get("orig_data_vers") or "").strip()
    orig_dal = (request.form.get("orig_dal") or "").strip()
    orig_al = (request.form.get("orig_al") or "").strip()
    orig_nome = (request.form.get("orig_nome") or "").strip()
    orig_tipo = (request.form.get("orig_tipo") or "").strip()
    orig_tess = (request.form.get("orig_tessera") or "").strip()
    orig_rif = (request.form.get("orig_riferimento") or "").strip()
    orig_val = (request.form.get("orig_valore") or "").strip()

    # Se non abbiamo ID (tabella senza colonna ID), proviamo a escludere il record tramite firma dei campi originali
    if not record_id:
        try:
            exclude_sig = {
                "data_versamento_iso": orig_data,
                "dal_iso": orig_dal,
                "al_iso": orig_al,
                "valore_key": orig_val,
                "nome_raw": orig_nome,
                "tipo_raw": orig_tipo,
            }
            covered_days, _overlaps = _overlap_days_with_existing_versamenti(
                store_code=str(store_code),
                dal_iso=new_dal,
                al_iso=new_al,
                exclude_signature=exclude_sig,
            )
            if covered_days:
                preview = ", ".join(covered_days[:6])
                if len(covered_days) > 6:
                    preview += f" (+{len(covered_days) - 6})"
                flash(f"Nel periodo selezionato ci sono giorni già versati: {preview}. Modifica il periodo.", "warning")
                return redirect(url_for("rendiconto.versamenti", ym=ym_norm))
        except Exception:
            pass

    old_foto = None
    try:
        old_foto = get_versamento_photo_file(
            store_code=str(store_code),
            record_id=str(record_id or ""),
            orig_data_vers_iso=orig_data,
            orig_dal_iso=orig_dal,
            orig_al_iso=orig_al,
            orig_nome=orig_nome,
            orig_tipo=orig_tipo,
            orig_tessera=orig_tess,
            orig_riferimento=orig_rif,
            orig_valore_key=orig_val,
        )
    except Exception:
        old_foto = None

    new_foto_file = None
    foto = request.files.get("foto")
    if foto is not None and getattr(foto, "filename", ""):
        try:
            new_foto_file = upload_versamento_photo(
                sb_jwt=str(session.get("sb_token") or ""),
                user_id=str(session.get("uid") or ""),
                store_code=str(store_code),
                file_storage=foto,
                data_iso=new_data_vers,
            )
        except Exception as e:
            current_app.logger.exception("Errore upload foto versamento (update)")
            flash(f"Foto: upload fallito ({e}). Il versamento verrà aggiornato senza cambiare la foto.", "warning")
            new_foto_file = None

    try:
        n = update_versamento(
            store_code=str(store_code),
            record_id=record_id,
            orig_data_vers_iso=orig_data,
            orig_dal_iso=orig_dal,
            orig_al_iso=orig_al,
            orig_nome=orig_nome,
            orig_tipo=orig_tipo,
            orig_tessera=orig_tess,
            orig_riferimento=orig_rif,
            orig_valore_key=orig_val,
            new_data_vers_iso=new_data_vers,
            new_dal_iso=new_dal,
            new_al_iso=new_al,
            new_nome=new_nome,
            new_tipo=new_tipo,
            new_tessera=tessera_digits,
            new_riferimento=new_rif,
            new_valore_euro=new_val,
            new_foto_file=new_foto_file,
        )
        if n > 0:
            flash("Versamento aggiornato.", "success")
            if new_foto_file and old_foto and old_foto != new_foto_file:
                try:
                    delete_versamento_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(old_foto),
                    )
                except Exception as e:
                    current_app.logger.warning("Delete vecchia foto versamento fallita: %s", e)
        else:
            flash("Nessuna riga aggiornata (record non trovato).", "warning")
            if new_foto_file:
                try:
                    delete_versamento_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=str(new_foto_file),
                    )
                except Exception:
                    pass
    except Exception as e:
        current_app.logger.exception("Errore aggiornamento versamento")
        if new_foto_file:
            try:
                delete_versamento_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    filename=str(new_foto_file),
                )
            except Exception:
                pass
        flash(f"Errore aggiornamento versamento: {e}", "danger")

    return redirect(url_for("rendiconto.versamenti", ym=ym_norm))


@rendiconto_bp.get("/versamenti/photo/<path:filename>")
def versamenti_photo(filename: str):
    """Serve la foto collegata ad un versamento (stream dal repository SharePoint)."""
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    if not store_code:
        abort(404)

    fname = (filename or "").strip()
    if "/" in fname or "\\" in fname:
        abort(404)

    try:
        content = download_versamento_photo(
            sb_jwt=str(session.get("sb_token") or ""),
            user_id=str(session.get("uid") or ""),
            store_code=str(store_code),
            filename=fname,
        )
        return send_file(
            io.BytesIO(content),
            mimetype="image/jpeg",
            download_name=fname,
            as_attachment=False,
            max_age=300,
        )
    except Exception as e:
        current_app.logger.warning("Foto versamento non disponibile: %s", e)
        abort(404)


@rendiconto_bp.get("/api/versamenti/compute")
def api_versamenti_compute():
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    dal = (request.args.get("dal") or "").strip()
    al = (request.args.get("al") or "").strip()
    valore_raw = (request.args.get("valore") or "").strip()
    exclude_id = (request.args.get("exclude_id") or "").strip()
    # Quando si modifica un versamento, il controllo overlap deve ignorare quel versamento stesso.
    # Se la tabella non ha un ID, usiamo una "signature" basata sui valori originali del record.
    exclude_sig_data = (request.args.get("exclude_sig_data_vers") or "").strip()
    exclude_sig_dal = (request.args.get("exclude_sig_dal") or "").strip()
    exclude_sig_al = (request.args.get("exclude_sig_al") or "").strip()
    exclude_sig_val = (request.args.get("exclude_sig_valore") or "").strip()

    exclude_signature = None
    if exclude_sig_data and exclude_sig_dal and exclude_sig_al and exclude_sig_val:
        exclude_signature = {
            "data_versamento_iso": exclude_sig_data,
            "dal_iso": exclude_sig_dal,
            "al_iso": exclude_sig_al,
            "valore_key": exclude_sig_val,
        }


    if not dal or not al:
        return jsonify(error="Periodo non valido"), 400

    if not (_is_valid_iso_date(dal) and _is_valid_iso_date(al)):
        return jsonify(error="Periodo non valido"), 400

    if not _same_month_iso(dal, al):
        return jsonify(error="Il periodo di competenza deve essere all'interno dello stesso mese."), 400

    if al < dal:
        return jsonify(error="Periodo non valido: la data 'Al' non può essere precedente alla data 'Dal'."), 400

    dal_iso = _parse_date_iso(dal)
    al_iso = _parse_date_iso(al)
    valore = _parse_float(valore_raw)

    # Non permettere di selezionare giorni già inclusi in altri versamenti.
    try:
        overlap_days, overlap_summaries = _overlap_days_with_existing_versamenti(
            store_code=str(store_code),
            dal_iso=str(dal_iso),
            al_iso=str(al_iso),
            exclude_id=exclude_id or None,
            exclude_signature=exclude_signature,
        )
        if overlap_days:
            preview_days = ", ".join(overlap_days[:6])
            if len(overlap_days) > 6:
                preview_days += f" (+{len(overlap_days) - 6})"
            msg = f"Nel periodo selezionato ci sono giorni già versati: {preview_days}. Modifica il periodo."
            return (
                jsonify(
                    {
                        "error": msg,
                        "error_code": "period_overlaps_existing",
                        "days": overlap_days,
                        "overlaps": overlap_summaries,
                    }
                ),
                409,
            )
    except Exception:
        # se fallisce la verifica overlap, non blocchiamo l'utente: continuerà il calcolo
        pass

    try:
        dist = float(
            sum_categoria_period(
                store_code=str(store_code),
                start_iso=dal_iso,
                end_iso=al_iso,
                categoria="Distinte",
            )
        )
    except Exception as e:
        current_app.logger.exception("Errore somma distinte periodo")
        msg = str(e or "")
        low = msg.lower()
        if ("hy000" in msg) or ("driver did not supply" in low):
            return jsonify(error="Database occupato o file in uso. Riprova.", error_code="db_busy"), 503
        return jsonify(error=msg), 500

    diff = float(valore) - dist
    return jsonify({"distinte": dist, "diff": diff})


_DISTINTE_TAGLI = [5, 10, 20, 50, 100, 200, 500]


def _distinte_payload_from_rows(rows: list[dict]) -> dict:
    d1_map = {t: 0 for t in _DISTINTE_TAGLI}
    d2_map = {t: 0 for t in _DISTINTE_TAGLI}
    d1_mon = 0.0
    d2_mon = 0.0
    total = 0.0

    for r in rows or []:
        voce = str(r.get("voce") or "").strip()
        try:
            val_f = float(r.get("valore") or 0.0)
        except Exception:
            val_f = 0.0
        total += val_f

        v_up = voce.upper()
        if v_up.startswith("D1|MONETE"):
            d1_mon = float(val_f)
            continue
        if v_up.startswith("D2|MONETE"):
            d2_mon = float(val_f)
            continue

        # D1|TAGLIO=5|QTA=2
        if "TAGLIO=" in v_up and "QTA=" in v_up:
            try:
                which = "D1" if v_up.startswith("D1|") else "D2" if v_up.startswith("D2|") else ""
                if not which:
                    continue
                parts = voce.split("|")
                taglio = None
                qta = None
                for p in parts:
                    p2 = p.strip()
                    if p2.upper().startswith("TAGLIO="):
                        taglio = float(p2.split("=", 1)[1].replace(",", "."))
                    elif p2.upper().startswith("QTA="):
                        qta = int(float(p2.split("=", 1)[1]))
                if taglio is None or qta is None:
                    continue
                taglio_i = int(round(taglio))
                if taglio_i not in _DISTINTE_TAGLI:
                    continue
                if which == "D1":
                    d1_map[taglio_i] = int(qta)
                else:
                    d2_map[taglio_i] = int(qta)
            except Exception:
                continue

    return {
        "distinte1": [{"taglio": t, "qta": int(d1_map.get(t, 0))} for t in _DISTINTE_TAGLI],
        "distinte2": [{"taglio": t, "qta": int(d2_map.get(t, 0))} for t in _DISTINTE_TAGLI],
        "d1_monete": d1_mon,
        "d2_monete": d2_mon,
        "total": float(total),
    }


@rendiconto_bp.get("/api/versamenti/days")
def api_versamenti_days():
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    dal = (request.args.get("dal") or "").strip()
    al = (request.args.get("al") or "").strip()
    valore_raw = (request.args.get("valore") or "").strip()

    if not dal or not al:
        return jsonify(error="Periodo non valido"), 400

    if not (_is_valid_iso_date(dal) and _is_valid_iso_date(al)):
        return jsonify(error="Periodo non valido"), 400

    if not _same_month_iso(dal, al):
        return jsonify(error="Il periodo di competenza deve essere all'interno dello stesso mese."), 400

    if al < dal:
        return jsonify(error="Periodo non valido: la data 'Al' non può essere precedente alla data 'Dal'."), 400

    dal_iso = _parse_date_iso(dal)
    al_iso = _parse_date_iso(al)
    valore_dec = _money_to_decimal(valore_raw)

    exclude = {
        "id": (request.args.get("exclude_id") or "").strip(),
        "orig_data_vers": (request.args.get("exclude_orig_data_vers") or "").strip(),
        "orig_dal": (request.args.get("exclude_orig_dal") or "").strip(),
        "orig_al": (request.args.get("exclude_orig_al") or "").strip(),
        "orig_nome": (request.args.get("exclude_orig_nome") or "").strip(),
        "orig_tipo": (request.args.get("exclude_orig_tipo") or "").strip(),
        "orig_tessera": (request.args.get("exclude_orig_tessera") or "").strip(),
        "orig_riferimento": (request.args.get("exclude_orig_riferimento") or "").strip(),
        "orig_valore": (request.args.get("exclude_orig_valore") or "").strip(),
    }

    # Totali distinte per giorno nel range
    try:
        by_day = sum_categoria_by_day_range(
            store_code=str(store_code),
            start_iso=dal_iso,
            end_iso=al_iso,
            categoria="Distinte",
        )
    except Exception as e:
        current_app.logger.exception("Errore somma distinte per giorno")
        return jsonify(error=str(e)), 500

    # Giorni lockati da altri versamenti
    locked = _locked_days_in_range(store_code=str(store_code), start_iso=dal_iso, end_iso=al_iso, exclude=exclude)

    days = []
    total_dist = Decimal("0")
    for day_iso in _daterange_iso(dal_iso, al_iso):
        tot = float(by_day.get(day_iso) or 0.0)
        total_dist += Decimal(str(tot))
        # Nota: manteniamo sia "total" (atteso dal popup) sia "distinte" (retro-compatibilità).
        days.append({
            "date": day_iso,
            "total": float(tot),
            "distinte": float(tot),
            "locked": day_iso in locked,
        })

    diff = _round2(valore_dec - _round2(total_dist))
    return jsonify({"days": days, "distinte_periodo": float(_round2(total_dist)), "diff": float(diff)})


@rendiconto_bp.get("/api/distinte/day")
def api_distinte_day():
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    d = (request.args.get("d") or "").strip()
    if not d:
        return jsonify(error="Data non valida"), 400
    d_iso = _parse_date_iso(d)

    try:
        rows = load_primanota_day(store_code=str(store_code), data_iso=d_iso, categories=["Distinte"])
    except Exception as e:
        current_app.logger.exception("Errore lettura distinte giorno")
        return jsonify(error=str(e)), 500

    payload = _distinte_payload_from_rows(rows)
    payload["d"] = d_iso
    return jsonify(payload)


@rendiconto_bp.post("/api/distinte/day/save")
def api_distinte_day_save():
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    try:
        data = request.get_json(force=True, silent=False) or {}
    except Exception:
        return jsonify(error="Payload non valido"), 400

    d_iso = _parse_date_iso(str(data.get("d") or ""))
    exclude = data.get("exclude") or {}

    # lock: giornata inclusa in altri versamenti (escludendo, se serve, quello in modifica)
    locked = d_iso in _locked_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso, exclude=exclude)
    if locked:
        return jsonify(error="Giornata bloccata da un versamento: non puoi modificare le distinte."), 409

    def _to_items(x):
        return x if isinstance(x, list) else []

    distinte1 = _to_items(data.get("distinte1"))
    distinte2 = _to_items(data.get("distinte2"))

    d1_mon = _parse_float(str(data.get("d1_monete") or ""))
    d2_mon = _parse_float(str(data.get("d2_monete") or ""))

    entries: List[Dict[str, Any]] = []

    def _add_distinte(which: str, items: List[Dict[str, Any]]):
        for it in items or []:
            try:
                taglio = float(it.get("taglio"))
                qta = int(it.get("qta"))
            except Exception:
                continue
            if taglio <= 0 or qta < 0:
                continue
            if qta == 0:
                continue
            voce = f"{which}|TAGLIO={taglio}|QTA={qta}"
            entries.append({"categoria": "Distinte", "voce": voce, "tipo": "SI", "valore": float(taglio * qta)})

    _add_distinte("D1", distinte1)
    _add_distinte("D2", distinte2)

    if d1_mon != 0.0 or str(data.get("d1_monete") or "").strip() != "":
        entries.append({"categoria": "Distinte", "voce": "D1|MONETE", "tipo": "SI", "valore": float(d1_mon)})
    if d2_mon != 0.0 or str(data.get("d2_monete") or "").strip() != "":
        entries.append({"categoria": "Distinte", "voce": "D2|MONETE", "tipo": "SI", "valore": float(d2_mon)})

    try:
        replace_primanota_day(store_code=str(store_code), data_iso=d_iso, entries=entries, categories=["Distinte"])
    except Exception as e:
        current_app.logger.exception("Errore salvataggio distinte giorno")
        return jsonify(error=str(e)), 500

    # ritorna nuovo totale giorno
    try:
        rows2 = load_primanota_day(store_code=str(store_code), data_iso=d_iso, categories=["Distinte"])
        payload = _distinte_payload_from_rows(rows2)
        return jsonify({"ok": True, "d": d_iso, "total": payload.get("total", 0.0)})
    except Exception:
        return jsonify({"ok": True, "d": d_iso, "total": 0.0})


@rendiconto_bp.post("/api/versamenti/commit")
def api_versamenti_commit():
    """Salva un versamento SOLO se la differenza è zero (dopo eventuale correzione distinte)."""
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    mode = (request.form.get("mode") or "insert").strip().lower()
    ym = (request.form.get("ym") or request.args.get("ym") or "").strip()
    _y, _m, ym_norm = _parse_ym(ym)

    data_versamento = (request.form.get("data_versamento") or "").strip()
    periodo_dal = (request.form.get("periodo_dal") or "").strip()
    periodo_al = (request.form.get("periodo_al") or "").strip()
    nome_cognome = (request.form.get("nome_cognome") or "").strip()
    tipo_versamento = (request.form.get("tipo_versamento") or "").strip()
    tessera = (request.form.get("tessera") or "").strip()
    riferimento = (request.form.get("riferimento") or "").strip()
    valore_raw = (request.form.get("valore") or "").strip()

    if not (data_versamento and periodo_dal and periodo_al and nome_cognome and tipo_versamento and valore_raw):
        return jsonify(error="Compila tutti i campi obbligatori."), 400

    if not (_is_valid_iso_date(data_versamento) and _is_valid_iso_date(periodo_dal) and _is_valid_iso_date(periodo_al)):
        return jsonify(error="Periodo non valido"), 400

    if not _same_month_iso(periodo_dal, periodo_al):
        return jsonify(error="Il periodo di competenza deve essere all'interno dello stesso mese."), 400

    if periodo_al < periodo_dal:
        return jsonify(error="Periodo non valido: la data 'Al' non può essere precedente alla data 'Dal'."), 400

    dv_iso = _parse_date_iso(data_versamento)
    dal_iso = _parse_date_iso(periodo_dal)
    al_iso = _parse_date_iso(periodo_al)
    valore_dec = _round2(_money_to_decimal(valore_raw))

    # Blocco: non permettere salvataggi su periodi già versati (anche tramite endpoint /api/versamenti/commit).
    # Nota: in update escludiamo il record corrente (se identificabile).
    try:
        ex_id = None
        ex_sig = None
        if mode == "update":
            _rid = (request.form.get("id") or "").strip()
            if _rid:
                ex_id = _rid
            else:
                _odv = (request.form.get("orig_data_vers") or "").strip()
                _odal = (request.form.get("orig_dal") or "").strip()
                _oal = (request.form.get("orig_al") or "").strip()
                _oval = (request.form.get("orig_valore") or "").strip()
                if _odv and _odal and _oal and _oval:
                    ex_sig = {
                        "data_versamento_iso": _odv,
                        "dal_iso": _odal,
                        "al_iso": _oal,
                        "valore_key": _oval,
                    }

        covered_days, _overlaps = _overlap_days_with_existing_versamenti(
            store_code=str(store_code),
            dal_iso=dal_iso,
            al_iso=al_iso,
            exclude_id=ex_id,
            exclude_signature=ex_sig,
        )
        if covered_days:
            preview = ", ".join(covered_days[:6])
            if len(covered_days) > 6:
                preview += f" (+{len(covered_days) - 6})"
            return jsonify(error=f"Nel periodo selezionato ci sono giorni già versati: {preview}. Modifica il periodo."), 409
    except Exception:
        current_app.logger.exception("Errore verifica giorni già versati (commit)")

    # Verifica differenza
    try:
        dist = Decimal(str(sum_categoria_period(store_code=str(store_code), start_iso=dal_iso, end_iso=al_iso, categoria="Distinte")))
    except Exception as e:
        current_app.logger.exception("Errore somma distinte periodo")
        return jsonify(error=str(e)), 500

    diff = _round2(valore_dec - _round2(dist))
    if diff != Decimal("0.00"):
        return jsonify({"needs_adjustment": True, "diff": float(diff), "distinte_periodo": float(_round2(dist))}), 409

    foto_file = None
    foto = request.files.get("foto")

    if mode == "update":
        record_id = (request.form.get("id") or "").strip()

        orig_data_vers = (request.form.get("orig_data_vers") or "").strip()
        orig_dal = (request.form.get("orig_dal") or "").strip()
        orig_al = (request.form.get("orig_al") or "").strip()
        orig_nome = (request.form.get("orig_nome") or "").strip()
        orig_tipo = (request.form.get("orig_tipo") or "").strip()
        orig_tessera = (request.form.get("orig_tessera") or "").strip()
        orig_riferimento = (request.form.get("orig_riferimento") or "").strip()
        orig_valore = (request.form.get("orig_valore") or "").strip()

        old_foto = None
        try:
            old_foto = get_versamento_photo_file(
                store_code=str(store_code),
                record_id=record_id,
                orig_data_vers_iso=_parse_date_iso(orig_data_vers),
                orig_dal_iso=_parse_date_iso(orig_dal),
                orig_al_iso=_parse_date_iso(orig_al),
                orig_nome=orig_nome,
                orig_tipo=orig_tipo,
                orig_tessera=orig_tessera,
                orig_riferimento=orig_riferimento,
                orig_valore_key=orig_valore,
            )
        except Exception:
            old_foto = None

        new_foto_file = None
        if foto is not None and getattr(foto, "filename", ""):
            try:
                new_foto_file = upload_versamento_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    file_storage=foto,
                    data_iso=dv_iso,
                )
            except Exception as e:
                return jsonify(error=f"Errore upload foto versamento: {e}"), 500

        try:
            update_versamento(
                store_code=str(store_code),
                record_id=record_id,
                orig_data_vers_iso=_parse_date_iso(orig_data_vers),
                orig_dal_iso=_parse_date_iso(orig_dal),
                orig_al_iso=_parse_date_iso(orig_al),
                orig_nome=orig_nome,
                orig_tipo=orig_tipo,
                orig_tessera=orig_tessera,
                orig_riferimento=orig_riferimento,
                orig_valore_key=orig_valore,
                new_data_vers_iso=dv_iso,
                new_dal_iso=dal_iso,
                new_al_iso=al_iso,
                new_nome=nome_cognome,
                new_tipo=tipo_versamento,
                new_tessera=tessera,
                new_riferimento=riferimento,
                new_valore_euro=str(valore_dec),
                new_foto_file=new_foto_file,
            )
        except Exception as e:
            if new_foto_file:
                try:
                    delete_versamento_photo(
                        sb_jwt=str(session.get("sb_token") or ""),
                        user_id=str(session.get("uid") or ""),
                        store_code=str(store_code),
                        filename=new_foto_file,
                    )
                except Exception:
                    pass
            return jsonify(error=f"Errore aggiornamento versamento: {e}"), 500

        if new_foto_file and old_foto and old_foto != new_foto_file:
            try:
                delete_versamento_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    filename=old_foto,
                )
            except Exception:
                pass

        return jsonify({"ok": True, "ym": ym_norm})

    # INSERT
    if foto is not None and getattr(foto, "filename", ""):
        try:
            foto_file = upload_versamento_photo(
                sb_jwt=str(session.get("sb_token") or ""),
                user_id=str(session.get("uid") or ""),
                store_code=str(store_code),
                file_storage=foto,
                data_iso=dv_iso,
            )
        except Exception as e:
            return jsonify(error=f"Errore upload foto versamento: {e}"), 500

    try:
        insert_versamento(
            store_code=str(store_code),
            data_versamento_iso=dv_iso,
            periodo_dal_iso=dal_iso,
            periodo_al_iso=al_iso,
            nome_cognome=nome_cognome,
            tipo_versamento=tipo_versamento,
            tessera=tessera,
            riferimento=riferimento,
            valore_euro=str(valore_dec),
            foto_file=foto_file,
        )
    except Exception as e:
        if foto_file:
            try:
                delete_versamento_photo(
                    sb_jwt=str(session.get("sb_token") or ""),
                    user_id=str(session.get("uid") or ""),
                    store_code=str(store_code),
                    filename=foto_file,
                )
            except Exception:
                pass
        return jsonify(error=f"Errore salvataggio versamento: {e}"), 500

    return jsonify({"ok": True, "ym": ym_norm})


# -------------------------
# Distinta cassa (Prima Nota)
# -------------------------


_CHIUSURA_FIELDS = [
    ("vendite_lorde", "VENDITE LORDE", "money"),
    ("fatture", "FATTURE", "money"),
    ("numero_fatture", "NUMERO FATTURE", "int"),
    # SCONTRINI = conteggio (intero), non valuta
    ("scontrini", "SCONTRINI", "int"),
    ("annullati", "ANNULLATI", "money"),
    ("omaggi", "OMAGGI", "money"),
    ("vendite_iva_4", "VENDITE IVA 4%", "money"),
    ("vendite_iva_22", "VENDITE IVA 22%", "money"),
    ("contanti", "CONTANTI", "money"),
    ("pos", "POS", "money"),
]


_VOICE_TO_KEY = {lbl: key for (key, lbl, _t) in _CHIUSURA_FIELDS}
_KEY_TO_TYPE = {key: t for (key, _lbl, t) in _CHIUSURA_FIELDS}


@rendiconto_bp.get("/distinta-cassa")
def distinta_cassa():
    _ensure_session_keys()
    store_code = session.get("store_code")
    store_name = session.get("store_name")

    d_iso = _parse_date_iso(request.args.get("d") or "")

    # Se la giornata è già inclusa nel periodo competenza di un versamento,
    # blocchiamo la modifica delle distinte contanti (tagli/monete).
    locked_distinte = False
    if store_code:
        locked_set = _locked_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso)
        locked_distinte = d_iso in locked_set

    options = {"tickets": [], "deliveries": [], "coupons": []}
    existing_rows: List[Dict[str, Any]] = []
    spese_info = {"total": 0.0, "note_credito": 0.0, "net": 0.0}

    init = {
        "chiusura": {k: "" for (k, _lbl, _t) in _CHIUSURA_FIELDS},
        "distinte1": [],
        "distinte2": [],
        "d1_monete": "",
        "d2_monete": "",
        "tickets": [],
        "deliveries": [],
        "coupons": [],
    }

    if store_code:
        try:
            options = get_elenchi_options(store_code=str(store_code))
        except Exception as e:
            current_app.logger.exception("Errore lettura ELENCHI")
            flash(f"Errore lettura ELENCHI: {e}", "danger")

        try:
            existing_rows = load_primanota_day(store_code=str(store_code), data_iso=d_iso)
        except Exception as e:
            current_app.logger.exception("Errore lettura DATIPRIMANOTA")
            flash(f"Errore lettura DATIPRIMANOTA: {e}", "danger")

        try:
            spese_info = sum_spese_day(store_code=str(store_code), data_iso=d_iso)
        except Exception as e:
            current_app.logger.exception("Errore calcolo spese giorno")
            flash(f"Errore calcolo spese giorno: {e}", "danger")

        # parse existing
        for r in existing_rows:
            cat = str(r.get("categoria") or "")
            voce = str(r.get("voce") or "")
            tipo = str(r.get("tipo") or "SI")
            val = r.get("valore")
            try:
                val_f = float(val)
            except Exception:
                val_f = 0.0

            if cat == "Dati chiusura":
                key = _VOICE_TO_KEY.get(voce)
                if key:
                    t = _KEY_TO_TYPE.get(key, "money")
                    if t == "int":
                        try:
                            init["chiusura"][key] = str(int(round(val_f)))
                        except Exception:
                            init["chiusura"][key] = "0"
                    else:
                        init["chiusura"][key] = str(val_f)

            elif cat == "Distinte":
                # Voce formato: D1|TAGLIO=20|QTA=3 oppure D1|MONETE
                v = voce.strip()
                if v.startswith("D1|MONETE"):
                    init["d1_monete"] = str(val_f)
                elif v.startswith("D2|MONETE"):
                    init["d2_monete"] = str(val_f)
                elif v.startswith("D1|TAGLIO="):
                    try:
                        parts = v.split("|")
                        taglio = float(parts[1].split("=", 1)[1])
                        qta = int(parts[2].split("=", 1)[1])
                        init["distinte1"].append({"taglio": taglio, "qta": qta})
                    except Exception:
                        pass
                elif v.startswith("D2|TAGLIO="):
                    try:
                        parts = v.split("|")
                        taglio = float(parts[1].split("=", 1)[1])
                        qta = int(parts[2].split("=", 1)[1])
                        init["distinte2"].append({"taglio": taglio, "qta": qta})
                    except Exception:
                        pass

            elif cat == "Ticket":
                init["tickets"].append({"voce": voce, "tipo": tipo, "valore": val_f})
            elif cat == "Delivery":
                init["deliveries"].append({"voce": voce, "tipo": tipo, "valore": val_f})
            elif cat == "Coupon":
                init["coupons"].append({"voce": voce, "tipo": tipo, "valore": val_f})

    has_saved_data = bool(existing_rows)

    return render_template(
        "rendiconto_distinta_cassa.html",
        store_code=store_code,
        store_name=store_name,
        d=d_iso,
        locked_distinte=locked_distinte,
        options=options,
        spese_info=spese_info,
        init=init,
        has_saved_data=has_saved_data,
        chiusura_fields=_CHIUSURA_FIELDS,
    )


@rendiconto_bp.post("/distinta-cassa/save")
def distinta_cassa_save():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.distinta_cassa"))

    d_iso = _parse_date_iso(request.form.get("d") or "")

    # campi chiusura
    required_keys = {"vendite_lorde", "scontrini", "contanti", "pos"}
    chiusura_vals: Dict[str, float] = {}
    for (key, _lbl, _t) in _CHIUSURA_FIELDS:
        raw = (request.form.get(key) or "").strip()
        if key in required_keys and raw == "":
            flash("Compila i campi obbligatori: Vendite lorde, Scontrini, Contanti, POS.", "warning")
            return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))
        if raw == "":
            chiusura_vals[key] = 0.0
            continue

        # parsing coerente: money -> float, int -> intero
        if _KEY_TO_TYPE.get(key) == "int":
            try:
                chiusura_vals[key] = float(int(round(_parse_float(raw))))
            except Exception:
                chiusura_vals[key] = 0.0
        else:
            chiusura_vals[key] = _parse_float(raw)

    # json lists
    def _load_json(name: str):
        s = (request.form.get(name) or "").strip()
        if not s:
            return []
        try:
            v = json.loads(s)
            return v if isinstance(v, list) else []
        except Exception:
            return []

    distinte1 = _load_json("distinte1_json")
    distinte2 = _load_json("distinte2_json")
    tickets = _load_json("tickets_json")
    deliveries = _load_json("deliveries_json")
    coupons = _load_json("coupons_json")

    d1_monete = _parse_float(request.form.get("d1_monete") or "")
    d2_monete = _parse_float(request.form.get("d2_monete") or "")

    # opzioni per tipo SI/NO (per sicurezza lato server)
    try:
        opt = get_elenchi_options(store_code=str(store_code))
    except Exception:
        opt = {"tickets": [], "deliveries": [], "coupons": []}

    ticket_tipo_map = {o.get("value"): (o.get("tipo") or "SI") for o in (opt.get("tickets") or [])}
    delivery_tipo_map = {o.get("value"): (o.get("tipo") or "SI") for o in (opt.get("deliveries") or [])}
    coupon_tipo_map = {o.get("value"): (o.get("tipo") or "SI") for o in (opt.get("coupons") or [])}

    entries: List[Dict[str, Any]] = []

    # Dati chiusura (solo manuali)
    for (key, lbl, _t) in _CHIUSURA_FIELDS:
        entries.append(
            {
                "categoria": "Dati chiusura",
                "voce": lbl,
                "tipo": "SI",
                "valore": float(chiusura_vals.get(key, 0.0)),
            }
        )

    # Distinte (contanti). Se la giornata è inclusa in un periodo competenza di un versamento,
    # NON consentiamo modifiche su tagli/monete e preserviamo quanto già salvato.
    locked_distinte = d_iso in _locked_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso)

    if locked_distinte:
        try:
            existing_dist = load_primanota_day(store_code=str(store_code), data_iso=d_iso, categories=["Distinte"])
        except Exception:
            existing_dist = []
        for e in existing_dist:
            entries.append(
                {
                    "categoria": "Distinte",
                    "voce": str(e.get("voce") or ""),
                    "tipo": str(e.get("tipo") or "SI"),
                    "valore": float(e.get("valore") or 0.0),
                }
            )
    else:
        def _add_distinte(which: str, items: List[Dict[str, Any]]):
            for it in items or []:
                try:
                    taglio = float(it.get("taglio"))
                    qta = int(it.get("qta"))
                except Exception:
                    continue
                if taglio <= 0 or qta <= 0:
                    continue
                voce = f"{which}|TAGLIO={taglio}|QTA={qta}"
                entries.append(
                    {
                        "categoria": "Distinte",
                        "voce": voce,
                        "tipo": "SI",
                        "valore": float(taglio * qta),
                    }
                )

        _add_distinte("D1", distinte1)
        _add_distinte("D2", distinte2)

        if d1_monete != 0.0 or (request.form.get("d1_monete") or "").strip() != "":
            entries.append({"categoria": "Distinte", "voce": "D1|MONETE", "tipo": "SI", "valore": float(d1_monete)})
        if d2_monete != 0.0 or (request.form.get("d2_monete") or "").strip() != "":
            entries.append({"categoria": "Distinte", "voce": "D2|MONETE", "tipo": "SI", "valore": float(d2_monete)})

    # Ticket / Delivery / Coupon
    def _add_list(cat: str, items: List[Dict[str, Any]], tipo_map: Dict[str, str]):
        for it in items or []:
            voce = str(it.get("voce") or "").strip()
            if not voce:
                continue
            val = it.get("valore")
            try:
                val_f = float(val)
            except Exception:
                val_f = 0.0
            tipo = tipo_map.get(voce, str(it.get("tipo") or "SI"))
            entries.append({"categoria": cat, "voce": voce, "tipo": tipo, "valore": float(val_f)})

    _add_list("Ticket", tickets, ticket_tipo_map)
    _add_list("Delivery", deliveries, delivery_tipo_map)
    _add_list("Coupon", coupons, coupon_tipo_map)

    # Calcoli per scrittura DatiDatabase
    giro_affari = float(chiusura_vals.get("vendite_lorde", 0.0)) - float(chiusura_vals.get("annullati", 0.0))
    totale_delivery_si = 0.0
    for e in entries:
        if str(e.get("categoria") or "") == "Delivery" and str(e.get("tipo") or "").strip().upper() == "SI":
            try:
                totale_delivery_si += float(e.get("valore") or 0.0)
            except Exception:
                pass
    try:
        scontrini_int = int(round(float(chiusura_vals.get("scontrini", 0.0) or 0.0)))
    except Exception:
        scontrini_int = 0

    try:
        replace_primanota_day(store_code=str(store_code), data_iso=d_iso, entries=entries)
        # Scrittura aggiuntiva su DatiDatabase (non deve bloccare la Distinta)
        try:
            upsert_datidatabase_from_distinta(
                store_code=str(store_code),
                data_iso=d_iso,
                giro_affari=giro_affari,
                totale_delivery=totale_delivery_si,
                scontrini=scontrini_int,
            )
        except Exception as ex2:
            current_app.logger.exception("Errore scrittura DatiDatabase da Distinta")
            flash(f"Distinta salvata, ma errore scrittura DatiDatabase: {ex2}", "warning")

        flash("Distinta cassa salvata.", "success")
    except Exception as e:
        current_app.logger.exception("Errore salvataggio Distinta cassa")
        flash(f"Errore salvataggio Distinta cassa: {e}", "danger")

    return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))




@rendiconto_bp.post("/distinta-cassa/delete")
def distinta_cassa_delete():
    _ensure_session_keys()
    r = _require_login()
    if r is not None:
        return r

    store_code = session.get("store_code")
    if not store_code:
        flash("Seleziona prima uno store.", "warning")
        return redirect(url_for("rendiconto.distinta_cassa"))

    d_iso = _parse_date_iso(request.form.get("d") or "")

    # Cancellazione non consentita se la giornata è inclusa nel periodo competenza di un versamento
    locked = d_iso in _locked_days_in_range(store_code=str(store_code), start_iso=d_iso, end_iso=d_iso)
    if locked:
        flash("Giornata bloccata da un versamento: non puoi eliminare la distinta di cassa.", "warning")
        return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))

    try:
        delete_primanota_day(
            store_code=str(store_code),
            data_iso=d_iso,
            categories=["Dati chiusura", "Distinte", "Ticket", "Delivery", "Coupon"],
        )

        # Cancellazione aggiuntiva su DatiDatabase (se presente)
        try:
            res = delete_datidatabase_day(store_code=str(store_code), data_iso=d_iso)
            if not res.get("ok"):
                raise Exception(res.get("error") or "Errore cancellazione DatiDatabase")
        except Exception as ex2:
            current_app.logger.exception("Errore cancellazione DatiDatabase da Distinta")
            flash(f"Distinta eliminata, ma errore cancellazione DatiDatabase: {ex2}", "warning")

        flash("Distinta cassa eliminata.", "success")
    except Exception as e:
        current_app.logger.exception("Errore eliminazione Distinta cassa")
        flash(f"Errore eliminazione Distinta cassa: {e}", "danger")

    return redirect(url_for("rendiconto.distinta_cassa", d=d_iso))


# -------------------------
# Dashboard Rendiconto (API)
# -------------------------


@rendiconto_bp.get("/api/dashboard/month")
def api_dashboard_month():
    """Ritorna mappa giorni del mese con {giro, diff} e summary per la dashboard, più lo stato versamenti."""
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    try:
        year = int(request.args.get("year") or 0)
        month = int(request.args.get("month") or 0)
        if year < 2000 or not (1 <= month <= 12):
            raise ValueError
    except Exception:
        today = _date.today()
        year, month = today.year, today.month

    try:
        prim_rows = load_primanota_month_agg(str(store_code), year=year, month=month)
    except Exception as e:
        current_app.logger.exception("Errore lettura DATIPRIMANOTA (month agg)")
        return jsonify(error=f"Errore lettura prima nota: {e}"), 500

    try:
        spese_by_day = sum_spese_month_by_day(store_code=str(store_code), year=year, month=month)
    except Exception:
        current_app.logger.exception("Errore lettura SPESE (month agg)")
        spese_by_day = {}

    # Aggregazione per giorno
    by_day = {}
    for r in prim_rows or []:
        d_iso = str(r.get("date") or "").strip()
        if not d_iso:
            continue
        cat = str(r.get("categoria") or "").strip()
        voce = str(r.get("voce") or "").strip()
        tipo = str(r.get("tipo") or "SI").strip().upper()
        try:
            s = float(r.get("sum") or 0)
        except Exception:
            s = 0.0

        agg = by_day.setdefault(
            d_iso,
            {
                "vendite_lorde": 0.0,
                "annullati": 0.0,
                "pos": 0.0,
                "scontrini": 0.0,
                "distinte": 0.0,
                "ticket_si": 0.0,
                "delivery_si": 0.0,
                "coupon_si": 0.0,
            },
        )

        if cat == "Dati chiusura":
            key = _VOICE_TO_KEY.get(voce)
            if key in {"vendite_lorde", "annullati", "pos", "scontrini"}:
                agg[key] += s
        elif cat == "Distinte":
            agg["distinte"] += s
        elif cat == "Ticket":
            if tipo == "SI":
                agg["ticket_si"] += s
        elif cat == "Delivery":
            if tipo == "SI":
                agg["delivery_si"] += s
        elif cat == "Coupon":
            if tipo == "SI":
                agg["coupon_si"] += s

    days = {}
    for d_iso, a in by_day.items():
        giro = float(a.get("vendite_lorde", 0.0)) - float(a.get("annullati", 0.0))
        spnet = float((spese_by_day.get(d_iso) or {}).get("net") or 0.0)
        diff = (
            float(a.get("distinte", 0.0))
            + float(a.get("ticket_si", 0.0))
            + float(a.get("delivery_si", 0.0))
            + float(a.get("coupon_si", 0.0))
            + float(a.get("pos", 0.0))
            + spnet
            - giro
        )
        days[d_iso] = {"giro": giro, "diff": diff}

    # Totali mese (riepilogo)
    tot_vendite_lorde = 0.0
    tot_annullati = 0.0
    tot_pos = 0.0
    tot_scontrini = 0.0
    tot_distinte = 0.0
    tot_ticket_si = 0.0
    tot_delivery_si = 0.0
    tot_coupon_si = 0.0

    for _d_iso, a in by_day.items():
        tot_vendite_lorde += float(a.get("vendite_lorde", 0.0))
        tot_annullati += float(a.get("annullati", 0.0))
        tot_pos += float(a.get("pos", 0.0))
        tot_scontrini += float(a.get("scontrini", 0.0))
        tot_distinte += float(a.get("distinte", 0.0))
        tot_ticket_si += float(a.get("ticket_si", 0.0))
        tot_delivery_si += float(a.get("delivery_si", 0.0))
        tot_coupon_si += float(a.get("coupon_si", 0.0))

    tot_spese_net = 0.0
    for _d_iso, srec in (spese_by_day or {}).items():
        try:
            tot_spese_net += float((srec or {}).get("net") or 0.0)
        except Exception:
            pass

    giro_mese = float(tot_vendite_lorde) - float(tot_annullati)
    diff_mese = (
        float(tot_distinte)
        + float(tot_ticket_si)
        + float(tot_delivery_si)
        + float(tot_coupon_si)
        + float(tot_pos)
        + float(tot_spese_net)
        - float(giro_mese)
    )

    try:
        scontrini_mese = int(round(float(tot_scontrini)))
    except Exception:
        scontrini_mese = 0

    summary = {
        "giro": giro_mese,
        "scontrini": scontrini_mese,
        "pos": float(tot_pos),
        "distinte": float(tot_distinte),
        "annullati": float(tot_annullati),
        "diff": float(diff_mese),
    }

        # ---- Versamenti: stato annuale (Distinte anno vs Versato) ----
    versamenti_status = {
        "ultimo_al_iso": "",
        "ultimo_al_disp": "",
        "giorni_da_ultimo": None,
        "distinte_non_versate": 0.0,  # usato in UI come "totale da versare"
        "giorni_non_versati": 0,
        "is_fragmented": False,
        "should_alert": False,
        # extra (non obbligatori per la UI attuale)
        "totale_distinte_anno": 0.0,
        "totale_versato_anno": 0.0,
        "totale_da_versare": 0.0,
    }

    try:
        today = _date.today()
        year_start = _date(year, 1, 1)
        year_end = _date(year, 12, 31)
        upper = min(today, year_end)

        # Totale distinte nell'anno (YTD)
        tot_distinte_anno = sum_categoria_period(
            store_code=str(store_code),
            start_iso=year_start.isoformat(),
            end_iso=upper.isoformat(),
            categoria="Distinte",
        )

        # Totale versato nell'anno: somma dei versamenti (campo valore) con data "AL" nel range
        vres = list_versamenti_periods_overlapping(
            store_code=str(store_code),
            start_iso=year_start.isoformat(),
            end_iso=upper.isoformat(),
        )
        vrows = (vres or {}).get("rows") or []

        total_versato = 0.0
        last_al = None

        for r in vrows:
            r = r or {}
            al_iso = str(r.get("al_iso") or "").strip()
            d_al = None
            if _is_valid_iso_date(al_iso):
                try:
                    d_al = _date.fromisoformat(al_iso)
                except Exception:
                    d_al = None

            # consideriamo nel totale solo i versamenti con AL nel range dell'anno (<= upper)
            if d_al is None or d_al < year_start or d_al > upper:
                continue

            if (last_al is None) or (d_al > last_al):
                last_al = d_al

            try:
                v_key = r.get("valore_key")
                v_dec = v_key if isinstance(v_key, Decimal) else Decimal(str(v_key))
            except Exception:
                v_dec = Decimal("0")

            total_versato += float(v_dec)

        totale_da_versare = float(tot_distinte_anno) - float(total_versato)

        versamenti_status["distinte_non_versate"] = float(totale_da_versare)
        versamenti_status["totale_distinte_anno"] = float(tot_distinte_anno)
        versamenti_status["totale_versato_anno"] = float(total_versato)
        versamenti_status["totale_da_versare"] = float(totale_da_versare)

        if last_al:
            versamenti_status["ultimo_al_iso"] = last_al.isoformat()
            versamenti_status["ultimo_al_disp"] = last_al.strftime("%d/%m/%Y")

            giorni_da_ultimo = (today - last_al).days
            if giorni_da_ultimo < 0:
                giorni_da_ultimo = 0

            versamenti_status["giorni_da_ultimo"] = int(giorni_da_ultimo)
            versamenti_status["giorni_non_versati"] = int(giorni_da_ultimo)

            versamenti_status["should_alert"] = bool(
                giorni_da_ultimo > 7 and abs(totale_da_versare) > 1e-9
            )
    except Exception:
        current_app.logger.exception("Errore calcolo stato versamenti dashboard")

    return jsonify(
        {
            "year": year,
            "month": month,
            "days": days,
            "summary": summary,
            "versamenti_status": versamenti_status,
        }
    )


@rendiconto_bp.get("/api/dashboard/day")
def api_dashboard_day():
    """Ritorna dettaglio del giorno per il popup rendiconto in dashboard."""
    _ensure_session_keys()

    if not session.get("uid"):
        return jsonify(error="Non autenticato"), 401

    store_code = session.get("store_code")
    if not store_code:
        return jsonify(error="Seleziona prima uno store"), 400

    d_iso = _parse_date_iso(request.args.get("date") or "")

    try:
        rows = load_primanota_day(store_code=str(store_code), data_iso=d_iso)
    except Exception as e:
        current_app.logger.exception("Errore lettura DATIPRIMANOTA (day)")
        return jsonify(error=f"Errore lettura prima nota: {e}"), 500

    try:
        spese = sum_spese_day(store_code=str(store_code), data_iso=d_iso)
    except Exception:
        current_app.logger.exception("Errore calcolo spese giorno")
        spese = {"total": 0.0, "note_credito": 0.0, "net": 0.0}

    chiusura = {k: 0.0 for (k, _lbl, _t) in _CHIUSURA_FIELDS}
    distinte = 0.0
    ticket_si = 0.0
    delivery_si = 0.0
    delivery_no = 0.0
    coupon_si = 0.0

    for r in rows or []:
        cat = str(r.get("categoria") or "").strip()
        voce = str(r.get("voce") or "").strip()
        tipo = str(r.get("tipo") or "SI").strip().upper()
        try:
            val = float(r.get("valore") or 0)
        except Exception:
            val = 0.0

        if cat == "Dati chiusura":
            key = _VOICE_TO_KEY.get(voce)
            if key:
                chiusura[key] = float(chiusura.get(key, 0.0)) + val
        elif cat == "Distinte":
            distinte += val
        elif cat == "Ticket":
            if tipo == "SI":
                ticket_si += val
        elif cat == "Delivery":
            if tipo == "SI":
                delivery_si += val
            else:
                delivery_no += val
        elif cat == "Coupon":
            if tipo == "SI":
                coupon_si += val

    vendite_lorde = float(chiusura.get("vendite_lorde", 0.0))
    annullati = float(chiusura.get("annullati", 0.0))
    pos = float(chiusura.get("pos", 0.0))

    giro = vendite_lorde - annullati
    spnet = float(spese.get("net") or 0.0)
    diff = distinte + ticket_si + delivery_si + coupon_si + pos + spnet - giro

    chiusura_rows = []
    for (key, lbl, t) in _CHIUSURA_FIELDS:
        v = chiusura.get(key, 0.0)
        if t == "int":
            try:
                v = int(round(float(v)))
            except Exception:
                v = 0
        chiusura_rows.append({"key": key, "label": lbl, "type": t, "value": v})

    return jsonify(
        {
            "date": d_iso,
            "giro": giro,
            "diff": diff,
            "distinte": distinte,
            "ticket_si": ticket_si,
            "delivery_si": delivery_si,
            "delivery_no": delivery_no,
            "coupon_si": coupon_si,
            "spese_net": float(spese.get("net") or 0.0),
            "spese_total": float(spese.get("total") or 0.0),
            "note_credito": float(spese.get("note_credito") or 0.0),
            "chiusura_rows": chiusura_rows,
        }
    )
# ------------------------------------------------------------
# RIEPILOGO (multi-store, mensile) - sezione Rendiconto
# ------------------------------------------------------------

def _parse_month_yyyy_mm(value: str):
    """Parse YYYY-MM -> (year, month, start_date, end_date)."""
    raw = (value or "").strip()
    y = 0
    m = 0
    if raw and "-" in raw:
        try:
            p = raw.split("-", 1)
            y = int(p[0])
            m = int(p[1])
        except Exception:
            y = 0
            m = 0

    if y < 2000 or not (1 <= m <= 12):
        today = _date.today()
        y, m = today.year, today.month

    start = _date(y, m, 1)
    # last day of month
    if m == 12:
        end = _date(y + 1, 1, 1) - timedelta(days=1)
    else:
        end = _date(y, m + 1, 1) - timedelta(days=1)
    return y, m, start, end


def _available_stores_for_user(user_id: str | None):
    """Lista store visibili all'utente (stessa logica Magazzino)."""
    try:
        role = str(session.get("role") or "").lower()

        # refresh role ogni ~10 minuti (o se admin)
        try:
            if user_id:
                now = int(time.time())
                cached_for = session.get("role_verified_for")
                cached_at = int(session.get("role_verified_at") or 0)
                need_check = (role == "admin") or (cached_for != user_id) or ((now - cached_at) > 600)
                if need_check:
                    from db_integration import get_profile_role_by_id

                    srv_role = get_profile_role_by_id(user_id)
                    if srv_role:
                        role = str(srv_role).lower()
                        session["role"] = srv_role
                        session["role_verified_for"] = user_id
                        session["role_verified_at"] = now
        except Exception:
            pass

        if role == "admin":
            stores = get_warehouse_stores() or []
        else:
            assigned = []
            if user_id:
                try:
                    assigned = get_user_warehouse_stores(str(user_id))
                except Exception:
                    assigned = []

            all_stores = get_warehouse_stores() or []
            allowed_codes = {str(row.get("store_code")) for row in assigned if row.get("store_code")}

            if allowed_codes:
                stores = [s for s in all_stores if str((s or {}).get("code")) in allowed_codes]
            else:
                cur_code = session.get("store_code")
                if cur_code:
                    stores = [s for s in all_stores if str((s or {}).get("code")) == str(cur_code)]
                else:
                    stores = []
    except Exception:
        stores = []

    try:
        stores = sorted(
            stores or [],
            key=lambda s: (
                str((s or {}).get("name") or "").strip().lower(),
                str((s or {}).get("code") or "").strip().lower(),
            ),
        )
    except Exception:
        pass

    return stores


_CHIUSURA_VOICE_TO_KEY = {
    "VENDITE LORDE": "vendite_lorde",
    "ANNULLATI": "annullati",
    "POS": "pos",
}


@rendiconto_bp.get("/riepilogo")
def riepilogo():
    _ensure_session_keys()
    return render_template("rendiconto_riepilogo.html")


@rendiconto_bp.get("/api/riepilogo/mensile")
def api_riepilogo_mensile():
    """Riepilogo mensile multi-store (Rendiconto)."""
    _ensure_session_keys()

    user_id = session.get("uid")
    if not user_id:
        return jsonify({"ok": False, "error": "not_logged_in"}), 401

    month = (request.args.get("month") or "").strip()  # YYYY-MM
    y, m, start_d, end_d = _parse_month_yyyy_mm(month)
    month_norm = f"{y:04d}-{m:02d}"

    try:
        stores = _available_stores_for_user(str(user_id)) or []
    except Exception:
        stores = []

    rows = []
    warnings_all = []

    for s in stores or []:
        code = (s or {}).get("code")
        name = (s or {}).get("name")
        if not code:
            continue

        row = {
            "store_code": str(code),
            "store_name": str(name or ""),
            "giro_affari": 0.0,
            "diff_cassa": 0.0,
            "distinte": 0.0,
            "pos": 0.0,
            "spese": 0.0,
            "versamenti": 0.0,
            "ticket_si": 0.0,
            "ticket_no": 0.0,
            "delivery_si": 0.0,
            "delivery_no": 0.0,
            "coupon_si": 0.0,
            "coupon_no": 0.0,
            "warnings": [],
        }

        # --- Prima nota (aggregati mese) ---
        try:
            prim_rows = load_primanota_month_agg_totals(str(code), year=y, month=m)

            tot_vendite_lorde = 0.0
            tot_annullati = 0.0
            tot_pos = 0.0
            tot_distinte = 0.0
            t_si = 0.0
            t_no = 0.0
            d_si = 0.0
            d_no = 0.0
            c_si = 0.0
            c_no = 0.0

            for r in prim_rows or []:
                cat = str(r.get("categoria") or "").strip()
                voce = str(r.get("voce") or "").strip()
                tipo = str(r.get("tipo") or "SI").strip().upper()
                try:
                    s_val = float(r.get("sum") or 0.0)
                except Exception:
                    s_val = 0.0

                if cat == "Dati chiusura":
                    k = _CHIUSURA_VOICE_TO_KEY.get(voce.upper())
                    if k == "vendite_lorde":
                        tot_vendite_lorde += s_val
                    elif k == "annullati":
                        tot_annullati += s_val
                    elif k == "pos":
                        tot_pos += s_val
                elif cat == "Distinte":
                    tot_distinte += s_val
                elif cat == "Ticket":
                    if tipo == "SI":
                        t_si += s_val
                    else:
                        t_no += s_val
                elif cat == "Delivery":
                    if tipo == "SI":
                        d_si += s_val
                    else:
                        d_no += s_val
                elif cat == "Coupon":
                    if tipo == "SI":
                        c_si += s_val
                    else:
                        c_no += s_val

            row["giro_affari"] = float(tot_vendite_lorde) - float(tot_annullati)
            row["pos"] = float(tot_pos)
            row["distinte"] = float(tot_distinte)
            row["ticket_si"] = float(t_si)
            row["ticket_no"] = float(t_no)
            row["delivery_si"] = float(d_si)
            row["delivery_no"] = float(d_no)
            row["coupon_si"] = float(c_si)
            row["coupon_no"] = float(c_no)
        except Exception as e:
            current_app.logger.exception("Errore riepilogo rendiconto (primanota) store %s", code)
            warnings_all.append(f"[{code}] Prima nota: {e}")
            row["warnings"].append("Prima nota non disponibile")

        # --- Spese (totale mese) ---
        try:
            row["spese"] = float(sum_spese_month_total_net(store_code=str(code), year=y, month=m))
        except Exception as e:
            current_app.logger.exception("Errore riepilogo rendiconto (spese) store %s", code)
            warnings_all.append(f"[{code}] Spese: {e}")
            row["warnings"].append("Spese non disponibili")

        # --- Versamenti (totale mese) ---
        try:
            row["versamenti"] = float(sum_versamenti_month_total(store_code=str(code), year=y, month=m))
        except Exception as e:
            current_app.logger.exception("Errore riepilogo rendiconto (versamenti) store %s", code)
            warnings_all.append(f"[{code}] Versamenti: {e}")
            row["warnings"].append("Versamenti non disponibili")

        # --- Differenza di cassa (mese) ---
        # Coerente con il calcolo usato nella dashboard: distinte + ticket_si + delivery_si + coupon_si + pos + spese_net - giro
        try:
            row["diff_cassa"] = (
                float(row.get("distinte") or 0.0)
                + float(row.get("ticket_si") or 0.0)
                + float(row.get("delivery_si") or 0.0)
                + float(row.get("coupon_si") or 0.0)
                + float(row.get("pos") or 0.0)
                + float(row.get("spese") or 0.0)
                - float(row.get("giro_affari") or 0.0)
            )
        except Exception:
            row["diff_cassa"] = 0.0

        rows.append(row)

    # Dedup warnings
    uniq_w = []
    seen = set()
    for w in warnings_all:
        if w not in seen:
            uniq_w.append(w)
            seen.add(w)

    # Sort by store name then code
    try:
        rows = sorted(rows, key=lambda r: (str(r.get("store_name") or "").lower(), str(r.get("store_code") or "")))
    except Exception:
        pass

    return jsonify(
        {
            "ok": True,
            "month": month_norm,
            "period": {"start": start_d.isoformat(), "end": end_d.isoformat()},
            "rows": rows,
            "warnings": uniq_w[:200],
        }
    )


@rendiconto_bp.get("/api/riepilogo/mensile.xlsx")
def api_riepilogo_mensile_xlsx():
    """Esporta in Excel la tabella del riepilogo mensile multi-store."""
    _ensure_session_keys()

    if Workbook is None:
        abort(500, description="Dipendenza mancante: openpyxl")

    user_id = session.get("uid")
    if not user_id:
        abort(401)

    month = (request.args.get("month") or "").strip()  # YYYY-MM
    y, m, start_d, end_d = _parse_month_yyyy_mm(month)
    month_norm = f"{y:04d}-{m:02d}"

    try:
        stores = _available_stores_for_user(str(user_id)) or []
    except Exception:
        stores = []

    # Costruiamo le stesse righe dell'API JSON (senza dipendenze dal frontend)
    rows = []
    for s in stores or []:
        code = (s or {}).get("code")
        name = (s or {}).get("name")
        if not code:
            continue

        r = {
            "store_code": str(code),
            "store_name": str(name or ""),
            "giro_affari": 0.0,
            "distinte": 0.0,
            "pos": 0.0,
            "spese": 0.0,
            "versamenti": 0.0,
            "ticket_si": 0.0,
            "ticket_no": 0.0,
            "delivery_si": 0.0,
            "delivery_no": 0.0,
            "coupon_si": 0.0,
            "coupon_no": 0.0,
            "diff_cassa": 0.0,
        }

        # Prima nota (aggregati mese)
        try:
            prim_rows = load_primanota_month_agg_totals(str(code), year=y, month=m)

            tot_vendite_lorde = 0.0
            tot_annullati = 0.0
            tot_pos = 0.0
            tot_distinte = 0.0
            t_si = 0.0
            t_no = 0.0
            d_si = 0.0
            d_no = 0.0
            c_si = 0.0
            c_no = 0.0

            for rr in prim_rows or []:
                cat = str(rr.get("categoria") or "").strip()
                voce = str(rr.get("voce") or "").strip()
                tipo = str(rr.get("tipo") or "SI").strip().upper()
                try:
                    s_val = float(rr.get("sum") or 0.0)
                except Exception:
                    s_val = 0.0

                if cat == "Dati chiusura":
                    k = _CHIUSURA_VOICE_TO_KEY.get(voce.upper())
                    if k == "vendite_lorde":
                        tot_vendite_lorde += s_val
                    elif k == "annullati":
                        tot_annullati += s_val
                    elif k == "pos":
                        tot_pos += s_val
                elif cat == "Distinte":
                    tot_distinte += s_val
                elif cat == "Ticket":
                    if tipo == "SI":
                        t_si += s_val
                    else:
                        t_no += s_val
                elif cat == "Delivery":
                    if tipo == "SI":
                        d_si += s_val
                    else:
                        d_no += s_val
                elif cat == "Coupon":
                    if tipo == "SI":
                        c_si += s_val
                    else:
                        c_no += s_val

            r["giro_affari"] = float(tot_vendite_lorde) - float(tot_annullati)
            r["pos"] = float(tot_pos)
            r["distinte"] = float(tot_distinte)
            r["ticket_si"] = float(t_si)
            r["ticket_no"] = float(t_no)
            r["delivery_si"] = float(d_si)
            r["delivery_no"] = float(d_no)
            r["coupon_si"] = float(c_si)
            r["coupon_no"] = float(c_no)
        except Exception:
            # Se fallisce, lasciamo 0 e continuiamo
            pass

        # Spese
        try:
            r["spese"] = float(sum_spese_month_total_net(store_code=str(code), year=y, month=m))
        except Exception:
            pass

        # Versamenti
        try:
            r["versamenti"] = float(sum_versamenti_month_total(store_code=str(code), year=y, month=m))
        except Exception:
            pass

        # Diff cassa
        try:
            r["diff_cassa"] = float(
                float(r.get("distinte") or 0.0)
                + float(r.get("ticket_si") or 0.0)
                + float(r.get("delivery_si") or 0.0)
                + float(r.get("coupon_si") or 0.0)
                + float(r.get("pos") or 0.0)
                + float(r.get("spese") or 0.0)
                - float(r.get("giro_affari") or 0.0)
            )
        except Exception:
            r["diff_cassa"] = 0.0

        rows.append(r)

    try:
        rows = sorted(rows, key=lambda rr: (str(rr.get("store_name") or "").lower(), str(rr.get("store_code") or "")))
    except Exception:
        pass

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Riepilogo"

    headers = [
        "Store",
        "Giro affari",
        "Differenza di cassa",
        "Distinte",
        "POS",
        "Spese",
        "Versamenti",
        "Ticket",
        "Delivery",
        "Coupon",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze: prima riga + prima colonna
    ws.freeze_panes = "B2"

    def _money(v):
        try:
            return float(v)
        except Exception:
            return 0.0

    for rr in rows:
        store_label = (str(rr.get("store_name") or "").strip() or str(rr.get("store_code") or "").strip())
        ticket = _money(rr.get("ticket_si")) + _money(rr.get("ticket_no"))
        delivery = _money(rr.get("delivery_si")) + _money(rr.get("delivery_no"))
        coupon = _money(rr.get("coupon_si")) + _money(rr.get("coupon_no"))
        ws.append(
            [
                store_label,
                _money(rr.get("giro_affari")),
                _money(rr.get("diff_cassa")),
                _money(rr.get("distinte")),
                _money(rr.get("pos")),
                _money(rr.get("spese")),
                _money(rr.get("versamenti")),
                _money(ticket),
                _money(delivery),
                _money(coupon),
            ]
        )

    # Formattazione numeri (Euro)
    eur_fmt = u"€ #,##0.00"
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(2, 11):
            ws.cell(row=row_idx, column=col_idx).number_format = eur_fmt
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="right")

    # Larghezze colonne
    widths = [32, 16, 18, 16, 14, 14, 16, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"riepilogo_{month_norm}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
